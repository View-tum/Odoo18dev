from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
from collections import defaultdict


import base64
from io import BytesIO
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')


class KSWarehouseReportMeasures(models.Model):
    _name = "ks.warehouse.report.measure"
    _description = "Stock Measures / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    kr_in_dates = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'opening_stock': 3, 'closing_stock': 4,
                   'qty_date': 5}

    ks_name = fields.Char(default='Stock Measure Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                 default=lambda self: self.env.company)

    ks_show_opening = fields.Boolean('Show Opening', default=True)
    ks_show_sales = fields.Boolean('Show Closing', default=True)
    ks_show_adjustment = fields.Boolean('Show Adjustment', default=True)
    ks_show_scrap_loss = fields.Boolean('Show Scrap/Loss', default=True)
    ks_show_purchase = fields.Boolean('Show Purchase', default=True)
    ks_show_internal = fields.Boolean('Show Internal', default=True)
    ks_show_current = fields.Boolean('Show Current', default=True)

    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_create_workbook_header(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=14)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=14)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=14)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=14)

        sheet['A8'] = "S.NO"
        sheet.merge_cells(start_row=8, end_row=9, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=9, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        sheet.merge_cells(start_row=8, end_row=9, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        sheet.merge_cells(start_row=8, end_row=9, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        sheet.merge_cells(start_row=8, end_row=9, start_column=5, end_column=5)

        sheet['F8'] = "Location"
        sheet.merge_cells(start_row=8, end_row=9, start_column=6, end_column=6)

        sheet['G8'] = "Company"
        sheet.merge_cells(start_row=8, end_row=9, start_column=7, end_column=7)

        self.ks_dynamic_sheet(sheet)

        sheet.freeze_panes = 'C10'

    def ks_generate_xlsx_report(self):
        report_name = self.ks_name
        workbook = openpyxl.Workbook()

        sheet = workbook.active

        self.ks_create_workbook_header(report_name, sheet)

        # get qty available
        self.env.cr.execute("""
        SELECT ks_product_code,ks_product_type,ks_product_categ_id,ks_product_name,ks_location_id,ks_company_id,
               ks_product_sales_price, ks_product_qty_available, ks_product_id 
        FROM ks_warehouse_report 
        WHERE ks_company_id = %s and ks_product_qty_available != 0 and (ks_usage = 'internal' or ks_usage = 'inventory')
            order by ks_location_id
        """ % self.ks_company_id.id)

        datas = self.env.cr.fetchall()
        if not datas:
            raise ValidationError(_("Opps! There are no data."))

        dates_in = self.ks_data_in_date()
        ks_adjusted_stock ={}
        ks_scrap_stock ={}
        ks_sale_product ={}
        ks_purchase_product ={}
        if self.ks_show_adjustment:
            ks_adjusted_stock = self.ks_adjusted_stock()
        if self.ks_show_scrap_loss:
            ks_scrap_stock = self.ks_scrap_stock()
        if self.ks_show_sales:
            ks_sale_product = self.ks_sale_product()
        if self.ks_show_purchase:
            ks_purchase_product = self.ks_purchase_product()

        datas = self.ks_merge_data(datas, dates_in, ks_adjusted_stock, ks_scrap_stock, ks_sale_product, ks_purchase_product)


        category_ids = list(set(data[2] for data in datas))
        location_ids = list(set(data[4] for data in datas))
        # Pre-load the data to avoid querying inside the loop

        categories = {c['id']: c['name'] for c in
                          self.env['product.category'].search_read([('id', 'in', category_ids)], ['name'])}
        locations = {l['id']: l['display_name'] for l in
                         self.env['stock.location'].search_read([('id', 'in', location_ids)], ['display_name'])}

        # Create a list of column indices based on conditions to reduce checks
        columns_to_show = []
        if self.ks_show_opening:
            columns_to_show.append(6)
        if self.ks_show_sales:
            columns_to_show.append(7)
        if self.ks_show_purchase:
            columns_to_show.append(8)
        if self.ks_show_internal:
            columns_to_show.append(9)
        if self.ks_show_adjustment:
            columns_to_show.append(10)
        if self.ks_show_scrap_loss:
            columns_to_show.append(11)
        if self.ks_show_current:
            columns_to_show.append(12)

        if datas:
            i = 1
            row = 10
            for data in datas:
                # Direct assignment to the cells
                sheet.cell(row, 1, i)
                sheet.cell(row, 2, data[0])
                sheet.cell(row, 3, 'Stockable' if data[1] == 'consu' else 'Consumable' if data[1] == 'service' else '')

                # Fetch category name if data[2] exists
                sheet.cell(row, 4, categories.get(data[2]) if categories.get(data[2]) else '')

                # Fetch values for column 5 (product info)
                sheet.cell(row, 5, list(data[3].values())[0] if data[3] else '')

                # Fetch location name if data[4] exists
                sheet.cell(row, 6, locations.get(data[4]) if locations.get(data[4]) else None)

                # Fetch company name if data[5] exists
                comp_id = self.ks_company_id
                sheet.cell(row, 7, comp_id.name if comp_id else '')

                # Add values for columns based on conditions
                c_1 = 8
                for col_idx in columns_to_show:
                    sheet.cell(row, c_1, data[col_idx - 1])  # Adjust for 0-based index of data
                    c_1 += 1

                # Move to the next row and increment index
                row += 1
                i += 1

        # Save workbook into an in-memory buffer to avoid filesystem permission issues
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        file_data = output.read()
        out = base64.encodebytes(file_data)

        # Files actions - create binary attachment record
        attach_vals = {
            'report_name': str(report_name) + '.xlsx',
            'datas': out,
        }

        act_id = self.env['ks.warehouse.report.measure.out'].create(attach_vals)
        _log.info("Workbook saved to memory and attachment created.")
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.measure.out',
            'res_id': act_id.id,
            'view_type': 'form',
            'view_mode': 'form',
            'context': self.env.context,
            'target': 'new',
        }

    def ks_purchase_product(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)

        self.env.cr.execute("""
            SELECT sm.product_id, sm.location_dest_id, sm.company_id,
                   SUM(sm.product_uom_qty) AS Purchase
            FROM stock_move AS sm
            LEFT JOIN stock_location AS sl ON sl.id = sm.location_id
            LEFT JOIN stock_location AS sld ON sld.id = sm.location_dest_id
            WHERE sm.state = 'done'
              AND sm.company_id = %s 
              AND sm.date >= %s 
              AND sm.date <= %s
              AND sl.usage = 'supplier' 
              AND sld.usage IN ('internal', 'transit')
            GROUP BY sm.product_id, sm.location_dest_id, sm.company_id
        """, (self.ks_company_id.id, ks_date_from, ks_date_to.strftime('%Y-%m-%d 23:59:59')))

        purchase_date = self.env.cr.fetchall()

        if not purchase_date:
            return {}
        else:
            ks_dict = dict()
            for ks in purchase_date:
                # Creating a dictionary with product_id + company_id as key
                ks_dict[str(ks[0]) + str(ks[2])] = ks[3]
            return ks_dict

    def ks_sale_product(self):
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)
        self.env.cr.execute("""
                    select sm.product_id, sm.location_id, sm.company_id,
                        sum(sm.product_uom_qty) as Sale
                    from stock_move as sm
                        left join stock_location as sl on sl.id = sm.location_id
                        left join stock_location as sld on sld.id = sm.location_dest_id
                    where sm.state = 'done' and sm.company_id = '%s' and sm.date >= '%s' and sm.date <= '%s'
                        and sl.usage in ('internal', 'transit') and sld.usage not in ('internal', 'transit')
                        and sld.scrap_location = False
                    group by sm.product_id, sm.location_id, sm.company_id
                """ % (self.ks_company_id.id, ks_date_from, ks_date_to.strftime('%Y-%m-%d 23:59:59')))

        sale_date = self.env.cr.fetchall()
        if not sale_date:
            return {}
        else:
            ks_dict = dict()
            for ks in sale_date:  # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0]) + str(ks[2])] = ks[3]
            sale_date = ks_dict
        return sale_date


    def ks_adjusted_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)
        self.env.cr.execute("""
            select sml.product_id, sml.location_id, sm.company_id, sum(sml.quantity),
                sum(case when spt.code= 'internal' then sml.quantity else 0 end)
            from stock_move_line as sml
                left join stock_move as sm on sm.id = sml.move_id
                left join stock_picking_type as spt on spt.id = sm.picking_type_id
                left join stock_location as sld on sld.id = sm.location_dest_id
            where sml.state = 'done' and sm.company_id = '%s' and sml.date >= '%s' and sml.date <= '%s'
                and sld.scrap_location = False
            group by sml.product_id, sml.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from,ks_date_to.strftime('%Y-%m-%d 23:59:59')))

        adjusted_date = self.env.cr.fetchall()
        if not adjusted_date:
            return {}
        else:
            ks_dict = dict()
            for ks in adjusted_date:  # product_id + location_id + company_id : Adjustment, internal(code ='internal')
                ks_dict[str(ks[0]) + str(ks[1]) + str(ks[2])] = [ks[3], ks[4]]
            adjusted_date = ks_dict
        return adjusted_date

    def ks_data_in_date(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)

        self.env.cr.execute("""
            select sq.product_id, sm.location_dest_id, sq.company_id,
                 sum(case when sq.create_date <= '%s' then sq.quantity else 0 end) as opening_stock,
                 sum(sq.quantity) as closing_stock, 
                 sum(case when sq.create_date >= '%s' then  sq.quantity else 0 end) as qty_date
            from stock_valuation_layer as sq
                left join stock_move as sm on sm.id = sq.stock_move_id
                left join stock_location as sld on sld.id = sm.location_dest_id
                left join stock_location as sl ON sl.id = sm.location_id 
            where sq.company_id = '%s'
            and sq.create_date >= '%s' 
            and sq.create_date <= '%s' 
            group by sq.product_id, sm.location_dest_id, sq.company_id
        """ % (
            ks_date_from,
            ks_date_from,
            self.ks_company_id.id,
            ks_date_from,
            ks_date_to.strftime('%Y-%m-%d 23:59:59')
        ))
        dates_in = self.env.cr.fetchall()
        if not dates_in:
            raise ValidationError(_("Opps! There are no data."))
        return dates_in


    def ks_scrap_stock(self):
        # get the stock_quant data via date in query
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)
        self.env.cr.execute("""
            select scrap.product_id, scrap.location_id, sm.company_id, sum(scrap.scrap_qty) 
            from stock_scrap as scrap
                left join stock_move as sm on sm.scrap_id = scrap.id
            where scrap.state = 'done' and sm.company_id = '%s' and scrap.date_done >= '%s' and scrap.date_done <= '%s'
            group by scrap.product_id, scrap.location_id, sm.company_id
        """ % (self.ks_company_id.id, ks_date_from,  ks_date_to.strftime('%Y-%m-%d 23:59:59')))

        scrap_date = self.env.cr.fetchall()
        if not scrap_date:
            return {}
        else:
            ks_dict = dict()
            for ks in scrap_date:  # product_id + location_id + company_id : qty_done(state=done)
                ks_dict[str(ks[0]) + str(ks[1]) + str(ks[2])] = ks[3]
            scrap_date = ks_dict
        return scrap_date

    
    def ks_dynamic_sheet(self, sheet):
        c_1 = 8
        sheet.cell(8, c_1, "Stock Measures")
        self.ks_apply_style(sheet.cell(8, c_1), True, True, False, False)
        if self.ks_show_opening:
            sheet.cell(9, c_1, "Opening")
            c_1 += 1
        if self.ks_show_sales:
            sheet.cell(9, c_1, "Closing")
            c_1 += 1
        if self.ks_show_purchase:
            sheet.cell(9, c_1, "Purchase")
            c_1 += 1
        if self.ks_show_internal:
            sheet.cell(9, c_1, "Internal")
            c_1 += 1
        if self.ks_show_adjustment:
            sheet.cell(9, c_1, "Adjustment")
            c_1 += 1
        if self.ks_show_scrap_loss:
            sheet.cell(9, c_1, "Scrap/Loss")
            c_1 += 1
        if self.ks_show_current:
            sheet.cell(9, c_1, "Current")
        sheet.merge_cells(start_row=8, end_row=8, start_column=8, end_column=c_1)

    def ks_merge_data(self, datas, dates_in, adjusted=None, scrap=None, sale=None, purchase=None):
        ks_list = []
        kr = self.ks_report
        kid = self.kr_in_dates

        # Use defaultdicts to avoid repetitive key checks and make code cleaner
        adjusted = adjusted or defaultdict(lambda: (0, 0))  # Default to (0, 0) if key is missing
        scrap = scrap or defaultdict(int)  # Default to 0 if key is missing
        sale = sale or defaultdict(int)  # Default to 0 if key is missing
        purchase = purchase or defaultdict(int)  # Default to 0 if key is missing

        # Precompute keys for the sale and purchase dictionaries
        sale_keys = {str(dp_id) + str(dc_id): sale.get(str(dp_id) + str(dc_id), 0) for dp_id, dc_id in
                     {(data[kr['product_id']], data[kr['company_id']]) for data in datas}}
        purchase_keys = {str(dp_id) + str(dc_id): purchase.get(str(dp_id) + str(dc_id), 0) for dp_id, dc_id in
                         {(data[kr['product_id']], data[kr['company_id']]) for data in datas}}

        # Create a dictionary for quick lookups of 'dates_in' data based on the key
        date_dict = {(date[kid['product_id']], date[kid['location_id']], date[kid['company_id']]): date for date in
                     dates_in}

        # Iterate over 'datas' and join with 'dates_in' efficiently
        for data in datas:
            dp_id, dl_id, dc_id = data[kr['product_id']], data[kr['location_id']], data[kr['company_id']]

            # Check if the date exists for the current product, location, and company
            date_key = (dp_id, dl_id, dc_id)
            if date_key in date_dict:
                date = date_dict[date_key]

                # Efficient dictionary lookups
                ks_adjusted = adjusted.get(str(dp_id) + str(dl_id) + str(dc_id),
                                           (0, 0))  # Safe lookup with default value
                ks_scrap = scrap.get(str(dp_id) + str(dl_id) + str(dc_id), 0)  # Safe lookup with default value
                ks_sale = sale_keys.get(str(dp_id) + str(dc_id), 0)  # Safe lookup with default value
                ks_purchase = purchase_keys.get(str(dp_id) + str(dc_id), 0)  # Safe lookup with default value

                # Remove items from sale and purchase if needed
                if ks_sale: del sale_keys[str(dp_id) + str(dc_id)]
                if ks_purchase: del purchase_keys[str(dp_id) + str(dc_id)]

                ks_list.append(
                    (
                        data[kr['product_id']], data[kr['product_type']], data[kr['product_categ_id']],
                        data[kr['product_name']], data[kr['location_id']], data[kr['company_id']],
                        date[kid['opening_stock']], ks_sale, ks_purchase, ks_adjusted[1], ks_adjusted[0],
                        ks_scrap, date[kid['closing_stock']]
                    )
                )

        if not ks_list:
            raise ValidationError(_("Opps! There are no data."))

        return ks_list

    class KSWarehouseReportMeasureOUT(models.Model):
        _name = "ks.warehouse.report.measure.out"
        _description = "Stock Measure report Out"

        datas = fields.Binary('File', readonly=True)
        report_name = fields.Char('Report Name', readonly=True)
