from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

import base64
from io import BytesIO
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import logging
import pathlib
import os

_log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    _log.debug('Can not `import openpyxl`.')

class KSWarehouseReportAgeingWithMovement(models.Model):
    _name = "ks.warehouse.report.ageing.with.movement"
    _description = "Stock Ageing With Movements / Stock Report"

    ks_report = {'product_code': 0, 'product_type': 1, 'product_categ_id': 2, 'product_name': 3, 'location_id': 4,
                 'company_id': 5, 'product_sales_price': 6, 'product_qty_available': 7, 'product_id': 8}
    kr_in_dates = {'product_id': 0, 'location_id': 1, 'company_id': 2, 'opening_stock': 3, 'closing_stock': 4,
                   'qty_date': 5}

    ks_name = fields.Char(default='Stock Aging With Movement Report')
    ks_date_from = fields.Date('Start Date', required=True)
    ks_date_to = fields.Date('End Date', required=True)
    ks_company_id = fields.Many2one('res.company', 'Company', required=True,
                                    default=lambda self: self.env.company)
    ks_duration = fields.Integer('Duration Range', required=True, default=30)

    def ks_action_generate_report(self):
        return True

    def ks_apply_style(self, ks_cell, kc='', vc='', sz=False, wp=False):
        ks_cell.alignment = Alignment(horizontal="center" if kc else '', vertical="center" if vc else '',
                                      wrap_text=wp)
        if sz: ks_cell.font = Font(b=True, size=sz)

    def ks_set_default_5_columns_left(self, report_name, sheet):
        sheet.title = str(report_name)

        sheet['A1'] = str(report_name)
        self.ks_apply_style(sheet['A1'], True, True, 20, True)
        sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=20)

        sheet['A3'] = "COMPANY : " + self.ks_company_id.name
        self.ks_apply_style(sheet['A3'], True, True, 14, True)
        sheet.merge_cells(start_row=3, end_row=3, start_column=1, end_column=20)

        sheet['A4'] = 'FROM : ' + str(self.ks_date_from) + ' | TO : ' + str(self.ks_date_to)
        self.ks_apply_style(sheet['A4'], True, True, 10, True)
        sheet.merge_cells(start_row=4, end_row=4, start_column=1, end_column=20)

        sheet['A6'] = "REPORT"
        self.ks_apply_style(sheet['A6'], True, True, 14, True)
        sheet.merge_cells(start_row=6, end_row=7, start_column=1, end_column=20)

        sheet['A8'] = "S.NO"
        self.ks_apply_style(sheet['A8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=1, end_column=1)

        sheet['B8'] = "Reference/Code"
        self.ks_apply_style(sheet['B8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=2, end_column=2)

        sheet['C8'] = "Type"
        self.ks_apply_style(sheet['C8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=3, end_column=3)

        sheet['D8'] = "Category"
        self.ks_apply_style(sheet['D8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=4, end_column=4)

        sheet['E8'] = "Product"
        self.ks_apply_style(sheet['E8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=5, end_column=5)

        sheet['F8'] = "Location"
        self.ks_apply_style(sheet['F8'], True, True, False, True)
        sheet.merge_cells(start_row=8, end_row=11, start_column=6, end_column=6)

        sheet.freeze_panes = 'C12'

    def ks_generate_xlsx_report(self):
        """Generate an Excel stock aging report."""
        if self.ks_duration <= 0:
            raise UserError(_('You must set a period length greater than 0.'))
        if not self.ks_date_from:
            raise UserError(_('You must set a start date.'))

        report_name = self.ks_name
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        self.ks_set_default_5_columns_left(report_name, sheet)

        # Convert date fields
        ks_date_from = fields.Datetime.to_datetime(self.ks_date_from)
        ks_date_to = fields.Datetime.to_datetime(self.ks_date_to)

        # Generate date ranges for periods
        date_ranges = self._compute_date_ranges(ks_date_from, ks_date_to, self.ks_duration)

        # Populate period headers in Excel
        self._write_period_headers(sheet, date_ranges)

        # Fetch warehouse stock data
        stock_data = self._fetch_stock_data()

        if not stock_data:
            raise ValidationError(_("No stock data found for the selected period."))

        # Fetch product prices in bulk
        product_prices = self._fetch_product_prices(stock_data)

        # Fetch stock movements and group by periods
        product_stock = self._fetch_stock_movements(stock_data, date_ranges)

        # Write product details and stock data to the Excel sheet
        self._write_stock_data(sheet, stock_data, product_prices, product_stock, date_ranges)

        # Save and return the report
        return self._save_and_return_report(workbook, report_name)

    def _compute_date_ranges(self, start_date, end_date, duration):
        """Compute the date periods based on duration."""
        date_ranges = []
        while start_date <= end_date:
            period_end = min(start_date + timedelta(days=duration - 1), end_date)
            date_ranges.append((start_date, period_end))
            start_date = period_end + timedelta(days=1)
        return date_ranges

    def _write_period_headers(self, sheet, date_ranges):
        """Write period headers in Excel."""
        ks_inc = 0
        for i, (start, end) in enumerate(date_ranges):
            period_str = f"{start.strftime('%d-%m-%Y')} : {end.strftime('%d-%m-%Y')}"
            sheet.cell(8, 7 + ks_inc, f"{start.day}-{end.day}")
            self.ks_apply_style(sheet.cell(8, 7 + ks_inc), True, True, False, True)
            sheet.cell(9, 7 + ks_inc, period_str)
            self.ks_apply_style(sheet.cell(9, 7 + ks_inc), True, True, False, True)
            sheet.merge_cells(start_row=8, end_row=8, start_column=7 + ks_inc, end_column=16 + ks_inc)
            sheet.merge_cells(start_row=9, end_row=9, start_column=7 + ks_inc, end_column=16 + ks_inc)
            sheet.cell(10, 7 + ks_inc, 'Sale')
            self.ks_apply_style(sheet.cell(10, 7 + ks_inc), True, True, False, True)
            sheet.cell(11, 7 + ks_inc, "Qty")
            sheet.cell(11, 8 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=7 + ks_inc, end_column=8 + ks_inc)
            sheet.cell(10, 9 + ks_inc, "Purchase")
            self.ks_apply_style(sheet.cell(10, 9 + ks_inc), True, True, False, True)
            sheet.cell(11, 9 + ks_inc, "Qty")
            sheet.cell(11, 10 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=9 + ks_inc, end_column=10 + ks_inc)
            sheet.cell(10, 11 + ks_inc, "Internal")
            self.ks_apply_style(sheet.cell(10, 11 + ks_inc), True, True, False, True)
            sheet.cell(11, 11 + ks_inc, "Qty")
            sheet.cell(11, 12 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=11 + ks_inc, end_column=12 + ks_inc)
            sheet.cell(10, 13 + ks_inc, "Adjustment")
            self.ks_apply_style(sheet.cell(10, 13 + ks_inc), True, True, False, True)
            sheet.cell(11, 13 + ks_inc, "Qty")
            sheet.cell(11, 14 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=13 + ks_inc, end_column=14 + ks_inc)
            sheet.cell(10, 15 + ks_inc, "Scrap")
            self.ks_apply_style(sheet.cell(10, 15 + ks_inc), True, True, False, True)
            sheet.cell(11, 15 + ks_inc, "Qty")
            sheet.cell(11, 16 + ks_inc, "Value")
            sheet.merge_cells(start_row=10, end_row=10, start_column=15 + ks_inc, end_column=16 + ks_inc)
            ks_inc += 10

    def _fetch_stock_data(self):
        """Retrieve stock data from the warehouse report."""
        self.env.cr.execute("""
               SELECT ks_product_id, ks_product_type, ks_product_categ_id, ks_product_name, 
                      ks_location_id, ks_company_id, ks_product_sales_price, ks_product_qty_available
               FROM ks_warehouse_report 
               WHERE ks_company_id = %s AND ks_product_qty_available != 0 AND ks_usage = 'internal'
               ORDER BY ks_location_id
           """, (self.ks_company_id.id,))

        return self.env.cr.fetchall()

    def _fetch_product_prices(self, stock_data):
        """Retrieve standard prices for products in bulk."""
        product_ids = list({data[0] for data in stock_data})  # Get unique product IDs

        # Fetch only the necessary data in one optimized query
        products = self.env['product.product'].search_read(
            [('id', 'in', product_ids)], ['id', 'standard_price']
        )

        # Create a dictionary for quick lookup
        return {product['id']: product['standard_price'] for product in products}

    def _fetch_stock_movements(self, stock_data, date_ranges):
        """Retrieve stock movements for given products within date ranges."""
        product_ids = {data[0] for data in stock_data}
        location_ids = {data[4] for data in stock_data}
        sql_conditions = []
        scrap_days = ""
        adjutsment_days = ""
        sale_days = ""
        purchase_days = ""
        main_days = ""
        groupby = ""
        for i, (start, end) in enumerate(date_ranges):
            scrap_days += f"""
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS scrap_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS scrap_period_{start.day}_{end.day}_end,\n
                SUM(CASE WHEN scrap.date_done BETWEEN '{start.strftime('%Y-%m-%d 00:00:00')}' AND '{end.strftime('%Y-%m-%d 23:59:59')}' THEN scrap.scrap_qty ELSE 0 END) AS scrap_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS adjust_period_{start.day}_{end.day}_end,
                0 AS adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS internal_adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS internal_adjust_period_{start.day}_{end.day}_end,
                0 AS internal_adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS sale_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS sale_period_{start.day}_{end.day}_end,
                0 AS sale_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS purchase_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS purchase_period_{start.day}_{end.day}_end,
                0 AS purchase_period_{start.day}_{end.day}_qty,
            """
            adjutsment_days += f"""
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS scrap_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS scrap_period_{start.day}_{end.day}_end,\n
                0 AS scrap_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS adjust_period_{start.day}_{end.day}_end,\n
                SUM(CASE WHEN sml.date BETWEEN '{start.strftime('%Y-%m-%d 00:00:00')}' AND '{end.strftime('%Y-%m-%d 23:59:59')}' THEN sml.quantity ELSE 0 END) AS adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS internal_adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS internal_adjust_period_{start.day}_{end.day}_end,
                SUM(CASE WHEN spt.code = 'internal' AND sml.date BETWEEN '{start.strftime('%Y-%m-%d 00:00:00')}' AND '{end.strftime('%Y-%m-%d 23:59:59')}' THEN sml.quantity ELSE 0 END) AS internal_adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS sale_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS sale_period_{start.day}_{end.day}_end,
                0 AS sale_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS purchase_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS purchase_period_{start.day}_{end.day}_end,
                0 AS purchase_period_{start.day}_{end.day}_qty,
            """
            sale_days += f"""
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS scrap_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS scrap_period_{start.day}_{end.day}_end,\n
                0 AS scrap_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS adjust_period_{start.day}_{end.day}_end,\n
                0 AS adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS internal_adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS internal_adjust_period_{start.day}_{end.day}_end,
                0 AS internal_adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS sale_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS sale_period_{start.day}_{end.day}_end,\n
                SUM(CASE WHEN sm.date BETWEEN '{start.strftime('%Y-%m-%d 00:00:00')}' AND '{end.strftime('%Y-%m-%d 23:59:59')}' THEN sm.product_uom_qty ELSE 0 END) AS sale_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS purchase_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS purchase_period_{start.day}_{end.day}_end,
                0 AS purchase_period_{start.day}_{end.day}_qty,
            """
            purchase_days += f"""
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS scrap_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS scrap_period_{start.day}_{end.day}_end,\n
                0 AS scrap_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS adjust_period_{start.day}_{end.day}_end,\n
                0 AS adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS internal_adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS internal_adjust_period_{start.day}_{end.day}_end,
                0 AS internal_adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS sale_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS sale_period_{start.day}_{end.day}_end,\n
                0 AS sale_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS purchase_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS purchase_period_{start.day}_{end.day}_end,\n
                SUM(CASE WHEN sm.date BETWEEN '{start.strftime('%Y-%m-%d 00:00:00')}' AND '{end.strftime('%Y-%m-%d 23:59:59')}' THEN sm.product_uom_qty ELSE 0 END) AS purchase_period_{start.day}_{end.day}_qty,
            """
            main_days += f"""
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS scrap_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS scrap_period_{start.day}_{end.day}_end,\n
                sum(data.scrap_period_{start.day}_{end.day}_qty) AS scrap_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS adjust_period_{start.day}_{end.day}_end,\n
                sum(data.adjust_period_{start.day}_{end.day}_qty) AS adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS internal_adjust_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS internal_adjust_period_{start.day}_{end.day}_end,
                sum(data.internal_adjust_period_{start.day}_{end.day}_qty) AS internal_adjust_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS sale_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS sale_period_{start.day}_{end.day}_end,\n
                sum(data.sale_period_{start.day}_{end.day}_qty) AS sale_period_{start.day}_{end.day}_qty,
                '{start.strftime('%Y-%m-%d 00:00:00')}' AS purchase_period_{start.day}_{end.day}_start,
                '{end.strftime('%Y-%m-%d 23:59:59')}' AS purchase_period_{start.day}_{end.day}_end,\n
                sum(data.purchase_period_{start.day}_{end.day}_qty) AS purchase_period_{start.day}_{end.day}_qty,
            """
            groupby += f"""
                data.scrap_period_{start.day}_{end.day}_start,
                data.scrap_period_{start.day}_{end.day}_end,
                data.adjust_period_{start.day}_{end.day}_start,
                data.adjust_period_{start.day}_{end.day}_end,
                data.internal_adjust_period_{start.day}_{end.day}_start,
                data.internal_adjust_period_{start.day}_{end.day}_end,
                data.sale_period_{start.day}_{end.day}_start,
                data.sale_period_{start.day}_{end.day}_end,
                data.purchase_period_{start.day}_{end.day}_start,
                data.purchase_period_{start.day}_{end.day}_end,
            """
        query = """
            SELECT 
                data.product_id,
                data.company_id,
                %s
                data.location_id
            FROM (
                (SELECT 
                    scrap.product_id, 
                    sm.company_id, 
                    %s
                    scrap.location_id
                FROM stock_scrap AS scrap
                LEFT JOIN stock_move AS sm ON sm.scrap_id = scrap.id
                WHERE scrap.state = 'done' 
                    AND sm.company_id = %s 
                    AND scrap.date_done BETWEEN '%s' AND '%s'
                    AND scrap.product_id IN %s 
                    AND scrap.location_id IN %s
                GROUP BY scrap.product_id, scrap.location_id, sm.company_id)
                
                UNION
                
                -- Adjustment and Internal Adjustment data
                (SELECT 
                    sml.product_id, 
                    sm.company_id, 
                    %s
                    sml.location_id
                FROM stock_move_line AS sml
                LEFT JOIN stock_move AS sm ON sm.id = sml.move_id
                LEFT JOIN stock_location AS sld ON sld.id = sm.location_dest_id
                LEFT JOIN stock_picking_type AS spt ON spt.id = sm.picking_type_id
                WHERE sml.state = 'done' 
                    AND sm.company_id = %s 
                    AND sml.date BETWEEN '%s' AND '%s'
                    AND sml.product_id IN %s 
                    AND sml.location_id IN %s
                    AND sld.scrap_location = FALSE
                GROUP BY sml.product_id, sml.location_id, sm.company_id)
                
                UNION
                
                -- Sale data
                (SELECT 
                    sm.product_id, 
                    sm.company_id, 
                    %s
                    sm.location_id
                FROM stock_move AS sm
                LEFT JOIN stock_location AS sl ON sl.id = sm.location_id
                LEFT JOIN stock_location AS sld ON sld.id = sm.location_dest_id
                WHERE sm.state = 'done' 
                    AND sm.company_id = %s 
                    AND sm.date BETWEEN '%s' AND '%s'
                    AND sl.usage IN ('internal', 'transit') 
                    AND sld.usage NOT IN ('internal', 'transit')
                    AND sm.product_id IN %s 
                    AND sm.location_id IN %s
                    AND sld.scrap_location = FALSE
                GROUP BY sm.product_id, sm.location_id, sm.company_id)
                
                UNION
                
                -- Purchase data
                (SELECT 
                    sm.product_id, 
                    sm.company_id, 
                    %s
                    sm.location_dest_id AS location_id
                FROM stock_move AS sm
                LEFT JOIN stock_location AS sl ON sl.id = sm.location_id
                LEFT JOIN stock_location AS sld ON sld.id = sm.location_dest_id
                WHERE sm.state = 'done' 
                    AND sm.company_id = %s 
                    AND sm.date BETWEEN '%s' AND '%s'
                    AND sl.usage = 'supplier' 
                    AND sld.usage IN ('internal', 'transit')
                    AND sm.product_id IN %s 
                    AND sm.location_dest_id IN %s
                GROUP BY sm.product_id, sm.location_dest_id, sm.company_id)
        ) AS data
            GROUP BY data.product_id,
                     data.location_id,
                     %s
                     data.company_id
            ORDER BY
                data.company_id,
                data.product_id,
                data.location_id
                     
        """ % (
            main_days, scrap_days, self.ks_company_id.id, self.ks_date_from, self.ks_date_to, tuple(product_ids),tuple(location_ids),
            adjutsment_days, self.ks_company_id.id, self.ks_date_from, self.ks_date_to, tuple(product_ids),tuple(location_ids),
            sale_days, self.ks_company_id.id, self.ks_date_from, self.ks_date_to, tuple(product_ids),tuple(location_ids),
            purchase_days, self.ks_company_id.id, self.ks_date_from, self.ks_date_to, tuple(product_ids),tuple(location_ids),
            groupby
        )
        self.env.cr.execute(query)
        
        #(tuple(product_ids), self.ks_company_id.id, self.ks_date_from, self.ks_date_to, tuple(location_ids))
        raw_stock_movements = self.env.cr.dictfetchall()
        # Organize stock data into {product_id: {location_id: {date_range: qty}}}
        product_stock = {}
        for raw_stock_movement in raw_stock_movements:
            # Assign movement to a period
            for start, end in date_ranges:
                period_start = raw_stock_movement.get(f"scrap_period_{start.day}_{end.day}_start", False)
                period_end = raw_stock_movement.get(f"scrap_period_{start.day}_{end.day}_end", False)
                
                scrap_period_qty = raw_stock_movement.get(f"scrap_period_{start.day}_{end.day}_qty", 0)
                adjust_period_qty = raw_stock_movement.get(f"adjust_period_{start.day}_{end.day}_qty", 0)
                internal_adjust_period_qty = raw_stock_movement.get(f"internal_adjust_period_{start.day}_{end.day}_qty", 0)
                sale_period_qty = raw_stock_movement.get(f"sale_period_{start.day}_{end.day}_qty", 0)
                purchase_period_qty = raw_stock_movement.get(f"purchase_period_{start.day}_{end.day}_qty", 0)
                
                product_id = raw_stock_movement.get('product_id', False)
                location_id = raw_stock_movement.get('location_id', False)
                
                if period_start and period_end:
                    period_start = datetime.strptime(period_start, "%Y-%m-%d 00:00:00")
                    period_end = datetime.strptime(period_end, "%Y-%m-%d 23:59:59")
                    if (start <= period_start <= end) and (start <= period_end <= end):
                        
                        scrap_period_key = f"scrap_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                        adjust_period_key = f"adjust_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                        internal_adjust_period_key = f"internal_adjust_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                        sale_period_key = f"sale_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                        purchase_period_key = f"purchase_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"

                        product_stock.setdefault(product_id, {}).setdefault(location_id, {})[scrap_period_key] = scrap_period_qty
                        product_stock.setdefault(product_id, {}).setdefault(location_id, {})[adjust_period_key] = adjust_period_qty
                        product_stock.setdefault(product_id, {}).setdefault(location_id, {})[internal_adjust_period_key] = internal_adjust_period_qty
                        product_stock.setdefault(product_id, {}).setdefault(location_id, {})[sale_period_key] = sale_period_qty
                        product_stock.setdefault(product_id, {}).setdefault(location_id, {})[purchase_period_key] = purchase_period_qty
        return product_stock

    def _write_stock_data(self, sheet, stock_data, product_prices, product_stock, date_ranges):
        """Write stock and product details into the sheet with optimized queries."""

        row = 12

        # Fetch all unique product, category, and location IDs
        product_ids = list(set(data[0] for data in stock_data))
        category_ids = list(set(data[2] for data in stock_data))
        location_ids = list(set(data[4] for data in stock_data))

        products = {p['id']: p['name'] for p in
                    self.env['product.product'].search_read([('id', 'in', product_ids)], ['name'])}
        categories = {c['id']: c['name'] for c in
                      self.env['product.category'].search_read([('id', 'in', category_ids)], ['name'])}
        locations = {l['id']: l['display_name'] for l in
                     self.env['stock.location'].search_read([('id', 'in', location_ids)], ['display_name'])}

        for idx, data in enumerate(stock_data, start=1):
            product_id, product_type, categ_id, _, location_id, *_ = data
            sheet.cell(row, 1, idx)  # Serial Number
            sheet.cell(row, 2, product_id)  # Product ID
            sheet.cell(row, 3, 'Stockable' if product_type == 'consu' else 'Consumable')
            sheet.cell(row, 4, categories.get(categ_id, ''))  # Fetch category name
            sheet.cell(row, 5, products.get(product_id, 'Unknown Product'))  # Fetch product name
            sheet.cell(row, 6, locations.get(location_id, ''))  # Fetch location name
            col = 7
            for col_idx, (start, end) in enumerate(date_ranges):

                scrap_period_key = f"scrap_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                adjust_period_key = f"adjust_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                internal_adjust_period_key = f"internal_adjust_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                sale_period_key = f"sale_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                purchase_period_key = f"purchase_period_{start.day}-{start.month}-{start.year}-{end.day}-{end.month}-{end.year}"
                
                scrap_period_qty = product_stock.get(product_id, {}).get(location_id, {}).get(scrap_period_key, 0)
                adjust_period_qty = product_stock.get(product_id, {}).get(location_id, {}).get(adjust_period_key, 0)
                internal_adjust_period_qty = product_stock.get(product_id, {}).get(location_id, {}).get(internal_adjust_period_key, 0)
                sale_period_qty = product_stock.get(product_id, {}).get(location_id, {}).get(sale_period_key, 0)
                purchase_period_qty = product_stock.get(product_id, {}).get(location_id, {}).get(purchase_period_key, 0)

                cost = product_prices.get(product_id, 0)
                sheet.cell(row, col, sale_period_qty)
                col += 1
                sheet.cell(row, col, sale_period_qty * cost)
                col += 1
                sheet.cell(row, col, purchase_period_qty)
                col += 1
                sheet.cell(row, col, purchase_period_qty * cost)
                col += 1
                sheet.cell(row, col, internal_adjust_period_qty)
                col += 1
                sheet.cell(row, col, internal_adjust_period_qty * cost)
                col += 1
                sheet.cell(row, col, adjust_period_qty)
                col += 1
                sheet.cell(row, col, adjust_period_qty * cost)
                col += 1
                sheet.cell(row, col, scrap_period_qty)
                col += 1
                sheet.cell(row, col, scrap_period_qty * cost)
                col += 1
            row += 1

    def _save_and_return_report(self, workbook, report_name):
        """Save the report and return it as an Odoo action."""
        # Save workbook into an in-memory buffer and create attachment without touching filesystem
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        file_data = base64.encodebytes(output.read())

        attachment = self.env['ks.warehouse.report.ageing.with.movement.out'].create({
            'report_name': f"{report_name}.xlsx",
            'datas': file_data,
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ks.warehouse.report.ageing.with.movement.out',
            'res_id': attachment.id,
            'view_type': 'form',
            'view_mode': 'form',
            'target': 'new',
        }
    


class KSWarehouseReportAgeingWithMovementOUT(models.Model):
    _name = "ks.warehouse.report.ageing.with.movement.out"
    _description = "Stock Ageing No Movement report Out"

    datas = fields.Binary('File', readonly=True)
    report_name = fields.Char('Report Name', readonly=True)
