from __future__ import annotations

import copy
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


GOLD_FILE = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU (1).xlsx")
UAT_FILE = Path(r"C:\Users\tumsu\Downloads\UAT_Test_Scenario.xlsx")
OUT_FILE = Path(r"C:\Users\tumsu\Downloads\UAT_Test_Scenario_LINKED.xlsx")

STANDARD_HEADERS = [
    "ลำดับ",
    "Case ID",
    "Backlog IDs",
    "ลำดับเหตุการณ์",
    "Scenario ทดสอบ",
    "บทบาท / หน่วยงาน",
    "Menu Path in local UAT (English)",
    "เงื่อนไขก่อนทดสอบ",
    "ข้อมูลทดสอบ",
    "ขั้นตอนทดสอบแบบละเอียด",
    "ผลลัพธ์ที่คาดหวัง",
    "แนวทางตรวจแก้ / Recovery",
    "สถานะที่รันใน local",
    "Review - คุณไอซ์",
    "Review - คุณติ๊ก",
    "สถานะจริง",
]

ACTUAL_SHEET_PREFIXES = ("01_", "02_", "03_", "04_", "05_", "06_", "07_")
STATUS_VALUES = ["Passed", "Failed", "Pending", "Under Testing", "Not Start", "Cancelled"]


def norm(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def split_backlogs(value):
    if value is None:
        return []
    parts = re.split(r"[,;/\n]+", str(value))
    result = []
    for part in parts:
        item = part.strip().upper()
        if not item:
            continue
        # Normalize common typo spacing without changing the visible workbook value.
        item = re.sub(r"\s+", "", item)
        result.append(item)
    return result


def find_header_row(ws):
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [norm(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "ลำดับ" in values and any("scenario" in value or "ทดสอบ" in value for value in values):
            return row
    return None


def source_header_map(ws, header_row):
    mapping = {}
    for col in range(1, ws.max_column + 1):
        key = norm(ws.cell(header_row, col).value)
        if key:
            mapping[key] = col
    return mapping


def pick_col(mapping, *names):
    normalized = [norm(name) for name in names]
    for key, col in mapping.items():
        for name in normalized:
            if key == name:
                return col
    for key, col in mapping.items():
        for name in normalized:
            if name and name in key:
                return col
    return None


def extract_actual_rows(ws):
    header_row = find_header_row(ws)
    if not header_row:
        return []
    mapping = source_header_map(ws, header_row)

    cols = {
        "seq": pick_col(mapping, "ลำดับ"),
        "case": pick_col(mapping, "Case ID"),
        "backlog": pick_col(mapping, "Backlog IDs", "Backlog ID", "Backlog"),
        "event": pick_col(mapping, "ลำดับเหตุการณ์"),
        "scenario": pick_col(mapping, "Scenario ทดสอบ"),
        "role": pick_col(mapping, "บทบาท / หน่วยงาน", "Role"),
        "menu": pick_col(mapping, "Menu Path in local UAT (English)", "Menu Path"),
        "pre": pick_col(mapping, "เงื่อนไขก่อนทดสอบ"),
        "data": pick_col(mapping, "ข้อมูลทดสอบ"),
        "steps": pick_col(mapping, "ขั้นตอนทดสอบแบบละเอียด"),
        "expected": pick_col(mapping, "ผลลัพธ์ที่คาดหวัง"),
        "recovery": pick_col(mapping, "แนวทางตรวจแก้ / Recovery", "Recovery"),
        "local_status": pick_col(mapping, "สถานะฟังก์ชันใน local UAT", "สถานะที่รันใน local"),
        "note": pick_col(mapping, "หมายเหตุ / หลักฐาน"),
        "ice": pick_col(mapping, "Review - คุณไอซ์"),
        "tik": pick_col(mapping, "Review - คุณติ๊ก"),
    }

    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        if not any(ws.cell(row, col).value is not None for col in range(1, min(ws.max_column, 18) + 1)):
            continue
        seq = ws.cell(row, cols["seq"]).value if cols["seq"] else None
        scenario = ws.cell(row, cols["scenario"]).value if cols["scenario"] else None
        backlog = ws.cell(row, cols["backlog"]).value if cols["backlog"] else None
        case_id = ws.cell(row, cols["case"]).value if cols["case"] else None
        if not case_id:
            row_text = " | ".join(
                str(ws.cell(row, col).value or "")
                for col in range(1, min(ws.max_column, 20) + 1)
            )
            match = re.search(r"\bMU\d{2}-\d{2}\b", row_text, flags=re.IGNORECASE)
            if match:
                case_id = match.group(0).upper()
        if not any([seq, scenario, backlog, case_id]):
            continue

        local_status = ws.cell(row, cols["local_status"]).value if cols["local_status"] else None
        note = ws.cell(row, cols["note"]).value if cols["note"] else None
        if note and local_status:
            local_status = f"{local_status} | {note}"
        elif note and not local_status:
            local_status = note

        rows.append(
            {
                "seq": seq,
                "case": case_id,
                "backlog": backlog,
                "event": ws.cell(row, cols["event"]).value if cols["event"] else None,
                "scenario": scenario,
                "role": ws.cell(row, cols["role"]).value if cols["role"] else None,
                "menu": ws.cell(row, cols["menu"]).value if cols["menu"] else None,
                "pre": ws.cell(row, cols["pre"]).value if cols["pre"] else None,
                "data": ws.cell(row, cols["data"]).value if cols["data"] else None,
                "steps": ws.cell(row, cols["steps"]).value if cols["steps"] else None,
                "expected": ws.cell(row, cols["expected"]).value if cols["expected"] else None,
                "recovery": ws.cell(row, cols["recovery"]).value if cols["recovery"] else None,
                "local_status": local_status,
                "ice": ws.cell(row, cols["ice"]).value if cols["ice"] else None,
                "tik": ws.cell(row, cols["tik"]).value if cols["tik"] else None,
            }
        )
    return rows


def style_standard_sheet(ws, data_start_row):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Angsana New", size=16, bold=True)
    body_font = Font(name="Angsana New", size=16)
    thin_gray = Side(style="thin", color="B7B7B7")
    border = Border(top=thin_gray, bottom=thin_gray, left=thin_gray, right=thin_gray)

    widths = [9, 14, 18, 24, 36, 22, 42, 36, 34, 58, 42, 42, 36, 18, 18, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for col in range(1, len(STANDARD_HEADERS) + 1):
        cell = ws.cell(data_start_row, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(data_start_row + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 58
        for col in range(1, len(STANDARD_HEADERS) + 1):
            cell = ws.cell(row, col)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A{data_start_row}:P{ws.max_row}"


def rebuild_actual_sheet(ws, rows):
    header_row = 6
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    if ws.max_column > len(STANDARD_HEADERS):
        ws.delete_cols(len(STANDARD_HEADERS) + 1, ws.max_column - len(STANDARD_HEADERS))

    for col, header in enumerate(STANDARD_HEADERS, start=1):
        ws.cell(header_row, col).value = header

    for idx, item in enumerate(rows, start=header_row + 1):
        values = [
            item["seq"],
            item["case"],
            item["backlog"],
            item["event"],
            item["scenario"],
            item["role"],
            item["menu"],
            item["pre"],
            item["data"],
            item["steps"],
            item["expected"],
            item["recovery"],
            item["local_status"],
            item["ice"],
            item["tik"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col).value = value
        ws.cell(idx, 16).value = (
            f'=IF(AND(N{idx}="Passed",O{idx}="Passed"),"Passed",'
            f'IF(OR(N{idx}="Failed",O{idx}="Failed"),"Failed","Pending"))'
        )

    dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_VALUES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    if rows:
        dv.add(f"N{header_row + 1}:O{header_row + len(rows)}")

    style_standard_sheet(ws, header_row)


def copy_sheet_between_workbooks(src_ws, dst_wb):
    title = src_ws.title
    if title in dst_wb.sheetnames:
        title = f"GM_{title}"[:31]
    dst_ws = dst_wb.create_sheet(title)
    dst_ws.sheet_format = copy.copy(src_ws.sheet_format)
    dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy.copy(src_ws.page_margins)
    dst_ws.page_setup = copy.copy(src_ws.page_setup)
    dst_ws.freeze_panes = src_ws.freeze_panes

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width
        dst_ws.column_dimensions[key].hidden = dim.hidden
    for key, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[key].height = dim.height
        dst_ws.row_dimensions[key].hidden = dim.hidden

    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(src_cell.row, src_cell.column, src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.fill = copy.copy(src_cell.fill)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy.copy(src_cell.protection)
            if src_cell.comment:
                dst_cell.comment = copy.copy(src_cell.comment)
    return dst_ws


def build_status_index(wb, actual_sheet_rows):
    for title in ["_UAT_Backlog_Index", "_Gold_Backlog_Link", "00_Coverage_Check"]:
        if title in wb.sheetnames:
            del wb[title]

    idx_ws = wb.create_sheet("_UAT_Backlog_Index")
    idx_ws.append(["Backlog ID", "Case ID", "Source Sheet", "สถานะจริง", "Scenario ทดสอบ"])
    idx_row = 2
    actual_backlog_to_cases = defaultdict(list)
    rows_missing_case = []
    for sheet_name, rows in actual_sheet_rows.items():
        for data_row_offset, item in enumerate(rows, start=7):
            case_id = item["case"]
            backlogs = split_backlogs(item["backlog"])
            if not case_id:
                rows_missing_case.append((sheet_name, data_row_offset, item["backlog"], item["scenario"]))
            for backlog in backlogs:
                idx_ws.cell(idx_row, 1).value = backlog
                idx_ws.cell(idx_row, 2).value = f"='{sheet_name}'!B{data_row_offset}"
                idx_ws.cell(idx_row, 3).value = sheet_name
                idx_ws.cell(idx_row, 4).value = f"='{sheet_name}'!P{data_row_offset}"
                idx_ws.cell(idx_row, 5).value = f"='{sheet_name}'!E{data_row_offset}"
                actual_backlog_to_cases[backlog].append((sheet_name, data_row_offset, case_id, item["scenario"]))
                idx_row += 1
    idx_ws.sheet_state = "hidden"

    gold_ws = wb["Scenarios"] if "Scenarios" in wb.sheetnames else None
    gold_backlogs = defaultdict(list)
    gold_row_backlogs = {}
    link_ws = wb.create_sheet("_Gold_Backlog_Link")
    link_ws.append(["Gold Sheet", "Gold Row", "Backlog ID", "UAT Linked Status"])
    link_row = 2
    if gold_ws:
        for row in range(9, gold_ws.max_row + 1):
            backlog_value = gold_ws.cell(row, 3).value
            backlogs = split_backlogs(backlog_value)
            if not backlogs:
                continue
            gold_row_backlogs[row] = backlogs
            for backlog in backlogs:
                link_ws.cell(link_row, 1).value = "Scenarios"
                link_ws.cell(link_row, 2).value = row
                link_ws.cell(link_row, 3).value = backlog
                link_ws.cell(link_row, 4).value = (
                    f'=IF(COUNTIF(_UAT_Backlog_Index!$A:$A,C{link_row})=0,"Missing",'
                    f'IF(COUNTIFS(_UAT_Backlog_Index!$A:$A,C{link_row},_UAT_Backlog_Index!$D:$D,"Passed")>0,"Passed",'
                    f'IF(COUNTIFS(_UAT_Backlog_Index!$A:$A,C{link_row},_UAT_Backlog_Index!$D:$D,"Failed")>0,"Failed","Pending")))'
                )
                gold_backlogs[backlog].append(row)
                link_row += 1

        for row in gold_row_backlogs:
            gold_ws.cell(row, 8).value = (
                f'=IF(COUNTIF(_Gold_Backlog_Link!$B:$B,ROW())=0,"",'
                f'IF(COUNTIFS(_Gold_Backlog_Link!$B:$B,ROW(),_Gold_Backlog_Link!$D:$D,"Missing")>0,"Missing",'
                f'IF(COUNTIFS(_Gold_Backlog_Link!$B:$B,ROW(),_Gold_Backlog_Link!$D:$D,"Failed")>0,"Failed",'
                f'IF(COUNTIFS(_Gold_Backlog_Link!$B:$B,ROW(),_Gold_Backlog_Link!$D:$D,"Pending")>0,"Pending","Passed"))))'
            )
    link_ws.sheet_state = "hidden"

    coverage_ws = wb.create_sheet("00_Coverage_Check", 1)
    coverage_headers = [
        "Backlog ID",
        "Coverage Status",
        "Actual Case IDs / Sheets",
        "GoldMints Scenario Rows",
        "หมายเหตุ",
    ]
    coverage_ws.append(coverage_headers)
    all_backlogs = sorted(set(actual_backlog_to_cases) | set(gold_backlogs))
    row_no = 2
    for backlog in all_backlogs:
        actual_items = actual_backlog_to_cases.get(backlog, [])
        gold_rows = gold_backlogs.get(backlog, [])
        if actual_items and gold_rows:
            status = "Covered"
            note = "พบทั้งใน UAT_Test_Scenario และ GoldMints Scenarios"
        elif gold_rows and not actual_items:
            status = "Missing in UAT_Test_Scenario"
            note = "GoldMints มี backlog นี้ แต่ actual flow ยังไม่มีเคสครอบคลุม"
        else:
            status = "Extra in UAT_Test_Scenario"
            note = "actual flow มี backlog นี้ แต่ไม่พบใน GoldMints Scenarios"
        coverage_ws.cell(row_no, 1).value = backlog
        coverage_ws.cell(row_no, 2).value = status
        coverage_ws.cell(row_no, 3).value = "; ".join(
            f"{case or '-'} ({sheet} R{row})" for sheet, row, case, _scenario in actual_items
        )
        coverage_ws.cell(row_no, 4).value = ", ".join(str(r) for r in gold_rows)
        coverage_ws.cell(row_no, 5).value = note
        row_no += 1

    if rows_missing_case:
        row_no += 1
        coverage_ws.cell(row_no, 1).value = "Rows without Case ID"
        coverage_ws.cell(row_no, 2).value = len(rows_missing_case)
        row_no += 1
        coverage_ws.append(["Sheet", "Row", "Backlog IDs", "Scenario ทดสอบ", "หมายเหตุ"])
        for sheet_name, source_row, backlog, scenario in rows_missing_case:
            coverage_ws.append([sheet_name, source_row, backlog, scenario, "ควรเติม Case ID ให้ครบเพื่อ trace ได้"])

    green = PatternFill("solid", fgColor="D9EAD3")
    red = PatternFill("solid", fgColor="F4CCCC")
    orange = PatternFill("solid", fgColor="FCE5CD")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Angsana New", size=16, bold=True, color="FFFFFF")
    body_font = Font(name="Angsana New", size=16)
    for cell in coverage_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, coverage_ws.max_row + 1):
        status = coverage_ws.cell(row, 2).value
        fill = green if status == "Covered" else red if status == "Missing in UAT_Test_Scenario" else orange if status == "Extra in UAT_Test_Scenario" else None
        for col in range(1, 6):
            cell = coverage_ws.cell(row, col)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill
    widths = [18, 30, 60, 22, 55]
    for idx, width in enumerate(widths, start=1):
        coverage_ws.column_dimensions[get_column_letter(idx)].width = width
    coverage_ws.freeze_panes = "A2"
    coverage_ws.auto_filter.ref = f"A1:E{coverage_ws.max_row}"

    return {
        "total_gold_backlogs": len(gold_backlogs),
        "total_actual_backlogs": len(actual_backlog_to_cases),
        "missing_in_uat": len([b for b in gold_backlogs if b not in actual_backlog_to_cases]),
        "extra_in_uat": len([b for b in actual_backlog_to_cases if b not in gold_backlogs]),
        "rows_missing_case": len(rows_missing_case),
    }


def add_link_summary(wb, summary):
    ws = wb["00_สรุป"] if "00_สรุป" in wb.sheetnames else wb.create_sheet("00_สรุป", 0)
    start_row = 16
    ws.cell(start_row, 1).value = "GoldMints Link / Coverage"
    ws.cell(start_row + 1, 1).value = "GoldMints backlog IDs"
    ws.cell(start_row + 1, 2).value = summary["total_gold_backlogs"]
    ws.cell(start_row + 2, 1).value = "Actual UAT backlog IDs"
    ws.cell(start_row + 2, 2).value = summary["total_actual_backlogs"]
    ws.cell(start_row + 3, 1).value = "Missing in UAT_Test_Scenario"
    ws.cell(start_row + 3, 2).value = summary["missing_in_uat"]
    ws.cell(start_row + 4, 1).value = "Extra in UAT_Test_Scenario"
    ws.cell(start_row + 4, 2).value = summary["extra_in_uat"]
    ws.cell(start_row + 5, 1).value = "Rows without Case ID"
    ws.cell(start_row + 5, 2).value = summary["rows_missing_case"]
    for row in range(start_row, start_row + 6):
        for col in range(1, 3):
            ws.cell(row, col).font = Font(name="Angsana New", size=16, bold=(row == start_row))
            ws.cell(row, col).alignment = Alignment(wrap_text=True)


def main():
    if not GOLD_FILE.exists():
        raise FileNotFoundError(GOLD_FILE)
    if not UAT_FILE.exists():
        raise FileNotFoundError(UAT_FILE)

    uat_wb = load_workbook(UAT_FILE)
    gold_wb = load_workbook(GOLD_FILE)

    actual_sheet_rows = {}
    for sheet_name in list(uat_wb.sheetnames):
        if not sheet_name.startswith(ACTUAL_SHEET_PREFIXES):
            continue
        ws = uat_wb[sheet_name]
        rows = extract_actual_rows(ws)
        actual_sheet_rows[sheet_name] = rows
        rebuild_actual_sheet(ws, rows)

    for src_ws in gold_wb.worksheets:
        copy_sheet_between_workbooks(src_ws, uat_wb)

    summary = build_status_index(uat_wb, actual_sheet_rows)
    add_link_summary(uat_wb, summary)

    uat_wb.calculation.fullCalcOnLoad = True
    uat_wb.calculation.forceFullCalc = True
    uat_wb.save(OUT_FILE)
    print(OUT_FILE)
    print(summary)


if __name__ == "__main__":
    main()
