from __future__ import annotations

from pathlib import Path

import openpyxl


SRC = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_full_backlog_coverage_20260407.xlsx")
OUT = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_full_backlog_coverage_th_20260407.xlsx")
SUMMARY = Path(r"C:\365_project\TheCool18e\Dev\reports\local_uat_manu_full_backlog_coverage_th_20260407.md")


HEADERS_TH = [
    "ลำดับ",
    "รหัสเคส",
    "Backlog ID",
    "ชื่อสถานการณ์ทดสอบ",
    "บทบาท / แผนก",
    "Menu Path",
    "เงื่อนไขก่อนทดสอบ / Setup",
    "ข้อมูลทดสอบ",
    "ขั้นตอนทดสอบอย่างละเอียด",
    "ผลที่คาดหวัง",
    "แนวทางแก้ไข / Recovery",
    "หลักฐานที่ต้องเก็บ",
    "ผลทดสอบจริง / หมายเหตุ",
    "ผู้ทดสอบ",
    "วันที่ทดสอบ",
    "Review - คุณไอซ์",
    "Review - คุณติ๊ก",
    "สถานะรวม",
    "หมายเหตุสำหรับผู้สอน",
]


TH_META = {
    "MU15": {
        "topic": "การจัดการวัตถุดิบ / ขนาดภาชนะ / คลังลอย / การกรองข้อมูล",
        "objective": "ให้ทีมวางแผนผลิต คลัง และหน้างานทดสอบเรื่องการเบิกตามขนาดภาชนะ การอัปเดตสต็อกแบบ Real-time โครงสร้างคลังลอย และการกรองรายการให้เห็นเฉพาะที่ต้องใช้",
        "scope": "เบิกตามภาชนะ, Real-time stock, โครงสร้างคลังลอย, filter ตามโรงงาน",
    },
    "MU16": {
        "topic": "การติดตามย้อนกลับ / QC / Sampling / เอกสารคุณภาพ",
        "objective": "ให้ทีมคุณภาพใช้ MU นี้ครอบคลุม requirement ที่เกี่ยวกับ traceability, sampling, WI, BMR, reject, log book และการรับ FG เข้าระบบ",
        "scope": "traceability, assay, ident, WI, BMR, reject, log book, QC approvals, package checks, FG issue, scale interface",
    },
    "MU17": {
        "topic": "การกำหนดเส้นทางการผลิต / การวางแผน / การแยกใบ / Backorder / การวางแผนรายสัปดาห์",
        "objective": "ให้ทีมวางแผนและ key user ใช้ MU นี้ทดสอบ routing, การ split หรือ merge, backorder, MPS, policy รายสัปดาห์ และการจำลอง OT",
        "scope": "routing, split MO, MPS, backorder, merge semi, MTO batch policy, weekly plan, OT simulation",
    },
    "MU18": {
        "topic": "การควบคุม Shop Floor / การปิดการผลิต / Breakdown / Line Clearance",
        "objective": "ให้ทีมหน้างานทดสอบ control เพิ่มเติมบน GMP Shop Floor เช่น labor รายบุคคล การแก้เวลา การรับของเกิน การบันทึก scrap แบบ draft การปิด production และ line clearance",
        "scope": "labor per person, edit duration, overproduction, scrap log, close production, breakdown, line clearance",
    },
    "MU19": {
        "topic": "การเชื่อมฝ่ายผลิตกับบัญชี / ต้นทุน / การรวมข้อมูลต้นทุน",
        "objective": "ให้ทีมผลิตและบัญชีใช้ MU นี้ทดสอบจุดเชื่อมด้านต้นทุน วิธีคิดต้นทุน และการรวม cost across backorders",
        "scope": "real-time accounting check, average vs FIFO, work center cost, total MO cost across backorders",
    },
    "MU20": {
        "topic": "รายงาน / แบบพิมพ์ / นโยบายหลังการผลิต",
        "objective": "ให้ทีมเอกสารและผู้ใช้งานปลายทางตรวจความพร้อมของรายงาน แบบพิมพ์ และ policy หลังการผลิตที่ยังไม่มี MU รองรับ",
        "scope": "post-production QC policy, MO print, WO print, scrap report, machine report, production report",
    },
}


CASE_TH = {
    "MU15-01": ("เบิกวัตถุดิบตามจำนวนเต็มภาชนะอัตโนมัติ", "Planner / คลัง / ฝ่ายผลิต", "Manufacturing > Operations > Manufacturing Orders"),
    "MU15-02": ("เบิกวัตถุดิบแล้วสต็อกต้องอัปเดตแบบ Real-time", "คลัง / Production Controller", "Manufacturing > Operations > Manufacturing Orders"),
    "MU15-03": ("ตรวจโครงสร้าง Sublocation ของคลังลอยและคลังผลิตแยก Plastic / Pharma", "Key User / Warehouse Lead", "Inventory > Configuration > Warehouse Management > Locations"),
    "MU15-04": ("กรองวัตถุดิบให้ user เห็นเฉพาะฝั่งโรงงานที่ต้องใช้", "Operator / Warehouse User", "Inventory > Operations > Transfers"),
    "MU16-01": ("ติดตาม FG และ Component แบบย้อนกลับได้ครบด้วย lot หรือ QR", "QA / QC / Warehouse", "Manufacturing > Products > Lots/Serial Numbers"),
    "MU16-02": ("คำนวณจำนวนตัวอย่าง Assay และ Ident ตามจำนวนภาชนะ", "QC Lab", "Quality > Quality Control > Quality Checks"),
    "MU16-03": ("ค้นหา แก้ไข และควบคุม revision ของ WI ในระบบ", "QA Document Control", "Documents > Documents"),
    "MU16-04": ("ทดสอบ flow BMR ในระบบจากเริ่มผลิตจนจบการอนุมัติ", "QA / Production Record Owner", "ยังไม่มีเมนูจริงใน local uat (ต้องออกแบบ flow เพิ่ม)"),
    "MU16-05": ("สืบต้นเหตุ Reject และ reject ทั้ง lot ได้ครบถ้วน", "QA / Warehouse / Production", "Quality > Quality Control > Quality Alerts"),
    "MU16-06": ("บันทึก Log Book เครื่องมือและหัวข้อการวิเคราะห์", "QC Lab", "ยังไม่มีเมนูจริงใน local uat (ต้องออกแบบ log book เพิ่ม)"),
    "MU16-07": ("กด Passed พร้อมระบุผู้วิเคราะห์และผู้อนุมัติ", "Analyst / Approver", "Quality > Quality Control > Quality Checks"),
    "MU16-08": ("กำหนดการตรวจสอบเป็น Routine Step ตาม Process", "QA / Process Owner", "Quality > Quality Control > Control Points"),
    "MU16-09": ("บันทึกและติดตามการสุ่มตัวอย่าง Package แบบ √n + 1", "QC Packaging", "ยังไม่มีเมนูจริงใน local uat (ต้องออกแบบ sampling worksheet เพิ่ม)"),
    "MU16-10": ("Checklist QC Package หน้าเครื่องวันละ 4 รอบ", "QC Packaging", "ยังไม่มีเมนูจริงใน local uat (ต้องออกแบบ checklist เพิ่ม)"),
    "MU16-11": ("รับ Finished Product จากฝ่ายผลิตและออก Issue เข้าระบบทันที", "Warehouse / Production Receiving", "Manufacturing > Operations > Manufacturing Orders"),
    "MU16-12": ("รับข้อมูลน้ำหนักจากเครื่องชั่งเข้าระบบโดยตรง", "QC / Production", "ยังไม่มีเมนูจริงใน local uat (ต้องทำ interface เพิ่ม)"),
    "MU17-01": ("Routing ต้องระบุเครื่องจักรและเวลาจริงของแต่ละขั้นตอน", "Planner / IE / Production Supervisor", "Manufacturing > Products > Bills of Materials"),
    "MU17-02": ("Parallel Printer ใน BOM ต้องแตกงานถูกต้อง", "Planner / Key User", "Manufacturing > Products > Bills of Materials"),
    "MU17-03": ("Split MO หลังเริ่มรันแล้วเมื่อมีงานแทรก", "Planner / Production Manager", "Manufacturing > Operations > Manufacturing Orders"),
    "MU17-04": ("วางแผน MPS จาก Finished Goods แล้วแตก demand ลง component", "Planner", "Manufacturing > Planning > MPS > Master Production Schedule"),
    "MU17-05": ("Backorder ต้องสมบูรณ์ทั้ง receipt, transfer และ MO", "Warehouse / Planner / Production", "Inventory > Operations > Transfers"),
    "MU17-06": ("Semi ที่มีอยู่แล้วกับ demand ใหม่ควรรวมเป็นใบเดียวหรือไม่", "Planner / Production Control", "Inventory > Operations > Procurement > Replenishment"),
    "MU17-07": ("MTO ฝั่งยาตาม order แต่ฝั่งพลาสติกตาม batch หรือ container", "Planner / Production Control", "Manufacturing > Operations > Manufacturing Orders"),
    "MU17-08": ("Planning เรียงลำดับงานจากเล็กไปใหญ่เพื่อใช้เครื่องให้คุ้ม", "Planner / Production Manager", "Manufacturing > Planning > Work Orders"),
    "MU17-09": ("ยืนยัน policy การเปิด MO เป็นรายสัปดาห์", "Planner / Management", "Manufacturing > Operations > Manufacturing Orders"),
    "MU17-10": ("จำลอง Working Hours และ OT จากโหลดการผลิต FG", "Planner / Production Manager / Finance", "Manufacturing > Planning > Work Orders"),
    "MU18-01": ("บันทึกชื่อพนักงานและต้นทุนแรงงานรายบุคคลบนงานผลิต", "Operator / Supervisor / HR Costing", "GMP Shop Floor"),
    "MU18-02": ("Manager แก้ไขเวลาปฏิบัติงานย้อนหลังจากหน้า MO", "Manager / Key User", "Manufacturing > Operations > Manufacturing Orders"),
    "MU18-03": ("ผลิตเกินแผนแล้วใช้ duplicate MO รับส่วนเกิน", "Production Control / Warehouse", "Manufacturing > Operations > Manufacturing Orders"),
    "MU18-04": ("Scrap ใน Shop Floor ต้องบันทึกก่อน validate และแก้จาก log ได้", "Operator / Supervisor / Warehouse", "GMP Shop Floor"),
    "MU18-05": ("Close Production ต้องทำให้ทุก Work Order เป็น Done", "Supervisor / Key User", "GMP Shop Floor"),
    "MU18-06": ("กด Close Production แล้วต้องกลับหน้าแรกของ Shop Floor", "Operator / Supervisor", "GMP Shop Floor"),
    "MU18-07": ("ป้องกันกรณี Close MO แต่ยังมี Work Order ค้างจน consumed ผิด", "Supervisor / Key User / Accounting Review", "GMP Shop Floor"),
    "MU18-08": ("เพิ่มสถานะ Work Center - Breakdown และผลต่อ queue", "Supervisor / Maintenance / Planner", "GMP Shop Floor"),
    "MU18-09": ("Line Clearance checkbox ต้องติ๊กก่อนจึงกด Start ได้", "Operator / Supervisor", "GMP Shop Floor"),
    "MU19-01": ("ตรวจ flow เชื่อมข้อมูลฝ่ายผลิตกับบัญชีแบบ Real-time", "Production Controller / Accounting", "Manufacturing > Operations > Manufacturing Orders"),
    "MU19-02": ("เปรียบเทียบผลกระทบของ Cost Method Average กับ FIFO", "Accounting / Key User", "Inventory > Configuration > Products"),
    "MU19-03": ("เพิ่มและตรวจต้นทุนบน Work Center ว่าส่งผลต่อ MO Cost", "IE / Accounting / Production Control", "Manufacturing > Configuration > Work Centers"),
    "MU19-04": ("ดู cost ของ MO รวมได้แม้มี backorder ต่อเนื่อง", "Accounting / Planner / Key User", "Manufacturing > Operations > Manufacturing Orders"),
    "MU20-01": ("ยืนยัน policy ว่าไม่มีขั้นตอน QC หลัง Post-Production", "QA / Production Manager", "Manufacturing > Operations > Manufacturing Orders"),
    "MU20-02": ("รายงานหรือพิมพ์ใบสั่งผลิต (Manufacturing Order)", "Production Admin / Planner", "Manufacturing > Operations > Manufacturing Orders"),
    "MU20-03": ("รายงานหรือพิมพ์ใบสั่งงาน (Work Order)", "Supervisor / Operator Lead", "Manufacturing > Operations > Work Orders"),
    "MU20-04": ("รายงานของเสีย (Scrap Report)", "Production / QA / Accounting", "Manufacturing > Operations > Scrap"),
    "MU20-05": ("รายงานเครื่องจักร (Machine Report)", "Maintenance / Production Manager", "Manufacturing > Reporting > Machine Report"),
    "MU20-06": ("รายงานการผลิต (Production Report)", "Management / Planner / Production Admin", "Manufacturing > Reporting > Production Analysis"),
}


def thai_note(phase: str, owner: str, requirement: str, future: bool) -> str:
    label = "Future Scope / ยังไม่มีใน local uat" if future else "ออกแบบ test case แล้ว"
    return f"{label} | Phase: {phase or '-'} | Owner: {owner or '-'} | Requirement: {requirement or '-'}"


def main():
    wb = openpyxl.load_workbook(SRC)
    pb = wb["Product_backlog"]
    backlog = {}
    for r in range(2, pb.max_row + 1):
        bid = pb.cell(r, 1).value
        if bid:
            backlog[bid] = {
                "phase": pb.cell(r, 4).value or "",
                "owner": pb.cell(r, 12).value or "",
                "asis": (pb.cell(r, 5).value or "").strip(),
                "tobe": (pb.cell(r, 6).value or "").strip(),
            }

    for sheet, meta in TH_META.items():
        ws = wb[sheet]
        ws["A1"] = f"ขั้นตอน UAT ฝ่ายผลิต - {sheet}"
        ws["B2"] = meta["topic"]
        ws["B3"] = meta["objective"]
        ws["B4"] = meta["scope"]
        for c, header in enumerate(HEADERS_TH, start=1):
            ws.cell(6, c).value = header

        r = 7
        while True:
            case_id = ws.cell(r, 2).value
            if not case_id:
                break
            backlog_id = ws.cell(r, 3).value
            info = backlog.get(backlog_id, {})
            title, role, menu = CASE_TH[case_id]
            future = "ยังไม่มีเมนูจริง" in menu or "Future Scope" in ws.cell(r, 19).value
            requirement = info.get("tobe") or info.get("asis")

            ws.cell(r, 4).value = title
            ws.cell(r, 5).value = role
            ws.cell(r, 6).value = menu
            ws.cell(r, 7).value = f"อ้างอิง requirement backlog {backlog_id}. ตรวจว่าหน้าจอและข้อมูลที่เกี่ยวข้องพร้อมสำหรับ use case นี้ก่อนเริ่มทดสอบ"
            ws.cell(r, 8).value = f"ใช้ข้อมูลตัวอย่างตาม requirement ของ {backlog_id} และเลือกเอกสาร/สินค้า/งานที่เกี่ยวข้องใน local uat"
            ws.cell(r, 9).value = (
                f"1) เปิดเมนู {menu}\n"
                f"2) เตรียมข้อมูลทดสอบตาม requirement ของ {backlog_id}\n"
                f"3) ทำขั้นตอนการทำงานจริงจากต้นทางถึงปลายทางทีละจุด\n"
                f"4) ตรวจจำนวน สถานะ เอกสารที่ระบบสร้าง และผลกระทบต่อ stock / cost / report ที่เกี่ยวข้อง\n"
                f"5) เปรียบเทียบผลที่ได้กับ requirement จริง\n"
                f"6) เก็บภาพหน้าจอและจด gap หรือ workaround ถ้ายังไม่ครบ"
            )
            ws.cell(r, 10).value = f"ผลลัพธ์ต้องตอบ requirement ของ {backlog_id}: {requirement}"
            ws.cell(r, 11).value = "หากผลไม่ตรง ให้ไล่จาก master data, route/rule, operation type, สิทธิ์ผู้ใช้, report template หรือ custom ที่เกี่ยวข้อง แล้วสรุปแนวทางแก้"
            ws.cell(r, 12).value = "เก็บภาพหน้าจอหลัก, เอกสารที่ระบบสร้าง, รายงาน/print ที่เกี่ยวข้อง, และ before/after ที่พิสูจน์ผลการทดสอบ"
            ws.cell(r, 19).value = thai_note(info.get("phase", ""), info.get("owner", ""), requirement, future)
            r += 1

    if "LOCAL_UAT_SUMMARY" in wb.sheetnames:
        ws = wb["LOCAL_UAT_SUMMARY"]
        ws["A1"] = "รายการ"
        ws["B1"] = "ค่า"
        labels = {
            "Total Product_backlog items": "จำนวน Product_backlog ทั้งหมด",
            "Backlog items linked to MU test cases": "จำนวน backlog ที่ถูกผูกกับ MU แล้ว",
            "Backlog items not linked to any MU case": "จำนวน backlog ที่ยังไม่มี MU",
            "Total MU test cases in workbook": "จำนวน MU test cases ทั้งหมดใน workbook",
            "MU test cases with local UAT evidence": "จำนวน MU test cases ที่มีหลักฐาน local uat จริง",
            "Coverage of Product_backlog by MU design (%)": "Coverage ของ Product_backlog โดย MU design (%)",
            "Coverage of MU cases by local evidence (%)": "Coverage ของ MU cases โดยหลักฐาน local uat (%)",
        }
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val in labels:
                ws.cell(r, 1).value = labels[val]

    if "LOCAL_UAT_COVERAGE" in wb.sheetnames:
        ws = wb["LOCAL_UAT_COVERAGE"]
        headers = ["Backlog ID", "ผูกกับ MU แล้วหรือไม่", "จำนวน MU ที่ผูก", "มีหลักฐาน local uat แล้วหรือไม่", "สรุป backlog"]
        for c, header in enumerate(headers, start=1):
            ws.cell(1, c).value = header
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value == "Yes":
                ws.cell(r, 2).value = "มี"
            elif ws.cell(r, 2).value == "No":
                ws.cell(r, 2).value = "ไม่มี"
            if ws.cell(r, 4).value == "Yes":
                ws.cell(r, 4).value = "มี"
            elif ws.cell(r, 4).value == "No":
                ws.cell(r, 4).value = "ไม่มี"

    wb.save(OUT)

    SUMMARY.write_text(
        "\n".join(
            [
                "# Local UAT Manufacturing Full Backlog Coverage (Thai)",
                "",
                f"- Source workbook: `{SRC}`",
                f"- Output workbook: `{OUT}`",
                "- MU15-MU20 translated to Thai",
                "- Menu Path updated to exact local-uat menu names where the menu exists",
                "- Cases without a real menu are marked clearly as `ยังไม่มีเมนูจริงใน local uat`",
            ]
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
