from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(r"C:\365_project\TheCool18e\Dev")
INPUT_XLSX = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_actual_flow_th_evidence_20260407.xlsx"
)
OUTPUT_XLSX = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_actual_flow_th_evidence_rerun_20260407.xlsx"
)
RERUN_JSON = ROOT / "reports" / "manu_actual_flow_functional_rerun_20260407.json"
SUMMARY_MD = ROOT / "reports" / "manu_actual_flow_functional_rerun_status_update_20260407.md"
SUMMARY_JSON = ROOT / "reports" / "manu_actual_flow_functional_rerun_status_update_20260407.json"


PHASE_SHEETS = {
    "01_ตรวจสต็อกและวางแผน",
    "02_เปิดงานผลิต",
    "03_โอนและBackorder",
    "04_ShopFloorและMold",
    "05_ปิดงานและแก้ไข",
    "06_คุณภาพเอกสาร",
    "07_รายงานต้นทุนUoM",
}


OVERRIDES = {
    "MU06-01": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: รัน replenishment ของ "
            "FG-PNC-TH-01001 แล้วระบบสร้าง MO GMP/MOPH/00044, ผลิตจบ 250 หน่วย "
            "และ stock/forecast เพิ่มจาก 100121 เป็น 100371 ตามแผน."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU06-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: ทดสอบสินค้า TMP MTS ENOUGH FG "
            "ที่มีของคงเหลือ 10 มากกว่าจุดสั่งเติม min/max 5 แล้วระบบไม่สร้าง MO หรือเอกสารเติมของเพิ่ม."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU06-03": {
        "status": "มีเมนูและฟังก์ชันจริง (ยังไม่ยืนยันเชิงฟังก์ชัน)",
        "note": (
            "rerun เชิงธุรกรรม 2026-04-07 แล้วยังไม่ปิดเคส: local UAT สร้างได้เฉพาะ MO ชั้นบน "
            "แต่ยังไม่สร้าง child MO / transfer chain ลึกตามที่คาดสำหรับ deep MTS chain."
        ),
        "round_status": "Not closed - local UAT ยังไม่สร้าง child chain ตามคาด",
    },
    "MU02-03": {
        "status": "มีเมนูและฟังก์ชันจริง (ยังไม่ยืนยันเชิงฟังก์ชัน)",
        "note": (
            "rerun เชิงธุรกรรม 2026-04-07 แล้วยังไม่ปิดเคส: local UAT สร้างได้เฉพาะ MO ชั้นบน "
            "แต่ยังไม่สร้าง child MO / transfer chain ลึกตามที่คาด จึงยังยืนยันการโอนวัตถุดิบหลายชั้นไม่ได้."
        ),
        "round_status": "Not closed - local UAT ยังไม่สร้าง child chain ตามคาด",
    },
    "MU02-01": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: สร้าง TMP ACTUAL MO, วัตถุดิบต้องใช้ 10 "
            "จองและใช้จริงครบ 10, work order จบ, MO ปิดสถานะ done ได้ครบ."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU07-01": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: manual MO full flow ปิดงานสำเร็จ, "
            "finished qty = 5 และ work order state = done."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU02-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: TMP PARTIAL MAT MO ต้องใช้วัตถุดิบ 10 "
            "แต่จองได้ 4, forecast_availability = -6 และสถานะวัตถุดิบเป็น partially_available/shortage ตามคาด."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU05-01": {
        "status": "ยืนยันเชิงฟังก์ชัน (มีหมายเหตุ)",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: SO SOD-263066 สำหรับ FG-MTK-IL-01001 "
            "confirm แล้ว complete picking GMP/PICK/03587 และ delivery GMP/OUT/02509 จนครบ 1 หน่วย. "
            "local UAT รอบนี้ fulfill จาก stock-first behavior จึงไม่เกิด fresh MO ในรอบทดสอบ."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07 (มีหมายเหตุ)",
    },
    "MU05-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: SO SOD-263067 สร้าง upstream MO GMP/MOPH/00045 "
            "และ PO P00026 สำหรับ component TMP MTO BUY COMP ตาม shortage path."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU03-01": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Plastic แบบโอนครบ ปิดเอกสาร "
            "GMP/TRPL/00022 ได้สำเร็จ."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU03-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Plastic partial สร้าง backorder มาตรฐาน "
            "จาก GMP/TRPL/00023 ไป GMP/TRPL/00024 จำนวนคงเหลือ 6 หน่วย."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU03-03": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Plastic partial ที่กด No Backorder "
            "สามารถกู้ late backorder ได้จาก GMP/TRPL/00025 เป็น GMP/TRPL/00026 จำนวน 6 หน่วย."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU04-01": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Pharma แบบโอนครบ ปิดเอกสาร "
            "GMP/TRPH/00041 ได้สำเร็จ."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU04-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Pharma partial สร้าง backorder มาตรฐาน "
            "จาก GMP/TRPH/00042 ไป GMP/TRPH/00043 จำนวนคงเหลือ 6 หน่วย."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU04-03": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: Transfer Pharma partial ที่กด No Backorder "
            "สามารถกู้ late backorder ได้จาก GMP/TRPH/00044 เป็น GMP/TRPH/00045 จำนวน 6 หน่วย."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU07-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: MO partial สร้าง backorder มาตรฐานได้ "
            "โดยเกิด backorder qty 4 หน่วย."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU07-04": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: overproduction sync ปรับ product_qty และ component demand "
            "ของ TMP OVER MO เป็น 12 หน่วยตาม output จริง."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU08-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: โอนของเหลือจาก staging กลับ stock สำเร็จ "
            "ผ่านเอกสาร GMP/TRPH/00046 จาก GMP/Stock/คลังลอย กลับ GMP/Stock."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU10-02": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: trace stock movement จาก MO TMP TRACE MO "
            "เห็น raw move done 6 หน่วย และ finished move done 2 หน่วย เชื่อมถึง flow manual MO ได้."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
    "MU10-03": {
        "status": "ยืนยันเชิงฟังก์ชัน",
        "note": (
            "ยืนยันเชิงฟังก์ชันจาก rerun 2026-04-07: BOM demand เท่ากับ actual component demand บน MO "
            "GMP/MOPH/00047 โดย expected 6 และ move component qty = 6."
        ),
        "round_status": "Passed - rerun เชิงธุรกรรม 2026-04-07",
    },
}


def load_rerun_data() -> dict:
    return json.loads(RERUN_JSON.read_text(encoding="utf-8"))


def update_workbook() -> dict:
    rerun = load_rerun_data()
    wb = load_workbook(INPUT_XLSX)

    touched = []
    functional_verified = 0
    functional_note = 0
    not_closed = 0

    for ws in wb.worksheets:
        if ws.title not in PHASE_SHEETS:
            continue
        for row in range(7, ws.max_row + 1):
            case_id = ws.cell(row, 2).value
            if case_id not in OVERRIDES:
                continue
            update = OVERRIDES[case_id]
            ws.cell(row, 13).value = update["status"]
            ws.cell(row, 14).value = update["note"]
            ws.cell(row, 20).value = update["round_status"]
            for col in (10, 11, 12, 14, 18, 20):
                ws.cell(row, col).alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="left"
                )
            touched.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "case_id": case_id,
                    "status": update["status"],
                    "round_status": update["round_status"],
                }
            )
            if update["status"] == "ยืนยันเชิงฟังก์ชัน":
                functional_verified += 1
            elif update["status"] == "ยืนยันเชิงฟังก์ชัน (มีหมายเหตุ)":
                functional_note += 1
            else:
                not_closed += 1

    ws = wb.worksheets[0]
    if ws["A15"].value == "ชุด flow ตามลำดับเหตุการณ์":
        ws.insert_rows(15, amount=5)
    ws["A13"] = "ผล rerun เชิงธุรกรรมรอบนี้"
    ws["B13"] = "ทดสอบธุรกรรม end-to-end ลึกบน local UAT แล้วอัปเดตสถานะเชิงฟังก์ชันตามผลจริง"
    ws["A14"] = "จำนวนเคสที่ rerun รอบนี้"
    ws["B14"] = rerun["summary"]["total_cases"]
    ws["A15"] = "จำนวนเคสที่ขยับเป็นยืนยันเชิงฟังก์ชัน"
    ws["B15"] = functional_verified
    ws["A16"] = "จำนวนเคสที่ยืนยันเชิงฟังก์ชัน (มีหมายเหตุ)"
    ws["B16"] = functional_note
    ws["A17"] = "จำนวนเคสที่ยังไม่ปิดจาก rerun รอบนี้"
    ws["B17"] = not_closed

    wb.save(OUTPUT_XLSX)

    summary = {
        "input_workbook": str(INPUT_XLSX),
        "output_workbook": str(OUTPUT_XLSX),
        "rerun_summary": rerun["summary"],
        "updated_cases": touched,
        "functional_verified": functional_verified,
        "functional_verified_with_note": functional_note,
        "not_closed": not_closed,
    }
    SUMMARY_JSON.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    lines = [
        "# อัปเดตสถานะ rerun เชิงธุรกรรมของชุด actual flow",
        "",
        f"- ไฟล์ต้นทาง: `{INPUT_XLSX}`",
        f"- ไฟล์ผลลัพธ์: `{OUTPUT_XLSX}`",
        f"- rerun รอบนี้: {rerun['summary']['total_cases']} เคส",
        f"- ขยับเป็น `ยืนยันเชิงฟังก์ชัน`: {functional_verified} เคส",
        f"- `ยืนยันเชิงฟังก์ชัน (มีหมายเหตุ)`: {functional_note} เคส",
        f"- ยังไม่ปิดจาก rerun รอบนี้: {not_closed} เคส",
        "",
        "## เคสที่อัปเดต",
        "",
    ]
    for item in touched:
        lines.append(
            f"- `{item['case_id']}` [{item['sheet']}] -> {item['status']} | {item['round_status']}"
        )
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")
    return summary


if __name__ == "__main__":
    summary = update_workbook()
    print(json.dumps(summary, ensure_ascii=False, indent=2))
