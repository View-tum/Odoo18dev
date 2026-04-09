from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANUV2.xlsx")
OUTPUT = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANUV2_detailed_teststeps.xlsx")

STATUSES = ["Not Start", "Pending", "Under Testing", "Passed", "Failed", "Cancelled"]

TOPICS = [
    {
        "sheet": "MU01",
        "title": "ตรวจ On Hand / Forecast / Replenishment",
        "objective": "ให้ planner และคลังเช็กสต็อกจริง, forecasted stock และรายการเติมของก่อนเริ่มผลิตได้อย่างถูกต้อง",
        "coverage": "On Hand, Forecast, Replenishment, shortage review, สินค้าพร้อม/ไม่พร้อม",
        "cases": [
            {
                "case_id": "MU01-01",
                "backlogs": "MA01, MA10",
                "name": "เช็ก On Hand และ Forecast ของ FG ก่อนวางแผนผลิต",
                "role": "Planner / คลัง",
                "path": "Inventory > Products > Products > เปิดสินค้า FG > On Hand / Forecasted",
                "pre": "มีสินค้า FG ตัวอย่าง เช่น FG-PNC-TH-01001 และมี stock อย่างน้อย 1 warehouse",
                "data": "Product = FG-PNC-TH-01001, Warehouse = GMP",
                "steps": "1) เปิดเมนู Inventory > Products > Products\\n2) ค้นหาสินค้า FG-PNC-TH-01001 แล้วกดเปิดฟอร์มสินค้า\\n3) ดูค่า On Hand ว่ามี stock ปัจจุบันอยู่เท่าไร\\n4) กด smart button Forecasted หรือเปิดเมนู Forecasted Report\\n5) ตรวจว่ามี incoming / outgoing / manufacturing demand อะไรผูกอยู่บ้าง\\n6) จดตัวเลข On Hand, Free to Use และ Forecasted เพื่อนำไปคุยกับ planner",
                "expected": "ผู้ใช้ต้องเห็นความต่างระหว่าง On Hand, Reserved, Forecasted ได้ และตอบได้ว่าสินค้าพร้อมขาย/พร้อมผลิตหรือไม่",
                "fix": "ถ้าเลขไม่ตรง ให้กด Product Moves และตรวจ stock move ล่าสุด, lot/location และ draft transfer ที่ยังไม่ done",
                "evidence": "Capture: product form, forecast report, product moves",
                "note": "ใช้เคสนี้อธิบายพื้นฐานก่อนเข้า Replenishment",
            },
            {
                "case_id": "MU01-02",
                "backlogs": "MA01, MA02, MA06",
                "name": "เช็ก Replenishment ว่ามีรายการตกหล่นหรือวัตถุดิบขาดหรือไม่",
                "role": "Planner / MRP",
                "path": "Inventory > Operations > Replenishment",
                "pre": "มีสินค้าและ orderpoint จริงในระบบ เช่น FG-PSS-TH-01005 หรือ FG-PNC-TH-01001",
                "data": "Warehouse = GMP, Product = FG-PSS-TH-01005",
                "steps": "1) เปิดเมนู Inventory > Operations > Replenishment\\n2) Filter เฉพาะ Warehouse = GMP\\n3) ค้นหาสินค้า FG-PSS-TH-01005\\n4) ตรวจค่า Forecasted, Min, Max และ Route ที่ใช้เติมของ\\n5) ถ้ามี shortage ให้เปิดเอกสารสินค้าและ BOM เพื่อดูว่าขาดที่ FG, Semi หรือ RM\\n6) ถ้ามีปุ่ม Order Once / Replenish ให้กดตามสิทธิ์และตรวจว่าระบบสร้าง procurement ต่อได้",
                "expected": "ผู้ใช้ต้องตอบได้ว่าระบบจะเติมของโดย route ไหน และ shortage เกิดที่จุดไหนของ chain",
                "fix": "ถ้าไม่ขึ้นใน Replenishment ให้ตรวจ orderpoint, route บนสินค้า, warehouse และ stock forecast",
                "evidence": "Capture: Replenishment screen, product route, orderpoint",
                "note": "เหมาะกับการสอนความต่างระหว่าง MTS กับสินค้าไม่มี trigger",
            },
            {
                "case_id": "MU01-03",
                "backlogs": "MA02, MA10, MA16",
                "name": "เช็กว่ารายการที่เติมของแล้วสร้างเอกสารถูกหรือไม่",
                "role": "Planner / Accounting Review",
                "path": "Inventory > Operations > Replenishment > เปิดเอกสารที่ระบบสร้าง",
                "pre": "มีรายการ Replenishment ที่สร้างเอกสารแล้ว หรือทำต่อจากเคส MU01-02",
                "data": "ใช้รายการที่ระบบสร้างจาก FG-PNC-TH-01001 หรือ FG-PSS-TH-01005",
                "steps": "1) เปิดรายการ Replenishment ที่เพิ่ง trigger\\n2) เปิดเอกสารปลายทาง เช่น MO, Transfer หรือ RFQ\\n3) ตรวจ Operation Type ว่าถูกฝั่ง Pharma / Plastic / Buy\\n4) ตรวจ source/destination location ของ move ว่าตรงตาม flow ที่ออกแบบไว้\\n5) ตรวจว่า origin และ procurement group ถูก propagate มาถูกใบงาน",
                "expected": "เอกสารปลายทางต้องตรง route/rule ที่ตั้งไว้ และ user อธิบายได้ว่าทำไมระบบสร้างเอกสารชนิดนั้น",
                "fix": "ถ้าเอกสารผิดฝั่ง ให้ย้อนตรวจ route ที่สินค้า, orderpoint route และ rule sequence",
                "evidence": "Capture: MO/Transfer/RFQ ที่ระบบสร้าง",
                "note": "ใช้ปิดบท Replenishment ก่อนเข้าเคสผลิตจริง",
            },
        ],
    },
    {
        "sheet": "MU02",
        "title": "สร้าง MO เองและเช็กความพร้อมวัตถุดิบ",
        "objective": "ให้ฝ่ายผลิตและ planner เปิด MO manual, ตรวจ material availability และแยกกรณีของครบ / ของขาดได้",
        "coverage": "Manual MO, component reservation, material availability, raw move review",
        "cases": [
            {
                "case_id": "MU02-01",
                "backlogs": "MA01, MA08, MA14",
                "name": "เปิด MO manual กรณีวัตถุดิบครบ",
                "role": "Planner / Production User",
                "path": "Manufacturing > Operations > Manufacturing Orders",
                "pre": "สินค้าและ BOM พร้อมใช้งาน และ RM/Semi ที่ต้องใช้มี stock เพียงพอ",
                "data": "Product ตัวอย่าง = SM-PLS-UP-01001 หรือ FG-PNC-TH-01001",
                "steps": "1) เปิดเมนู Manufacturing > Operations > Manufacturing Orders\\n2) กด New\\n3) เลือก Product, BOM และ Quantity\\n4) ตรวจ Operation Type ว่าเป็นฝั่งที่ถูกต้อง\\n5) กด Confirm\\n6) ตรวจแท็บ Components ว่าขึ้น Available หรือ Reserved ครบทุกบรรทัด\\n7) เปิด Raw Moves และจดจำนวนที่ระบบจองให้",
                "expected": "MO ต้อง confirm ได้ทันที, raw material reserve ครบ และไม่มี shortage warning",
                "fix": "ถ้า reserve ไม่ครบ ให้เช็ก stock by location, lot, owner และ route ของวัตถุดิบ",
                "evidence": "Capture: MO form, Components tab, Raw Moves",
                "note": "ใช้เป็น baseline ก่อนเทสกรณีของขาด",
            },
            {
                "case_id": "MU02-02",
                "backlogs": "MA01, MA08, MA14",
                "name": "เปิด MO manual กรณีมีของบางส่วน",
                "role": "Planner / Production User",
                "path": "Manufacturing > Operations > Manufacturing Orders",
                "pre": "มีบาง component เพียงพอ แต่มีอย่างน้อย 1 รายการที่ไม่พอ",
                "data": "ใช้สินค้าเดียวกับ MU02-01 แต่ปรับ stock ให้มี shortage 1 รายการ",
                "steps": "1) เปิด MO ใหม่ด้วยสินค้าเดียวกับ baseline\\n2) กด Confirm\\n3) ตรวจ Components ว่ามีบางบรรทัด Available/Reserved และบางบรรทัด Waiting\\n4) เปิด Forecasted ของ component ที่ขาด\\n5) ตัดสินใจว่าจะรอซื้อ, trigger replenishment หรือใช้ stock float/stock อื่น",
                "expected": "ระบบต้องแยก component ที่พร้อมกับที่ไม่พร้อมได้ชัดเจน และ user ต้องรู้ว่าต้องไปแก้ที่ component ตัวใด",
                "fix": "ถ้าทุกบรรทัดขึ้นผิดสถานะ ให้ตรวจ BOM line UoM, warehouse และ source location ของ rule",
                "evidence": "Capture: component lines ที่พร้อม/ไม่พร้อม",
                "note": "ใช้สอนการอ่านสถานะ component อย่างละเอียด",
            },
            {
                "case_id": "MU02-03",
                "backlogs": "MA04, MA06, MA07",
                "name": "ตรวจการเบิกของเข้าคลังลอย / staging ของแต่ละแผนก",
                "role": "Production Plastic / Production Pharma / Warehouse",
                "path": "MO > Components / smart button Transfers",
                "pre": "มี MO ที่ confirm แล้วและมี route transfer ของ Semi/Raw อยู่จริง",
                "data": "MO ฝั่ง Plastic หรือ Pharma ที่มี component ต้องโอนเข้าคลังลอย",
                "steps": "1) เปิด MO ที่ confirm แล้ว\\n2) กด smart button Transfers หรือดู component move ที่สร้าง transfer\\n3) แยกให้ได้ว่างานนี้ใช้ Transfer Plastic หรือ Transfer Pharma\\n4) เปิด transfer แล้วตรวจ source/destination ว่ามาจาก stock หรือ semi และปลายทางเป็นคลังลอย/staging\\n5) ตอบให้ได้ว่าแผนกไหนเป็นคนทำเอกสารนี้ และทำก่อนหรือหลัง Work Order",
                "expected": "ผู้ใช้งานต้องแยก flow การเบิกของของแต่ละแผนกได้ และรู้ว่า transfer นี้เป็นของ Plastic หรือ Pharma",
                "fix": "ถ้า transfer ไม่ถูกสร้าง ให้ย้อนดู route Auto Transfer Semi และ rule source/destination",
                "evidence": "Capture: MO + linked transfer",
                "note": "หัวใจของการสอนเรื่องเบิกของเข้าผลิตแต่ละแผนก",
            },
        ],
    },
    {
        "sheet": "MU03",
        "title": "Transfer Plastic และ Backorder",
        "objective": "ให้คลังและฝ่ายผลิตรู้ผลของการกด Create Backorder / No Backorder ในเอกสาร Transfer Plastic",
        "coverage": "Internal transfer plastic, partial done, backorder, no backorder, recovery",
        "cases": [
            {
                "case_id": "MU03-01",
                "backlogs": "MA04, MA07",
                "name": "Transfer Plastic กรณีโอนของครบ",
                "role": "Warehouse / Plastic",
                "path": "Inventory > Operations > Transfers > Operation Type = Transfer Plastic",
                "pre": "มีเอกสาร Transfer Plastic draft/ready ที่มี stock พร้อมครบ",
                "data": "Transfer Plastic ที่ผูกจาก MO จริง หรือสร้าง manual เพื่อทดสอบ",
                "steps": "1) เปิดเมนู Transfers และ filter Operation Type = Transfer Plastic\\n2) เปิดเอกสารที่พร้อมทำงาน\\n3) ใส่ Done Qty เท่ากับ Demand ทุกบรรทัด\\n4) กด Validate\\n5) ตรวจว่าเอกสารปิด done โดยไม่ถาม backorder",
                "expected": "เอกสาร done ใบเดียว ไม่มี backorder และ stock ย้ายไป location ปลายทางครบ",
                "fix": "ถ้ายังถาม backorder แปลว่า done qty ไม่ครบหรือมีบรรทัดตกหล่น ให้ตรวจ move line ทุกบรรทัด",
                "evidence": "Capture: transfer before/after validate",
                "note": "ใช้เทียบกับกรณี partial ด้านล่าง",
            },
            {
                "case_id": "MU03-02",
                "backlogs": "MA04, MA07, MA16",
                "name": "Transfer Plastic กรณีของมาไม่ครบและเลือก Create Backorder",
                "role": "Warehouse / Plastic",
                "path": "Inventory > Operations > Transfers > Validate partial",
                "pre": "มี transfer ที่ต้องการ 100 แต่มีของพร้อมจริงน้อยกว่า",
                "data": "ตั้ง Done Qty บางบรรทัดต่ำกว่าความต้องการ",
                "steps": "1) เปิดเอกสาร Transfer Plastic\\n2) ใส่ Done Qty เฉพาะจำนวนที่หยิบได้จริง\\n3) กด Validate\\n4) เมื่อ popup ถาม backorder ให้เลือก Create Backorder\\n5) เปิดเอกสารเดิมและเอกสาร backorder ใหม่เพื่อตรวจ remaining qty",
                "expected": "เอกสารเดิมต้อง done เฉพาะที่โอนได้จริง และระบบสร้าง backorder ใหม่สำหรับค้างโอน",
                "fix": "ถ้าไม่สร้าง backorder ให้ตรวจ Operation Type ว่าตั้ง Create Backorder = Ask/Always หรือไม่",
                "evidence": "Capture: popup backorder, original transfer, new backorder",
                "note": "ให้ผู้ทดสอบบันทึกเลขเอกสารคู่เดิม/ใหม่",
            },
            {
                "case_id": "MU03-03",
                "backlogs": "MA04, MA07, MA16",
                "name": "Transfer Plastic กรณีกด No Backorder ผิด แล้วกู้กลับ",
                "role": "Warehouse / Plastic",
                "path": "Inventory > Operations > Transfers > Done transfer",
                "pre": "ใช้เคส partial เช่น MU03-02 แต่ตั้งใจเลือก No Backorder",
                "data": "Done Qty น้อยกว่าความต้องการ และ Operation Type = Ask Backorder",
                "steps": "1) เปิด Transfer Plastic ที่ partial\\n2) กด Validate\\n3) เมื่อ popup ถาม backorder ให้เลือก No Backorder\\n4) เปิดเอกสารที่ done แล้วอีกครั้ง\\n5) ตรวจว่าปุ่ม Create Backorder แสดงขึ้นหรือไม่\\n6) กดปุ่ม Create Backorder\\n7) เปิดเอกสาร backorder ที่ระบบสร้างขึ้นใหม่",
                "expected": "ถ้า user กด No Backorder ผิด ระบบต้องให้กด Create Backorder ย้อนกลับได้ และเมื่อสร้างแล้วปุ่มต้องหาย",
                "fix": "ถ้าปุ่มไม่ขึ้น ให้ตรวจว่ามี remaining qty จริง, operation type เป็น Ask และเอกสารยังไม่มี backorder มาก่อน",
                "evidence": "Capture: done transfer, recovery button, backorder created",
                "note": "ใช้ยืนยัน custom late_backorder_recovery",
            },
        ],
    },
    {
        "sheet": "MU04",
        "title": "Transfer Pharma และ Backorder",
        "objective": "ให้ผู้ใช้ฝั่งยาเข้าใจการโอนของ partial และผลของการเลือก backorder/no backorder",
        "coverage": "Transfer Pharma, partial transfer, recovery path",
        "cases": [
            {
                "case_id": "MU04-01",
                "backlogs": "MA04, MA07",
                "name": "Transfer Pharma กรณีโอนครบ",
                "role": "Warehouse / Pharma",
                "path": "Inventory > Operations > Transfers > Operation Type = Transfer Pharma",
                "pre": "มีเอกสาร Transfer Pharma ที่พร้อมและ stock ครบ",
                "data": "Transfer Pharma จาก MO ฝั่งยา",
                "steps": "1) เปิด transfer ฝั่ง Pharma\\n2) ใส่ Done Qty ครบตาม Demand\\n3) กด Validate\\n4) ตรวจ source/destination location และ lot/serial ถ้ามี",
                "expected": "เอกสาร done ได้ทันที ไม่มีค้าง backorder",
                "fix": "ถ้า validate ไม่ผ่าน ให้ดู lot requirement หรือ stock location ต้นทาง",
                "evidence": "Capture: completed transfer pharma",
                "note": "ใช้เทียบกับ partial transfer",
            },
            {
                "case_id": "MU04-02",
                "backlogs": "MA04, MA07, MA16",
                "name": "Transfer Pharma partial แล้วเลือก Create Backorder",
                "role": "Warehouse / Pharma",
                "path": "Inventory > Operations > Transfers",
                "pre": "มี transfer pharma ที่โอนไม่ครบจริง",
                "data": "Done Qty ต่ำกว่าความต้องการอย่างน้อย 1 บรรทัด",
                "steps": "1) เปิด transfer\\n2) ใส่ Done Qty ตามที่มีจริง\\n3) กด Validate\\n4) เลือก Create Backorder\\n5) เปิด backorder ที่สร้างและเช็ก remaining qty",
                "expected": "เอกสารเดิม done บางส่วน และมี backorder ใหม่สำหรับยอดค้าง",
                "fix": "ถ้าเอกสารใหม่ไม่เกิด ให้ตรวจ create_backorder policy ของ operation type",
                "evidence": "Capture: popup + backorder transfer pharma",
                "note": "ให้ user จดเลขเอกสารคู่เพื่อ trace",
            },
            {
                "case_id": "MU04-03",
                "backlogs": "MA04, MA07, MA16",
                "name": "Transfer Pharma partial แล้วกด No Backorder จากนั้นกู้กลับ",
                "role": "Warehouse / Pharma",
                "path": "Inventory > Operations > Transfers > done transfer",
                "pre": "ตั้งใจสร้างเคส partial transfer pharma",
                "data": "เหมือน MU04-02 แต่เลือก No Backorder",
                "steps": "1) ทำ partial transfer\\n2) ตอน popup ให้เลือก No Backorder\\n3) เปิดเอกสารที่ done\\n4) ตรวจว่ามีปุ่ม Create Backorder บนเอกสาร\\n5) กดปุ่มและเปิด backorder ที่สร้างใหม่",
                "expected": "ระบบกู้ backorder ได้จากเอกสารเดิม และหลังสร้างแล้วปุ่ม recovery ต้องไม่แสดงซ้ำ",
                "fix": "ถ้า recovery ใช้ไม่ได้ ให้เช็ก custom module และเงื่อนไข remaining qty",
                "evidence": "Capture: no backorder path + recovery path",
                "note": "ใช้สอนความต่างระหว่างการเลือกสองปุ่มนี้โดยตรง",
            },
        ],
    },
    {
        "sheet": "MU05",
        "title": "MTO จาก Sales Order",
        "objective": "ให้ทีมขายและ MRP เข้าใจว่า SO แบบ MTO เริ่ม demand อย่างไรและแตกเอกสารอะไรบ้าง",
        "coverage": "SO confirm, auto MO, shortage handling, priority MTO",
        "cases": [
            {
                "case_id": "MU05-01",
                "backlogs": "MA02, MA12",
                "name": "SO ต่างประเทศ MTO กรณีวัตถุดิบครบ",
                "role": "Sales / Planner",
                "path": "Sales > Orders > Quotations > Confirm > trace MO",
                "pre": "FG ใช้ route MTO + Manufacture และ RM/Semi ที่ต้องใช้มี stock พร้อม",
                "data": "Product ตัวอย่าง = FG-MTK-IL-01001",
                "steps": "1) เปิด Sales > Quotations แล้วสร้าง SO ใหม่\\n2) ใส่ลูกค้า, Product = FG-MTK-IL-01001, Qty ตาม test\\n3) กด Confirm\\n4) เปิด smart button Delivery / Manufacturing หรือเปิดเอกสารผ่าน origin\\n5) ตรวจว่าเกิด MO จาก SO จริง\\n6) เปิด Components และตรวจว่า reserve วัตถุดิบได้ครบ",
                "expected": "SO confirm แล้วเกิด MO ตาม demand จากออเดอร์ และไม่ควรเกิด PO ถ้าของครบ",
                "fix": "ถ้าไม่เกิด MO ให้ตรวจ route MTO, BOM, manufacturing route และ stock policy ของสินค้า",
                "evidence": "Capture: SO, linked MO, component status",
                "note": "ใช้สอน order-driven supply",
            },
            {
                "case_id": "MU05-02",
                "backlogs": "MA01, MA02, MA12",
                "name": "SO ต่างประเทศ MTO กรณีวัตถุดิบขาดบางตัว",
                "role": "Sales / Planner / Purchase",
                "path": "Sales > Orders > Quotations > Confirm > trace shortage",
                "pre": "FG ใช้ route MTO แต่มี RM บางรายการไม่พอ และ RM route = Buy",
                "data": "ใช้สินค้าเดียวกับ MU05-01 แต่ลด stock RM บางรายการ",
                "steps": "1) สร้างและ Confirm SO\\n2) เปิด MO ที่ระบบสร้าง\\n3) ตรวจ Components ว่ามีรายการ Waiting/Not Available\\n4) เปิด component ที่ขาดแล้วดู route / vendor\\n5) ตรวจ procurement ว่าไป RFQ/PO หรือ Replenishment list หรือไม่",
                "expected": "ระบบต้องสร้าง MO จาก SO และแยกให้เห็นชัดว่ามี component ตัวไหนต้องซื้อ/เติมเพิ่ม",
                "fix": "ถ้าไม่ไป PO ให้ตรวจ Buy route, vendor, reordering policy และ company/warehouse บนสินค้า",
                "evidence": "Capture: MO shortage, product route, RFQ/PO/replenishment",
                "note": "สอนความเชื่อมระหว่าง MTO และ Buy",
            },
            {
                "case_id": "MU05-03",
                "backlogs": "MA03, MA12",
                "name": "มีงาน MTS ค้างอยู่ แล้วมี MTO เร่งด่วนแทรก",
                "role": "Planner / Production Supervisor",
                "path": "Manufacturing > Operations > Manufacturing Orders / Work Orders",
                "pre": "มี MO ฝั่ง MTS ค้างอย่างน้อย 1 ใบ และมี SO ใหม่แบบ MTO",
                "data": "ใช้ MTO ของ FG-MTK-IL-01001 เทียบกับ MTS ของ FG-PNC-TH-01001",
                "steps": "1) เปิด MTS MO ที่กำลังรอผลิต/กำลังผลิต\\n2) สร้าง SO MTO ใหม่และกด Confirm\\n3) เปิด MO ใหม่ที่เกิดจาก SO\\n4) ตรวจ priority, source demand, raw move และการจองเครื่อง/วัตถุดิบ\\n5) บันทึกแนวทางการตัดสินใจว่าจะแทรกคิวหรือไม่",
                "expected": "ทีมต้องมองเห็นงาน MTO แยกจาก MTS ชัดเจน และตัดสินใจจัดลำดับได้โดยไม่ทำ flow เดิมเสีย",
                "fix": "ถ้าข้อมูล priority ไม่ชัด ให้ใช้ manual priority / sequence / planned start time ตามนโยบายหน้างาน",
                "evidence": "Capture: MTO MO, MTS MO, priority fields",
                "note": "เคสนี้ใช้คุย business rule แม้ระบบยังให้ manual decision",
            },
        ],
    },
    {
        "sheet": "MU06",
        "title": "MTS / Reordering Rule / Min-Max",
        "objective": "ให้ planner เข้าใจการเติม stock แบบ MTS และผลเมื่อ stock พอหรือไม่พอ",
        "coverage": "orderpoint, replenish, no action, auto MO",
        "cases": [
            {
                "case_id": "MU06-01",
                "backlogs": "MA02, MA06",
                "name": "MTS กรณี stock ต่ำกว่าเกณฑ์และระบบต้องสร้าง MO",
                "role": "Planner",
                "path": "Inventory > Operations > Replenishment / Manufacturing > MO",
                "pre": "มี orderpoint สำหรับ FG เช่น FG-PNC-TH-01001 และ forecast ต่ำกว่าจุด trigger",
                "data": "Product = FG-PNC-TH-01001, Orderpoint = GMP/Stock",
                "steps": "1) เปิด Replenishment\\n2) ค้นหา FG-PNC-TH-01001\\n3) ตรวจว่า forecast ต่ำและมี action ให้เติมของ\\n4) กด Replenish หรือรันตาม process ที่ทีมใช้\\n5) เปิด MO ที่สร้างขึ้นใหม่",
                "expected": "ระบบต้องสร้าง MO ฝั่ง Manufacturing Pharma ตาม route ของสินค้า",
                "fix": "ถ้าไม่สร้าง MO ให้ตรวจ orderpoint, route บนสินค้า และ warehouse context",
                "evidence": "Capture: replenishment row + created MO",
                "note": "ใช้สอน stock-driven manufacturing",
            },
            {
                "case_id": "MU06-02",
                "backlogs": "MA02, MA10",
                "name": "MTS กรณี stock ยังพอ ไม่ควรสร้างเอกสารเพิ่ม",
                "role": "Planner",
                "path": "Inventory > Replenishment / Forecasted",
                "pre": "สินค้าเดียวกับ MU06-01 แต่ stock เพียงพอ",
                "data": "FG เดียวกันแต่เติม stock ให้ forecast ไม่ติดลบ",
                "steps": "1) เปิด product forecast หรือ replenishment row\\n2) ตรวจค่าคงเหลือและ forecast\\n3) ยืนยันว่าไม่มี action ให้เติมของ\\n4) ตรวจว่าไม่มี MO/RFQ ใหม่เกิดเพิ่ม",
                "expected": "ระบบต้องไม่สร้าง MO ถ้า stock ยังพอ",
                "fix": "ถ้าเกิด MO ทั้งที่ stock พอ ให้ย้อนดู orderpoint min/max, manual replenish ที่ค้าง และ move ที่ยังไม่ done",
                "evidence": "Capture: no-action state",
                "note": "ใช้เปรียบเทียบกับ MU06-01",
            },
            {
                "case_id": "MU06-03",
                "backlogs": "MA01, MA02, MA07",
                "name": "MTS ที่แตก child chain ไป Semi/Plastic/Pharma",
                "role": "Planner / Production",
                "path": "MO > Components / child MO / linked transfers",
                "pre": "ใช้สินค้า MTS ที่มี BOM หลายชั้น เช่น FG-PSS-TH-01005",
                "data": "Product = FG-PSS-TH-01005",
                "steps": "1) Trigger replenishment ของ FG-PSS-TH-01005\\n2) เปิด MO แม่ที่ระบบสร้าง\\n3) ไล่ดู child requirements ว่ามีทั้ง FG ย่อย, semi plastic, solution pharma หรือไม่\\n4) เปิด transfer/MO ลูกที่ระบบสร้างต่อ\\n5) สรุปให้ได้ว่า chain นี้ผ่าน Plastic, Pharma และ Buy ตรงไหนบ้าง",
                "expected": "ทีมต้องเห็น chain จริงและ trace ได้ว่าของแต่ละชั้นถูกจัดหาด้วย flow แบบใด",
                "fix": "ถ้า chain ไม่ครบ ให้ตรวจ BOM level, route ของลูก และ stock availability ของแต่ละชั้น",
                "evidence": "Capture: MO แม่ + child docs",
                "note": "ใช้ผูกกับ training route/rule",
            },
        ],
    },
    {
        "sheet": "MU07",
        "title": "ปิดงานผลิต / MO Backorder / Recovery",
        "objective": "ให้ฝ่ายผลิตเข้าใจผลของการผลิตครบ, ผลิตไม่ครบ, ผลิตเกิน และการกู้ backorder หลังเลือกผิด",
        "coverage": "workorder done, MO done, backorder ask, no backorder recovery",
        "cases": [
            {
                "case_id": "MU07-01",
                "backlogs": "MA08, MA16",
                "name": "ปิด MO ได้ครบตามแผน",
                "role": "Production User",
                "path": "Manufacturing > Work Orders / MO > Produce All",
                "pre": "มี MO ที่ raw materials พร้อมและ workorder พร้อมทำงาน",
                "data": "MO ตัวอย่าง = GMP/MOPH/... หรือ GMP/MOPL/...",
                "steps": "1) เปิด Work Order หรือ MO\\n2) Start งานและบันทึก good qty ตามแผน\\n3) ใส่ consume ตามจริงถ้ามีขั้นตอน manual\\n4) กด Mark as Done / Validate\\n5) ตรวจ finished move และ stock ที่เพิ่มเข้า FG/Semi",
                "expected": "MO ปิด done ได้ใบเดียว ไม่มี backorder และ stock finished เพิ่มถูกต้อง",
                "fix": "ถ้า done ไม่ได้ ให้ตรวจ lot/serial, workorder state, raw move และ quantity produced",
                "evidence": "Capture: MO done, finished moves",
                "note": "baseline ก่อนเทส partial production",
            },
            {
                "case_id": "MU07-02",
                "backlogs": "MA08, MA16, MA17, MA18",
                "name": "ผลิตไม่ครบและเลือก Create Backorder",
                "role": "Production User / Planner",
                "path": "MO / Work Order > partial done",
                "pre": "มี MO ที่ตั้ง qty สูงกว่า output ที่จะผลิตจริง",
                "data": "MO partial test",
                "steps": "1) เปิด MO หรือ Work Order\\n2) ผลิตจริงน้อยกว่าปริมาณแผน\\n3) กด Done/Validate\\n4) เมื่อ popup ถาม backorder ให้เลือก Create Backorder\\n5) เปิด MO เดิมและ MO backorder ใหม่เพื่อตรวจ remaining qty",
                "expected": "MO เดิม done ตาม qty จริง และมี backorder MO สำหรับยอดค้าง",
                "fix": "ถ้าไม่เกิด MO ใหม่ ให้ตรวจ backorder policy ของ operation type และ qty produced ว่ากรอกต่ำกว่าตามแผนจริง",
                "evidence": "Capture: popup + backorder MO",
                "note": "ใช้คุยผลกระทบต่อ accounting/month-end ด้วย",
            },
            {
                "case_id": "MU07-03",
                "backlogs": "MA08, MA16, MA17, MA18",
                "name": "ผลิตไม่ครบและกด No Backorder แล้วสร้างกลับภายหลัง",
                "role": "Production User / Planner",
                "path": "MO done > recovery button",
                "pre": "มี MO partial และ operation ตั้ง Create Backorder = Ask",
                "data": "Partial MO เช่น MU07-02 แต่เลือก No Backorder",
                "steps": "1) ทำ partial production\\n2) ตอน popup ให้เลือก No Backorder\\n3) เปิด MO ที่ done แล้วอีกครั้ง\\n4) ตรวจว่ามีปุ่ม Create Backorder บนฟอร์ม\\n5) กดปุ่มเพื่อสร้าง backorder ใหม่\\n6) เปิด MO backorder และตรวจ remaining qty",
                "expected": "ระบบกู้ backorder ให้ได้จาก MO ใบเดิม และหลังสร้างแล้วปุ่ม recovery หาย",
                "fix": "ถ้าปุ่มไม่ขึ้น ให้ตรวจว่ามี qty ค้างจริงและยังไม่เคยสร้าง backorder มาก่อน",
                "evidence": "Capture: original MO, recovery button, new backorder MO",
                "note": "ใช้เทส custom late backorder recovery ฝั่ง MRP",
            },
            {
                "case_id": "MU07-04",
                "backlogs": "MA08, MA10, MA16",
                "name": "ผลิตเกินแผนและตรวจ actual component consumption",
                "role": "Production User / Accounting Review",
                "path": "MO > Components / Costing",
                "pre": "มี MO ที่อนุญาตให้ผลิตเกินหรือปรับ quantity produced สูงกว่าแผนได้",
                "data": "MO over-production test",
                "steps": "1) เปิด MO\\n2) ใส่ quantity produced มากกว่า planned qty ตาม policy ที่อนุญาต\\n3) กด Done\\n4) เปิด component moves และ valuation/costing\\n5) เทียบ BOM theoretical กับ actual ที่ใช้จริง",
                "expected": "MO ต้องปิดได้ตาม policy และทีมต้อง trace actual usage เทียบกับ BOM ได้",
                "fix": "ถ้าปิดไม่ได้ ให้ตรวจ setting overproduction/tolerance และ UoM rounding",
                "evidence": "Capture: component moves + cost view",
                "note": "ใช้คุยเรื่อง BOM vs actual",
            },
        ],
    },
    {
        "sheet": "MU08",
        "title": "Scrap / ของเหลือ / คลังลอย",
        "objective": "ให้ทีมผลิตและคลังแยกกรณี scrap, reject และของเหลือกลับคลังได้",
        "coverage": "scrap, leftover return, float stock handling",
        "cases": [
            {
                "case_id": "MU08-01",
                "backlogs": "MA08, MA13",
                "name": "บันทึก Scrap ระหว่างการผลิต",
                "role": "Production User / Warehouse",
                "path": "Work Order / Inventory > Operations > Scrap",
                "pre": "มี WO/MO ที่อยู่ระหว่างหรือหลังการผลิต และมีของเสียจริงตาม test",
                "data": "เลือก product/component ที่จะ scrap",
                "steps": "1) เปิด Work Order หรือเมนู Scrap\\n2) เลือก Product และ Source Location ที่ถูกต้อง\\n3) ใส่ Qty ที่ scrap\\n4) Validate Scrap\\n5) กลับไปตรวจ stock card / product moves ของสินค้านั้น",
                "expected": "qty ใน source location ลดลง, มีเอกสาร scrap, และ trace ย้อนหลังได้ว่า scrap จากงานไหน",
                "fix": "ถ้า stock ไม่ลด ให้ตรวจ source location และสถานะ scrap ว่า done หรือยัง",
                "evidence": "Capture: scrap form, product moves",
                "note": "ย้ำว่าของเสียไม่ใช่แค่จดโน้ต ต้องมี stock movement",
            },
            {
                "case_id": "MU08-02",
                "backlogs": "MA06, MA07",
                "name": "คืนของเหลือจากคลังลอย / staging กลับ stock",
                "role": "Warehouse / Production",
                "path": "Inventory > Transfers / MO leftover handling",
                "pre": "มีวัตถุดิบเหลือในคลังลอยหลังปิดงาน",
                "data": "ใช้ component ที่ค้างใน GMP/Stock/คลังลอย หรือ staging",
                "steps": "1) เปิด location report ของคลังลอย\\n2) ระบุวัตถุดิบที่เหลือหลังจบงาน\\n3) สร้าง Internal Transfer คืนเข้าคลังหลักหรือ location ที่กำหนด\\n4) Validate transfer\\n5) ตรวจว่า stock move trace กลับไป location ที่ถูกต้อง",
                "expected": "ของเหลือไม่ค้างในคลังลอยเกินจริง และ stock กลับเข้าคลังหลักอย่าง traceable",
                "fix": "ถ้าของเหลือหายจากระบบ ให้เทียบ quantity ใน location report กับ transfer history",
                "evidence": "Capture: location report before/after",
                "note": "เคสนี้ผูกกับ backlog คลังลอย",
            },
            {
                "case_id": "MU08-03",
                "backlogs": "MA13",
                "name": "ของเสียที่นำกลับมาใช้ใหม่ (ถ้ามี setup by-product/reuse)",
                "role": "Production / Warehouse / Planner",
                "path": "MO > By-Product / Inventory moves",
                "pre": "มีสินค้า/ของเสียที่ตั้งให้รับกลับใช้ใหม่ได้",
                "data": "ตัวอย่างไม้แขวนเสื้อหรือ by-product ตาม setup จริง",
                "steps": "1) เปิด MO ที่ควรมีของเสียรับกลับ\\n2) ตรวจว่ามี by-product หรือ scrap-to-stock flow หรือไม่\\n3) ปิดงานและดู stock move ของของเสียที่รับกลับ\\n4) ตรวจว่าของดังกล่าวถูกใช้เป็น RM ในรอบถัดไปได้หรือไม่",
                "expected": "ถ้ามี setup รองรับ ระบบต้อง trace ของเสียที่รับกลับเป็น stock ใหม่ได้",
                "fix": "ถ้า flow นี้ยังไม่รองรับ ให้บันทึกเป็น gap และยืนยัน process manual ปัจจุบัน",
                "evidence": "Capture: by-product / related stock move",
                "note": "เคสนี้เป็น coverage เชิง process แม้บางฐานอาจยังไม่เปิดใช้เต็ม",
            },
        ],
    },
    {
        "sheet": "MU09",
        "title": "Mold / Workcenter / Shopfloor",
        "objective": "ให้หัวหน้างานเข้าใจการ assign mold, workcenter และการบันทึกใน shopfloor",
        "coverage": "mold matrix, parallel workorder, shopfloor logging",
        "cases": [
            {
                "case_id": "MU09-01",
                "backlogs": "MA11, MA14",
                "name": "ระบบ assign mold และ workcenter อัตโนมัติ",
                "role": "Production Supervisor",
                "path": "MO > Work Orders / Work Center / Mold",
                "pre": "มีสินค้าที่ map mold/workcenter แล้ว เช่น SM-PLS-UP-01001 หรือ SM-JOI-PK-02001",
                "data": "ใช้สินค้าในกลุ่ม plastic ที่มี mold mapping",
                "steps": "1) เปิด MO ของสินค้า plastic ที่มี mold mapping\\n2) เปิด Work Orders\\n3) ตรวจ Work Center ที่ระบบ assign มา\\n4) เปิดรายละเอียด WO และดู mold ที่ถูก assign\\n5) เทียบกับ matrix ว่าตรงกับที่ setup ไว้",
                "expected": "ระบบต้องเลือก machine/mold ที่เข้ากันได้ตาม matrix",
                "fix": "ถ้า mold ไม่ขึ้น ให้ตรวจ Compatible Machines, Mold Matrix และ Is Mold?",
                "evidence": "Capture: WO + mold assignment",
                "note": "ใช้สอนคนหน้างานดูผล ไม่ใช่แค่ admin setup",
            },
            {
                "case_id": "MU09-02",
                "backlogs": "MA11",
                "name": "Parallel workorder ต้องไม่ใช้ mold ซ้ำกัน",
                "role": "Production Supervisor",
                "path": "MO > Work Orders",
                "pre": "มี MO ที่แตก parallel WO บนหลายเครื่อง",
                "data": "ใช้ MO plastic ที่มี parallel scenario",
                "steps": "1) เปิด MO ที่มีหลาย WO\\n2) ตรวจว่าแต่ละ WO active/cancel เป็นอย่างไร\\n3) ตรวจ mold บน WO แต่ละใบ\\n4) ยืนยันว่า mold เดียวกันไม่ถูก assign ให้สอง WO active พร้อมกัน",
                "expected": "มี active WO เพียงตัวที่ถือ mold จริง และ sibling ที่เกินต้องไม่ใช้ mold ซ้ำ",
                "fix": "ถ้า mold ซ้ำ ให้ตรวจ custom guard และ workorder state/cancel logic",
                "evidence": "Capture: list WO + mold column",
                "note": "ใช้ย้ำ bug fix ล่าสุด",
            },
            {
                "case_id": "MU09-03",
                "backlogs": "MA08, MA11, MA14",
                "name": "บันทึก good qty / reject / labor ผ่าน shopfloor",
                "role": "Production User / Supervisor",
                "path": "GMP Shop Floor / Work Order tablet view",
                "pre": "มี WO ที่พร้อมเริ่มงานและมี operator ทดสอบ",
                "data": "WO ฝั่ง plastic/pharma 1 ใบ",
                "steps": "1) เปิดหน้าจอ Shop Floor หรือ Tablet view ของ WO\\n2) Start งาน\\n3) บันทึก good qty, reject qty และเวลาทำงานตามจริง\\n4) ถ้ามี mold/labor fields ให้กรอกตามหน้างาน\\n5) กด Done\\n6) กลับไปตรวจ MO cost/labor/mold postings",
                "expected": "ผู้ใช้ปิดงานจาก shopfloor ได้ และข้อมูลผลผลิต/ของเสีย/เวลาไปโผล่บน WO และ MO ถูกต้อง",
                "fix": "ถ้าหน้าจอไม่รับค่า ให้ตรวจ WO state, permissions และ configuration ของ shopfloor console",
                "evidence": "Capture: shopfloor before/after, WO result",
                "note": "เชื่อม execution เข้ากับ cost review",
            },
        ],
    },
    {
        "sheet": "MU10",
        "title": "Reports / Stock Movement / BOM vs Actual / UoM",
        "objective": "ให้ key user ตรวจรายงานการผลิต, stock movement, BOM vs actual และความถูกต้องของ UoM conversion ได้",
        "coverage": "manufacturing reports, stock moves, BOM variance, UoM conversion",
        "cases": [
            {
                "case_id": "MU10-01",
                "backlogs": "MA16",
                "name": "ดูรายงานการผลิตรายเดือน / ครึ่งปี",
                "role": "Planner / Management",
                "path": "Manufacturing > Reporting > Manufacturing Orders / Pivot / Graph",
                "pre": "มีข้อมูล MO done ในระบบอย่างน้อย 1 เดือน",
                "data": "กรอง state = Done, Finished Date ตามช่วงเวลา",
                "steps": "1) เปิดเมนู Manufacturing > Reporting > Manufacturing Orders\\n2) Filter State = Done\\n3) ใส่วันที่ตามช่วงที่ต้องการ เช่น 6 เดือนย้อนหลัง\\n4) เปลี่ยนเป็น Pivot\\n5) Group by Finished Date > Month และ Product\\n6) อ่านปริมาณผลิตรายเดือน และ export หากต้องใช้ต่อ",
                "expected": "ผู้ใช้ต้องดึงรายงานว่าผลิตอะไร เท่าไร ในแต่ละเดือนได้โดยไม่ต้อง query เพิ่ม",
                "fix": "ถ้าตัวเลขไม่ตรง ให้ตรวจ filter state/date และว่ารายงานใช้ planned qty หรือ qty produced จริง",
                "evidence": "Capture: pivot by month",
                "note": "ใช้ใน training เรื่อง production report",
            },
            {
                "case_id": "MU10-02",
                "backlogs": "MA10, MA16",
                "name": "เช็ก Stock Movement ย้อนจาก MO ถึง transfer และสินค้า",
                "role": "Production / Warehouse / Accounting",
                "path": "MO > smart buttons / Inventory > Reporting > Product Moves",
                "pre": "มี MO done อย่างน้อย 1 ใบ",
                "data": "MO ตัวอย่าง 1 ใบที่มี transfer และ finished goods",
                "steps": "1) เปิด MO ที่ done แล้ว\\n2) เปิด smart button Moves / Transfers / Traceability ถ้ามี\\n3) ตรวจ raw move, finished move และ internal transfer ที่เกี่ยวข้อง\\n4) เปิด Product Moves ของ component และ FG เพื่อ trace ย้อน\\n5) สรุปให้ได้ว่าของเข้า/ออก location ไหนบ้าง",
                "expected": "ทีมต้อง trace การเคลื่อนไหว stock จาก MO ไปถึง transfer/finished ได้ครบ",
                "fix": "ถ้า trace ไม่ครบ ให้ตรวจว่า move ไหนยัง draft/cancel หรือใช้คนละ warehouse/location",
                "evidence": "Capture: product moves + traceability",
                "note": "หัวข้อที่ตกหล่นในคู่มือ inventory เดิม",
            },
            {
                "case_id": "MU10-03",
                "backlogs": "MA10, MA16",
                "name": "เทียบ BOM กับ actual consumption บน MO",
                "role": "Production / Accounting Review",
                "path": "MO > Components / Cost / Reporting",
                "pre": "มี MO done และมีการใช้วัตถุดิบจริง",
                "data": "MO ที่มี actual ต่างจาก BOM ตาม test",
                "steps": "1) เปิด MO ที่ done\\n2) เปิดแท็บ Components และอ่าน planned qty เทียบ done qty\\n3) ตรวจ stock moves จริงของ component\\n4) บันทึกว่ามี over/under consumption ตัวไหน\\n5) ถ้าระบบมี report cost/variance ให้เปิดประกอบ",
                "expected": "ผู้ใช้ต้องตอบได้ว่าของจริงต่างจาก BOM ตรงไหน และ trace ผ่าน move line ได้",
                "fix": "ถ้าอ่านไม่ออก ให้ export component moves และเทียบ UoM/rounding",
                "evidence": "Capture: components tab + stock moves",
                "note": "เชื่อมกับ backlog ตรวจ BOM vs actual",
            },
            {
                "case_id": "MU10-04",
                "backlogs": "MA01",
                "name": "ซื้อกับใช้คนละหน่วย แล้วระบบแปลง UoM ถูกต้อง",
                "role": "Purchase / Warehouse / Production",
                "path": "Product > UoM / PO / Receipt / MO Components",
                "pre": "มีสินค้าที่ตั้ง Purchase UoM ต่างจาก Stock UoM และอยู่ใน category เดียวกัน",
                "data": "ตัวอย่างเช่น ซื้อ 1 ถุง = 1000 Kg หรือ setup เทียบเท่าใน UAT",
                "steps": "1) เปิดสินค้าและตรวจ UoM กับ Purchase UoM\\n2) สร้าง PO ตาม Purchase UoM\\n3) รับของเข้า stock\\n4) เปิด On Hand ว่าระบบแปลงเป็น stock UoM ถูกหรือไม่\\n5) เปิด MO หรือ component move แล้วตรวจว่าระบบตัดใช้ด้วย stock UoM ได้หรือไม่",
                "expected": "ระบบต้องรับของและตัดใช้ได้แม้หน่วยซื้อกับหน่วยใช้ต่างกัน ถ้า UoM อยู่ใน category เดียวกัน",
                "fix": "ถ้า convert ไม่ได้ ให้ตรวจว่า UoM อยู่คนละ category หรือ conversion factor ผิด",
                "evidence": "Capture: product UoM setup, receipt qty, stock qty, component consumption",
                "note": "ใช้ตอบคำถามหน่วยซื้อไม่เท่าหน่วยใช้",
            },
        ],
    },
]


def last_nonempty_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1


def apply_header(cell, fill="1F4E78"):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_body(cell, wrap=False, fill=None):
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    thin = Side(style="thin", color="E6E6E6")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def build_topic_sheet(wb, topic, status_validation):
    if topic["sheet"] in wb.sheetnames:
        del wb[topic["sheet"]]
    ws = wb.create_sheet(topic["sheet"])
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"Manufacturing UAT Test Step - {topic['sheet']}"
    ws["A1"].font = Font(size=16, bold=True, color="1F1F1F")
    ws["A2"] = "หัวข้อ"
    ws["B2"] = topic["title"]
    ws["A3"] = "วัตถุประสงค์"
    ws["B3"] = topic["objective"]
    ws["A4"] = "ขอบเขต"
    ws["B4"] = topic["coverage"]
    for cell in ("A2", "A3", "A4"):
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ("B2", "B3", "B4"):
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")

    headers = [
        "No.",
        "Case ID",
        "Backlog IDs",
        "Scenario Name",
        "Role / Department",
        "Menu Path",
        "Pre-Condition / Setup",
        "Test Data",
        "Detailed Test Steps",
        "Expected Result",
        "Fix / Recovery Path",
        "Evidence to Capture",
        "Actual Result / Notes",
        "Executor",
        "Test Date",
        "Review - คุณไอซ์",
        "Review - คุณติ๊ก",
        "Overall Status",
        "Trainer Note",
    ]
    header_row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        apply_header(cell)

    widths = {
        1: 6, 2: 12, 3: 18, 4: 34, 5: 22, 6: 28, 7: 28, 8: 24, 9: 58,
        10: 38, 11: 34, 12: 28, 13: 30, 14: 18, 15: 14, 16: 16, 17: 16,
        18: 16, 19: 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    start_row = 7
    map_rows = []
    for idx, case in enumerate(topic["cases"], start=1):
        r = start_row + idx - 1
        values = [
            idx, case["case_id"], case["backlogs"], case["name"], case["role"], case["path"],
            case["pre"], case["data"], case["steps"], case["expected"], case["fix"],
            case["evidence"], None, None, None, None, None, None, case["note"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(r, col, value)
            apply_body(cell, wrap=col >= 3)

        ws.cell(r, 18).value = (
            f'=IF(AND(P{r}="Passed",Q{r}="Passed"),"Passed",'
            f'IF(OR(P{r}="Failed",Q{r}="Failed"),"Failed",'
            f'IF(COUNTA(P{r}:Q{r})=0,"Not Start",'
            f'IF(OR(P{r}="Pending",Q{r}="Pending",P{r}="Cancelled",Q{r}="Cancelled"),"Pending","Under Testing"))))'
        )
        apply_body(ws.cell(r, 18), wrap=True, fill="FFF2CC")
        status_validation.add(ws.cell(r, 16))
        status_validation.add(ws.cell(r, 17))
        ws.row_dimensions[r].height = 92

        for backlog in [x.strip() for x in case["backlogs"].split(",") if x.strip()]:
            map_rows.append((backlog, case["case_id"], topic["sheet"], r))

    ws.freeze_panes = "A7"
    return map_rows


def main():
    wb = load_workbook(SOURCE)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for name in [x["sheet"] for x in TOPICS] + ["_MANU_UAT_MAP", "_MANU_UAT_LISTS", "MANU_UAT_README"]:
        if name in wb.sheetnames:
            del wb[name]

    lists_ws = wb.create_sheet("_MANU_UAT_LISTS")
    lists_ws.sheet_state = "hidden"
    for idx, status in enumerate(STATUSES, start=1):
        lists_ws.cell(idx, 1, status)

    status_validation = DataValidation(
        type="list",
        formula1="'_MANU_UAT_LISTS'!$A$1:$A$6",
        allow_blank=True,
    )

    all_map_rows = []
    for topic in TOPICS:
        rows = build_topic_sheet(wb, topic, status_validation)
        all_map_rows.extend(rows)
    for topic in TOPICS:
        wb[topic["sheet"]].add_data_validation(status_validation)

    sc_ws = wb["Scenarios"]
    sc_ws["I8"] = "Review - คุณไอซ์"
    sc_ws["J8"] = "Review - คุณติ๊ก"
    sc_ws["K8"] = "Scenario Sheet"
    sc_ws["L8"] = "Menu Path"
    sc_ws["M8"] = "Fix / Recovery Path"
    for cell_ref in ("I8", "J8", "K8", "L8", "M8"):
        apply_header(sc_ws[cell_ref], fill="4F81BD")

    sc_ws["F1"] = '=COUNTIF(H9:H5000, "Passed")'
    sc_ws["F2"] = '=COUNTIF(H9:H5000, "Failed")'
    sc_ws["F3"] = '=COUNTIF(H9:H5000, E3)'
    sc_ws["F4"] = '=COUNTIF(H9:H5000, E4)'
    sc_ws["F5"] = '=COUNTIF(H9:H5000, E5)'
    sc_ws["F6"] = '=COUNTIF(H9:H5000, "Not Start")'
    sc_ws["F7"] = '=SUM(F1:F6)'
    sc_ws["G1"] = '=F1/F7'
    sc_ws["G2"] = '=F2/F7'
    sc_ws["G3"] = '=F3/F7'
    sc_ws["G4"] = '=F4/F7'
    sc_ws["G5"] = '=F5/F7'
    sc_ws["G6"] = '=F6/F7'
    sc_ws["G7"] = '=SUM(G1:G6)'

    start_row = last_nonempty_row(sc_ws) + 1
    seq_no = 1
    for topic in TOPICS:
        ws = wb[topic["sheet"]]
        for local_idx, _case in enumerate(topic["cases"], start=1):
            r = 7 + local_idx - 1
            row = start_row
            sc_ws.cell(row, 1, seq_no)
            sc_ws.cell(row, 2, f"={topic['sheet']}!B{r}")
            sc_ws.cell(row, 3, f"={topic['sheet']}!C{r}")
            sc_ws.cell(row, 4, f"={topic['sheet']}!D{r}")
            sc_ws.cell(row, 5, f"={topic['sheet']}!E{r}")
            sc_ws.cell(row, 6, f"={topic['sheet']}!O{r}")
            sc_ws.cell(row, 7, f"={topic['sheet']}!N{r}")
            sc_ws.cell(row, 8, f"={topic['sheet']}!R{r}")
            sc_ws.cell(row, 9, f"={topic['sheet']}!P{r}")
            sc_ws.cell(row, 10, f"={topic['sheet']}!Q{r}")
            sc_ws.cell(row, 11, topic["sheet"])
            sc_ws.cell(row, 12, f"={topic['sheet']}!F{r}")
            sc_ws.cell(row, 13, f"={topic['sheet']}!K{r}")
            for col in range(1, 14):
                apply_body(sc_ws.cell(row, col), wrap=col >= 3)
            sc_ws.row_dimensions[row].height = 48
            start_row += 1
            seq_no += 1

    sc_ws.column_dimensions["D"].width = 42
    sc_ws.column_dimensions["E"].width = 22
    sc_ws.column_dimensions["I"].width = 16
    sc_ws.column_dimensions["J"].width = 16
    sc_ws.column_dimensions["K"].width = 14
    sc_ws.column_dimensions["L"].width = 28
    sc_ws.column_dimensions["M"].width = 28

    map_ws = wb.create_sheet("_MANU_UAT_MAP")
    map_ws.sheet_state = "hidden"
    map_headers = ["Backlog ID", "Case ID", "Scenario Sheet", "Case Row", "Review Ice", "Review Tik", "Overall"]
    for c, header in enumerate(map_headers, 1):
        map_ws.cell(1, c, header)
    for idx, (backlog, case_id, sheet_name, row_num) in enumerate(all_map_rows, start=2):
        map_ws.cell(idx, 1, backlog)
        map_ws.cell(idx, 2, case_id)
        map_ws.cell(idx, 3, sheet_name)
        map_ws.cell(idx, 4, row_num)
        map_ws.cell(idx, 5, f"='{sheet_name}'!P{row_num}")
        map_ws.cell(idx, 6, f"='{sheet_name}'!Q{row_num}")
        map_ws.cell(idx, 7, f"='{sheet_name}'!R{row_num}")

    pb_ws = wb["Product_backlog"]
    pb_ws.insert_rows(1)
    headers = [
        "Backlog ID", "Module", "Sub Module", "Phase", "As-Is", "To-Be", "Solution / Workaround",
        "Requirement Type", "Requirement Date", "Ref 1", "Ref 2", "Owner", "Project Stage",
        "Issue Status", "Internal Review", "Legacy UAT Status", "Legacy User Review",
        "Review - คุณไอซ์", "Review - คุณติ๊ก", "Detailed UAT Overall",
    ]
    for idx, header in enumerate(headers, start=1):
        apply_header(pb_ws.cell(1, idx, header), fill="5B9BD5")
    pb_ws.freeze_panes = "A2"

    last_pb = last_nonempty_row(pb_ws)
    for r in range(2, last_pb + 1):
        if not pb_ws.cell(r, 1).value:
            continue
        pb_ws.cell(r, 18, f'=IF(COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r})=0,"N/A",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$E:$E,"Failed")>0,"Failed",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$E:$E,"Passed")=COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r}),"Passed","Pending")))')
        pb_ws.cell(r, 19, f'=IF(COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r})=0,"N/A",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$F:$F,"Failed")>0,"Failed",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$F:$F,"Passed")=COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r}),"Passed","Pending")))')
        pb_ws.cell(r, 20, f'=IF(COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r})=0,"N/A",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$G:$G,"Failed")>0,"Failed",IF(COUNTIFS(_MANU_UAT_MAP!$A:$A,$A{r},_MANU_UAT_MAP!$G:$G,"Passed")=COUNTIF(_MANU_UAT_MAP!$A:$A,$A{r}),"Passed","Pending")))')
        for col in (18, 19, 20):
            apply_body(pb_ws.cell(r, col), wrap=True, fill="E2F0D9")

    for col in range(1, 21):
        if col <= 4:
            pb_ws.column_dimensions[get_column_letter(col)].width = 14
        elif col <= 8:
            pb_ws.column_dimensions[get_column_letter(col)].width = 22
        elif col <= 12:
            pb_ws.column_dimensions[get_column_letter(col)].width = 14
        elif col <= 17:
            pb_ws.column_dimensions[get_column_letter(col)].width = 18
        else:
            pb_ws.column_dimensions[get_column_letter(col)].width = 18

    readme = wb.create_sheet("MANU_UAT_README", 0)
    lines = [
        ["หัวข้อ", "คำอธิบาย"],
        ["ไฟล์นี้คืออะไร", "ชุด UAT Test Step ฝ่ายผลิตแบบละเอียด สำหรับให้ผู้ใช้งานหน้างานกดตามใน Odoo ได้จริง"],
        ["วิธีใช้ Scenario Sheet", "เปิดแต่ละ sheet ตั้งแต่ MU01-MU10 แล้วทำตาม Detailed Test Steps ทีละข้อ"],
        ["Review ผู้ใช้", "คุณไอซ์กรอกคอลัมน์ Review - คุณไอซ์ และคุณติ๊กกรอกคอลัมน์ Review - คุณติ๊ก"],
        ["สูตร Passed", "ถ้าทั้งสอง review = Passed ระบบจะสรุป Overall Status = Passed อัตโนมัติ"],
        ["Scenarios sheet", "สรุปทุก test case พร้อมสถานะรวมและลิงก์กลับไปที่ sheet ต้นทาง"],
        ["Product_backlog sheet", "เพิ่มคอลัมน์ Review - คุณไอซ์ / คุณติ๊ก / Detailed UAT Overall เพื่อ aggregate สถานะจากทุก test case ที่อ้าง backlog นั้น"],
        ["Backorder recovery", "มีเคสทดสอบทั้งเลือก Create Backorder และ No Backorder แล้วกู้กลับด้วยปุ่ม Create Backorder"],
        ["ข้อควรทำก่อนเริ่ม", "เช็กสิทธิ์ user, สินค้าตัวอย่าง, route/BOM, stock ต้นทาง, และใช้ฐาน UAT เท่านั้น"],
    ]
    for row_idx, row_values in enumerate(lines, start=1):
        for col_idx, value in enumerate(row_values, start=1):
            readme.cell(row_idx, col_idx, value)
            if row_idx == 1:
                apply_header(readme.cell(row_idx, col_idx))
            else:
                apply_body(readme.cell(row_idx, col_idx), wrap=True)
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 110

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
