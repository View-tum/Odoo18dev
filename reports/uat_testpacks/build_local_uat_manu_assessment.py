from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(r"C:/365_project/TheCool18e/Dev")
WORKBOOK = Path(r"C:/Users/tumsu/Downloads/UAT_GoldMints_Test Scenario_MANU.xlsx")
OUTPUT_WORKBOOK = Path(
    r"C:/Users/tumsu/Downloads/UAT_GoldMints_Test Scenario_MANU_local_uat_assessment_20260407.xlsx"
)
OUTPUT_MD = ROOT / "reports" / "local_uat_manu_assessment_20260407.md"


@dataclass(frozen=True)
class CaseResult:
    case_id: str
    status: str
    note: str
    evidence: str
    tester: str = "Codex (local uat)"
    test_date: str = "2026-04-07"


CASE_RESULTS: list[CaseResult] = [
    CaseResult(
        "MU01-01",
        "Passed",
        "Verified from local UAT product form + forecast report for FG-PNC-TH-01001. User can read On Hand, Forecasted, and linked moves.",
        "reports/manu_uat_20260406_mu01_mu04_images/MU01_01_product_search.png; MU01_01_product_form.png; MU01_01_forecast_report.png",
    ),
    CaseResult(
        "MU01-02",
        "Passed with note",
        "Verified from local UAT replenishment and product route screenshots for FG-PSS-TH-01005. Review used existing live records and did not trigger new replenishment documents.",
        "reports/manu_uat_20260406_mu01_mu04_images/MU01_02_replenishment_fg_pss.png; MU01_02_product_route_fg_pss.png",
    ),
    CaseResult(
        "MU01-03",
        "Passed with note",
        "Verified from existing local UAT MO trace for FG-PNC-TH-01001. Review confirms replenishment path reaches Manufacturing Pharma. This was traced from existing document, not by generating a brand new procurement run.",
        "reports/manu_uat_20260406_mu01_mu04_images/MU01_03_mo_created_fg_pnc.png",
    ),
    CaseResult(
        "MU04-03",
        "Passed",
        "Late backorder recovery verified on local UAT for Transfer Pharma. After user selected No Backorder, Create Backorder became available and recreated the missing picking.",
        "reports/late_backorder_recovery_uat_test_20260403_final.json",
    ),
    CaseResult(
        "MU07-03",
        "Passed",
        "Late backorder recovery verified on local UAT for Manufacturing Pharma. MO remained done, and Create Backorder recreated the remaining production quantity.",
        "reports/late_backorder_recovery_uat_test_20260403_final.json",
    ),
    CaseResult(
        "MU08-01",
        "Passed with note",
        "Scrap flow is verified on local UAT through shell-driven scenarios: scrap wizard product restriction, same-location replenish, internal-transfer replenish, and scrap landed cost finalization. Coverage is technical/business-flow level rather than a pure manual UI walkthrough.",
        "reports/shopfloor_auto_uat_suite_20260406.json; reports/mrp_scrap_landed_cost_uat_20260406.json",
    ),
    CaseResult(
        "MU09-01",
        "Passed",
        "Auto assignment of mold and workcenter verified in local UAT. Mold matrix assigned compatible mold and computed expected duration.",
        "reports/shopfloor_auto_uat_suite_20260406.json",
    ),
    CaseResult(
        "MU09-02",
        "Passed",
        "Parallel mold guard verified in local UAT. Duplicate mold usage across sibling workorders was prevented.",
        "reports/shopfloor_auto_uat_suite_20260406.json",
    ),
    CaseResult(
        "MU09-03",
        "Passed with note",
        "Local UAT verifies qty logs aggregation and labor cost calculation. Reject-qty capture was not separately signed off in this evidence set, so this case is treated as pass with note.",
        "reports/shopfloor_auto_uat_suite_20260406.json",
    ),
    CaseResult(
        "MU11-04",
        "Passed",
        "Plastic GMP Shop Floor flow with mold and workcenter verified directly in UI on local UAT.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_full_20260406.json",
    ),
    CaseResult(
        "MU14-01",
        "Passed",
        "Card shows mold name and shot counter.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/02_mold_card.png",
    ),
    CaseResult(
        "MU14-02",
        "Passed",
        "Change mold from card before start.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/06_change_mold_dialog.png; reports/mold_shopfloor_uat_20260406_images/07_change_mold_result.png",
    ),
    CaseResult(
        "MU14-03",
        "Passed",
        "Full mold warning with Continue Anyway.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/03_full_mold_warning.png; reports/mold_shopfloor_uat_20260406_images/04_continue_anyway_started.png",
    ),
    CaseResult(
        "MU14-04",
        "Passed",
        "Full mold warning with Change Mold then Start.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/08_change_from_warning_result.png; reports/mold_shopfloor_uat_20260406_images/09_changed_mold_started.png",
    ),
    CaseResult(
        "MU14-05",
        "Passed",
        "Reset mold life after maintenance.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/10_reset_life_result.png",
    ),
    CaseResult(
        "MU14-06",
        "Passed with note",
        "Breakdown recovery path is usable from Shop Floor via More > Report Issue and stop/recovery behavior. There is still no dedicated breakdown button on the card.",
        "reports/mold_shopfloor_uat_test_20260406.json; reports/mold_shopfloor_uat_20260406_images/11_more_menu_report_issue.png",
    ),
    CaseResult(
        "MU14-07",
        "Not closed",
        "No suitable local UAT sample was available to sign off done/cancel sibling queue behavior end-to-end in UI.",
        "reports/mold_shopfloor_uat_test_20260406.json",
    ),
]


STATUS_FILLS = {
    "Passed": PatternFill("solid", fgColor="C6EFCE"),
    "Passed with note": PatternFill("solid", fgColor="FFF2CC"),
    "Not closed": PatternFill("solid", fgColor="F4CCCC"),
}


def _sheet_case_row_map(workbook) -> dict[str, tuple[str, int]]:
    mapping: dict[str, tuple[str, int]] = {}
    for ws in workbook.worksheets:
        if not ws.title.startswith("MU"):
            continue
        for row in range(7, ws.max_row + 1):
            case_id = ws.cell(row, 2).value
            if isinstance(case_id, str) and case_id.startswith("MU"):
                mapping[case_id] = (ws.title, row)
    return mapping


def _scenario_case_row_map(workbook) -> dict[str, int]:
    ws = workbook["Scenarios"]
    mapping: dict[str, int] = {}
    for row in range(9, ws.max_row + 1):
        case_id = ws.cell(row, 2).value
        if isinstance(case_id, str) and case_id.startswith("MU"):
            mapping[case_id] = row
    return mapping


def _load_coverage_frames() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    backlog = pd.read_excel(WORKBOOK, sheet_name="Product_backlog")
    scenarios = pd.read_excel(WORKBOOK, sheet_name="Scenarios", skiprows=7)
    uat_map = pd.read_excel(WORKBOOK, sheet_name="_MANU_UAT_MAP")
    for frame in (backlog, scenarios, uat_map):
        for col in frame.columns:
            if frame[col].dtype == object:
                frame[col] = frame[col].astype(str).str.strip()
    return backlog, scenarios, uat_map


def _apply_case_results(workbook, case_results: Iterable[CaseResult]) -> None:
    case_row_map = _sheet_case_row_map(workbook)
    scenario_row_map = _scenario_case_row_map(workbook)
    scenarios_ws = workbook["Scenarios"]

    for result in case_results:
        if result.case_id in case_row_map:
            sheet_name, row = case_row_map[result.case_id]
            ws = workbook[sheet_name]
            ws.cell(row, 13).value = result.note
            ws.cell(row, 14).value = result.tester
            ws.cell(row, 15).value = result.test_date
            ws.cell(row, 19).value = result.evidence
            for col in (13, 14, 15):
                ws.cell(row, col).fill = STATUS_FILLS.get(result.status, PatternFill())
            ws.cell(row, 13).font = Font(color="000000")

        if result.case_id in scenario_row_map:
            row = scenario_row_map[result.case_id]
            scenarios_ws.cell(row, 6).value = result.test_date
            scenarios_ws.cell(row, 7).value = result.tester
            # Keep the workbook's primary status vocabulary stable.
            scenarios_status = {
                "Passed": "Passed",
                "Passed with note": "Under Testing",
                "Not closed": "Not Start",
            }[result.status]
            scenarios_ws.cell(row, 8).value = scenarios_status


def _write_summary_sheets(workbook, backlog: pd.DataFrame, scenarios: pd.DataFrame, uat_map: pd.DataFrame) -> None:
    for name in ["LOCAL_UAT_SUMMARY", "LOCAL_UAT_COVERAGE"]:
        if name in workbook.sheetnames:
            del workbook[name]

    summary = workbook.create_sheet("LOCAL_UAT_SUMMARY")
    summary.append(["Local UAT MANU Assessment", "2026-04-07"])
    summary.append([])

    total_backlog = sorted(set(backlog["Backlog ID"].dropna().astype(str).str.strip()))
    mapped_backlog = sorted(set(uat_map["Backlog ID"].dropna().astype(str).str.strip()))
    missing_backlog = [x for x in total_backlog if x not in mapped_backlog]
    total_cases = scenarios["Scenario ID"].dropna().astype(str).str.startswith("MU").sum()
    tested_case_ids = [c.case_id for c in CASE_RESULTS]

    summary.append(["Metric", "Value"])
    summary.append(["Total Product_backlog items", len(total_backlog)])
    summary.append(["Backlog items linked to MU test cases", len(mapped_backlog)])
    summary.append(["Backlog items not linked to any MU test case", len(missing_backlog)])
    summary.append(["Total MU test cases in workbook", int(total_cases)])
    summary.append(["MU test cases with local UAT evidence", len(tested_case_ids)])
    summary.append(["Coverage of Product_backlog by MU design", round(len(mapped_backlog) / len(total_backlog) * 100, 2)])
    summary.append(["Coverage of MU cases by local evidence", round(len(tested_case_ids) / int(total_cases) * 100, 2)])
    summary.append([])

    summary.append(["Case ID", "Assessment Status", "Evidence / Note"])
    for item in CASE_RESULTS:
        summary.append([item.case_id, item.status, f"{item.note} | {item.evidence}"])

    coverage = workbook.create_sheet("LOCAL_UAT_COVERAGE")
    coverage.append(["Backlog ID", "Mapped to MU?", "Mapped MU Count", "Observed local UAT evidence?", "Backlog Description"])

    tested_case_set = {c.case_id for c in CASE_RESULTS}
    case_counts = uat_map.groupby("Backlog ID")["Case ID"].nunique().to_dict()
    tested_backlog = set(
        uat_map.loc[uat_map["Case ID"].isin(tested_case_set), "Backlog ID"].dropna().astype(str).str.strip()
    )
    backlog_desc = backlog.set_index("Backlog ID")["To-Be"].to_dict()
    for backlog_id in total_backlog:
        coverage.append(
            [
                backlog_id,
                "Yes" if backlog_id in mapped_backlog else "No",
                case_counts.get(backlog_id, 0),
                "Yes" if backlog_id in tested_backlog else "No",
                backlog_desc.get(backlog_id, ""),
            ]
        )

    for ws in (summary, coverage):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A3"


def _write_markdown(backlog: pd.DataFrame, scenarios: pd.DataFrame, uat_map: pd.DataFrame) -> None:
    total_backlog = sorted(set(backlog["Backlog ID"].dropna().astype(str).str.strip()))
    mapped_backlog = sorted(set(uat_map["Backlog ID"].dropna().astype(str).str.strip()))
    missing_backlog = [x for x in total_backlog if x not in mapped_backlog]
    total_cases = int(scenarios["Scenario ID"].dropna().astype(str).str.startswith("MU").sum())
    tested_case_ids = [c.case_id for c in CASE_RESULTS]
    tested_case_set = set(tested_case_ids)
    tested_backlog = sorted(
        set(uat_map.loc[uat_map["Case ID"].isin(tested_case_set), "Backlog ID"].dropna().astype(str).str.strip())
    )

    lines: list[str] = []
    lines.append("# Local UAT MANU Assessment")
    lines.append("")
    lines.append("Workbook: `UAT_GoldMints_Test Scenario_MANU.xlsx`")
    lines.append("Database: `uat` on local Odoo (`localhost:8811`) ")
    lines.append("")
    lines.append("## Coverage Summary")
    lines.append("")
    lines.append(f"- Total `Product_backlog` items: `{len(total_backlog)}`")
    lines.append(f"- Backlog items linked to MU test cases: `{len(mapped_backlog)}`")
    lines.append(f"- Backlog items not linked to any MU test case: `{len(missing_backlog)}`")
    lines.append(f"- Total MU test cases in workbook: `{total_cases}`")
    lines.append(f"- MU test cases with local UAT evidence: `{len(tested_case_ids)}`")
    lines.append(f"- Product_backlog design coverage: `{len(mapped_backlog) / len(total_backlog) * 100:.2f}%`")
    lines.append(f"- Case-level local evidence coverage: `{len(tested_case_ids) / total_cases * 100:.2f}%`")
    lines.append("")
    lines.append("## Key Finding")
    lines.append("")
    lines.append("The workbook does **not** fully cover `Product_backlog`. Only 15 of 60 backlog IDs are mapped into MU cases. The remaining 45 backlog IDs currently have no detailed MU test step linked to them.")
    lines.append("")
    lines.append("## MU Cases With Local UAT Evidence")
    lines.append("")
    for item in CASE_RESULTS:
        lines.append(f"- `{item.case_id}`: `{item.status}`")
        lines.append(f"  Note: {item.note}")
        lines.append(f"  Evidence: {item.evidence}")
    lines.append("")
    lines.append("## Backlog IDs Not Covered By Any MU Case")
    lines.append("")
    lines.append(", ".join(f"`{x}`" for x in missing_backlog))
    lines.append("")
    lines.append("## Backlog IDs Already Touched By Local UAT Evidence")
    lines.append("")
    lines.append(", ".join(f"`{x}`" for x in tested_backlog))
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    backlog, scenarios, uat_map = _load_coverage_frames()
    workbook = load_workbook(WORKBOOK)
    _apply_case_results(workbook, CASE_RESULTS)
    _write_summary_sheets(workbook, backlog, scenarios, uat_map)
    workbook.save(OUTPUT_WORKBOOK)
    _write_markdown(backlog, scenarios, uat_map)
    print(OUTPUT_WORKBOOK)
    print(OUTPUT_MD)


if __name__ == "__main__":
    main()
