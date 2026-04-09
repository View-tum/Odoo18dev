from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


INPUT_XLSX = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_actual_flow_th_evidence_rerun_20260407.xlsx"
)
OUTPUT_XLSX = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_actual_flow_th_evidence_rerun_menu_en_20260408.xlsx"
)
SUMMARY_MD = Path(
    r"C:\365_project\TheCool18e\Dev\reports\manu_actual_flow_menu_en_update_20260408.md"
)


PHASE_PREFIXES = ("01_", "02_", "03_", "04_", "05_", "06_", "07_")


MENU_MAP = {
    "สินค้าคงคลัง > สินค้า > สินค้า > เปิดสินค้า FG > Smart Button: คงเหลือ / คาดการณ์":
        "Inventory > Products > Products > Open FG Product > Smart Button: On Hand / Forecasted",
    "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า":
        "Inventory > Operations > Procurement > Replenishment",
    "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า > เปิดเอกสารที่ระบบสร้าง":
        "Inventory > Operations > Procurement > Replenishment > Open Document Created by System",
    "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า > เปิดสินค้า > Smart Button: คาดการณ์":
        "Inventory > Operations > Procurement > Replenishment > Open Product > Smart Button: Forecasted",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO แม่ > เอกสารลูก / การโอนที่เกี่ยวข้อง":
        "Manufacturing > Operations > Manufacturing Orders > Open Parent MO > Child Documents / Related Transfers",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต":
        "Manufacturing > Operations > Manufacturing Orders",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO > Smart Button: การโอน":
        "Manufacturing > Operations > Manufacturing Orders > Open MO > Smart Button: Transfers",
    "การขาย > คำสั่ง > ใบเสนอราคา > ยืนยัน > trace ไปใบสั่งผลิต":
        "Sales > Orders > Quotations > Confirm > Trace to Manufacturing Order",
    "การขาย > คำสั่ง > ใบเสนอราคา > ยืนยัน > trace shortage ไป MO / RFQ / การเติมสินค้า":
        "Sales > Orders > Quotations > Confirm > Trace Shortage to MO / RFQ / Replenishment",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน > กรองประเภทการดำเนินการ = Transfer Plastic":
        "Inventory > Operations > Transfers > Filter Operation Type = Transfer Plastic",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิด Transfer Plastic > Validate แบบ partial":
        "Inventory > Operations > Transfers > Open Transfer Plastic > Validate Partial",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิดเอกสารที่ Done > ปุ่ม Create Backorder":
        "Inventory > Operations > Transfers > Open Done Transfer > Create Backorder",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน > กรองประเภทการดำเนินการ = Transfer Pharma":
        "Inventory > Operations > Transfers > Filter Operation Type = Transfer Pharma",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิด Transfer Pharma > Validate แบบ partial":
        "Inventory > Operations > Transfers > Open Transfer Pharma > Validate Partial",
    "GMP Shop Floor":
        "GMP Shop Floor",
    "GMP Shop Floor > เปิดคำสั่งงาน":
        "GMP Shop Floor > Open Work Order",
    "GMP Shop Floor > เปิดคำสั่งงาน > Done":
        "GMP Shop Floor > Open Work Order > Done",
    "GMP Shop Floor > เปิดงานฝั่ง Plastic":
        "GMP Shop Floor > Open Plastic Job",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > คำสั่งงาน / การผลิต > การกำหนดค่า > ศูนย์งาน":
        "Manufacturing > Operations > Manufacturing Orders > Work Orders / Manufacturing > Configuration > Work Centers",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > คำสั่งงาน":
        "Manufacturing > Operations > Manufacturing Orders > Work Orders",
    "การผลิต > การปฏิบัติการ > คำสั่งงาน":
        "Manufacturing > Operations > Work Orders",
    "GMP Shop Floor > การ์ดคำสั่งงาน > ปุ่ม Mold":
        "GMP Shop Floor > Work Order Card > Mold Button",
    "GMP Shop Floor > การ์ดคำสั่งงาน > Start":
        "GMP Shop Floor > Work Order Card > Start",
    "GMP Shop Floor > Start > Popup เตือนอายุแม่พิมพ์ > Change Mold":
        "GMP Shop Floor > Start > Mold Life Warning Popup > Change Mold",
    "การผลิต > การกำหนดค่า > ศูนย์งาน > เปิดแม่พิมพ์":
        "Manufacturing > Configuration > Work Centers > Open Mold",
    "GMP Shop Floor > การ์ดคำสั่งงาน > More > Report Issue":
        "GMP Shop Floor > Work Order Card > More > Report Issue",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต หรือ การผลิต > การปฏิบัติการ > คำสั่งงาน":
        "Manufacturing > Operations > Manufacturing Orders or Manufacturing > Operations > Work Orders",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > ทำ partial แล้ว Validate/Done":
        "Manufacturing > Operations > Manufacturing Orders > Partial Produce and Validate/Done",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > ปุ่ม Create Backorder":
        "Manufacturing > Operations > Manufacturing Orders > Open Done MO > Create Backorder",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > แท็บ Components / Cost":
        "Manufacturing > Operations > Manufacturing Orders > Components / Cost Tab",
    "สินค้าคงคลัง > การปฏิบัติการ > Scrap หรือ การผลิต > การปฏิบัติการ > คำสั่งงาน > Scrap":
        "Inventory > Operations > Scrap or Manufacturing > Operations > Work Orders > Scrap",
    "สินค้าคงคลัง > การปฏิบัติการ > การโอน":
        "Inventory > Operations > Transfers",
    "การผลิต > การปฏิบัติการ > คำสั่งรื้อ":
        "Manufacturing > Operations > Unbuild Orders",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > Smart Button / สินค้าคงคลัง > การรายงาน > ประวัติการย้าย":
        "Manufacturing > Operations > Manufacturing Orders > Open Done MO > Smart Button / Inventory > Reporting > Moves History",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > ปุ่ม Create Backorder หรือเปิด MO ใหม่":
        "Manufacturing > Operations > Manufacturing Orders > Open Done MO > Create Backorder or Open New MO",
    "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > Smart Button: การประเมินมูลค่า / รายการบัญชี":
        "Manufacturing > Operations > Manufacturing Orders > Open Done MO > Smart Button: Valuation / Journal Entries",
    "สินค้าคงคลัง > การกำหนดค่า > การบริหารสินค้าคงคลัง > ตำแหน่ง":
        "Inventory > Configuration > Warehouse Management > Locations",
    "การผลิต > สินค้า > ล็อต/หมายเลขซีเรียล":
        "Manufacturing > Products > Lots/Serial Numbers",
    "เอกสาร > เอกสาร":
        "Documents > Documents",
    "คุณภาพ > การควบคุมคุณภาพ > การแจ้งเตือนการจัดการคุณภาพ":
        "Quality > Quality Control > Quality Alerts",
    "คุณภาพ > การควบคุมคุณภาพ > จุดควบคุม":
        "Quality > Quality Control > Control Points",
    "การผลิต > การรายงาน > วิเคราะห์การผลิต":
        "Manufacturing > Reporting > Production Analysis",
    "สินค้าคงคลัง > การรายงาน > ประวัติการย้าย หรือ เปิดจาก MO ผ่าน Smart Button":
        "Inventory > Reporting > Moves History or Open from MO via Smart Button",
    "สั่งซื้อ > สินค้า > สินค้า / สั่งซื้อ > คำสั่ง > ใบแจ้งขอใบเสนอราคา หรือ คำสั่งซื้อ / การผลิต > ใบสั่งผลิต":
        "Purchase > Products > Products / Purchase > Orders > Requests for Quotation or Purchase Orders / Manufacturing > Operations > Manufacturing Orders",
    "คำขอซื้อ > คำขอซื้อ > คำขอซื้อ หรือ สั่งซื้อ > คำสั่ง > คำสั่งซื้อ > แท็บรายการ":
        "Purchase Requests > Purchase Requests > Purchase Requests or Purchase > Orders > Purchase Orders > Order Lines Tab",
    "สั่งซื้อ > สินค้า > สินค้า > เปิดสินค้า > แท็บการสั่งซื้อ":
        "Purchase > Products > Products > Open Product > Purchase Tab",
    "การผลิต > สินค้า > บิลวัสดุ":
        "Manufacturing > Products > Bills of Materials",
    "การผลิต > การวางแผน > MPS > กำหนดการการผลิตหลัก":
        "Manufacturing > Planning > MPS > Master Production Schedule",
    "การผลิต > การกำหนดค่า > ศูนย์งาน":
        "Manufacturing > Configuration > Work Centers",
    "การผลิต > การปฏิบัติการ > Scrap":
        "Manufacturing > Operations > Scrap",
    "การผลิต > การรายงาน > Machine Report":
        "Manufacturing > Reporting > Machine Report",
}


def main() -> None:
    wb = load_workbook(INPUT_XLSX)
    updated = 0
    unknown = []

    for ws in wb.worksheets:
        if not ws.title.startswith(PHASE_PREFIXES):
            continue
        ws.cell(6, 7).value = "Menu Path in local UAT (English)"
        for row in range(7, ws.max_row + 1):
            value = ws.cell(row, 7).value
            if not value:
                continue
            mapped = MENU_MAP.get(value)
            if mapped:
                ws.cell(row, 7).value = mapped
                updated += 1
            else:
                unknown.append((ws.title, row, value))

    wb.save(OUTPUT_XLSX)

    lines = [
        "# Menu Path Update to English",
        "",
        f"- Input: `{INPUT_XLSX}`",
        f"- Output: `{OUTPUT_XLSX}`",
        f"- Updated menu path cells: `{updated}`",
        f"- Unmapped values: `{len(unknown)}`",
        "",
    ]
    if unknown:
        lines.append("## Unmapped Values")
        lines.append("")
        for sheet, row, value in unknown:
            lines.append(f"- [{sheet}] row {row}: `{value}`")
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")

    print(f"updated={updated}")
    print(f"unknown={len(unknown)}")
    for sheet, row, value in unknown:
        print(f"{sheet}|{row}|{value}")


if __name__ == "__main__":
    main()
