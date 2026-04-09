from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANUV2_detailed_teststeps.xlsx")
OUTPUT = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANUV2_detailed_teststeps_v2.xlsx")

FONT_NAME = "Aptos"
STATUSES = ["Not Start", "Pending", "Under Testing", "Passed", "Failed", "Cancelled"]

NEW_TOPICS = [
    {
        "sheet": "MU11",
        "title": "GMP Shop Floor - ทุก Scenario หน้างาน",
        "objective": "ให้ operator และหัวหน้างานใช้ GMP Shop Floor ได้ครบตั้งแต่รับงาน, เริ่มงาน, บันทึก good/reject, หยุดงาน, ปิดงาน และตรวจผลย้อนหลัง",
        "coverage": "queue, start, pause, continue, good qty, reject qty, scrap link, labor time, งาน cancel/done, handover ระหว่างแผนก",
        "cases": [
            {
                "case_id": "MU11-01",
                "backlogs": "MA08, MA14",
                "name": "เปิดคิวงานใน GMP Shop Floor และเลือกงานของแผนกตัวเอง",
                "role": "Operator / Supervisor",
                "path": "Manufacturing > GMP Shop Floor",
                "pre": "มี Work Order ในสถานะ Ready หรือ In Progress อย่างน้อย 1 ใบของแผนก Plastic หรือ Pharma",
                "data": "เลือก WO ที่มีเลขอ้างอิงชัดเจน และรู้แผนกปลายทาง",
                "steps": "1) เปิดเมนู GMP Shop Floor\\n2) ดูรายการงานทั้งหมดใน queue\\n3) Filter หรือเลือกเฉพาะงานของแผนกตัวเอง\\n4) เปิดการ์ดงานหรือบรรทัดงานที่ต้องทำ\\n5) ตรวจข้อมูลหลักก่อนเริ่ม เช่น Product, Operation, Work Center, Planned Qty, Mold/Employee ถ้ามี",
                "expected": "ผู้ใช้ต้องเลือกงานของตัวเองได้ถูก และอ่านข้อมูลสำคัญจากหน้าจอ Shop Floor ได้ก่อนเริ่มงาน",
                "fix": "ถ้าไม่เห็นงาน ให้ตรวจ WO state, work center assignment, permission ของ user และว่ามีการ cancel/done ไปแล้วหรือไม่",
                "evidence": "Capture: queue list ก่อนเปิดงาน",
                "note": "ใช้เป็นจุดเริ่มของทุกการสอน shop floor",
            },
            {
                "case_id": "MU11-02",
                "backlogs": "MA08, MA14",
                "name": "เริ่มงาน, หยุดพัก, กลับมาทำต่อ บน Shop Floor",
                "role": "Operator",
                "path": "GMP Shop Floor > Work Order",
                "pre": "มี WO ที่พร้อมเริ่มงานจริง",
                "data": "WO 1 ใบสำหรับทดสอบ start/pause/resume",
                "steps": "1) เปิด WO จาก Shop Floor\\n2) กด Start\\n3) ตรวจว่าเวลาทำงานเริ่มนับหรือ state เปลี่ยนเป็น In Progress\\n4) กด Pause/Hold ถ้าหน้าจอมี\\n5) ตรวจว่าระบบหยุดเวลาหรือเก็บสถานะพักงาน\\n6) กด Resume หรือ Start ต่ออีกครั้ง\\n7) ตรวจว่าเวลารวมยังต่อเนื่องและไม่สร้างงานซ้ำ",
                "expected": "ระบบต้องเก็บเวลาทำงานได้ต่อเนื่อง และ operator สามารถพัก/กลับมาทำต่อโดยไม่ทำให้ quantity หรือเวลาเพี้ยน",
                "fix": "ถ้าปุ่มไม่ขึ้นหรือเวลานับผิด ให้ตรวจสิทธิ์ user, state ของ WO และ console/workcenter setup",
                "evidence": "Capture: before start, pause, resume",
                "note": "เหมาะกับ operator ใหม่ที่ยังไม่คุ้นกับจังหวะการกดหน้าจอ",
            },
            {
                "case_id": "MU11-03",
                "backlogs": "MA08, MA10, MA14",
                "name": "บันทึก Good Qty, Reject Qty และปิด Operation จาก Shop Floor",
                "role": "Operator / Supervisor",
                "path": "GMP Shop Floor > Work Order > Done",
                "pre": "มี WO ที่พร้อมกรอกผลผลิตและมีของเสียตัวอย่างสำหรับทดสอบ",
                "data": "Good Qty ตามจริง + Reject Qty อย่างน้อย 1 หน่วย",
                "steps": "1) เปิด WO บน Shop Floor\\n2) กรอก Good Qty ตามที่ผลิตได้จริง\\n3) กรอก Reject หรือ Scrap Qty ถ้ามีช่องให้กรอก\\n4) ตรวจค่า Remaining Qty/Completion % ถ้ามี\\n5) กด Done/Close Operation\\n6) กลับไปดู WO และ MO ว่าระบบบันทึกผลผลิต/ของเสียถูกต้อง",
                "expected": "WO ต้องปิดได้, qty ดี/เสียต้องถูกบันทึก และไปสะท้อนใน MO หรือ scrap flow ตามที่ระบบออกแบบไว้",
                "fix": "ถ้าตัวเลขไม่ตรง ให้ตรวจหน่วยนับ, rounding, lot/serial และว่าผู้ใช้กรอก good/reject สลับช่องหรือไม่",
                "evidence": "Capture: shop floor result + WO result",
                "note": "เป็นเคสหลักของการปิดงานหน้างานจริง",
            },
            {
                "case_id": "MU11-04",
                "backlogs": "MA08, MA11, MA14",
                "name": "Shop Floor ฝั่ง Plastic ที่มี Mold และ Workcenter",
                "role": "Operator Plastic / Supervisor",
                "path": "GMP Shop Floor > Plastic Work Order",
                "pre": "มี WO พลาสติกที่ map mold/workcenter แล้ว",
                "data": "ใช้ WO ของสินค้า plastic ที่มี mold mapping จริง",
                "steps": "1) เปิด WO ฝั่ง Plastic จาก Shop Floor\\n2) ตรวจ Work Center และ Mold ที่ระบบเลือกให้\\n3) Start งาน\\n4) บันทึก Good Qty ตามจริง\\n5) Done งาน\\n6) กลับไปดู WO detail เพื่อยืนยัน mold, labor และ output",
                "expected": "ระบบต้องแสดง workcenter/mold ถูกตัว และเมื่อปิดงานแล้วข้อมูล mold/labor/output ต้องตามไปที่ WO/MO",
                "fix": "ถ้า mold ไม่ขึ้น ให้ตรวจ matrix, compatible machine, is_mold และ workcenter ที่ assign",
                "evidence": "Capture: shop floor card + WO detail",
                "note": "เชื่อมกับ training mold และ cost",
            },
            {
                "case_id": "MU11-05",
                "backlogs": "MA08, MA14, MA16",
                "name": "งานที่ Done หรือ Cancel แล้วต้องไม่ให้ operator กดซ้ำ",
                "role": "Supervisor / IT Review",
                "path": "GMP Shop Floor / Work Order",
                "pre": "มี WO done และ WO cancel อย่างน้อยอย่างละ 1 ใบ",
                "data": "ใช้เอกสารจริงที่ปิดงานหรือยกเลิกแล้ว",
                "steps": "1) เปิด queue ใน Shop Floor\\n2) ค้นหา WO ที่ done แล้ว\\n3) ยืนยันว่าไม่ควรกดเริ่ม/ปิดซ้ำได้\\n4) ค้นหา WO ที่ cancel แล้ว\\n5) ยืนยันว่าไม่ควรแสดงให้ operator ทำงานต่อ หรือถ้าแสดงต้องไม่ให้บันทึก qty",
                "expected": "งาน done/cancel ต้องไม่ถูกใช้งานซ้ำ และ queue หน้างานต้องสะอาดพอให้ operator ไม่กดผิด",
                "fix": "ถ้ายังเห็นหรือกดได้ ให้ตรวจ state filter ของ queue และ custom guard ฝั่ง shop floor",
                "evidence": "Capture: queue + disabled/hidden behavior",
                "note": "ใช้ยืนยัน behavior หลังแก้ parallel/cancel logic",
            },
        ],
    },
    {
        "sheet": "MU12",
        "title": "ปิด MO ไปแล้ว แก้ไขยังไง / Unbuild / Reverse Flow",
        "objective": "ให้ทีมรู้แนวทางแก้ไขเมื่อ MO ถูกปิดไปแล้วแต่ข้อมูลหรือผลผลิตผิด โดยแยกว่าเคสไหนใช้ Unbuild, เคสไหนใช้ปรับสต็อก, และเคสไหนต้องสร้างเอกสารใหม่",
        "coverage": "MO done correction, unbuild, reverse moves, stock adjustment, re-open strategy",
        "cases": [
            {
                "case_id": "MU12-01",
                "backlogs": "MA10, MA16, MA17, MA18",
                "name": "ปิด MO แล้วจำนวน FG ผิด ต้องแก้ด้วย Unbuild",
                "role": "Production / Accounting Review",
                "path": "Manufacturing > Operations > Unbuild Orders",
                "pre": "มี MO done และ FG ยังอยู่ใน stock/ยังไม่ได้ขายหรือใช้ต่อ",
                "data": "เลือก MO ที่ปิดแล้วและ FG ยังย้อนกลับได้จริง",
                "steps": "1) เปิด MO ที่ done แล้ว\\n2) ตรวจว่าของสำเร็จรูปยังอยู่ใน stock location ที่ระบบหาเจอ\\n3) ไปเมนู Unbuild Orders\\n4) กด New แล้วเลือก Product, Quantity, MO ต้นทาง และ Location\\n5) Confirm/Validate Unbuild\\n6) กลับไปดู stock move ว่า FG ถูกดึงออก และ components ถูกคืนกลับตาม logic ระบบ",
                "expected": "ถ้า FG ยังไม่ถูกใช้ต่อ ระบบต้อง unbuild ได้ และ trace ย้อนกลับไป component ได้",
                "fix": "ถ้า unbuild ไม่ได้ ให้ตรวจว่ามี FG on hand จริงหรือไม่, lot/location ถูกหรือไม่ และเอกสารถูกใช้ downstream ไปแล้วหรือยัง",
                "evidence": "Capture: Unbuild order + before/after stock",
                "note": "ใช้เมื่อของสำเร็จรูปผิดและยังย้อนกลับได้ทาง stock",
            },
            {
                "case_id": "MU12-02",
                "backlogs": "MA10, MA16, MA17",
                "name": "ปิด MO แล้วใช้ component จริงผิด แต่ FG ถูกแล้ว",
                "role": "Production / Warehouse / Accounting",
                "path": "MO done > Product Moves / Inventory Adjustment or Scrap/Return",
                "pre": "มี MO done ที่ finished goods ถูกต้อง แต่ component issue ไม่ตรงของจริง",
                "data": "MO ตัวอย่างที่ actual consumption ต่างจาก BOM",
                "steps": "1) เปิด MO ที่ done แล้ว\\n2) ตรวจ component moves ที่ใช้จริงเทียบกับของจริงหน้างาน\\n3) แยกให้ได้ว่าผิดเพราะ issue น้อยเกิน, มากเกิน หรือหยิบผิดตัว\\n4) ถ้าน้อยเกินให้ทำ stock adjustment/issue เพิ่มตาม policy\\n5) ถ้ามากเกินให้ทำ return to stock หรือ adjustment กลับตาม policy\\n6) ใส่ note อ้างอิงเลข MO ทุกครั้ง",
                "expected": "ทีมต้องเลือกวิธีแก้ที่ไม่ทำให้ finished goods พัง และ stock ต้องกลับมาตรงความจริงพร้อมมีร่องรอย audit",
                "fix": "ถ้าทีมสับสนว่าใช้ Unbuild หรือ adjustment ให้ใช้หลักว่า FG ถูกต้องแล้วไม่ควร unbuild ทั้งใบ",
                "evidence": "Capture: MO moves + adjustment/return document",
                "note": "เคสนี้สำคัญมากในงานหน้างานจริง",
            },
            {
                "case_id": "MU12-03",
                "backlogs": "MA16, MA17, MA18",
                "name": "ปิด MO แล้วพบว่าควรเหลือ backorder แต่ user กดจบไปแล้ว",
                "role": "Production / Planner",
                "path": "MO done > Create Backorder / New MO",
                "pre": "มี MO done แบบ partial และ operation type ตั้ง Ask Backorder",
                "data": "ใช้ MO partial ที่กด No Backorder ไปแล้ว",
                "steps": "1) เปิด MO ที่ done แล้ว\\n2) ตรวจว่ามี remaining qty ที่ควรผลิตต่อจริง\\n3) ถ้ามีปุ่ม Create Backorder ให้กดสร้าง backorder\\n4) ถ้าไม่มีปุ่ม ให้ประเมินว่าต้องสร้าง MO ใหม่/manual continuation ตาม policy\\n5) บันทึก origin และอธิบายในหมายเหตุว่ามาจาก MO ใบเดิม",
                "expected": "ทีมต้องรู้ว่าถ้ากู้ backorder ได้ให้กู้จากใบเดิมก่อน และถ้ากู้ไม่ได้ต้องแตกเอกสารใหม่อย่าง traceable",
                "fix": "ถ้าปุ่มไม่ขึ้น ให้ตรวจเงื่อนไข remaining qty, policy Ask Backorder และว่ามี backorder ไปแล้วหรือไม่",
                "evidence": "Capture: done MO + recovery path/new MO",
                "note": "ผูกกับฟีเจอร์ backorder recovery ที่ทำไว้",
            },
            {
                "case_id": "MU12-04",
                "backlogs": "MA16, MA17",
                "name": "ปิด MO แล้วโพสต์บัญชีไปแล้ว ควรประสานยังไง",
                "role": "Production / Accounting",
                "path": "MO > Valuation / Journal Entries",
                "pre": "มี MO done และเกิด valuation/journal แล้ว",
                "data": "เลือก MO ที่มี stock valuation เกิดแล้ว",
                "steps": "1) เปิด MO แล้วดู smart button ต้นทุน/valuation/journal ถ้ามี\\n2) ตรวจว่าเอกสารบัญชีถูกสร้างไปแล้วหรือยัง\\n3) ก่อนแก้ stock หรือ unbuild ให้ประสานฝ่ายบัญชีเพื่อยืนยันแนวทาง\\n4) หลังแก้ไข ให้กลับมาตรวจ stock และ valuation อีกครั้ง\\n5) บันทึกเหตุผลการแก้ใน chatter หรือเอกสารกำกับ",
                "expected": "การแก้หลังปิด MO ต้องไม่ทำแบบเงียบ ๆ ถ้ามีผลกับบัญชี และต้องมีร่องรอยให้ทีมบัญชีตรวจย้อนหลังได้",
                "fix": "ถ้าไม่แน่ใจว่ากระทบบัญชีหรือไม่ ให้เช็ก valuation layer/journal ก่อนทุกครั้ง",
                "evidence": "Capture: valuation/journal + final correction docs",
                "note": "เหมาะกับ training ร่วม production + accounting",
            },
        ],
    },
    {
        "sheet": "MU13",
        "title": "UoM - เปลี่ยนหน่วยซื้อ / หน่วยใช้ / หน่วยใน BoM",
        "objective": "ให้ผู้ใช้เข้าใจทั้งการเปลี่ยน UoM แบบ manual ที่เอกสารซื้อ, การตั้ง default ให้ดึงอัตโนมัติ และการเปลี่ยน UoM ใน BOM ที่ต้องสะท้อนลง MO/เอกสารเบิก",
        "coverage": "purchase uom manual, purchase uom auto default, bom uom conversion",
        "cases": [
            {
                "case_id": "MU13-01",
                "backlogs": "MA01",
                "name": "เปลี่ยนหน่วยซื้อแบบ Manual บน PR/PO Order Line",
                "role": "Purchase User",
                "path": "Purchase > PR/PO > Order Lines",
                "pre": "มีสินค้าในระบบและ user มีสิทธิ์แก้ UoM บน order line",
                "data": "เลือกสินค้าที่มีหลายหน่วยซื้อใน UoM category เดียวกัน",
                "steps": "1) เปิด PR หรือ PO\\n2) ไปที่ Order Lines\\n3) เลือก Product ที่ต้องการซื้อ\\n4) ไปที่คอลัมน์ UoM / Unit of Measure\\n5) เปลี่ยนเป็นหน่วยที่ต้องการด้วยมือ\\n6) ตรวจ Quantity และราคาอีกครั้งว่าถูกตามหน่วยใหม่\\n7) Save เอกสาร",
                "expected": "ผู้ใช้ต้องเปลี่ยนหน่วยซื้อได้เองที่บรรทัดเอกสาร และระบบต้องเก็บหน่วยนั้นให้กับเอกสารซื้อใบนั้น",
                "fix": "ถ้าเลือกหน่วยไม่ได้ ให้ตรวจว่า UoM อยู่ category เดียวกันหรือสิทธิ์ field ถูกจำกัด",
                "evidence": "Capture: PO line ก่อน/หลังเปลี่ยน UoM",
                "note": "ใช้สอนกรณีซื้อเฉพาะครั้ง ไม่ต้องเปลี่ยน master data",
            },
            {
                "case_id": "MU13-02",
                "backlogs": "MA01",
                "name": "เปลี่ยนหน่วยซื้อให้ดึงอัตโนมัติจาก Product Master",
                "role": "Purchase Master Data / Buyer",
                "path": "Inventory > Product Master > Purchase tab > Vendors/Purchase Unit",
                "pre": "มีสินค้าและ vendor line บนสินค้า",
                "data": "เลือกสินค้าที่ต้องการให้ PO ดึง UoM ใหม่อัตโนมัติ",
                "steps": "1) เปิด Inventory > Products > Products\\n2) เปิดสินค้าที่ต้องการ\\n3) ไปที่แท็บ Purchase\\n4) แก้ Vendor line หรือ Purchase Unit ให้เป็นหน่วยที่ต้องการ\\n5) Save\\n6) กลับไปสร้าง PR/PO ใหม่\\n7) เลือก Product เดิม\\n8) ตรวจว่า UoM บน order line เปลี่ยนตาม master อัตโนมัติ",
                "expected": "เมื่อ master data ถูกตั้งแล้ว เอกสารซื้อใหม่ต้อง default หน่วยซื้อให้อัตโนมัติโดยไม่ต้องแก้ทุกครั้ง",
                "fix": "ถ้า UoM ไม่ดึงตาม ให้ตรวจ vendor line sequence, company, purchase UoM และ cache ของ form",
                "evidence": "Capture: product purchase tab + new PO line",
                "note": "เหมาะกับสินค้าที่ซื้อหน่วยเดิมประจำ",
            },
            {
                "case_id": "MU13-03",
                "backlogs": "MA01, MA10",
                "name": "เปลี่ยน UoM บน BOM แล้วให้ MO และเอกสารเบิกเปลี่ยนอัตโนมัติ",
                "role": "Manufacturing Master Data / Planner",
                "path": "Manufacturing > Products > Bills of Materials",
                "pre": "มี BOM ที่ใช้จริงและ component มี UoM หลายหน่วยใน category เดียวกัน",
                "data": "เลือก BOM ที่ต้องการแก้ component UoM",
                "steps": "1) เปิด Manufacturing > Products > Bills of Materials\\n2) เลือก BOM ที่ต้องการ\\n3) ไปที่ Components\\n4) เลือก component ที่ต้องการเปลี่ยนหน่วย\\n5) เปลี่ยน UoM เป็นหน่วยที่ต้องการ\\n6) Save BOM\\n7) ไปที่ Manufacturing Orders\\n8) กด New และเลือก BOM เดิม\\n9) Confirm MO\\n10) เปิด Components/Raw Moves\\n11) ตรวจว่าระบบใช้ UoM ใหม่ตามที่ตั้งไว้ และเอกสารเบิกใช้หน่วยนี้ตาม BOM",
                "expected": "เมื่อ BOM เปลี่ยน UoM แล้ว MO ใหม่ต้องสะท้อนหน่วยใหม่อัตโนมัติ และเอกสารเบิก/consume ต้องใช้หน่วยเดียวกับ BOM",
                "fix": "ถ้า MO ยังใช้หน่วยเดิม ให้ตรวจว่าเปิด MO ใหม่จริง, BOM version ถูกต้อง และ UoM อยู่ category เดียวกัน",
                "evidence": "Capture: BOM component line + MO component line",
                "note": "ย้ำว่าเอกสารเก่าจะไม่วิ่งตาม ต้องสร้าง MO ใหม่หลังแก้ BOM",
            },
        ],
    },
]


def thin_border(color="E6E6E6"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header(cell, fill="1F4E78"):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border("D9E2F3")


def style_body(cell, wrap=False, fill=None, size=10):
    cell.font = Font(name=FONT_NAME, size=size, color="1F1F1F")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = thin_border()


def last_nonempty_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1


def style_existing_workbook(wb):
    for ws in wb.worksheets:
        if ws.title.startswith("MU"):
            ws.sheet_view.showGridLines = False
            ws.sheet_view.zoomScale = 85
            ws.freeze_panes = "A7"
            ws.auto_filter.ref = f"A6:S{ws.max_row}"
            for c in range(1, 20):
                style_header(ws.cell(6, c))
            ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color="1F1F1F")
            for ref in ("A2", "A3", "A4"):
                ws[ref].font = Font(name=FONT_NAME, size=11, bold=True, color="1F1F1F")
            for ref in ("B2", "B3", "B4"):
                ws[ref].font = Font(name=FONT_NAME, size=10, color="1F1F1F")
                ws[ref].alignment = Alignment(wrap_text=True, vertical="top")
            for r in range(7, ws.max_row + 1):
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 100)
                for c in range(1, 20):
                    style_body(ws.cell(r, c), wrap=(c >= 3), fill="FFF2CC" if c == 18 else None)

        elif ws.title == "Scenarios":
            ws.sheet_view.zoomScale = 85
            ws.freeze_panes = "A9"
            ws.auto_filter.ref = f"A8:M{ws.max_row}"
            for c in range(1, 14):
                style_header(ws.cell(8, c), fill="4F81BD")
            for r in range(9, ws.max_row + 1):
                for c in range(1, 14):
                    style_body(ws.cell(r, c), wrap=(c >= 3))
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, 42)
        elif ws.title == "Product_backlog":
            ws.sheet_view.zoomScale = 80
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:T{ws.max_row}"
            for c in range(1, 21):
                style_header(ws.cell(1, c), fill="5B9BD5")
            for r in range(2, ws.max_row + 1):
                for c in range(1, 21):
                    fill = "E2F0D9" if c >= 18 else None
                    style_body(ws.cell(r, c), wrap=True, fill=fill)
        elif ws.title == "MANU_UAT_README":
            ws.sheet_view.zoomScale = 95
            style_header(ws["A1"])
            style_header(ws["B1"])
            for r in range(2, ws.max_row + 1):
                style_body(ws.cell(r, 1), wrap=True)
                style_body(ws.cell(r, 2), wrap=True)


def normalize_workbook_text(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "\\n" in cell.value:
                    cell.value = cell.value.replace("\\n", "\n")
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical or "top",
                        wrap_text=True,
                    )

    if "MANU_UAT_README" in wb.sheetnames:
        readme = wb["MANU_UAT_README"]
        readme["B3"] = "เปิดแต่ละ sheet ตั้งแต่ MU01-MU13 แล้วทำตาม Detailed Test Steps ทีละข้อ"
        readme["B3"].alignment = Alignment(vertical="top", wrap_text=True)


def improve_readability(wb):
    for ws in wb.worksheets:
        if ws.title.startswith("MU"):
            widths = {
                "A": 6,
                "B": 12,
                "C": 18,
                "D": 38,
                "E": 22,
                "F": 34,
                "G": 34,
                "H": 28,
                "I": 78,
                "J": 48,
                "K": 44,
                "L": 32,
                "M": 34,
                "N": 18,
                "O": 14,
                "P": 16,
                "Q": 16,
                "R": 16,
                "S": 30,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for r in range(7, ws.max_row + 1):
                step_text = ws[f"I{r}"].value if isinstance(ws[f"I{r}"].value, str) else ""
                line_count = max(1, step_text.count("\n") + 1)
                ws.row_dimensions[r].height = max(88, 22 * line_count)

        elif ws.title == "Scenarios":
            widths = {
                "A": 8,
                "B": 20,
                "C": 18,
                "D": 42,
                "E": 24,
                "F": 16,
                "G": 22,
                "H": 16,
                "I": 16,
                "J": 16,
                "K": 14,
                "L": 34,
                "M": 36,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for r in range(9, ws.max_row + 1):
                ws.row_dimensions[r].height = max(42, ws.row_dimensions[r].height or 0)

        elif ws.title == "Product_backlog":
            widths = {
                "A": 12,
                "B": 16,
                "C": 18,
                "D": 14,
                "E": 24,
                "F": 24,
                "G": 28,
                "H": 18,
                "I": 16,
                "J": 16,
                "K": 16,
                "L": 18,
                "M": 18,
                "N": 18,
                "O": 18,
                "P": 18,
                "Q": 18,
                "R": 16,
                "S": 16,
                "T": 18,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = max(36, ws.row_dimensions[r].height or 0)

        elif ws.title == "MANU_UAT_README":
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 120


def build_topic_sheet(wb, topic, status_validation):
    if topic["sheet"] in wb.sheetnames:
        del wb[topic["sheet"]]
    ws = wb.create_sheet(topic["sheet"])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws["A1"] = f"Manufacturing UAT Test Step - {topic['sheet']}"
    ws["A2"] = "หัวข้อ"
    ws["B2"] = topic["title"]
    ws["A3"] = "วัตถุประสงค์"
    ws["B3"] = topic["objective"]
    ws["A4"] = "ขอบเขต"
    ws["B4"] = topic["coverage"]
    ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True)
    for ref in ("A2", "A3", "A4"):
        ws[ref].font = Font(name=FONT_NAME, size=11, bold=True)
        ws[ref].fill = PatternFill("solid", fgColor="D9EAD3")
        ws[ref].border = thin_border()
    for ref in ("B2", "B3", "B4"):
        style_body(ws[ref], wrap=True, size=10)

    headers = [
        "No.", "Case ID", "Backlog IDs", "Scenario Name", "Role / Department", "Menu Path",
        "Pre-Condition / Setup", "Test Data", "Detailed Test Steps", "Expected Result",
        "Fix / Recovery Path", "Evidence to Capture", "Actual Result / Notes", "Executor",
        "Test Date", "Review - คุณไอซ์", "Review - คุณติ๊ก", "Overall Status", "Trainer Note",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(6, col, header)
        style_header(ws.cell(6, col))
    widths = {
        1: 6, 2: 12, 3: 18, 4: 34, 5: 22, 6: 28, 7: 30, 8: 26, 9: 60,
        10: 40, 11: 36, 12: 30, 13: 32, 14: 18, 15: 14, 16: 16, 17: 16,
        18: 16, 19: 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    mappings = []
    row = 7
    for idx, case in enumerate(topic["cases"], start=1):
        values = [
            idx, case["case_id"], case["backlogs"], case["name"], case["role"], case["path"], case["pre"],
            case["data"], case["steps"], case["expected"], case["fix"], case["evidence"], None, None, None,
            None, None, None, case["note"],
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row, col, value)
            style_body(ws.cell(row, col), wrap=(col >= 3), fill="FFF2CC" if col == 18 else None)
        ws.row_dimensions[row].height = 110
        ws.cell(row, 18, f'=IF(AND(P{row}="Passed",Q{row}="Passed"),"Passed",IF(OR(P{row}="Failed",Q{row}="Failed"),"Failed",IF(COUNTA(P{row}:Q{row})=0,"Not Start",IF(OR(P{row}="Pending",Q{row}="Pending",P{row}="Cancelled",Q{row}="Cancelled"),"Pending","Under Testing"))))')
        status_validation.add(ws.cell(row, 16))
        status_validation.add(ws.cell(row, 17))
        for backlog in [x.strip() for x in case["backlogs"].split(",") if x.strip()]:
            mappings.append((backlog, case["case_id"], topic["sheet"], row))
        row += 1
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:S{row-1}"
    ws.add_data_validation(status_validation)
    return mappings


def append_scenarios(wb):
    sc = wb["Scenarios"]
    start_row = last_nonempty_row(sc) + 1
    seq = max([sc.cell(r, 1).value for r in range(9, sc.max_row + 1) if isinstance(sc.cell(r, 1).value, int)], default=0) + 1
    for topic in NEW_TOPICS:
        for idx, _case in enumerate(topic["cases"], start=1):
            src_row = 6 + idx
            r = start_row
            formulas = {
                1: seq,
                2: f"={topic['sheet']}!B{src_row}",
                3: f"={topic['sheet']}!C{src_row}",
                4: f"={topic['sheet']}!D{src_row}",
                5: f"={topic['sheet']}!E{src_row}",
                6: f"={topic['sheet']}!O{src_row}",
                7: f"={topic['sheet']}!N{src_row}",
                8: f"={topic['sheet']}!R{src_row}",
                9: f"={topic['sheet']}!P{src_row}",
                10: f"={topic['sheet']}!Q{src_row}",
                11: topic["sheet"],
                12: f"={topic['sheet']}!F{src_row}",
                13: f"={topic['sheet']}!K{src_row}",
            }
            for c, value in formulas.items():
                sc.cell(r, c, value)
                style_body(sc.cell(r, c), wrap=(c >= 3))
            sc.row_dimensions[r].height = 48
            start_row += 1
            seq += 1


def append_map_rows(wb, mappings):
    map_ws = wb["_MANU_UAT_MAP"]
    start = last_nonempty_row(map_ws) + 1
    for idx, (backlog, case_id, sheet_name, row_num) in enumerate(mappings, start=start):
        map_ws.cell(idx, 1, backlog)
        map_ws.cell(idx, 2, case_id)
        map_ws.cell(idx, 3, sheet_name)
        map_ws.cell(idx, 4, row_num)
        map_ws.cell(idx, 5, f"='{sheet_name}'!P{row_num}")
        map_ws.cell(idx, 6, f"='{sheet_name}'!Q{row_num}")
        map_ws.cell(idx, 7, f"='{sheet_name}'!R{row_num}")


def update_readme(wb):
    ws = wb["MANU_UAT_README"]
    row = last_nonempty_row(ws) + 2
    extra = [
        ("รอบ v2 เพิ่มอะไร", "เพิ่ม MU11-MU13 สำหรับ GMP Shop Floor, แก้ไขหลังปิด MO/Unbuild และ UoM change"),
        ("ฟอนต์/ความอ่านง่าย", "ปรับฟอนต์ทั้ง workbook เป็น Aptos, เพิ่มความกว้างคอลัมน์, row height, wrap text และ autofilter"),
        ("เคส GMP Shop Floor", "ครอบคลุม queue, start/pause/resume, good/reject, mold/workcenter, และงาน done/cancel"),
        ("เคส Close MO", "ครอบคลุม Unbuild, ปรับ stock หลังปิดงาน, recover backorder และการประสานบัญชี"),
        ("เคส UoM", "เพิ่มทั้งเปลี่ยนหน่วยซื้อแบบ manual, auto default จาก product master และเปลี่ยนหน่วยใน BOM"),
    ]
    for title, desc in extra:
        ws.cell(row, 1, title)
        ws.cell(row, 2, desc)
        style_body(ws.cell(row, 1), wrap=True)
        style_body(ws.cell(row, 2), wrap=True)
        row += 1


def main():
    wb = load_workbook(SOURCE)
    list_ws = wb["_MANU_UAT_LISTS"]
    list_ws.sheet_state = "hidden"
    status_validation = DataValidation(type="list", formula1="'_MANU_UAT_LISTS'!$A$1:$A$6", allow_blank=True)

    new_mappings = []
    for topic in NEW_TOPICS:
        new_mappings.extend(build_topic_sheet(wb, topic, status_validation))

    append_scenarios(wb)
    append_map_rows(wb, new_mappings)
    update_readme(wb)
    style_existing_workbook(wb)
    normalize_workbook_text(wb)
    improve_readability(wb)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    main()
