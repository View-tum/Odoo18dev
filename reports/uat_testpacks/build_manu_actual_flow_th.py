from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(r"C:\365_project\TheCool18e\Dev")
SOURCE_PATH = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_full_backlog_coverage_th_20260407.xlsx"
)
ASSESSMENT_PATH = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_local_uat_assessment_20260407.xlsx"
)
OUTPUT_PATH = Path(
    r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_actual_flow_th_20260407.xlsx"
)
SUMMARY_MD = ROOT / "reports" / "local_uat_manu_actual_flow_th_20260407.md"


CASE_HEADERS = [
    "ลำดับ",
    "Case ID",
    "Backlog IDs",
    "ลำดับเหตุการณ์",
    "Scenario ทดสอบ",
    "บทบาท / หน่วยงาน",
    "Menu Path จริงใน local UAT",
    "เงื่อนไขก่อนทดสอบ",
    "ข้อมูลทดสอบ",
    "ขั้นตอนทดสอบแบบละเอียด",
    "ผลลัพธ์ที่คาดหวัง",
    "แนวทางตรวจแก้ / Recovery",
    "สถานะฟังก์ชันใน local UAT",
    "หมายเหตุ / หลักฐาน",
    "Review - คุณไอซ์",
    "Review - คุณติ๊ก",
    "สถานะรีวิว",
]


PHASES = [
    {
        "sheet": "01_ตรวจสต็อกและวางแผน",
        "topic": "ตรวจสต็อกและวางแผนก่อนเริ่มผลิต",
        "objective": "ให้ planner และคลังตรวจ stock, forecast, replenishment และ child chain ได้ก่อนเปิดงานผลิตจริง",
        "scope": "On Hand, Forecast, Replenishment, MTS / Min-Max, child chain",
        "event": "ขั้นที่ 1 ตรวจ stock และวางแผน",
        "cases": ["MU01-01", "MU01-02", "MU01-03", "MU06-01", "MU06-02", "MU06-03"],
    },
    {
        "sheet": "02_เปิดงานผลิต",
        "topic": "สร้างความต้องการและเปิดใบสั่งผลิต",
        "objective": "ให้ user เปิด MO ได้ทั้ง manual, MTO และดูความพร้อมวัตถุดิบ/เอกสารที่เกี่ยวข้องได้ต่อเนื่อง",
        "scope": "Manual MO, MTO, component availability, linked transfer",
        "event": "ขั้นที่ 2 สร้าง demand และเปิด MO",
        "cases": ["MU02-01", "MU02-02", "MU02-03", "MU05-01", "MU05-02"],
    },
    {
        "sheet": "03_โอนและBackorder",
        "topic": "โอนวัตถุดิบและจัดการ Backorder",
        "objective": "ให้คลังและฝ่ายผลิตทำ transfer ครบ, partial, backorder และ recovery ได้จริง",
        "scope": "Transfer Plastic, Transfer Pharma, Create Backorder, No Backorder, Recovery",
        "event": "ขั้นที่ 3 โอนวัตถุดิบและแก้กรณีโอนไม่ครบ",
        "cases": ["MU03-01", "MU03-02", "MU03-03", "MU04-01", "MU04-02", "MU04-03"],
    },
    {
        "sheet": "04_ShopFloorและMold",
        "topic": "ทำงานจริงบน GMP Shop Floor และ Mold",
        "objective": "ให้ operator และ supervisor ใช้ Shop Floor, mold, workcenter และ labor ได้จากหน้าเดียว",
        "scope": "Queue, start, done, good/reject, mold auto, mold warning, reset mold",
        "event": "ขั้นที่ 4 ทำงานหน้างานจริง",
        "cases": [
            "MU11-01",
            "MU11-02",
            "MU11-03",
            "MU11-04",
            "MU09-01",
            "MU09-02",
            "MU09-03",
            "MU14-01",
            "MU14-02",
            "MU14-03",
            "MU14-04",
            "MU14-05",
            "MU14-06",
        ],
    },
    {
        "sheet": "05_ปิดงานและแก้ไข",
        "topic": "ปิดงาน Scrap และแก้ไขหลังปิด",
        "objective": "ให้ฝ่ายผลิตและบัญชีแก้กรณี partial, scrap, close ผิด, unbuild และ trace valuation ได้",
        "scope": "MO completion, MO backorder, scrap, leftover return, unbuild, valuation trace",
        "event": "ขั้นที่ 5 ปิดงานและแก้ไขหลังปิด",
        "cases": ["MU07-01", "MU07-02", "MU07-03", "MU07-04", "MU08-01", "MU08-02", "MU12-01", "MU12-02", "MU12-03", "MU12-04"],
    },
    {
        "sheet": "06_คุณภาพเอกสาร",
        "topic": "คุณภาพ เอกสาร และการติดตามย้อนหลัง",
        "objective": "ให้ QA, QC และ document control ใช้เมนูจริงในระบบเพื่อติดตาม lot, quality check, quality alert และเอกสารประกอบได้",
        "scope": "Lots/Serial, Documents, Quality Checks, Quality Alerts, Control Points, Locations",
        "event": "ขั้นที่ 6 ตรวจคุณภาพและเอกสาร",
        "cases": ["MU15-03", "MU16-01", "MU16-03", "MU16-05", "MU16-08"],
    },
    {
        "sheet": "07_รายงานต้นทุนUoM",
        "topic": "รายงาน ต้นทุน และ UoM",
        "objective": "ให้ key user ใช้ routing, MPS, reports, cost, UoM และแบบพิมพ์จากเมนูจริงของ local UAT ได้ต่อเนื่อง",
        "scope": "Routing, MPS, Backorder coverage, Production Analysis, Machine Report, UoM, Cost",
        "event": "ขั้นที่ 7 ตรวจรายงาน ต้นทุน และ master data ปลายทาง",
        "cases": [
            "MU10-01",
            "MU10-02",
            "MU10-03",
            "MU10-04",
            "MU13-01",
            "MU13-02",
            "MU13-03",
            "MU17-01",
            "MU17-02",
            "MU17-04",
            "MU17-05",
            "MU19-03",
            "MU19-04",
            "MU20-02",
            "MU20-03",
            "MU20-04",
            "MU20-05",
            "MU20-06",
        ],
    },
]


MENU_OVERRIDES = {
    "MU01-01": "สินค้าคงคลัง > สินค้า > สินค้า > เปิดสินค้า FG > Smart Button: คงเหลือ / คาดการณ์",
    "MU01-02": "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า",
    "MU01-03": "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า > เปิดเอกสารที่ระบบสร้าง",
    "MU02-01": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต",
    "MU02-02": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต",
    "MU02-03": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO > Smart Button: การโอน",
    "MU03-01": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > กรองประเภทการดำเนินการ = Transfer Plastic",
    "MU03-02": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิด Transfer Plastic > Validate แบบ partial",
    "MU03-03": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิดเอกสารที่ Done > ปุ่ม Create Backorder",
    "MU04-01": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > กรองประเภทการดำเนินการ = Transfer Pharma",
    "MU04-02": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิด Transfer Pharma > Validate แบบ partial",
    "MU04-03": "สินค้าคงคลัง > การปฏิบัติการ > การโอน > เปิดเอกสารที่ Done > ปุ่ม Create Backorder",
    "MU05-01": "การขาย > คำสั่ง > ใบเสนอราคา > ยืนยัน > trace ไปใบสั่งผลิต",
    "MU05-02": "การขาย > คำสั่ง > ใบเสนอราคา > ยืนยัน > trace shortage ไป MO / RFQ / การเติมสินค้า",
    "MU06-01": "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า",
    "MU06-02": "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า > เปิดสินค้า > Smart Button: คาดการณ์",
    "MU06-03": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO แม่ > เอกสารลูก / การโอนที่เกี่ยวข้อง",
    "MU07-01": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต หรือ การผลิต > การปฏิบัติการ > คำสั่งงาน",
    "MU07-02": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > ทำ partial แล้ว Validate/Done",
    "MU07-03": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > ปุ่ม Create Backorder",
    "MU07-04": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > แท็บ Components / Cost",
    "MU08-01": "สินค้าคงคลัง > การปฏิบัติการ > Scrap หรือ การผลิต > การปฏิบัติการ > คำสั่งงาน > Scrap",
    "MU08-02": "สินค้าคงคลัง > การปฏิบัติการ > การโอน",
    "MU09-01": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > คำสั่งงาน / การผลิต > การกำหนดค่า > ศูนย์งาน",
    "MU09-02": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > คำสั่งงาน",
    "MU09-03": "GMP Shop Floor > เปิดคำสั่งงาน",
    "MU10-01": "การผลิต > การรายงาน > วิเคราะห์การผลิต",
    "MU10-02": "สินค้าคงคลัง > การรายงาน > ประวัติการย้าย หรือ เปิดจาก MO ผ่าน Smart Button",
    "MU10-03": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > แท็บ Components / Cost",
    "MU10-04": "สั่งซื้อ > สินค้า > สินค้า / สั่งซื้อ > คำสั่ง > ใบแจ้งขอใบเสนอราคา หรือ คำสั่งซื้อ / การผลิต > ใบสั่งผลิต",
    "MU11-01": "GMP Shop Floor",
    "MU11-02": "GMP Shop Floor > เปิดคำสั่งงาน",
    "MU11-03": "GMP Shop Floor > เปิดคำสั่งงาน > Done",
    "MU11-04": "GMP Shop Floor > เปิดงานฝั่ง Plastic",
    "MU12-01": "การผลิต > การปฏิบัติการ > คำสั่งรื้อ",
    "MU12-02": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > Smart Button / สินค้าคงคลัง > การรายงาน > ประวัติการย้าย",
    "MU12-03": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > ปุ่ม Create Backorder หรือเปิด MO ใหม่",
    "MU12-04": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต > เปิด MO ที่ Done > Smart Button: การประเมินมูลค่า / รายการบัญชี",
    "MU13-01": "คำขอซื้อ > คำขอซื้อ > คำขอซื้อ หรือ สั่งซื้อ > คำสั่ง > คำสั่งซื้อ > แท็บรายการ",
    "MU13-02": "สั่งซื้อ > สินค้า > สินค้า > เปิดสินค้า > แท็บการสั่งซื้อ",
    "MU13-03": "การผลิต > สินค้า > บิลวัสดุ",
    "MU14-01": "GMP Shop Floor",
    "MU14-02": "GMP Shop Floor > การ์ดคำสั่งงาน > ปุ่ม Mold",
    "MU14-03": "GMP Shop Floor > การ์ดคำสั่งงาน > Start",
    "MU14-04": "GMP Shop Floor > Start > Popup เตือนอายุแม่พิมพ์ > Change Mold",
    "MU14-05": "การผลิต > การกำหนดค่า > ศูนย์งาน > เปิดแม่พิมพ์",
    "MU14-06": "GMP Shop Floor > การ์ดคำสั่งงาน > More > Report Issue",
    "MU15-03": "สินค้าคงคลัง > การกำหนดค่า > การบริหารสินค้าคงคลัง > ตำแหน่ง",
    "MU16-01": "การผลิต > สินค้า > ล็อต/หมายเลขซีเรียล",
    "MU16-03": "เอกสาร > เอกสาร",
    "MU16-05": "คุณภาพ > การควบคุมคุณภาพ > การแจ้งเตือนการจัดการคุณภาพ",
    "MU16-08": "คุณภาพ > การควบคุมคุณภาพ > จุดควบคุม",
    "MU17-01": "การผลิต > สินค้า > บิลวัสดุ",
    "MU17-02": "การผลิต > สินค้า > บิลวัสดุ",
    "MU17-04": "การผลิต > การวางแผน > MPS > กำหนดการการผลิตหลัก",
    "MU17-05": "สินค้าคงคลัง > การปฏิบัติการ > การโอน",
    "MU19-03": "การผลิต > การกำหนดค่า > ศูนย์งาน",
    "MU19-04": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต",
    "MU20-02": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต",
    "MU20-03": "การผลิต > การปฏิบัติการ > คำสั่งงาน",
    "MU20-04": "การผลิต > การปฏิบัติการ > Scrap",
    "MU20-05": "การผลิต > การรายงาน > Machine Report",
    "MU20-06": "การผลิต > การรายงาน > วิเคราะห์การผลิต",
}


REPLACEMENTS = {
    "Inventory > Products > Products": "สินค้าคงคลัง > สินค้า > สินค้า",
    "Inventory > Operations > Replenishment": "สินค้าคงคลัง > การปฏิบัติการ > การจัดซื้อ > การเติมสินค้า",
    "Inventory > Operations > Transfers": "สินค้าคงคลัง > การปฏิบัติการ > การโอน",
    "Inventory > Operations > Scrap": "สินค้าคงคลัง > การปฏิบัติการ > Scrap",
    "Inventory > Reporting > Product Moves": "สินค้าคงคลัง > การรายงาน > ประวัติการย้าย",
    "Inventory > Reporting > Moves History": "สินค้าคงคลัง > การรายงาน > ประวัติการย้าย",
    "Manufacturing > Operations > Manufacturing Orders": "การผลิต > การปฏิบัติการ > ใบสั่งผลิต",
    "Manufacturing > Operations > Work Orders": "การผลิต > การปฏิบัติการ > คำสั่งงาน",
    "Manufacturing > Operations > Unbuild Orders": "การผลิต > การปฏิบัติการ > คำสั่งรื้อ",
    "Manufacturing > Products > Bills of Materials": "การผลิต > สินค้า > บิลวัสดุ",
    "Manufacturing > Configuration > Work Centers": "การผลิต > การกำหนดค่า > ศูนย์งาน",
    "Manufacturing > Reporting > Manufacturing Orders": "การผลิต > การรายงาน > วิเคราะห์การผลิต",
    "Manufacturing > Reporting > Production Analysis": "การผลิต > การรายงาน > วิเคราะห์การผลิต",
    "Manufacturing > Reporting > Machine Report": "การผลิต > การรายงาน > Machine Report",
    "Sales > Orders > Quotations": "การขาย > คำสั่ง > ใบเสนอราคา",
    "Purchase > PR/PO": "คำขอซื้อ / สั่งซื้อ",
    "Purchase > Product Master": "สั่งซื้อ > สินค้า > สินค้า",
    "Purchase > Orders > Requests for Quotation": "สั่งซื้อ > คำสั่ง > ใบแจ้งขอใบเสนอราคา",
    "Purchase > Orders > Purchase Orders": "สั่งซื้อ > คำสั่ง > คำสั่งซื้อ",
    "Purchase > Orders > Purchase Requests": "คำขอซื้อ > คำขอซื้อ > คำขอซื้อ",
    "GMP Shop Floor > Work Order": "GMP Shop Floor > เปิดคำสั่งงาน",
    "GMP Shop Floor > Work Order Card": "GMP Shop Floor > การ์ดคำสั่งงาน",
}


EXCLUDED_CASES = {
    "MU05-03": "ตัดออกจากเวอร์ชันนี้เพราะเป็นเคสเชิงนโยบายการจัดลำดับงานมากกว่าปุ่ม/ฟังก์ชันที่กดตามได้ตรง ๆ",
    "MU08-03": "ตัดออกจากเวอร์ชันนี้เพราะขึ้นกับการตั้งค่า by-product / reuse เฉพาะกรณี",
    "MU11-05": "ตัดออกจากเวอร์ชันนี้เพราะยังไม่ได้ sign-off end-to-end รอบล่าสุด แม้ฟังก์ชัน queue มีอยู่จริง",
    "MU14-07": "ตัดออกจากเวอร์ชันนี้เพราะยังไม่ได้ปิดเคส UAT end-to-end ใน queue จริง",
}


def replace_text(value: str | None) -> str | None:
    if not isinstance(value, str):
        return value
    result = value
    for old, new in REPLACEMENTS.items():
        result = result.replace(old, new)
    return result


def load_source_cases() -> Dict[str, Dict[str, str]]:
    wb = load_workbook(SOURCE_PATH)
    cases: Dict[str, Dict[str, str]] = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("MU"):
            continue
        ws = wb[sheet_name]
        row = 7
        while row <= ws.max_row and ws.cell(row, 2).value:
            case_id = ws.cell(row, 2).value
            cases[case_id] = {
                "source_sheet": sheet_name,
                "no": ws.cell(row, 1).value,
                "case_id": case_id,
                "backlog_ids": ws.cell(row, 3).value or "",
                "scenario": replace_text(ws.cell(row, 4).value) or "",
                "role": replace_text(ws.cell(row, 5).value) or "",
                "menu": replace_text(ws.cell(row, 6).value) or "",
                "pre": replace_text(ws.cell(row, 7).value) or "",
                "data": replace_text(ws.cell(row, 8).value) or "",
                "steps": replace_text(ws.cell(row, 9).value) or "",
                "expected": replace_text(ws.cell(row, 10).value) or "",
                "fix": replace_text(ws.cell(row, 11).value) or "",
                "evidence": replace_text(ws.cell(row, 12).value) or "",
            }
            row += 1
    return cases


def load_assessment_map() -> Dict[str, Dict[str, str]]:
    wb = load_workbook(ASSESSMENT_PATH)
    ws = wb["LOCAL_UAT_SUMMARY"]
    result: Dict[str, Dict[str, str]] = {}
    row = 13
    while row <= ws.max_row and ws.cell(row, 1).value:
        result[ws.cell(row, 1).value] = {
            "status": ws.cell(row, 2).value,
            "note": ws.cell(row, 3).value,
        }
        row += 1
    return result


def map_local_status(case_id: str, assessment_map: Dict[str, Dict[str, str]]) -> tuple[str, str]:
    assessment = assessment_map.get(case_id)
    if not assessment:
        return "มีเมนูและฟังก์ชันจริง", "ยังไม่ได้ rerun เก็บหลักฐานรอบล่าสุด แต่เมนูและฟังก์ชันมีอยู่จริงใน local UAT"
    status_map = {
        "Passed": "ยืนยันแล้ว",
        "Passed with note": "ยืนยันแล้ว (มีหมายเหตุ)",
        "Not closed": "มีฟังก์ชันจริง แต่ยังไม่ปิดเคส UAT",
    }
    return status_map.get(assessment["status"], assessment["status"]), assessment["note"] or ""


def collect_included_cases(source_cases: Dict[str, Dict[str, str]], assessment_map: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    included: List[Dict[str, str]] = []
    phase_map = {case_id: phase["event"] for phase in PHASES for case_id in phase["cases"]}
    for phase in PHASES:
        for case_id in phase["cases"]:
            base = dict(source_cases[case_id])
            base["event"] = phase_map[case_id]
            base["menu"] = MENU_OVERRIDES.get(case_id, base["menu"])
            status, note = map_local_status(case_id, assessment_map)
            base["local_status"] = status
            base["local_note"] = note
            included.append(base)
    return included


def backlog_set(backlog_text: str) -> set[str]:
    return {part.strip() for part in str(backlog_text).split(",") if part and part.strip()}


def write_section_sheet(ws, phase: Dict[str, str], cases: List[Dict[str, str]]):
    ws["A1"] = "UAT Manufacturing (เฉพาะฟังก์ชันที่มีจริงใน local UAT)"
    ws["A2"] = "หัวข้อ"
    ws["B2"] = phase["topic"]
    ws["A3"] = "วัตถุประสงค์"
    ws["B3"] = phase["objective"]
    ws["A4"] = "ขอบเขต"
    ws["B4"] = phase["scope"]

    for col, header in enumerate(CASE_HEADERS, start=1):
        ws.cell(6, col).value = header

    row = 7
    for index, case in enumerate(cases, start=1):
        values = [
            index,
            case["case_id"],
            case["backlog_ids"],
            case["event"],
            case["scenario"],
            case["role"],
            case["menu"],
            case["pre"],
            case["data"],
            case["steps"],
            case["expected"],
            case["fix"],
            case["local_status"],
            case["local_note"],
            None,
            None,
            '=IF(AND(O{0}="Passed",P{0}="Passed"),"Passed",IF(OR(O{0}="Failed",P{0}="Failed"),"Failed","Pending"))'.format(row),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col).value = value
        row += 1


def style_sheet(ws):
    dark_fill = PatternFill("solid", fgColor="1F4E78")
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = Font(name="Tahoma", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = dark_fill
    ws.merge_cells("A1:Q1")
    ws["A1"].alignment = Alignment(horizontal="center")

    for cell in ("A2", "A3", "A4"):
        ws[cell].font = Font(name="Tahoma", bold=True)
        ws[cell].fill = title_fill
        ws[cell].border = border
    for cell in ("B2", "B3", "B4"):
        ws[cell].alignment = Alignment(wrap_text=True, vertical="top")
        ws[cell].border = border

    for col in range(1, 18):
        cell = ws.cell(6, col)
        cell.font = Font(name="Tahoma", bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 8,
        "B": 14,
        "C": 20,
        "D": 22,
        "E": 34,
        "F": 20,
        "G": 42,
        "H": 30,
        "I": 24,
        "J": 60,
        "K": 34,
        "L": 36,
        "M": 24,
        "N": 42,
        "O": 16,
        "P": 16,
        "Q": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    review_validation = DataValidation(type="list", formula1='"Passed,Failed,Pending,N/A"', allow_blank=True)
    ws.add_data_validation(review_validation)
    review_validation.add(f"O7:P1000")

    for row in range(7, ws.max_row + 1):
        for col in range(1, 18):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Tahoma", size=10)
        ws.row_dimensions[row].height = 54

    ws.freeze_panes = "A7"
    ws.sheet_view.showGridLines = False


def build_summary_sheet(ws, included_cases: List[Dict[str, str]], source_cases: Dict[str, Dict[str, str]], coverage_rows: List[Dict[str, str]]):
    ws["A1"] = "สรุปชุด UAT Manufacturing (เฉพาะฟังก์ชันที่มีจริงใน local UAT)"
    ws["A3"] = "เกณฑ์การคัดเคส"
    ws["B3"] = "คัดเฉพาะฟังก์ชันที่มีเมนู/หน้าจอจริงใน local UAT และสามารถกดตาม flow ได้จริง ไม่ใส่ future scope หรือเคสที่เป็นนโยบายล้วน"

    verified = sum(1 for case in included_cases if case["local_status"].startswith("ยืนยันแล้ว"))
    real_menu = sum(1 for case in included_cases if case["local_status"] == "มีเมนูและฟังก์ชันจริง")

    metrics = [
        ("จำนวนเคสทั้งหมดใน workbook ต้นทาง", len(source_cases)),
        ("จำนวนเคสที่คัดเข้าชุด actual flow", len(included_cases)),
        ("จำนวนเคสที่ตัดออก", len(source_cases) - len(included_cases)),
        ("จำนวนเคสที่มีหลักฐานทดสอบ local UAT แล้ว", verified),
        ("จำนวนเคสที่มีเมนูจริงแต่ยังไม่ได้ rerun รอบล่าสุด", real_menu),
        ("จำนวน backlog ทั้งหมด", len(coverage_rows)),
        ("จำนวน backlog ที่ถูกครอบคลุมในชุด actual flow", sum(1 for row in coverage_rows if row["covered"] == "ใช่")),
        ("จำนวน backlog ที่ยังไม่ถูกครอบคลุมในชุด actual flow", sum(1 for row in coverage_rows if row["covered"] == "ไม่ใช่")),
    ]
    start = 5
    for idx, (label, value) in enumerate(metrics, start=start):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = value

    ws["A15"] = "ชุด flow ตามลำดับเหตุการณ์"
    for offset, phase in enumerate(PHASES, start=16):
        ws.cell(offset, 1).value = phase["event"]
        ws.cell(offset, 2).value = phase["topic"]
        ws.cell(offset, 3).value = len(phase["cases"])

    ws["A24"] = "เคสที่ตัดออกจากเวอร์ชันนี้"
    row = 25
    for case_id, reason in EXCLUDED_CASES.items():
        ws.cell(row, 1).value = case_id
        ws.cell(row, 2).value = reason
        row += 1

    for col in ("A", "B", "C"):
        ws.column_dimensions[col].width = 42 if col == "B" else 28
    ws.column_dimensions["B"].width = 80
    for row in range(1, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, col).font = Font(name="Tahoma", size=10)
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True)
    ws.sheet_view.showGridLines = False


def build_coverage_rows(product_backlog_ws, included_cases: List[Dict[str, str]]) -> List[Dict[str, str]]:
    covered_backlogs = set()
    case_by_backlog: Dict[str, List[str]] = {}
    for case in included_cases:
        for backlog_id in backlog_set(case["backlog_ids"]):
            covered_backlogs.add(backlog_id)
            case_by_backlog.setdefault(backlog_id, []).append(case["case_id"])

    rows = []
    r = 2
    while r <= product_backlog_ws.max_row and product_backlog_ws.cell(r, 1).value:
        backlog_id = product_backlog_ws.cell(r, 1).value
        phase = product_backlog_ws.cell(r, 4).value
        covered = "ใช่" if backlog_id in covered_backlogs else "ไม่ใช่"
        reason = "ถูกคัดเข้า actual flow แล้ว" if covered == "ใช่" else "ยังเป็น future scope / ยังต้องออกแบบเพิ่ม / ยังไม่คัดเข้าชุดของจริงพร้อมใช้"
        rows.append(
            {
                "backlog_id": backlog_id,
                "phase": phase,
                "covered": covered,
                "cases": ", ".join(case_by_backlog.get(backlog_id, [])),
                "reason": reason,
            }
        )
        r += 1
    return rows


def write_coverage_sheet(ws, rows: List[Dict[str, str]]):
    ws["A1"] = "Coverage เทียบกับ Product_backlog"
    headers = ["Backlog ID", "Phase", "อยู่ในชุด actual flow", "Case IDs ที่ครอบคลุม", "หมายเหตุ"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(3, idx).value = header
    row = 4
    for item in rows:
        ws.cell(row, 1).value = item["backlog_id"]
        ws.cell(row, 2).value = item["phase"]
        ws.cell(row, 3).value = item["covered"]
        ws.cell(row, 4).value = item["cases"]
        ws.cell(row, 5).value = item["reason"]
        row += 1

    widths = {"A": 16, "B": 16, "C": 18, "D": 36, "E": 72}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(1, ws.max_row + 1):
        for c in range(1, 6):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).font = Font(name="Tahoma", size=10)
    ws["A1"].font = Font(name="Tahoma", size=14, bold=True)
    ws.sheet_view.showGridLines = False


def copy_sheet_order(wb: Workbook):
    desired_order = ["00_สรุป"] + [phase["sheet"] for phase in PHASES] + ["99_Coverage_Backlog"]
    wb._sheets.sort(key=lambda ws: desired_order.index(ws.title))


def write_summary_md(included_cases: List[Dict[str, str]], coverage_rows: List[Dict[str, str]]):
    verified = sum(1 for case in included_cases if case["local_status"].startswith("ยืนยันแล้ว"))
    lines = [
        "# สรุปชุด UAT Manufacturing (เฉพาะฟังก์ชันที่มีจริงใน local UAT)",
        "",
        f"- ไฟล์หลัก: `{OUTPUT_PATH}`",
        f"- จำนวนเคสในชุด actual flow: `{len(included_cases)}`",
        f"- จำนวนเคสที่มีหลักฐานทดสอบ local UAT แล้ว: `{verified}`",
        f"- จำนวน backlog ทั้งหมด: `{len(coverage_rows)}`",
        f"- จำนวน backlog ที่ถูกครอบคลุมในชุด actual flow: `{sum(1 for row in coverage_rows if row['covered'] == 'ใช่')}`",
        "",
        "## ลำดับเหตุการณ์ใน workbook",
    ]
    for phase in PHASES:
        lines.append(f"- `{phase['event']}`: {phase['topic']} ({len(phase['cases'])} เคส)")
    lines.extend(
        [
            "",
            "## หลักการคัดเคส",
            "- คัดเฉพาะฟังก์ชันที่มีเมนู/หน้าจอจริงใน local UAT",
            "- ไม่ใส่ future scope หรือเคสที่ยังไม่มีเมนูจริง",
            "- ไม่ใส่เคสที่เป็นนโยบายหรือแนวทางปฏิบัติอย่างเดียวโดยไม่มีปุ่ม/หน้าจอรองรับ",
            "",
            "## เคสที่ตัดออกจากเวอร์ชันนี้",
        ]
    )
    for case_id, reason in EXCLUDED_CASES.items():
        lines.append(f"- `{case_id}`: {reason}")
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")


def main():
    source_cases = load_source_cases()
    assessment_map = load_assessment_map()
    included_cases = collect_included_cases(source_cases, assessment_map)

    source_wb = load_workbook(SOURCE_PATH)
    product_backlog_ws = source_wb["Product_backlog"]
    coverage_rows = build_coverage_rows(product_backlog_ws, included_cases)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    summary_ws = wb.create_sheet("00_สรุป")
    build_summary_sheet(summary_ws, included_cases, source_cases, coverage_rows)

    case_map = {case["case_id"]: case for case in included_cases}
    for phase in PHASES:
        ws = wb.create_sheet(phase["sheet"])
        phase_cases = [case_map[case_id] for case_id in phase["cases"]]
        write_section_sheet(ws, phase, phase_cases)
        style_sheet(ws)

    coverage_ws = wb.create_sheet("99_Coverage_Backlog")
    write_coverage_sheet(coverage_ws, coverage_rows)

    copy_sheet_order(wb)
    wb.save(OUTPUT_PATH)
    write_summary_md(included_cases, coverage_rows)
    print(f"Created: {OUTPUT_PATH}")
    print(f"Summary: {SUMMARY_MD}")


if __name__ == "__main__":
    main()
