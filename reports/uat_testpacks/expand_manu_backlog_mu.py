from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


BASE = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_local_uat_assessment_20260407.xlsx")
OUT = Path(r"C:\Users\tumsu\Downloads\UAT_GoldMints_Test Scenario_MANU_full_backlog_coverage_20260407.xlsx")
SUMMARY = Path(r"C:\365_project\TheCool18e\Dev\reports\local_uat_manu_full_backlog_coverage_20260407.md")

FILL_DESIGN = PatternFill(fill_type="solid", fgColor="FFF2CC")
FILL_FUTURE = PatternFill(fill_type="solid", fgColor="F4CCCC")


def overall_formula(row: int) -> str:
    return (
        f'=IF(AND(P{row}="Passed",Q{row}="Passed"),"Passed",'
        f'IF(OR(P{row}="Failed",Q{row}="Failed"),"Failed",'
        f'IF(COUNTA(P{row}:Q{row})=0,"Not Start",'
        f'IF(OR(P{row}="Pending",Q{row}="Pending",P{row}="Cancelled",Q{row}="Cancelled"),"Pending","Under Testing"))))'
    )


def case_steps(menu: str, focus: str) -> str:
    return (
        f"1) Open {menu}\n"
        f"2) Prepare the sample data needed for {focus}\n"
        f"3) Execute the business flow step by step and record each document created\n"
        f"4) Verify quantities, status changes, source/destination locations, and side effects on related records\n"
        f"5) Compare the system result against the requirement and note any gap or workaround\n"
        f"6) Capture screenshots or exports that prove the result"
    )


def scenario_refs(sheet: str, row: int):
    return [
        f"='{sheet}'!B{row}",
        f"='{sheet}'!C{row}",
        f"='{sheet}'!D{row}",
        f"='{sheet}'!E{row}",
        f"='{sheet}'!O{row}",
        f"='{sheet}'!N{row}",
        f"='{sheet}'!R{row}",
        f"='{sheet}'!P{row}",
        f"='{sheet}'!Q{row}",
        sheet,
        f"='{sheet}'!F{row}",
        f"='{sheet}'!K{row}",
    ]


GROUPS = {
    "MU15": {
        "topic": "Material Handling / Container / Float Stock / Filter",
        "objective": "Cover container-based issue quantity, real-time stock update, floating stock sublocations, and plant-specific filtering.",
        "scope": "container issue, real-time stock, floating stock location, plant filter",
        "cases": [
            ("MA05", "Issue component by full container quantity", "Planner / Warehouse / Production", "Manufacturing > Manufacturing Orders > Components", "component issue by container", "Component with unit-per-container rule", "MO requesting less than one full container", "System rounds issue demand to the defined container size.", "Review UoM, package size, rounding logic, and raw move generation.", False),
            ("MA30", "Stock must update in real time after issue", "Warehouse / Production Controller", "Manufacturing > MO > Check Availability / Inventory > Product Moves", "real-time stock update after issue", "MO with available raw stock", "One raw move before and after issue", "On Hand / Reserved / Forecasted values update immediately after issue.", "Review picking state, move line done qty, and inventory cache.", False),
            ("MA39", "Validate detailed floating-stock sublocations", "Key User / Warehouse Lead", "Inventory > Configuration > Locations", "location structure for floating stock", "Existing GMP location tree", "Plastic and Pharma location branches", "Location tree clearly separates Plastic, Pharma, in-factory, and external production areas.", "Extend location tree and align route source/destination plus putaway rules.", False),
            ("MA49", "Filter stock views so users only see relevant plant items", "Operator / Warehouse User", "Inventory > Operations > Transfers", "filtered visibility by plant", "Transfer list with mixed Plastic and Pharma items", "Favorite filters or search panel settings", "User sees only the RM or transfers relevant to the assigned plant.", "Add search filters, default favorites, or role-specific views.", False),
        ],
    },
    "MU16": {
        "topic": "Traceability / QC / Sampling / Documents",
        "objective": "Cover future-scope QA, QC, traceability, and document-control requirements that were not yet mapped to MU cases.",
        "scope": "traceability, assay, ident, WI, BMR, reject, log book, QC approvals, package checks, FG issue, scale interface",
        "cases": [
            ("MA09", "Trace FG and components end to end with lot or QR reference", "QA / QC / Warehouse", "Inventory > Lots/Serial Numbers / Manufacturing > Traceability", "end-to-end component traceability", "Linked FG and component lots", "One FG lot and all upstream component lots", "System can trace from FG back to component lots and from a component lot forward to affected FG.", "Review lot propagation, traceability links, and intermediate container handling.", True),
            ("MA19", "Calculate assay and ident sample counts from receipt quantity", "QC Lab", "Quality / Receipt Inspection", "assay and ident sampling rules", "Receipt with multiple containers", "Container count n and expected sqrt(n)+1 sample count", "System records assay sample quantity and ident 100 percent rule correctly.", "Add quality worksheet or computed helper for sampling rules.", True),
            ("MA20", "Store, search, and revise Work Instructions in one place", "QA Document Control", "Documents / Knowledge", "WI repository and revision control", "At least two WI versions", "WI code, revision, effective date", "Users can search, open, revise, and identify the active WI version.", "Centralize WI in one document repository with revision history.", True),
            ("MA21", "Run BMR flow in system from production start to approval", "QA / Production Record Owner", "Manufacturing > MO > Documents / Quality", "BMR workflow inside ERP", "MO plus BMR form/checklist", "One production batch record", "BMR data and approvals stay attached to the production record.", "Use document workflow, approvals, or custom BMR forms tied to MO.", True),
            ("MA22", "Investigate reject root cause and reject the full lot", "QA / Warehouse / Production", "Inventory > Lots/Serial Numbers / Scrap / Quality Alert", "full-lot rejection flow", "Lot with detected defect", "Lot, defect note, reject decision", "System stores root cause and can reject or quarantine the whole lot consistently.", "Add lot status, quarantine, and rejection workflow if only partial scrap exists.", True),
            ("MA23", "Record QC instrument log book by tool and analysis topic", "QC Lab", "Maintenance / Quality / Instrument Log", "instrument usage log", "Instrument master exists", "Instrument code and analysis topic", "Users can log tool usage and search history by instrument or test topic.", "Use maintenance log or custom QC instrument log.", True),
            ("MA24", "Pass QC result with analyst and approver audit trail", "Analyst / Approver", "Quality > Test Result / Quality Check", "pass and approval trace", "Pending QC result", "Analyst user and approver user", "System stores who analyzed, who approved, and when the result passed.", "Add analyst and approver fields with transition buttons and audit trail.", True),
            ("MA25", "Enforce routine QC steps by process sequence", "QA / Process Owner", "Quality > Control Points / Manufacturing > Operations", "process-sequenced checks", "A process with multiple inspection points", "Operation steps and control points", "Checks occur in the intended order and do not skip required routine steps.", "Adjust control-point triggers and sequencing or add state gates.", True),
            ("MA26", "Track package sampling with sqrt(n)+1 rule and quarantine state", "QC Packaging", "Quality > Sampling / Inventory > Receipts", "package sampling registration", "Package receipt with multiple boxes", "Package count and sampled boxes", "System records package count, sampled boxes, and quarantine or in-test status.", "Add structured package-sampling worksheet and box-level tracking.", True),
            ("MA27", "Run packaging QC checklist four times per day", "QC Packaging", "Quality > Checklists / Shop Floor QC", "time-based package QC checklist", "Machine and four planned rounds", "07:30 / 10:00 / 14:00 / 16:00 schedule", "System shows four daily checkpoints and captures results and responsible user.", "Use scheduled checklist generation or recurring activities.", True),
            ("MA28", "Issue finished product into system immediately after production handover", "Warehouse / Production Receiving", "Manufacturing > MO / Inventory > Transfers", "FG issue after production", "FG received from production", "FG quantity received and issued", "FG stock updates immediately and output report reflects the handover.", "Add receiving or issue step tied to production completion.", True),
            ("MA29", "Send scale output into system without paper re-entry", "QC / Production", "Scale Interface / Quality Result", "scale-to-system data capture", "Scale output or mock payload", "One weighing result", "Weight data is captured directly in the system with source and timestamp.", "Implement printer or scale integration or import bridge.", True),
        ],
    },
    "MU17": {
        "topic": "Routing / Planning / Split / Backorder / Weekly Scheduling",
        "objective": "Cover planning and routing gaps around routing detail, split or merge decisions, backorders, MPS, weekly MO policy, and overtime simulation.",
        "scope": "routing, parallel printer, split MO, MPS, backorder, merge semi, MTO batch policy, planning order, weekly MO, OT simulation",
        "cases": [
            ("MA31", "Routing must identify machine and realistic duration", "Planner / IE / Production Supervisor", "Manufacturing > BOM > Operations / Work Centers", "routing by machine and time", "BOM with multiple operations", "One BOM and its work centers", "Operations identify the right machine and realistic execution time.", "Maintain routing, work-center capacities, and standard durations.", False),
            ("MA34", "Parallel printer BOM logic must split work correctly", "Planner / Key User", "Manufacturing > BOM > Operations / GMP Shop Floor", "parallel printer split logic", "MO with printer-parallel operation", "One product using parallel printers", "Parallel work orders are created and planned quantity is split across printers.", "Review custom split logic and work-center capacities.", False),
            ("MA35", "Split MO after partial execution when urgent job interrupts", "Planner / Production Manager", "Manufacturing > Manufacturing Orders", "split MO after work started", "MO partly consumed or partly produced", "Original MO plus urgent demand", "Team can decide and document the right split or duplicate path without breaking stock and cost.", "Use split, duplicate, or backorder process consistently.", False),
            ("MA36", "Run MPS from finished goods and push demand downstream", "Planner", "Manufacturing > Planning > Master Production Schedule", "MPS from FG", "FG with complete BOM and routes", "One weekly or monthly plan", "MPS creates or suggests replenishment from finished-goods demand.", "Review planning horizon, lead times, and FG routes.", False),
            ("MA37", "Backorder flow must be complete for transfer and MO", "Warehouse / Planner / Production", "Inventory / Manufacturing partial validation flow", "complete backorder behavior", "Partial transfer and partial MO", "One stock case and one manufacturing case", "Backorder create, no-backorder, and recovery paths all behave consistently.", "Review operation type backorder policy and late recovery custom.", False),
            ("MA38", "Review whether existing semi MO and new demand should merge", "Planner / Production Control", "Manufacturing > Replenishment / Manufacturing Orders", "semi MO merge decision", "Existing semi MO plus new demand", "Semi product with open MO and extra demand", "Team documents whether merge is possible and what policy should be followed.", "Use planning policy or custom merge wizard if standard merge is not available.", False),
            ("MA50", "MTO pharma quantity follows order while plastic follows batch size", "Planner / Production Control", "Sales > SO / Manufacturing > MO", "MTO order quantity versus plastic batch size", "SO with pharma-plastic chain", "SO quantity and plastic batch size", "Pharma MO follows order quantity while plastic MO follows container or batch policy.", "Review batch rounding logic and semi-product routes.", False),
            ("MA53", "Planning should help sequence work from small to large", "Planner / Production Manager", "Manufacturing > Planning / Shop Floor Queue", "planning sequence logic", "Multiple jobs with different quantities", "Small and large MO set", "Planner can compare sequencing strategies and identify the fields needed to make the decision.", "Add sorting, grouping, or planning dashboard if current view is weak.", False),
            ("MA54", "Confirm policy of opening MO weekly", "Planner / Management", "Manufacturing > Manufacturing Orders / Weekly Planning", "weekly MO policy review", "Weekly demand picture", "One week of demand and MO list", "UAT documents whether a weekly MO opening policy is workable.", "Split policy by family if weekly opening does not fit all products.", False),
            ("MA55", "Simulate working hours and OT from production load", "Planner / Production Manager / Finance", "Manufacturing > Planning / Work Center Capacity", "OT simulation from FG load", "Calendar and capacity data available", "One week load on key work centers", "Team can estimate OT need from planned load and expected durations.", "Use exported capacity model if live simulation is still limited.", False),
        ],
    },
    "MU18": {
        "topic": "Shop Floor Control / Close Production / Breakdown / Line Clearance",
        "objective": "Cover shop-floor control requirements still missing from MU: labor by person, manager correction, overproduction, scrap logging, close-production behavior, breakdown, and line clearance.",
        "scope": "employee labor, edit duration, overproduction, scrap draft log, close production, breakdown state, line clearance gate",
        "cases": [
            ("MA32", "Record operator name and labor cost per person", "Operator / Supervisor / HR Costing", "GMP Shop Floor / Manufacturing > Work Orders > Productivity", "labor per person", "Employee master and labor rates exist", "WO with two operators", "System stores operator names and supports person-level labor analysis.", "Extend productivity logging and employee-cost rollup if still aggregated.", False),
            ("MA40", "Manager can correct operation time from MO", "Manager / Key User", "Manufacturing > Manufacturing Orders > Work Orders", "manager time correction", "WO time was entered incorrectly", "MO with incorrect duration", "Manager can correct time with audit trail and updated reports.", "Open manager-only edit path on productivity lines or MO action.", False),
            ("MA41", "Handle overproduction with duplicate MO for excess FG", "Production Control / Warehouse", "Manufacturing > MO > Duplicate / Close Production", "excess FG handling", "Overproduced FG case", "MO with excess quantity", "Team documents the duplicate-MO path for receiving excess FG without breaking stock and cost.", "Define SOP or custom overproduction receiving flow.", False),
            ("MA43", "Scrap on Shop Floor should save draft log before final validation", "Operator / Supervisor / Warehouse", "GMP Shop Floor > Scrap popup", "draft scrap log then validate on close", "WO with scrap entry", "One scrap line", "Scrap can be saved, reviewed, edited, and only finalized at production close.", "Split draft logging from final scrap validation.", False),
            ("MA45", "Close Production should complete all work orders", "Supervisor / Key User", "GMP Shop Floor > Close Production / Manufacturing > MO", "close production closes all WOs", "MO with several WOs", "One MO with unfinished work orders", "Close Production does not leave stray open work orders.", "Force-close or block close until all WOs are handled.", False),
            ("MA46", "Close Production should return user to first Shop Floor page", "Operator / Supervisor", "GMP Shop Floor > Close Production", "close production navigation", "Ready-to-close job", "One WO or MO", "UI returns to queue or dashboard after successful close.", "Add redirect or refresh after close action.", False),
            ("MA47", "Prevent MO close while work orders remain open", "Supervisor / Key User / Accounting", "Manufacturing > MO / Close Production", "guard against wrong consumed qty on close", "MO with open WO", "One MO with mixed WO statuses", "System blocks or resolves open WOs before allowing final close.", "Add validation or force-close sequence with audit note.", False),
            ("MA48", "Work Center breakdown status should affect queue and execution", "Supervisor / Maintenance / Planner", "GMP Shop Floor / Maintenance / Work Center", "breakdown state on work center", "A work center used by active jobs", "One work center and active queue", "Breakdown state is visible and influences planning or execution decisions.", "Add work-center state and queue behavior integration.", False),
            ("MAXX", "Line clearance checkbox must be required before Start", "Operator / Supervisor", "GMP Shop Floor > Work Order Card", "line clearance gate", "Ready job card", "One WO with line-clearance check", "Start remains disabled until line clearance is confirmed and logged.", "Add front-end and back-end validation to prevent bypass.", False),
        ],
    },
    "MU19": {
        "topic": "Production / Accounting / Cost Integration",
        "objective": "Cover production-accounting integration and costing design gaps that remain unmapped in MU.",
        "scope": "real-time accounting review, average versus FIFO, work-center cost, MO total cost across backorders",
        "cases": [
            ("MA33", "Review production-accounting check flow in real time", "Production Controller / Accounting", "Manufacturing > MO / Inventory Valuation / Journal Entries", "production-accounting real-time alignment", "MO done with valuation impact", "One MO with linked valuation and journal entries", "Users can trace MO cost impact into valuation and accounting without exporting everything.", "Add smart buttons or report links if traceability is weak.", False),
            ("MA42", "Compare FIFO and Average costing impact on manufacturing", "Accounting / Key User", "Inventory > Product Categories / Reporting > Valuation", "average versus FIFO review", "Category examples or comparison model", "Receipts with different unit cost", "Accounting can evaluate the effect of FIFO versus Average before changing policy.", "Plan migration and re-test if cost method is changed.", False),
            ("MA51", "Work Center cost should feed MO cost analysis", "IE / Accounting / Production Control", "Manufacturing > Configuration > Work Centers", "work-center cost effect", "Work center cost per hour configured", "MO using that work center", "MO cost analysis reflects the cost configured on the work center.", "Review work-center cost fields, productivity logs, and cost rollup.", False),
            ("MA52", "Show total MO cost across backorder chain", "Accounting / Planner / Key User", "Manufacturing > MO / Backorders / Cost Analysis", "aggregate cost across backorders", "MO with backorder chain", "Original MO and child MO", "Users can see or reconstruct total cost for one job even if split by backorders.", "Add chain-level aggregation report if standard view stays per document.", False),
        ],
    },
    "MU20": {
        "topic": "Reports / Printouts / Post-Production Policy",
        "objective": "Cover print forms and reports still missing from MU, plus the post-production QC policy check.",
        "scope": "post-production QC policy, MO print, WO print, scrap report, machine report, production report",
        "cases": [
            ("MA44", "Confirm no QC step is required after post-production", "QA / Production Manager", "Manufacturing > MO > Quality Checks", "post-production QC policy", "MO with related quality checks", "One completed MO", "UAT confirms whether post-production QC is removed as planned.", "Disable or move control points if the policy changes.", False),
            ("RP41", "Print and review Manufacturing Order form", "Production Admin / Planner", "Manufacturing > MO > Print", "MO print form", "One MO ready to print", "Printed MO sample", "Printed MO form includes the information needed on the shop floor.", "Adjust QWeb report or print action if fields are missing.", False),
            ("RP42", "Print and review Work Order form", "Supervisor / Operator Lead", "Manufacturing > Work Orders > Print", "WO print form", "One WO ready to print", "Printed WO sample", "Printed WO form includes operation, machine, instructions, and quantity details.", "Adjust report template and WO source fields.", False),
            ("RP68", "Review Scrap report output", "Production / QA / Accounting", "Inventory / Manufacturing > Reporting > Scrap", "scrap report", "Existing scrap transactions", "One date range or product family", "Scrap report can be filtered, reviewed, and exported for business use.", "Create pivot, saved filters, or custom report layout if needed.", False),
            ("RP69", "Review Machine report output", "Maintenance / Production Manager", "Manufacturing > Reporting / Maintenance", "machine report", "Existing machine usage or downtime data", "One work center over one month", "Machine report shows usable runtime, downtime, and productivity indicators.", "Build combined WO + maintenance report if current view is too weak.", False),
            ("RP70", "Review Production report output by month and product", "Management / Planner / Production Admin", "Manufacturing > Reporting > Production Analysis", "production report", "Done MO over a sample period", "Monthly grouping by FG", "Production report can summarize monthly output and drill down to source MO.", "Add saved pivot or dashboard for management view.", False),
        ],
    },
}


def load_backlog(wb):
    ws = wb["Product_backlog"]
    data = {}
    for r in range(2, ws.max_row + 1):
        backlog_id = ws.cell(r, 1).value
        if backlog_id:
            data[backlog_id] = {
                "phase": ws.cell(r, 4).value or "",
                "owner": ws.cell(r, 12).value or "",
                "asis": ws.cell(r, 5).value or "",
                "tobe": ws.cell(r, 6).value or "",
            }
    return data


def clone_sheet(wb, template_name: str, new_name: str):
    ws = wb.copy_worksheet(wb[template_name])
    ws.title = new_name
    for r in range(7, 1001):
        for c in range(1, 20):
            ws.cell(r, c).value = None
    return ws


def copy_row_style(src_ws, src_row: int, dst_ws, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def refresh_summary(wb, evidence_case_ids: set[str]):
    for name in ["LOCAL_UAT_SUMMARY", "LOCAL_UAT_COVERAGE"]:
        if name in wb.sheetnames:
            del wb[name]

    pb = wb["Product_backlog"]
    map_ws = wb["_MANU_UAT_MAP"]
    backlog_rows = []
    for r in range(2, pb.max_row + 1):
        bid = pb.cell(r, 1).value
        if bid:
            backlog_rows.append((bid, pb.cell(r, 5).value or ""))
    mapped = {map_ws.cell(r, 1).value for r in range(2, map_ws.max_row + 1) if map_ws.cell(r, 1).value}
    case_ids = {map_ws.cell(r, 2).value for r in range(2, map_ws.max_row + 1) if map_ws.cell(r, 2).value}

    ws = wb.create_sheet("LOCAL_UAT_SUMMARY")
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    rows = [
        ("Total Product_backlog items", len(backlog_rows)),
        ("Backlog items linked to MU test cases", len(mapped)),
        ("Backlog items not linked to any MU case", len([1 for bid, _ in backlog_rows if bid not in mapped])),
        ("Total MU test cases in workbook", len(case_ids)),
        ("MU test cases with local UAT evidence", len(evidence_case_ids)),
        ("Coverage of Product_backlog by MU design (%)", round(len(mapped) * 100 / len(backlog_rows), 2) if backlog_rows else 0),
        ("Coverage of MU cases by local evidence (%)", round(len(evidence_case_ids) * 100 / len(case_ids), 2) if case_ids else 0),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(i, 1, label)
        ws.cell(i, 2, value)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 18

    cov = wb.create_sheet("LOCAL_UAT_COVERAGE")
    headers = ["Backlog ID", "Mapped to MU?", "Mapped MU Count", "Observed Local UAT Evidence?", "Backlog Summary"]
    for c, header in enumerate(headers, start=1):
        cov.cell(1, c, header)
    out_row = 2
    for bid, summary in backlog_rows:
        mapped_rows = [r for r in range(2, map_ws.max_row + 1) if map_ws.cell(r, 1).value == bid]
        evidence = "Yes" if any(map_ws.cell(r, 2).value in evidence_case_ids for r in mapped_rows) else "No"
        cov.cell(out_row, 1, bid)
        cov.cell(out_row, 2, "Yes" if mapped_rows else "No")
        cov.cell(out_row, 3, len(mapped_rows))
        cov.cell(out_row, 4, evidence)
        cov.cell(out_row, 5, summary)
        out_row += 1
    for col, width in {"A": 12, "B": 14, "C": 16, "D": 24, "E": 120}.items():
        cov.column_dimensions[col].width = width


def main():
    wb = openpyxl.load_workbook(BASE)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    backlog = load_backlog(wb)
    evidence_case_ids = {
        "MU01-01", "MU01-02", "MU01-03",
        "MU04-03", "MU07-03", "MU08-01",
        "MU09-01", "MU09-02", "MU09-03",
        "MU11-04",
        "MU14-01", "MU14-02", "MU14-03", "MU14-04", "MU14-05", "MU14-06", "MU14-07",
    }

    for name in list(wb.sheetnames):
        if name.startswith("MU") and name[2:].isdigit() and int(name[2:]) >= 15:
            del wb[name]

    template_sheet = wb["MU01"]
    map_ws = wb["_MANU_UAT_MAP"]
    sc_ws = wb["Scenarios"]
    map_last = max(r for r in range(1, map_ws.max_row + 1) if any(map_ws.cell(r, c).value is not None for c in range(1, 8)))
    sc_last = max(r for r in range(1, sc_ws.max_row + 1) if any(sc_ws.cell(r, c).value is not None for c in range(1, 14)))
    scenario_no = max(sc_ws.cell(r, 1).value for r in range(1, sc_ws.max_row + 1) if isinstance(sc_ws.cell(r, 1).value, int))

    new_case_count = 0
    for sheet_name, meta in GROUPS.items():
        ws = clone_sheet(wb, "MU01", sheet_name)
        ws["A1"] = f"Manufacturing UAT Test Step - {sheet_name}"
        ws["B2"] = meta["topic"]
        ws["B3"] = meta["objective"]
        ws["B4"] = meta["scope"]
        for idx, case in enumerate(meta["cases"], start=1):
            backlog_id, title, role, menu, focus, pre, data, expected, fix, future = case
            row = 6 + idx
            copy_row_style(template_sheet, 7, ws, row, 19)
            ws.cell(row, 1, idx)
            ws.cell(row, 2, f"{sheet_name}-{idx:02d}")
            ws.cell(row, 3, backlog_id)
            ws.cell(row, 4, title)
            ws.cell(row, 5, role)
            ws.cell(row, 6, menu)
            ws.cell(row, 7, pre)
            ws.cell(row, 8, data)
            ws.cell(row, 9, case_steps(menu, focus))
            ws.cell(row, 10, expected)
            ws.cell(row, 11, fix)
            ws.cell(row, 12, f"Capture the main screen, changed record, related document, and report/export proving {focus}.")
            ws.cell(row, 13, "")
            ws.cell(row, 14, "")
            ws.cell(row, 15, "")
            ws.cell(row, 18, overall_formula(row))
            phase = backlog.get(backlog_id, {}).get("phase", "")
            owner = backlog.get(backlog_id, {}).get("owner", "")
            detail = backlog.get(backlog_id, {}).get("tobe") or backlog.get(backlog_id, {}).get("asis") or ""
            trainer = f"{'Later Phase / Future Scope' if future else 'Designed only'} | Phase: {phase} | Owner: {owner} | Requirement: {detail}"
            ws.cell(row, 19, trainer)
            ws.cell(row, 19).fill = FILL_FUTURE if future else FILL_DESIGN

            map_last += 1
            copy_row_style(map_ws, 2, map_ws, map_last, 7)
            map_ws.cell(map_last, 1, backlog_id)
            map_ws.cell(map_last, 2, f"{sheet_name}-{idx:02d}")
            map_ws.cell(map_last, 3, sheet_name)
            map_ws.cell(map_last, 4, row)
            map_ws.cell(map_last, 5, f"='{sheet_name}'!P{row}")
            map_ws.cell(map_last, 6, f"='{sheet_name}'!Q{row}")
            map_ws.cell(map_last, 7, f"='{sheet_name}'!R{row}")

            sc_last += 1
            scenario_no += 1
            copy_row_style(sc_ws, 104, sc_ws, sc_last, 13)
            sc_ws.cell(sc_last, 1, scenario_no)
            for offset, value in enumerate(scenario_refs(sheet_name, row), start=2):
                sc_ws.cell(sc_last, offset, value)
            new_case_count += 1

    refresh_summary(wb, evidence_case_ids)
    wb.save(OUT)

    mapped = {wb["_MANU_UAT_MAP"].cell(r, 1).value for r in range(2, wb["_MANU_UAT_MAP"].max_row + 1) if wb["_MANU_UAT_MAP"].cell(r, 1).value}
    backlog_ids = {wb["Product_backlog"].cell(r, 1).value for r in range(2, wb["Product_backlog"].max_row + 1) if wb["Product_backlog"].cell(r, 1).value}
    remaining = sorted(backlog_ids - mapped)
    lines = [
        "# Local UAT Manufacturing Full Backlog Coverage",
        "",
        f"- Source workbook: `{BASE}`",
        f"- Output workbook: `{OUT}`",
        f"- New MU sheets added: {', '.join(GROUPS.keys())}",
        f"- New MU cases added: `{new_case_count}`",
        f"- Total backlog IDs: `{len(backlog_ids)}`",
        f"- Backlog IDs mapped after expansion: `{len(mapped)}`",
        f"- Remaining unmapped backlog IDs: `{len(remaining)}`",
        "",
        "## Added MU Themes",
    ]
    for key, meta in GROUPS.items():
        lines.append(f"- `{key}`: {meta['topic']} ({len(meta['cases'])} cases)")
    if remaining:
        lines.extend(["", "## Remaining Unmapped IDs", ", ".join(remaining)])
    else:
        lines.extend(["", "## Remaining Unmapped IDs", "- None"])
    SUMMARY.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
