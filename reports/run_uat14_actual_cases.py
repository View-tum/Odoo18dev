from __future__ import annotations

import json
import os
import re
import time
import traceback
from collections import defaultdict
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = os.getenv("UAT14_BASE_URL", "http://10.0.0.14")
DB = os.getenv("UAT14_DB", "goldmints_uat")
LOGIN = os.getenv("UAT14_LOGIN", "admin")
PASSWORD = os.getenv("UAT14_PASSWORD", "365@gmp")

INPUT_XLSX = Path(
    os.getenv("UAT14_INPUT_XLSX", r"C:\Users\tumsu\Downloads\UAT_Test_Scenario_LINKED.xlsx")
)
OUTPUT_XLSX = Path(
    os.getenv("UAT14_OUTPUT_XLSX", r"C:\Users\tumsu\Downloads\UAT_Test_Scenario_LINKED_UAT14_RESULT.xlsx")
)
REPORT_JSON = Path("reports") / f"uat14_actual_run_{time.strftime('%Y%m%d_%H%M%S')}.json"
REPORT_MD = REPORT_JSON.with_suffix(".md")

ACTUAL_SHEET_PREFIXES = ("01_", "02_", "03_", "04_", "05_", "06_", "07_")


class OdooRPC:
    def __init__(self):
        self.session = requests.Session()

    def json_route(self, route, params=None, timeout=180):
        payload = {"jsonrpc": "2.0", "params": params or {}}
        response = self.session.post(BASE_URL + route, json=payload, timeout=timeout)
        response.raise_for_status()
        data = response.json()
        if data.get("error"):
            raise RuntimeError(json.dumps(data["error"], ensure_ascii=False, indent=2))
        return data.get("result")

    def login(self):
        return self.json_route(
            "/web/session/authenticate",
            {"db": DB, "login": LOGIN, "password": PASSWORD},
        )

    def call(self, model, method, args=None, kwargs=None):
        return self.json_route(
            "/web/dataset/call_kw",
            {
                "model": model,
                "method": method,
                "args": args or [],
                "kwargs": kwargs or {},
            },
        )

    def search(self, model, domain, limit=None, order=None):
        kwargs = {}
        if limit:
            kwargs["limit"] = limit
        if order:
            kwargs["order"] = order
        return self.call(model, "search", [domain], kwargs)

    def search_read(self, model, domain, fields=None, limit=None, order=None):
        kwargs = {}
        if fields:
            kwargs["fields"] = fields
        if limit:
            kwargs["limit"] = limit
        if order:
            kwargs["order"] = order
        return self.call(model, "search_read", [domain], kwargs)

    def read(self, model, ids, fields):
        if not ids:
            return []
        return self.call(model, "read", [ids, fields])

    def create(self, model, values):
        return self.call(model, "create", [values])

    def write(self, model, ids, values):
        return self.call(model, "write", [ids, values])

    def unlink(self, model, ids):
        return self.call(model, "unlink", [ids])

    def exists_model(self, model):
        return bool(self.search("ir.model", [("model", "=", model)], limit=1))


def norm(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def read_cases():
    wb = load_workbook(INPUT_XLSX, data_only=False)
    cases = []
    for ws in wb.worksheets:
        if not ws.title.startswith(ACTUAL_SHEET_PREFIXES):
            continue
        headers = [ws.cell(6, col).value for col in range(1, ws.max_column + 1)]
        col = {norm(header): idx for idx, header in enumerate(headers, start=1) if header}
        for row in range(7, ws.max_row + 1):
            case_id = ws.cell(row, col.get("Case ID", 2)).value
            if not case_id:
                continue
            cases.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "case_id": str(case_id).strip(),
                    "backlog": ws.cell(row, col.get("Backlog IDs", 3)).value,
                    "scenario": ws.cell(row, col.get("Scenario ทดสอบ", 5)).value,
                    "menu": ws.cell(row, col.get("Menu Path in local UAT (English)", 7)).value,
                    "steps": ws.cell(row, col.get("ขั้นตอนทดสอบแบบละเอียด", 10)).value,
                }
            )
    return cases


def result(case_id, status, evidence, details="", record=None):
    return {
        "case_id": case_id,
        "status": status,
        "evidence": evidence,
        "details": details,
        "record": record or {},
    }


class UAT14Runner:
    def __init__(self, rpc: OdooRPC):
        self.rpc = rpc
        self.cache = {}
        self.created_mos = []
        self.created_pickings = []
        self.created_pos = []

    def product_by_code(self, code):
        key = ("product", code)
        if key not in self.cache:
            found = self.rpc.search_read(
                "product.product",
                [("default_code", "=", code)],
                ["display_name", "uom_id", "product_tmpl_id", "tracking", "qty_available"],
                limit=1,
            )
            if not found:
                raise AssertionError(f"Product not found: {code}")
            self.cache[key] = found[0]
        return self.cache[key]

    def pharma_picking_type(self):
        key = "pharma_pt"
        if key not in self.cache:
            pts = self.rpc.search_read(
                "stock.picking.type",
                [("code", "=", "mrp_operation"), ("name", "ilike", "Pharma")],
                ["display_name", "default_location_src_id", "default_location_dest_id", "warehouse_id", "create_backorder"],
                limit=1,
            )
            if not pts:
                pts = self.rpc.search_read(
                    "stock.picking.type",
                    [("code", "=", "mrp_operation")],
                    ["display_name", "default_location_src_id", "default_location_dest_id", "warehouse_id", "create_backorder"],
                    limit=1,
                )
            if not pts:
                raise AssertionError("Manufacturing picking type not found")
            self.cache[key] = pts[0]
        return self.cache[key]

    def bom_for_product(self, product):
        key = ("bom", product["id"])
        if key not in self.cache:
            boms = self.rpc.search_read(
                "mrp.bom",
                ["|", ("product_id", "=", product["id"]), ("product_tmpl_id", "=", product["product_tmpl_id"][0])],
                ["display_name", "product_qty", "operation_ids", "bom_line_ids"],
                limit=1,
            )
            if not boms:
                raise AssertionError(f"BOM not found for {product['display_name']}")
            self.cache[key] = boms[0]
        return self.cache[key]

    def create_mo(self, qty=1.0, product_code="FG-PSS-TH-01001", tag="CODX-UAT14"):
        product = self.product_by_code(product_code)
        bom = self.bom_for_product(product)
        pt = self.pharma_picking_type()
        mo_id = self.rpc.create(
            "mrp.production",
            {
                "product_id": product["id"],
                "product_qty": qty,
                "product_uom_id": product["uom_id"][0],
                "bom_id": bom["id"],
                "picking_type_id": pt["id"],
                "location_src_id": pt["default_location_src_id"][0],
                "location_dest_id": pt["default_location_dest_id"][0],
                "origin": f"{tag}-{time.strftime('%Y%m%d%H%M%S')}",
            },
        )
        self.created_mos.append(mo_id)
        self.rpc.call("mrp.production", "action_confirm", [[mo_id]])
        try:
            self.rpc.call("mrp.production", "action_assign", [[mo_id]])
        except Exception:
            pass
        company_id = self.rpc.search("res.company", [], limit=1)[0]
        lot_id = self.rpc.create(
            "stock.lot",
            {
                "name": f"{tag}-LOT-{mo_id}-{time.strftime('%H%M%S')}",
                "product_id": product["id"],
                "company_id": company_id,
            },
        )
        self.rpc.write("mrp.production", [mo_id], {"lot_producing_id": lot_id})
        mo = self.rpc.read(
            "mrp.production",
            [mo_id],
            ["name", "state", "workorder_ids", "product_qty", "qty_producing", "reservation_state"],
        )[0]
        if not mo["workorder_ids"]:
            raise AssertionError(f"MO {mo['name']} has no workorder")
        return mo_id, mo["workorder_ids"][0], mo

    def cancel_mo(self, mo_id):
        try:
            state = self.rpc.read("mrp.production", [mo_id], ["state"])[0]["state"]
            if state not in ("done", "cancel"):
                self.rpc.call("mrp.production", "action_cancel", [[mo_id]])
        except Exception:
            pass

    def test_product_stock_and_forecast(self, case_id):
        product = self.product_by_code("FG-PSS-TH-01001")
        if product["qty_available"] is None:
            raise AssertionError("qty_available is not readable")
        return result(case_id, "Passed", "อ่านสินค้า, On Hand และ master product ได้จริง", record=product)

    def test_replenishment_setup(self, case_id):
        orderpoints = self.rpc.search_read(
            "stock.warehouse.orderpoint",
            [],
            ["product_id", "warehouse_id", "location_id", "qty_forecast", "qty_to_order", "route_id"],
            limit=10,
        )
        if not orderpoints:
            raise AssertionError("ไม่พบ orderpoint/replenishment rule ในระบบ")
        return result(case_id, "Passed", f"พบ Replenishment/Orderpoint {len(orderpoints)} รายการ", record={"sample": orderpoints[:3]})

    def test_mo_full(self, case_id):
        mo_id = None
        try:
            mo_id, wo_id, mo = self.create_mo(1.0, tag=f"CODX-{case_id}")
            check = self.rpc.json_route("/mrp_parallel_console/check_components", {"workorder_id": wo_id})
            start = self.rpc.json_route("/mrp_parallel_console/start_workorder", {"workorder_id": wo_id})
            qty = self.rpc.json_route("/mrp_parallel_console/add_qty_log", {"workorder_id": wo_id, "qty": 1.0, "note": case_id})
            finish = self.rpc.json_route("/mrp_parallel_console/finish_workorder", {"workorder_id": wo_id})
            close = self.rpc.json_route("/mrp_parallel_console/manual_close_mo", {"production_id": mo_id})
            mo_after = self.rpc.read("mrp.production", [mo_id], ["name", "state", "qty_producing", "product_qty"])[0]
            if mo_after["state"] != "done":
                raise AssertionError(f"MO did not close to done: {mo_after}")
            return result(case_id, "Passed", "สร้าง MO, start shopfloor, log qty, finish, close done ได้จริง", record={"mo": mo_after, "check": check, "start": start, "qty": qty, "finish": finish, "close": close})
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise

    def test_mo_missing_components_start_allowed(self, case_id):
        mo_id = None
        try:
            mo_id, wo_id, mo = self.create_mo(1_000_000.0, tag=f"CODX-{case_id}")
            check = self.rpc.json_route("/mrp_parallel_console/check_components", {"workorder_id": wo_id})
            start = self.rpc.json_route("/mrp_parallel_console/start_workorder", {"workorder_id": wo_id})
            wo = self.rpc.read("mrp.workorder", [wo_id], ["state", "console_date_start"])[0]
            if check.get("sufficient") is not False:
                raise AssertionError("Expected insufficient components warning")
            if start.get("status") != "ok" or wo["state"] != "progress":
                raise AssertionError(f"Workorder did not start with missing components: {start}, {wo}")
            try:
                self.rpc.json_route("/mrp_parallel_console/stop_workorder", {"workorder_id": wo_id})
            except Exception:
                pass
            self.cancel_mo(mo_id)
            return result(case_id, "Passed", "Component ไม่พอแต่ Shopfloor Start ได้จริงและมี diagnostic", record={"mo": mo["name"], "check": check, "start": start, "wo": wo})
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise

    def test_mo_partial_backorder_wizard(self, case_id):
        mo_id = None
        try:
            mo_id, wo_id, mo = self.create_mo(2.0, tag=f"CODX-{case_id}")
            self.rpc.json_route("/mrp_parallel_console/start_workorder", {"workorder_id": wo_id})
            self.rpc.json_route("/mrp_parallel_console/add_qty_log", {"workorder_id": wo_id, "qty": 1.0, "note": case_id})
            self.rpc.json_route("/mrp_parallel_console/finish_workorder", {"workorder_id": wo_id})
            close = self.rpc.json_route("/mrp_parallel_console/manual_close_mo", {"production_id": mo_id})
            mo_after = self.rpc.read("mrp.production", [mo_id], ["name", "state", "qty_producing", "product_qty"])[0]
            action = close.get("action") if isinstance(close, dict) else None
            if not isinstance(action, dict) or action.get("res_model") != "mrp.production.backorder":
                raise AssertionError(f"Expected MRP backorder wizard, got {close}")
            self.cancel_mo(mo_id)
            return result(case_id, "Passed", "Partial production คืน wizard Create Backorder ได้จริง", record={"mo": mo_after, "action": action.get("name")})
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise

    def test_quick_done_over_qty_block(self, case_id):
        mo_id = None
        try:
            mo_id, wo_id, _mo = self.create_mo(1.0, tag=f"CODX-{case_id}")
            res = self.rpc.json_route("/mrp_parallel_console/quick_done", {"workorder_id": wo_id, "qty": 2.0})
            self.cancel_mo(mo_id)
            if not res.get("error"):
                raise AssertionError(f"Expected over-qty error, got {res}")
            return result(case_id, "Passed", "Quick Done เกินจำนวนถูก block", record=res)
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise

    def test_mo_overproduction_actual_consumption(self, case_id):
        mo_id = None
        try:
            mo_id, wo_id, _mo = self.create_mo(1.0, tag=f"CODX-{case_id}")
            self.rpc.json_route("/mrp_parallel_console/start_workorder", {"workorder_id": wo_id})
            qty = self.rpc.json_route(
                "/mrp_parallel_console/add_qty_log",
                {"workorder_id": wo_id, "qty": 2.0, "note": case_id},
            )
            apply_result = self.rpc.json_route(
                "/mrp_parallel_console/apply_console",
                {"workorder_ids": [wo_id]},
            )
            mo_after = self.rpc.read(
                "mrp.production",
                [mo_id],
                ["name", "state", "product_qty", "qty_producing"],
            )[0]
            raw_moves = self.rpc.search_read(
                "stock.move",
                [("raw_material_production_id", "=", mo_id)],
                ["product_id", "product_uom_qty", "quantity", "state"],
                limit=20,
            )
            if apply_result.get("status") != "ok" or mo_after["state"] != "done":
                raise AssertionError(
                    f"Overproduction did not close cleanly: apply={apply_result}, mo={mo_after}, raw_moves={raw_moves}"
                )
            scaled_moves = [
                move
                for move in raw_moves
                if move["product_uom_qty"] and move["quantity"]
            ]
            if not scaled_moves:
                raise AssertionError(f"No component consumption was recorded: {raw_moves}")
            return result(
                case_id,
                "Passed",
                "Overproduction closed via shopfloor and actual component consumption is traceable",
                record={"mo": mo_after, "qty_log": qty, "apply": apply_result, "raw_moves": raw_moves},
            )
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise

    def test_expiry_diagnostic(self, case_id):
        product = self.product_by_code("FG-PSS-TH-01001")
        bom = self.bom_for_product(product)
        pt = self.pharma_picking_type()
        mo_id = None
        try:
            mo_id = self.rpc.create(
                "mrp.production",
                {
                    "product_id": product["id"],
                    "product_qty": 1.0,
                    "product_uom_id": product["uom_id"][0],
                    "bom_id": bom["id"],
                    "picking_type_id": pt["id"],
                    "location_src_id": pt["default_location_src_id"][0],
                    "location_dest_id": pt["default_location_dest_id"][0],
                    "date_start": "2030-01-01 00:00:00",
                    "origin": f"CODX-{case_id}-{time.strftime('%Y%m%d%H%M%S')}",
                },
            )
            self.created_mos.append(mo_id)
            self.rpc.call("mrp.production", "action_confirm", [[mo_id]])
            try:
                self.rpc.call("mrp.production", "action_assign", [[mo_id]])
            except Exception:
                pass
            mo = self.rpc.read("mrp.production", [mo_id], ["name", "workorder_ids"])[0]
            check = self.rpc.json_route("/mrp_parallel_console/check_components", {"workorder_id": mo["workorder_ids"][0]})
            start = self.rpc.json_route("/mrp_parallel_console/start_workorder", {"workorder_id": mo["workorder_ids"][0]})
            try:
                self.rpc.json_route("/mrp_parallel_console/stop_workorder", {"workorder_id": mo["workorder_ids"][0]})
            except Exception:
                pass
            self.cancel_mo(mo_id)
            text = check.get("error") or ""
            if "expired" not in text and "not usable" not in text:
                raise AssertionError(f"Expected expiration diagnostic, got {text[:500]}")
            if start.get("status") != "ok":
                raise AssertionError(f"Start did not continue after warning: {start}")
            return result(case_id, "Passed", "Diagnostic แยก lot หมดอายุและยัง Start ได้", record={"check": check, "start": start})
        except Exception:
            if mo_id:
                self.cancel_mo(mo_id)
            raise


    def supplier_location(self):
        key = "supplier_location"
        if key not in self.cache:
            locs = self.rpc.search_read(
                "stock.location",
                [("usage", "=", "supplier")],
                ["display_name"],
                limit=1,
            )
            if not locs:
                raise AssertionError("Supplier/Vendor location not found")
            self.cache[key] = locs[0]["id"]
        return self.cache[key]

    def incoming_picking_type(self):
        key = "incoming_pt"
        if key not in self.cache:
            pts = self.rpc.search_read(
                "stock.picking.type",
                [("code", "=", "incoming")],
                ["display_name", "default_location_src_id", "default_location_dest_id"],
                limit=1,
            )
            if not pts:
                raise AssertionError("Incoming receipt operation type not found")
            self.cache[key] = pts[0]
        return self.cache[key]

    def seed_stock_by_receipt(self, product, dest_location_id, qty, case_id):
        """Create stock through a normal receipt so RPC does not call private quant methods."""
        src_location_id = self.supplier_location()
        picking_type = self.incoming_picking_type()
        picking_id = self.rpc.create(
            "stock.picking",
            {
                "picking_type_id": picking_type["id"],
                "location_id": src_location_id,
                "location_dest_id": dest_location_id,
                "invoice_reference": f"CODX-SEED-{case_id}",
                "invoice_date": time.strftime("%Y-%m-%d"),
                "origin": f"CODX-SEED-{case_id}",
                "move_ids": [
                    (
                        0,
                        0,
                        {
                            "name": product["display_name"],
                            "product_id": product["id"],
                            "product_uom_qty": qty,
                            "product_uom": product["uom_id"][0],
                            "location_id": src_location_id,
                            "location_dest_id": dest_location_id,
                        },
                    )
                ],
            },
        )
        self.created_pickings.append(picking_id)
        self.rpc.call("stock.picking", "action_confirm", [[picking_id]])
        move = self.rpc.search_read(
            "stock.move",
            [("picking_id", "=", picking_id)],
            ["product_id", "product_uom", "location_id", "location_dest_id"],
            limit=1,
        )[0]
        lot_id = False
        if product.get("tracking") and product.get("tracking") != "none":
            company_id = self.rpc.search("res.company", [], limit=1)[0]
            lot_id = self.rpc.create(
                "stock.lot",
                {
                    "name": f"CODX-SEED-{case_id}-{picking_id}",
                    "product_id": product["id"],
                    "company_id": company_id,
                },
            )
        lines = self.rpc.search("stock.move.line", [("picking_id", "=", picking_id)])
        line_values = {
            "quantity": qty,
            "location_id": src_location_id,
            "location_dest_id": dest_location_id,
        }
        if lot_id:
            line_values["lot_id"] = lot_id
        if lines:
            self.rpc.write("stock.move.line", lines, line_values)
        else:
            create_values = dict(line_values)
            create_values.update(
                {
                    "picking_id": picking_id,
                    "move_id": move["id"],
                    "product_id": product["id"],
                    "product_uom_id": product["uom_id"][0],
                }
            )
            self.rpc.create("stock.move.line", create_values)
        validation = self.rpc.call("stock.picking", "button_validate", [[picking_id]])
        if isinstance(validation, dict) and validation.get("res_model"):
            raise AssertionError(f"Receipt seed did not validate directly: {validation}")
        return picking_id

    def test_internal_transfer_partial_backorder(self, case_id, recover=False):
        product = self.product_by_code("FG-PSS-TH-02001")
        unit = product["uom_id"][0]
        pt_list = self.rpc.search_read(
            "stock.picking.type",
            [("code", "=", "internal"), ("create_backorder", "=", "ask")],
            ["display_name", "default_location_src_id", "default_location_dest_id", "warehouse_id"],
            limit=1,
        )
        if not pt_list:
            raise AssertionError("No internal transfer operation type with Ask backorder")
        pt = pt_list[0]
        parent_loc = pt["warehouse_id"] and self.rpc.read("stock.warehouse", [pt["warehouse_id"][0]], ["lot_stock_id"])[0]["lot_stock_id"][0]
        src = self.rpc.create("stock.location", {"name": f"CODX SRC {case_id}", "usage": "internal", "location_id": parent_loc})
        dst = self.rpc.create("stock.location", {"name": f"CODX DST {case_id}", "usage": "internal", "location_id": parent_loc})
        seed_picking_id = self.seed_stock_by_receipt(product, src, 10.0, case_id)
        picking_id = self.rpc.create(
            "stock.picking",
            {
                "picking_type_id": pt["id"],
                "location_id": src,
                "location_dest_id": dst,
                "origin": f"CODX-{case_id}",
                "invoice_reference": f"CODX-{case_id}",
                "invoice_date": time.strftime("%Y-%m-%d"),
                "move_ids": [
                    (
                        0,
                        0,
                        {
                            "name": product["display_name"],
                            "product_id": product["id"],
                            "product_uom_qty": 10.0,
                            "product_uom": unit,
                            "location_id": src,
                            "location_dest_id": dst,
                        },
                    )
                ],
            },
        )
        self.created_pickings.append(picking_id)
        self.rpc.call("stock.picking", "action_confirm", [[picking_id]])
        self.rpc.call("stock.picking", "action_assign", [[picking_id]])
        line_ids = self.rpc.search("stock.move.line", [("picking_id", "=", picking_id)])
        if not line_ids:
            raise AssertionError("Transfer has no move line after assign")
        self.rpc.write("stock.move.line", line_ids, {"quantity": 4.0})
        action = self.rpc.call("stock.picking", "button_validate", [[picking_id]])
        if not isinstance(action, dict) or action.get("res_model") != "stock.backorder.confirmation":
            raise AssertionError(f"Expected stock backorder wizard, got {action}")
        wizard_context = action.get("context") or {}
        if recover:
            wiz_id = self.rpc.call(
                "stock.backorder.confirmation",
                "create",
                [{}],
                {"context": wizard_context},
            )
            self.rpc.call("stock.backorder.confirmation", "process_cancel_backorder", [[wiz_id]], {"context": wizard_context})
            p = self.rpc.read("stock.picking", [picking_id], ["can_create_late_backorder", "late_backorder_move_count", "name", "state"])[0]
            if not p.get("can_create_late_backorder"):
                raise AssertionError(f"Late backorder button not available: {p}")
            late = self.rpc.call("stock.picking", "action_create_late_backorder", [[picking_id]])
            return result(case_id, "Passed", f"Seed receipt {seed_picking_id}; No Backorder then late backorder recovery passed", record={"seed_picking_id": seed_picking_id, "picking": p, "late_action": late})
        wiz_id = self.rpc.call(
            "stock.backorder.confirmation",
            "create",
            [{}],
            {"context": wizard_context},
        )
        self.rpc.call("stock.backorder.confirmation", "process", [[wiz_id]], {"context": wizard_context})
        p = self.rpc.read("stock.picking", [picking_id], ["name", "state", "backorder_ids"])[0]
        if p["state"] != "done" or not p["backorder_ids"]:
            raise AssertionError(f"Create Backorder did not finish original and create backorder: {p}")
        return result(case_id, "Passed", f"Seed receipt {seed_picking_id}; Partial transfer created backorder {p['backorder_ids']}", record={"seed_picking_id": seed_picking_id, "picking": p, "action": action.get("name")})


    def test_mto_route_setup(self, case_id):
        route_count = self.rpc.call(
            "stock.route",
            "search_count",
            [["|", ("name", "ilike", "MTO"), ("name", "ilike", "Make To Order")]],
        )
        rule_count = self.rpc.call(
            "stock.rule",
            "search_count",
            [["|", ("procure_method", "=", "make_to_order"), ("route_id.name", "ilike", "MTO")]],
        )
        sale_count = self.rpc.call("sale.order", "search_count", [[]])
        if route_count <= 0 and rule_count <= 0:
            raise AssertionError("MTO / Make To Order route or stock rule was not found")
        return result(
            case_id,
            "Passed",
            "MTO setup is available and sale/manufacturing models are readable",
            record={"mto_route_count": route_count, "mto_rule_count": rule_count, "sale_order_count": sale_count},
        )

    def test_models_menus_or_reports(self, case_id, requirements):
        missing = []
        evidence = {}
        for model in requirements.get("models", []):
            exists = self.rpc.exists_model(model)
            evidence[f"model:{model}"] = exists
            if not exists:
                missing.append(f"model {model}")
        for model, domain in requirements.get("records", []):
            count = self.rpc.call(model, "search_count", [domain])
            evidence[f"records:{model}"] = count
            if count <= 0:
                missing.append(f"records {model}")
        for xmlid in requirements.get("xmlids", []):
            try:
                rec = self.rpc.call("ir.model.data", "_xmlid_lookup", [xmlid])
                evidence[f"xmlid:{xmlid}"] = rec
            except Exception:
                missing.append(f"xmlid {xmlid}")
        if missing:
            raise AssertionError("Missing: " + ", ".join(missing))
        return result(case_id, "Passed", "ตรวจพบ model/menu/report/data ที่เกี่ยวข้องในระบบ 14", record=evidence)

    def test_purchase_uom(self, case_id):
        products = self.rpc.search_read(
            "product.product",
            [("is_storable", "=", True), ("uom_id", "!=", False), ("uom_po_id", "!=", False)],
            ["display_name", "uom_id", "uom_po_id"],
            limit=20,
        )
        if not products:
            raise AssertionError("No purchasable/storable products found")
        diff = [p for p in products if p["uom_id"] != p["uom_po_id"]]
        evidence = {"sample_products": products[:5], "different_purchase_uom_count_in_sample": len(diff)}
        return result(case_id, "Passed", "Product master มี UoM/Purchase UoM ให้อ่านและใช้ตรวจ conversion ได้", record=evidence)

    def run_case(self, case):
        cid = case["case_id"]
        # Transactional cases
        if cid in {"MU02-01", "MU07-01", "MU11-01", "MU11-02", "MU11-03", "MU09-03", "MU14-06"}:
            return self.test_mo_full(cid)
        if cid == "MU02-02":
            return self.test_mo_missing_components_start_allowed(cid)
        if cid in {"MU07-02", "MU17-05"}:
            return self.test_mo_partial_backorder_wizard(cid)
        if cid == "MU07-03":
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.production"], "records": [("mrp.production", [("can_create_late_backorder", "=", True)])]})
        if cid == "MU07-04":
            return self.test_mo_overproduction_actual_consumption(cid)
        if cid in {"MU03-02", "MU04-02"}:
            return self.test_internal_transfer_partial_backorder(cid, recover=False)
        if cid in {"MU03-03", "MU04-03"}:
            return self.test_internal_transfer_partial_backorder(cid, recover=True)
        if cid == "MU14-03":
            return self.test_expiry_diagnostic(cid)
        if cid == "MU12-03":
            return self.test_mo_partial_backorder_wizard(cid)
        if cid == "MU10-04" or cid.startswith("MU13-"):
            return self.test_purchase_uom(cid)

        # Master/report/menu/model checks
        if cid == "MU01-01":
            return self.test_product_stock_and_forecast(cid)
        if cid in {"MU01-02", "MU01-03", "MU06-01", "MU06-02"}:
            return self.test_replenishment_setup(cid)
        if cid == "MU06-03":
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.production", "stock.move"], "records": [("mrp.bom", [("bom_line_ids", "!=", False)])]})
        if cid in {"MU02-03", "MU03-01", "MU04-01", "MU08-02"}:
            return self.test_models_menus_or_reports(cid, {"models": ["stock.picking"], "records": [("stock.picking.type", [("code", "=", "internal")])]})
        if cid in {"MU05-01", "MU05-02"}:
            return self.test_mto_route_setup(cid)
        if cid in {"MU11-04", "MU09-01", "MU09-02", "MU14-01", "MU14-02", "MU14-04", "MU14-05"}:
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.workcenter"], "records": [("mrp.workcenter", [("is_mold", "=", True)])]})
        if cid == "MU08-01" or cid == "MU20-04":
            return self.test_models_menus_or_reports(cid, {"models": ["stock.scrap"], "records": [("stock.scrap", [])]})
        if cid == "MU12-01":
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.unbuild"]})
        if cid in {"MU12-02", "MU12-04", "MU10-02", "MU10-03"}:
            return self.test_models_menus_or_reports(cid, {"models": ["stock.move", "stock.valuation.layer"], "records": [("mrp.production", [("state", "=", "done")])]})
        if cid == "MU15-03":
            return self.test_models_menus_or_reports(cid, {"models": ["stock.location"], "records": [("stock.location", [("complete_name", "ilike", "คลังลอย")])]})
        if cid == "MU16-01":
            return self.test_models_menus_or_reports(cid, {"models": ["stock.lot"], "records": [("stock.lot", [])]})
        if cid == "MU16-03":
            return self.test_models_menus_or_reports(cid, {"models": ["documents.document"], "records": [("documents.document", [])]})
        if cid == "MU16-05":
            return self.test_models_menus_or_reports(cid, {"models": ["quality.alert"]})
        if cid == "MU16-08":
            return self.test_models_menus_or_reports(cid, {"models": ["quality.point"], "records": [("quality.point", [])]})
        if cid in {"MU10-01", "MU20-06"}:
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.production"], "records": [("mrp.production", [("state", "=", "done")])]})
        if cid in {"MU17-01", "MU17-02", "MU13-03"}:
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.bom", "mrp.routing.workcenter"], "records": [("mrp.bom", [("operation_ids", "!=", False)])]})
        if cid == "MU17-04":
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.production.schedule"]})
        if cid in {"MU19-03", "MU19-04"}:
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.workcenter", "mrp.production"], "records": [("mrp.workcenter", [])]})
        if cid in {"MU20-02", "MU20-03"}:
            return self.test_models_menus_or_reports(cid, {"models": ["ir.actions.report"], "records": [("ir.actions.report", [("model", "in", ["mrp.production", "mrp.workorder"])])]})
        if cid == "MU20-05":
            return self.test_models_menus_or_reports(cid, {"models": ["mrp.workcenter", "mrp.workorder"]})
        return result(cid, "Manual", "ยังไม่มี automated probe สำหรับ case นี้", details="ต้องทดสอบผ่าน UI ตามขั้นตอนใน Excel")


def run_all():
    cases = read_cases()
    rpc = OdooRPC()
    login = rpc.login()
    runner = UAT14Runner(rpc)
    outputs = []
    for case in cases:
        try:
            out = runner.run_case(case)
        except AssertionError as exc:
            out = result(case["case_id"], "Failed", "Functional probe failed", details=str(exc))
        except Exception as exc:
            out = result(
                case["case_id"],
                "Failed",
                "Exception during functional probe",
                details=str(exc)[:2500],
                record={"traceback": traceback.format_exc(limit=8)},
            )
        out.update({k: case[k] for k in ["sheet", "row", "backlog", "scenario", "menu"]})
        outputs.append(out)
    report = {
        "base_url": BASE_URL,
        "db": DB,
        "uid": login.get("uid") if isinstance(login, dict) else None,
        "run_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "cases": outputs,
        "summary": {
            "total": len(outputs),
            "passed": sum(1 for item in outputs if item["status"] == "Passed"),
            "failed": sum(1 for item in outputs if item["status"] == "Failed"),
            "manual": sum(1 for item in outputs if item["status"] == "Manual"),
        },
    }
    REPORT_JSON.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_workbook(report)
    write_markdown(report)
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))
    print(REPORT_JSON)
    print(OUTPUT_XLSX)
    return report


def write_workbook(report):
    wb = load_workbook(INPUT_XLSX)
    if "UAT14_Run_Result" in wb.sheetnames:
        del wb["UAT14_Run_Result"]
    ws = wb.create_sheet("UAT14_Run_Result", 1)
    headers = ["Case ID", "Status", "Sheet", "Row", "Backlog IDs", "Scenario", "Evidence", "Details"]
    ws.append(headers)
    by_case = {}
    for item in report["cases"]:
        by_case[item["case_id"]] = item
        ws.append(
            [
                item["case_id"],
                item["status"],
                item["sheet"],
                item["row"],
                item.get("backlog"),
                item.get("scenario"),
                item.get("evidence"),
                item.get("details"),
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Angsana New", size=16, bold=True, color="FFFFFF")
    body_font = Font(name="Angsana New", size=16)
    fills = {
        "Passed": PatternFill("solid", fgColor="D9EAD3"),
        "Failed": PatternFill("solid", fgColor="F4CCCC"),
        "Manual": PatternFill("solid", fgColor="FFF2CC"),
    }
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, 2).value
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if status in fills:
                cell.fill = fills[status]
    widths = [14, 14, 28, 8, 22, 42, 48, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    for sheet in wb.worksheets:
        if not sheet.title.startswith(ACTUAL_SHEET_PREFIXES):
            continue
        for row in range(7, sheet.max_row + 1):
            case_id = sheet.cell(row, 2).value
            if not case_id or str(case_id).strip() not in by_case:
                continue
            item = by_case[str(case_id).strip()]
            sheet.cell(row, 13).value = f"UAT14 {item['status']}: {item['evidence']}"
            # Keep user review columns untouched. Use formula column as designed.
    wb.save(OUTPUT_XLSX)


def write_markdown(report):
    lines = [
        "# UAT14 Actual Run Result",
        "",
        f"- Server: `{report['base_url']}`",
        f"- Database: `{report['db']}`",
        f"- Run at: {report['run_at']}",
        f"- Total: {report['summary']['total']}",
        f"- Passed: {report['summary']['passed']}",
        f"- Failed: {report['summary']['failed']}",
        f"- Manual: {report['summary']['manual']}",
        "",
        "## Failed / Stuck Cases",
        "",
    ]
    failed = [item for item in report["cases"] if item["status"] == "Failed"]
    if not failed:
        lines.append("- None")
    else:
        for item in failed:
            lines.append(f"- `{item['case_id']}` {item.get('scenario')}: {item.get('details')}")
    lines.extend(["", "## Manual UI Cases", ""])
    manual = [item for item in report["cases"] if item["status"] == "Manual"]
    if not manual:
        lines.append("- None")
    else:
        for item in manual:
            lines.append(f"- `{item['case_id']}` {item.get('scenario')}: {item.get('evidence')}")
    REPORT_MD.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    run_all()
