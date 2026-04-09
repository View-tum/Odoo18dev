from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INBOUND_RECEIPT = "GMP/IN/00055"
MO_NAME = "GMP/MO/00015"
SO_NAME = "SOB-261945"
INVOICE_NAME = "INV-D/26/03/00015"
PAYMENT_NAME = "PRKBK1/2026/00545"
REPORT_NAME = "สรุปบัญชีขาเข้า_ขาออก_ไทย.xlsx"


def report_path() -> Path:
    return Path(__file__).resolve().parents[3] / "reports" / REPORT_NAME


def style_sheet(ws):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ws.columns:
        lengths = [len(str(cell.value)) for cell in column if cell.value is not None]
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(lengths, default=10) + 2, 48)
    ws.freeze_panes = "A2"


def append_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)
    return ws


def get_data(env):
    so = env["sale.order"].search([("name", "=", SO_NAME)], limit=1)
    invoice = env["account.move"].search([("name", "=", INVOICE_NAME)], limit=1)
    payment = env["account.payment"].search([("name", "=", PAYMENT_NAME)], limit=1)
    receipt = env["stock.picking"].search([("name", "=", INBOUND_RECEIPT)], limit=1)
    mo = env["mrp.production"].search([("name", "=", MO_NAME)], limit=1)
    if not all([so, invoice, payment, receipt, mo]):
        raise ValueError("หาเอกสารที่ต้องใช้ไม่ครบ")
    outgoing = so.picking_ids.filtered(lambda p: p.picking_type_id.code == "outgoing")[:1]
    internal_pick = so.picking_ids.filtered(lambda p: p.picking_type_id.code == "internal")[:1]
    inbound_account_moves = (
        receipt.move_ids.mapped("account_move_ids")
        | mo.move_raw_ids.mapped("account_move_ids")
        | mo.move_finished_ids.mapped("account_move_ids")
        | env["account.move"].search([("ref", "ilike", mo.name)])
    ).sorted(lambda m: m.id)
    outbound_account_moves = (
        outgoing.move_ids.mapped("account_move_ids") | invoice | payment.move_id
    ).sorted(lambda m: m.id)
    return {
        "so": so,
        "invoice": invoice,
        "payment": payment,
        "receipt": receipt,
        "mo": mo,
        "outgoing": outgoing,
        "internal_pick": internal_pick,
        "inbound_account_moves": inbound_account_moves,
        "outbound_account_moves": outbound_account_moves,
    }


def inbound_rows(data):
    rows = []
    receipt = data["receipt"]
    mo = data["mo"]
    receipt_am_names = set(receipt.move_ids.mapped("account_move_ids.name"))
    raw_am_names = set(mo.move_raw_ids.mapped("account_move_ids.name"))
    fin_am_names = set(mo.move_finished_ids.mapped("account_move_ids.name"))
    for move in data["inbound_account_moves"]:
        if move.name in receipt_am_names:
            origin_doc = receipt.name
            stage = "รับเข้า Receipt"
            auto_note = "ผู้ใช้สร้าง/กด Validate Receipt แล้วระบบ post stock journal อัตโนมัติ"
        elif move.name in raw_am_names:
            origin_doc = mo.name
            stage = "เบิกวัตถุดิบเข้า WIP"
            auto_note = "เมื่อปิด MO ระบบ post รายการเบิกวัตถุดิบเข้า WIP อัตโนมัติ"
        elif move.name in fin_am_names:
            origin_doc = mo.name
            stage = "รับสินค้าสำเร็จรูปจาก WIP"
            auto_note = "เมื่อปิด MO ระบบ post รับ FG และตัด WIP อัตโนมัติ"
        else:
            origin_doc = mo.name
            stage = "บันทึกค่าแรง/Overhead"
            auto_note = "เมื่อปิด MO ระบบ post ค่าแรง/Overhead เข้างานระหว่างทำอัตโนมัติ"
        for line in move.line_ids.filtered(lambda l: l.account_id):
            rows.append([
                stage,
                origin_doc,
                move.name,
                move.ref,
                line.account_id.code or "",
                line.account_id.name or "",
                float(line.debit or 0),
                float(line.credit or 0),
                float(line.balance or 0),
                "ระบบ post อัตโนมัติ",
                auto_note,
            ])
    return rows


def outbound_rows(data):
    rows = []
    outgoing = data["outgoing"]
    invoice = data["invoice"]
    payment = data["payment"]
    outgoing_am_names = set(outgoing.move_ids.mapped("account_move_ids.name"))
    for move in data["outbound_account_moves"]:
        if move.name in outgoing_am_names:
            stage = "ส่งของ / ตัดสต๊อก"
            origin_doc = outgoing.name
            auto_note = "ผู้ใช้กด Validate Delivery แล้วระบบ post stock journal ตัดสต๊อกอัตโนมัติ"
        elif move.id == invoice.id:
            stage = "ใบแจ้งหนี้ขาย"
            origin_doc = invoice.name
            auto_note = "ผู้ใช้สั่ง Create Invoice และ custom ระบบ post invoice อัตโนมัติ"
        elif move.id == payment.move_id.id:
            stage = "รับชำระเงิน"
            origin_doc = payment.name
            auto_note = "ผู้ใช้กด Register Payment แล้วระบบสร้าง payment entry และ reconcile ลูกหนี้อัตโนมัติ"
        else:
            stage = "อื่น ๆ"
            origin_doc = move.name
            auto_note = "ระบบบันทึกอัตโนมัติ"
        for line in move.line_ids.filtered(lambda l: l.account_id):
            rows.append([
                stage,
                origin_doc,
                move.name,
                move.ref,
                line.account_id.code or "",
                line.account_id.name or "",
                float(line.debit or 0),
                float(line.credit or 0),
                float(line.balance or 0),
                "ระบบ post อัตโนมัติ" if stage != "อื่น ๆ" else "อัตโนมัติ",
                auto_note,
            ])
    return rows


def document_flow_rows(data):
    so = data["so"]
    internal_pick = data["internal_pick"]
    outgoing = data["outgoing"]
    invoice = data["invoice"]
    payment = data["payment"]
    receipt = data["receipt"]
    mo = data["mo"]
    rows = [
        [
            "ขาเข้า",
            "1",
            "Receipt รับวัตถุดิบ",
            receipt.name,
            receipt.state,
            "ผู้ใช้สร้างและกด Validate",
            "ระบบสร้าง stock move / valuation / journal อัตโนมัติ",
            receipt.origin or "",
        ],
        [
            "ขาเข้า",
            "2",
            "MO ผลิตสินค้า",
            mo.name,
            mo.state,
            "ผู้ใช้/สคริปต์สร้าง MO และสั่งปิด",
            "ระบบสร้าง raw move / finished move / workorder / stock journal อัตโนมัติ",
            mo.origin or "",
        ],
        [
            "ขาออก",
            "1",
            "Sale Order",
            so.name,
            so.state,
            "ผู้ใช้สร้างและกด Confirm",
            "ระบบสร้าง Pick และ Delivery อัตโนมัติ",
            so.client_order_ref or "",
        ],
        [
            "ขาออก",
            "2",
            "Internal Pick",
            internal_pick.name if internal_pick else "",
            internal_pick.state if internal_pick else "",
            "ผู้ใช้กด Validate",
            "ไม่มี journal บัญชีใน step นี้",
            so.name,
        ],
        [
            "ขาออก",
            "3",
            "Delivery",
            outgoing.name if outgoing else "",
            outgoing.state if outgoing else "",
            "ผู้ใช้กด Validate",
            "ระบบตัดสต๊อกและ post stock journal อัตโนมัติ",
            so.name,
        ],
        [
            "ขาออก",
            "4",
            "Invoice",
            invoice.name,
            invoice.state,
            "ผู้ใช้สั่ง Create Invoice",
            "custom ระบบ post invoice อัตโนมัติ",
            so.name,
        ],
        [
            "ขาออก",
            "5",
            "Payment",
            payment.name,
            payment.state,
            "ผู้ใช้กด Register Payment",
            "ระบบสร้าง payment move และ reconcile อัตโนมัติ",
            invoice.name,
        ],
    ]
    return rows


def summary_rows(data):
    inbound = inbound_rows(data)
    outbound = outbound_rows(data)
    def total(rows, idx):
        return round(sum(r[idx] for r in rows), 6)
    return [
        ["ฝั่งขาเข้า", data["receipt"].name, data["mo"].name, total(inbound, 6), total(inbound, 7), "รวมเดบิต/เครดิตของ Receipt + MO + Overhead"],
        ["ฝั่งขาออก", data["so"].name, data["invoice"].name, total(outbound, 6), total(outbound, 7), "รวมเดบิต/เครดิตของ Delivery + Invoice + Payment"],
    ]


def auto_map_rows(data):
    return [
        ["Receipt", "สร้าง Receipt + กด Validate", "ผู้ใช้", "stock.move", "สร้างโดยระบบเมื่อบันทึกเอกสาร", "บางส่วนอัตโนมัติ"],
        ["Receipt", "post stock journal ขาเข้า", "ระบบ", "account.move / account.move.line", "อัตโนมัติหลัง Validate Receipt", "อัตโนมัติ"],
        ["MO", "สร้าง raw move / finished move / workorder", "ระบบ", "mrp.production / stock.move / mrp.workorder", "อัตโนมัติหลัง Confirm MO", "อัตโนมัติ"],
        ["MO", "post เบิกวัตถุดิบเข้า WIP", "ระบบ", "account.move / account.move.line", "อัตโนมัติเมื่อปิด MO", "อัตโนมัติ"],
        ["MO", "post รับ FG จาก WIP", "ระบบ", "account.move / account.move.line", "อัตโนมัติเมื่อปิด MO", "อัตโนมัติ"],
        ["MO", "post Overhead", "ระบบ", "account.move / account.move.line", "อัตโนมัติเมื่อปิด MO", "อัตโนมัติ"],
        ["SO", "สร้าง Pick/Delivery", "ระบบ", "stock.picking", "อัตโนมัติหลัง Confirm SO", "อัตโนมัติ"],
        ["Delivery", "post stock journal ขาออก", "ระบบ", "account.move / account.move.line", "อัตโนมัติหลัง Validate Delivery", "อัตโนมัติ"],
        ["Invoice", "post invoice", "ระบบ", "account.move / account.move.line", "custom module auto-post หลัง Create Invoice", "อัตโนมัติ"],
        ["Payment", "สร้าง payment entry + reconcile", "ระบบ", "account.payment / account.move", "อัตโนมัติหลัง Register Payment", "อัตโนมัติ"],
    ]


def build_workbook(env):
    data = get_data(env)
    path = report_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "สรุปภาพรวม"
    ws.append(["หัวข้อ", "เอกสารต้นทาง", "เอกสารปลายทาง", "เดบิตรวม", "เครดิตรวม", "คำอธิบาย"])
    for row in summary_rows(data):
        ws.append(row)
    style_sheet(ws)

    append_sheet(
        wb,
        "ผังเอกสาร",
        ["ฝั่ง", "ลำดับ", "ขั้นตอน", "เลขเอกสาร", "สถานะ", "ผู้ใช้ทำอะไร", "ระบบทำอะไรอัตโนมัติ", "อ้างอิง"],
        document_flow_rows(data),
    )
    append_sheet(
        wb,
        "บัญชีขาเข้า",
        ["ขั้นตอน", "เอกสาร", "เลขสมุดรายวัน", "อ้างอิง", "รหัสบัญชี", "ชื่อบัญชี", "เดบิต", "เครดิต", "คงเหลือสุทธิ", "สถานะการ post", "รายละเอียดการเกิดรายการ"],
        inbound_rows(data),
    )
    append_sheet(
        wb,
        "บัญชีขาออก",
        ["ขั้นตอน", "เอกสาร", "เลขสมุดรายวัน", "อ้างอิง", "รหัสบัญชี", "ชื่อบัญชี", "เดบิต", "เครดิต", "คงเหลือสุทธิ", "สถานะการ post", "รายละเอียดการเกิดรายการ"],
        outbound_rows(data),
    )
    append_sheet(
        wb,
        "ระบบอัตโนมัติ",
        ["โมดูล/เอกสาร", "เหตุการณ์", "ใครเป็นคนกด", "โมเดลที่เกิด", "ระบบสร้างอะไร", "สรุป"],
        auto_map_rows(data),
    )

    wb.save(path)
    return path


def run(env):
    path = build_workbook(env)
    return {"report_path": str(path), "generated_at": datetime.now().isoformat()}
