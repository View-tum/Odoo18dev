from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_REPORT = "view_uat_account_knock_detail_th.xlsx"
OUTPUT_REPORT = "view_uat_account_knock_onepage_th.xlsx"


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _reports_dir() -> Path:
    return _root_path() / "reports"


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _border() -> Border:
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_title(ws, row: int, columns: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = _header_fill("1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def _style_section(ws, row: int, columns: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row, 1)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = _header_fill("4F81BD")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _border()


def _style_table_header(ws, row: int):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _header_fill("1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()


def _style_body(ws, start_row: int, amount_cols=None):
    amount_cols = set(amount_cols or [])
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in amount_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"


def _auto_width(ws, max_width: int = 80):
    for column in ws.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(values, default=8) + 2, max_width)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def _sheet_rows(ws):
    return list(ws.iter_rows(values_only=True))


def _append_table(ws, row: int, section_title: str, headers, rows, amount_cols=None) -> int:
    columns = len(headers)
    ws.cell(row, 1, section_title)
    _style_section(ws, row, columns)
    row += 1

    for index, header in enumerate(headers, start=1):
        ws.cell(row, index, header)
    _style_table_header(ws, row)
    row += 1

    start_body = row
    for data_row in rows:
        for index, value in enumerate(data_row, start=1):
            ws.cell(row, index, value)
        row += 1

    _style_body(ws, start_body, amount_cols=amount_cols)
    return row + 1


def build_onepage():
    reports_dir = _reports_dir()
    source_path = reports_dir / SOURCE_REPORT
    output_path = reports_dir / OUTPUT_REPORT

    source_wb = load_workbook(source_path, data_only=True)
    docs = _sheet_rows(source_wb["เลขเอกสาร"])
    timeline = _sheet_rows(source_wb["ไทม์ไลน์ธุรกรรม"])
    knocks = _sheet_rows(source_wb["ไทม์ไลน์_KNOCK"])
    summary = _sheet_rows(source_wb["สรุป_KNOCK"])
    reconcile = _sheet_rows(source_wb["RECONCILE_AR"])

    wb = Workbook()
    ws = wb.active
    ws.title = "หน้าเดียว"

    title = "สรุปบัญชีขาเข้า ขาออก และจังหวะ Knock แบบหน้าเดียว"
    ws.cell(1, 1, title)
    _style_title(ws, 1, 10)
    ws.cell(2, 1, "ฐานข้อมูล")
    ws.cell(2, 2, "view")
    ws.cell(3, 1, "นิยาม knock")
    ws.cell(3, 2, "บัญชีพักหรือบัญชีคั่นถูก offset/เคลียร์ด้วยรายการถัดไป จนยอดวิ่งกลับเป็นศูนย์หรือใกล้ศูนย์")
    ws.cell(4, 1, "วิธีอ่าน")
    ws.cell(4, 2, "ไล่จากบนลงล่าง: เอกสาร -> สิ่งที่ระบบ post -> เดบิต/เครดิต -> บัญชีที่ถูก knock -> ผลหลังรายการ")
    _style_body(ws, 2)

    row = 6
    row = _append_table(
        ws,
        row,
        "1. เลขเอกสารที่ใช้จริง",
        docs[0],
        docs[1:],
    )

    flow_headers = [
        "Step",
        "ฝั่ง",
        "หน้าจอ Odoo",
        "ประเภทเอกสาร",
        "เลขที่เอกสาร",
        "ผู้ใช้ทำอะไร",
        "ระบบทำอะไรอัตโนมัติ",
        "JE",
        "ผลต่อการ knock",
    ]
    flow_rows = [
        [r[0], r[1], r[2], r[3], r[4], r[7], r[8], r[10], r[11]]
        for r in timeline[1:]
    ]
    row = _append_table(ws, row, "2. Flow จากหน้าจอ Odoo ลงมาถึงบัญชี", flow_headers, flow_rows)

    knock_headers = [
        "ลำดับ JE",
        "วันที่",
        "JE",
        "Ref",
        "ฝั่ง",
        "เอกสารต้นทาง",
        "เดบิต/เครดิตที่เกิด",
        "มาจากตรงไหน",
        "ไป knock ตรงไหน",
        "ผลหลังรายการ",
    ]
    row = _append_table(ws, row, "3. เจาะทีละ JE ว่ามาจากตรงไหนและไป knock ตรงไหน", knock_headers, knocks[1:])

    row = _append_table(
        ws,
        row,
        "4. สรุปว่าบัญชีไหน knock แล้วหรือยัง",
        summary[0],
        summary[1:],
        amount_cols=[6],
    )

    row = _append_table(
        ws,
        row,
        "5. Payment Reconcile ที่ปิดลูกหนี้",
        reconcile[0],
        reconcile[1:],
        amount_cols=[2],
    )

    ws.freeze_panes = "A7"
    _auto_width(ws, 88)
    wb.save(output_path)
    return output_path


report_path = build_onepage()
print(report_path)
