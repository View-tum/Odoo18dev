from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DB_NAME = "view"
RECEIPT_NAME = "GMP/IN/00072"
REPORT_NAME = "view_uat_vendor_bill_clearance_th.xlsx"
TAX_INVOICE_NUMBER = "SIM-TAX-GMP-IN-00072"
INPUT_ACCOUNT_CODE = "116901"


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _report_path() -> Path:
    return _root_path() / "reports" / REPORT_NAME


def _header_style():
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _cell_border():
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_sheet(ws, amount_cols=None):
    amount_cols = set(amount_cols or [])
    fill, font = _header_style()
    border = _cell_border()
    for cell in ws[1]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in amount_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    for column in ws.columns:
        lengths = [len(str(cell.value)) for cell in column if cell.value is not None]
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(lengths, default=10) + 2, 48)
    ws.freeze_panes = "A2"


def ensure_bill(env, receipt):
    journal = env["account.journal"].search(
        [("type", "=", "purchase"), ("company_id", "=", env.company.id)],
        limit=1,
    )
    if not journal:
        raise ValueError("ไม่พบ purchase journal")

    bill = env["account.move"].search(
        [
            ("move_type", "=", "in_invoice"),
            ("partner_id", "=", receipt.partner_id.id),
            "|",
            ("invoice_origin", "=", receipt.origin or receipt.name),
            ("ref", "=", receipt.name),
        ],
        order="id desc",
        limit=1,
    )
    if bill:
        if bill.state == "draft":
            if bill.tax_invoice_ids:
                bill.tax_invoice_ids.write(
                    {
                        "tax_invoice_number": TAX_INVOICE_NUMBER,
                        "tax_invoice_date": bill.invoice_date,
                    }
                )
            bill.action_post()
        return bill

    line_cmds = []
    for move in receipt.move_ids_without_package.sorted("id"):
        qty = sum(abs(qty) for qty in move.stock_valuation_layer_ids.mapped("quantity")) or 1.0
        value = sum(abs(val) for val in move.stock_valuation_layer_ids.mapped("value"))
        taxes = move.product_id.supplier_taxes_id.filtered(lambda tax: tax.company_id == env.company)
        line_cmds.append(
            (
                0,
                0,
                {
                    "product_id": move.product_id.id,
                    "name": move.product_id.display_name,
                    "quantity": qty,
                    "price_unit": value / qty if qty else 0.0,
                    "tax_ids": [(6, 0, taxes.ids)],
                },
            )
        )

    bill = env["account.move"].create(
        {
            "move_type": "in_invoice",
            "partner_id": receipt.partner_id.id,
            "journal_id": journal.id,
            "invoice_date": receipt.date_done.date() if receipt.date_done else env.company.period_lock_date,
            "invoice_origin": receipt.origin or receipt.name,
            "ref": receipt.name,
            "invoice_line_ids": line_cmds,
        }
    )
    if bill.tax_invoice_ids:
        bill.tax_invoice_ids.write(
            {
                "tax_invoice_number": TAX_INVOICE_NUMBER,
                "tax_invoice_date": bill.invoice_date,
            }
        )
    bill.action_post()
    return bill


def get_receipt_input_lines(receipt, input_account):
    lines = []
    for move in receipt.move_ids_without_package.sorted("id"):
        for journal in move.account_move_ids.sorted("id"):
            for aml in journal.line_ids.filtered(lambda line: line.account_id == input_account).sorted("id"):
                lines.append(
                    {
                        "step": "ก่อน Vendor Bill",
                        "screen": "Inventory > Receipts > Validate",
                        "doc_type": "Receipt",
                        "document": receipt.name,
                        "journal_entry": journal.name,
                        "ref": journal.ref or "",
                        "account_code": aml.account_id.code or "",
                        "account_name": aml.account_id.name or "",
                        "debit": float(aml.debit or 0),
                        "credit": float(aml.credit or 0),
                        "balance": float(aml.balance or 0),
                        "source": f"รับสินค้า {move.product_id.default_code or move.product_id.display_name}",
                        "knock_note": "เปิดยอดค้างของบัญชีพักรับเข้า 116901",
                    }
                )
    return lines


def get_bill_lines(bill, input_account):
    lines = []
    for aml in bill.line_ids.sorted("id"):
        if not aml.account_id:
            continue
        knock_note = ""
        if aml.account_id == input_account:
            knock_note = "เคลียร์ยอดค้างจาก Receipt เดิม"
        elif aml.account_id.code == "117001":
            knock_note = "ภาษีซื้อจาก Vendor Bill"
        elif aml.account_id.code == "212001":
            knock_note = "เปิดยอดเจ้าหนี้การค้า รอจ่ายชำระ"
        lines.append(
            {
                "step": "หลัง Vendor Bill",
                "screen": "Accounting > Vendors > Bills > Post",
                "doc_type": "Vendor Bill",
                "document": bill.name,
                "journal_entry": bill.name,
                "ref": bill.ref or "",
                "account_code": aml.account_id.code or "",
                "account_name": aml.account_id.name or "",
                "debit": float(aml.debit or 0),
                "credit": float(aml.credit or 0),
                "balance": float(aml.balance or 0),
                "source": aml.name or "",
                "knock_note": knock_note,
            }
        )
    return lines


def build_workbook(env, receipt, bill, input_account):
    receipt_lines = get_receipt_input_lines(receipt, input_account)
    bill_lines = get_bill_lines(bill, input_account)

    before_credit = round(sum(line["credit"] - line["debit"] for line in receipt_lines), 2)
    bill_debit = round(
        sum(line["debit"] - line["credit"] for line in bill_lines if line["account_code"] == input_account.code),
        2,
    )
    after_net = round(before_credit - bill_debit, 2)

    wb = Workbook()
    ws = wb.active
    ws.title = "SUMMARY"
    ws.append(
        [
            "ฐานข้อมูล",
            "Receipt",
            "Vendor Bill",
            "Tax Invoice No.",
            "บัญชีพักรับเข้า",
            "ค้างก่อน Bill",
            "Dr จาก Bill",
            "คงเหลือหลัง Bill",
            "สรุป",
        ]
    )
    ws.append(
        [
            DB_NAME,
            receipt.name,
            bill.name,
            TAX_INVOICE_NUMBER,
            f"{input_account.code} {input_account.name}",
            before_credit,
            bill_debit,
            after_net,
            "116901 เคลียร์จบที่ระดับ scenario นี้แล้ว" if abs(after_net) < 0.005 else "ยังมียอดค้าง",
        ]
    )
    style_sheet(ws, amount_cols=[6, 7, 8])

    ws2 = wb.create_sheet("STEP_BY_STEP")
    ws2.append(
        [
            "ลำดับ",
            "ช่วงเวลา",
            "หน้าจอ Odoo",
            "ประเภทเอกสาร",
            "เลขเอกสาร",
            "JE",
            "Ref",
            "Account Code",
            "Account Name",
            "Debit",
            "Credit",
            "Balance",
            "มาจากตรงไหน",
            "อธิบายการ knock",
        ]
    )
    detail_rows = receipt_lines + bill_lines
    for index, line in enumerate(detail_rows, start=1):
        ws2.append(
            [
                index,
                line["step"],
                line["screen"],
                line["doc_type"],
                line["document"],
                line["journal_entry"],
                line["ref"],
                line["account_code"],
                line["account_name"],
                line["debit"],
                line["credit"],
                line["balance"],
                line["source"],
                line["knock_note"],
            ]
        )
    style_sheet(ws2, amount_cols=[10, 11, 12])

    ws3 = wb.create_sheet("MATCH_116901")
    ws3.append(
        [
            "คู่เทียบ",
            "ต้นทาง",
            "เอกสารต้นทาง",
            "JE ต้นทาง",
            "ยอด Cr 116901",
            "ปลายทาง",
            "เอกสารปลายทาง",
            "JE ปลายทาง",
            "ยอด Dr 116901",
            "ต่างกัน",
            "ผล",
        ]
    )
    paired = list(zip(receipt_lines, [line for line in bill_lines if line["account_code"] == input_account.code]))
    for index, (before, after) in enumerate(paired, start=1):
        diff = round(before["credit"] - after["debit"], 2)
        ws3.append(
            [
                index,
                before["source"],
                before["document"],
                before["journal_entry"],
                before["credit"],
                after["source"],
                after["document"],
                after["journal_entry"],
                after["debit"],
                diff,
                "settle แล้ว" if abs(diff) < 0.005 else "ยังต่าง",
            ]
        )
    style_sheet(ws3, amount_cols=[5, 8, 9, 10])

    ws4 = wb.create_sheet("BILL_JE")
    ws4.append(
        [
            "JE",
            "วันที่",
            "Partner",
            "Invoice Origin",
            "Ref",
            "Account Code",
            "Account Name",
            "Debit",
            "Credit",
            "Balance",
            "คำอธิบาย",
        ]
    )
    for aml in bill.line_ids.sorted("id"):
        if not aml.account_id:
            continue
        ws4.append(
            [
                bill.name,
                str(bill.date or ""),
                bill.partner_id.display_name,
                bill.invoice_origin or "",
                bill.ref or "",
                aml.account_id.code or "",
                aml.account_id.name or "",
                float(aml.debit or 0),
                float(aml.credit or 0),
                float(aml.balance or 0),
                aml.name or "",
            ]
        )
    style_sheet(ws4, amount_cols=[8, 9, 10])

    report_path = _report_path()
    wb.save(report_path)
    return report_path, before_credit, bill_debit, after_net


def main():
    receipt = env["stock.picking"].search([("name", "=", RECEIPT_NAME)], limit=1)
    if not receipt:
        raise ValueError(f"ไม่พบ Receipt {RECEIPT_NAME}")
    input_account = env["account.account"].search([("code", "=", INPUT_ACCOUNT_CODE)], limit=1)
    if not input_account:
        raise ValueError(f"ไม่พบบัญชี {INPUT_ACCOUNT_CODE}")

    bill = ensure_bill(env, receipt)
    report_path, before_credit, bill_debit, after_net = build_workbook(env, receipt, bill, input_account)
    env.cr.commit()

    print(f"REPORT={report_path}")
    print(f"BILL={bill.name}")
    print(f"BILL_ID={bill.id}")
    print(f"BEFORE={before_credit:.2f}")
    print(f"BILL_DR={bill_debit:.2f}")
    print(f"AFTER={after_net:.2f}")


main()
