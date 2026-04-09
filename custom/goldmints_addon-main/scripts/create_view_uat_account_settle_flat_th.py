from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_REPORT = "view_uat_account_settle_checklist_th.xlsx"
OUTPUT_REPORT = "view_uat_account_settle_line_by_line_th.xlsx"


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _reports_dir() -> Path:
    return _root_path() / "reports"


def _border() -> Border:
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _status_fill(status: str) -> PatternFill:
    return {
        "ค้าง / ยังไม่ settle": PatternFill("solid", fgColor="F4B084"),
        "settle แล้ว": PatternFill("solid", fgColor="C6E0B4"),
        "settle แบบกลุ่ม": PatternFill("solid", fgColor="D9EAD3"),
        "ปลายทางงบ / ไม่ต้อง settle": PatternFill("solid", fgColor="D9E2F3"),
        "คงเหลือ stock": PatternFill("solid", fgColor="D9E2F3"),
    }.get(status, PatternFill())


def _find_header_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == "ลำดับ" and ws.cell(row, 10).value == "Account Code":
            return row
    raise ValueError("Detail header row not found")


def build_flat_report() -> Path:
    source_path = _reports_dir() / SOURCE_REPORT
    output_path = _reports_dir() / OUTPUT_REPORT

    source_wb = load_workbook(source_path, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    header_row = _find_header_row(source_ws)

    wb = Workbook()
    ws = wb.active
    ws.title = "แต่ละบรรทัด"

    headers = [source_ws.cell(header_row, col).value for col in range(1, 21)]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)

    out_row = 2
    for row in range(header_row + 1, source_ws.max_row + 1):
        values = [source_ws.cell(row, col).value for col in range(1, 21)]
        if not any(value not in (None, "") for value in values):
            continue
        for col, value in enumerate(values, start=1):
            ws.cell(out_row, col, value)
        out_row += 1

    border = _border()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        if cell.value is None:
            continue
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (12, 13, 14, 15) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
        ws.cell(row, 2).fill = yellow_fill
        ws.cell(row, 3).fill = yellow_fill
        ws.cell(row, 16).fill = _status_fill(str(ws.cell(row, 16).value or ""))

    widths = {
        1: 8,
        2: 10,
        3: 11,
        4: 12,
        5: 10,
        6: 28,
        7: 18,
        8: 20,
        9: 44,
        10: 12,
        11: 28,
        12: 12,
        13: 12,
        14: 14,
        15: 15,
        16: 18,
        17: 42,
        18: 28,
        19: 50,
        20: 24,
    }
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path


report = build_flat_report()
print(report)
