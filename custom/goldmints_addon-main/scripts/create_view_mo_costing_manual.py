from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
import sys

from openpyxl import Workbook

if globals().get("__file__"):
    SERVER_PATH = Path(__file__).resolve().parents[3] / "server"
else:
    SERVER_PATH = Path.cwd() / "server"
if str(SERVER_PATH) not in sys.path:
    sys.path.insert(0, str(SERVER_PATH))

import odoo
import odoo.service.server
from odoo import SUPERUSER_ID, api
from odoo.modules.registry import Registry
from odoo.tools import config
PRODUCT_CODE = "FG-PSS-TH-01001"
BOM_ID = 2980
PRODUCE_QTY = 200
MANUAL_DURATION_MINUTES = 60.0
REPORT_NAME = "view_mo_costing_overview_flow_6_1.xlsx"
MANUAL_NAME = "MFG-30_6.1_Overview_Flow_MO_Costing_TH.md"
TAG_PREFIX = "SIM-VIEW-MO-COST"
VENDOR_NAME_PREFIX = "SIM VIEW MO COST VENDOR"
RELEVANT_MODULES = [
    "mrp_account",
    "mrp_workorder_hr_account",
    "cost_sheet",
    "custom_landed_cost_account",
    "mrp_parallel_console",
]


def q2(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _report_path() -> Path:
    return _root_path() / "reports" / REPORT_NAME


def _manual_path() -> Path:
    return _root_path() / "reports" / MANUAL_NAME



def _get_env_and_cr():
    existing_env = globals().get("env")
    if existing_env is not None:
        return existing_env, None
    config.parse_config(["-c", str(_root_path() / "server" / "odoo.conf"), "-d", "view"])
    odoo.service.server.load_server_wide_modules()
    registry = Registry("view")
    cr = registry.cursor()
    return api.Environment(cr, SUPERUSER_ID, {}), cr
def _helper_functions():
    helper_path = _root_path() / "custom" / "goldmints_addon-main" / "scripts" / "create_view_uat_existing_product_trace.py"
    source = helper_path.read_text(encoding="utf-8-sig")
    source = re.sub(r"\nmain\(\)\s*$", "\n", source)
    namespace = {"__file__": str(helper_path), "__name__": "view_uat_trace_helper"}
    exec(compile(source, str(helper_path), "exec"), namespace)
    namespace["TAG_PREFIX"] = TAG_PREFIX
    namespace["VENDOR_NAME_PREFIX"] = VENDOR_NAME_PREFIX
    return namespace


HELPER = _helper_functions()
style_header = HELPER["style_header"]
style_grid = HELPER["style_grid"]
auto_width = HELPER["auto_width"]
ensure_schema = HELPER["ensure_schema"]
get_core_setup = HELPER["get_core_setup"]
create_inbound_receipt = HELPER["create_inbound_receipt"]
pin_raw_lots = HELPER["pin_raw_lots"]
_move_line_qty_field = HELPER["_move_line_qty_field"]


def _installed_modules(env):
    return env["ir.module.module"].search(
        [("name", "in", RELEVANT_MODULES), ("state", "=", "installed")],
        order="name",
    )


def _stock_doc_type(move):
    if move.picking_id:
        code = move.picking_id.picking_type_id.code
        if code == "incoming":
            return "Inbound Receipt"
        if code == "internal":
            return "Internal Pick"
        if code == "outgoing":
            return "Delivery"
        return f"Picking ({code})"
    if move.raw_material_production_id:
        return "MO Raw Consumption"
    if move.production_id:
        return "MO Finished Receipt"
    return "Other"


def _stock_doc_number(move):
    if move.picking_id:
        return move.picking_id.name
    if move.raw_material_production_id:
        return move.raw_material_production_id.name
    if move.production_id:
        return move.production_id.name
    return ""


def _move_qty_done(move):
    qty = sum(move.move_line_ids.mapped("quantity")) or sum(move.move_line_ids.mapped("qty_done")) or 0.0
    return float(qty)


def _ensure_finished_move_line(move, fg_lot, qty):
    qty_field = _move_line_qty_field(move.env["stock.move.line"])
    if move.move_line_ids:
        primary = move.move_line_ids[0]
        vals = {"lot_id": fg_lot.id}
        if qty_field:
            vals[qty_field] = qty
        primary.write(vals)
        if len(move.move_line_ids) > 1:
            move.move_line_ids[1:].unlink()
    else:
        vals = move._prepare_move_line_vals(quantity=qty, reserved_quant=False)
        vals.update({"lot_id": fg_lot.id})
        if qty_field:
            vals[qty_field] = qty
        move.env["stock.move.line"].create(vals)
    if "quantity" in move._fields:
        move.quantity = qty


def create_and_close_mo(env, setup, tag, lots_by_product):
    mo = env["mrp.production"].create(
        {
            "product_id": setup["product"].id,
            "product_qty": PRODUCE_QTY,
            "product_uom_id": setup["product"].uom_id.id,
            "bom_id": setup["bom"].id,
            "origin": tag,
            "location_src_id": setup["floating_location"].id,
            "location_dest_id": setup["warehouse"].lot_stock_id.id,
        }
    )
    mo.action_confirm()
    mo.action_assign()
    mo = env["mrp.production"].browse(mo.id)

    fg_lot = env["stock.lot"].create(
        {"name": f"{tag}-FG", "product_id": setup["product"].id, "company_id": setup["company"].id}
    )
    mo.lot_producing_id = fg_lot.id
    mo.qty_producing = mo.product_qty
    mo._console_fill_move_quantities_for_close({mo.id: mo.product_qty})
    pin_raw_lots(mo, lots_by_product)

    done_at = datetime.now().replace(microsecond=0)
    started_at = done_at - timedelta(minutes=MANUAL_DURATION_MINUTES)
    for workorder in mo.workorder_ids:
        if not workorder.time_ids:
            vals = workorder._prepare_timeline_vals(MANUAL_DURATION_MINUTES, started_at, done_at)
            env["mrp.workcenter.productivity"].create(vals)
        vals = {
            "state": "done",
            "qty_produced": mo.product_qty,
            "date_start": started_at,
            "date_finished": done_at,
            "duration": MANUAL_DURATION_MINUTES,
            "duration_unit": round(MANUAL_DURATION_MINUTES / max(mo.product_qty, 1), 4),
        }
        if "costs_hour" in workorder._fields:
            vals["costs_hour"] = workorder.workcenter_id.costs_hour
        workorder.with_context(bypass_duration_calculation=True).write(vals)
        workorder.end_all()

    raw_moves = mo.move_raw_ids.filtered(lambda move: move.state not in ("done", "cancel"))
    raw_moves.with_context(skip_mo_check=True)._action_done(cancel_backorder=True)
    mo._cal_price(raw_moves)

    finished_move = mo.move_finished_ids.filtered(
        lambda move: move.product_id == mo.product_id and move.state not in ("done", "cancel")
    )
    finished_move.ensure_one()
    _ensure_finished_move_line(finished_move, fg_lot, mo.product_qty)
    finished_move._action_done(cancel_backorder=True)
    finished_move.move_line_ids.consume_line_ids = [(6, 0, raw_moves.mapped("move_line_ids").ids)]

    mo.write({"date_finished": done_at, "priority": "0", "is_locked": True, "state": "done"})
    if hasattr(mo, "_post_labour"):
        mo._post_labour()
    return mo, finished_move, fg_lot


def _labour_moves(env, mo, stock_moves):
    stock_account_moves = stock_moves.mapped("account_move_ids")
    return env["account.move"].search(
        [("ref", "ilike", mo.name), ("id", "not in", stock_account_moves.ids)],
        order="id",
    )

def build_trace(env, setup, receipt, mo, finished_move, fg_lot, tag):
    raw_moves = mo.move_raw_ids.sorted(lambda move: move.id)
    finished_moves = mo.move_finished_ids.sorted(lambda move: move.id)
    stock_moves = (receipt.move_ids | raw_moves | finished_moves).sorted(lambda move: move.id)
    svls = env["stock.valuation.layer"].search([("stock_move_id", "in", stock_moves.ids)], order="id")
    labour_moves = _labour_moves(env, mo, stock_moves)
    account_moves = (stock_moves.mapped("account_move_ids") | labour_moves).sorted(lambda move: move.id)
    aml = account_moves.line_ids.filtered(lambda line: line.account_id).sorted(lambda line: (line.move_id.id, line.id))

    receipt_cost = q2(sum(sum(move.stock_valuation_layer_ids.mapped("value")) for move in receipt.move_ids))
    raw_cost = q2(-sum(sum(move.stock_valuation_layer_ids.mapped("value")) for move in raw_moves))
    fg_value = q2(sum(sum(move.stock_valuation_layer_ids.mapped("value")) for move in finished_moves))
    workorder_cost = q2(sum(wo._cal_cost() for wo in mo.workorder_ids))
    extra_cost_total = q2((mo.extra_cost or 0.0) * mo.product_qty)
    fg_unit_cost = q2((Decimal(str(fg_value)) / Decimal(str(mo.product_qty))) if mo.product_qty else 0)

    workorder_rows = []
    for wo in mo.workorder_ids:
        workorder_rows.append(
            {
                "wo_name": wo.name,
                "operation": wo.operation_id.name or "",
                "workcenter": wo.workcenter_id.display_name,
                "duration_min": float(wo.duration or 0.0),
                "costs_hour": float(wo.workcenter_id.costs_hour or 0.0),
                "cal_cost": float(wo._cal_cost() or 0.0),
                "expense_account": f"{wo.workcenter_id.expense_account_id.code or ''} {wo.workcenter_id.expense_account_id.name or ''}".strip(),
                "analytic": ", ".join(wo.workcenter_id.costs_hour_account_ids.mapped("name")),
            }
        )

    receipt_lots = {move.product_id.id: ", ".join(move.move_line_ids.mapped("lot_id.name")) for move in receipt.move_ids}
    component_rows = []
    for move in raw_moves:
        component_rows.append(
            {
                "product": move.product_id.display_name,
                "planned_qty": float(move.product_uom_qty or 0.0),
                "done_qty": _move_qty_done(move),
                "lot": ", ".join(move.move_line_ids.mapped("lot_id.name")),
                "receipt_lot": receipt_lots.get(move.product_id.id, ""),
                "consumption_value": float(-sum(move.stock_valuation_layer_ids.mapped("value")) or 0.0),
                "uom": move.product_uom.name,
            }
        )

    docs = [
        {"step": 1, "doc_type": "Inbound Receipt", "name": receipt.name, "state": receipt.state, "menu": "Inventory > Operations > Receipts", "what": "Validate รับวัตถุดิบตาม BOM เข้าคลังลอย", "auto": "ระบบสร้าง stock move, valuation layer, journal entry อัตโนมัติเมื่อ Validate"},
        {"step": 2, "doc_type": "Manufacturing Order", "name": mo.name, "state": mo.state, "menu": "Manufacturing > Operations > Manufacturing Orders", "what": "เปิด MO, confirm, assign, ระบุ FG lot และปิดงาน", "auto": "ระบบสร้าง raw consumption, FG receipt, SVL และ JE อัตโนมัติเมื่อปิด MO"},
        {"step": 3, "doc_type": "Work Orders", "name": ", ".join(mo.workorder_ids.mapped("name")), "state": ", ".join(sorted(set(mo.workorder_ids.mapped("state")))), "menu": "MO > Work Orders tab", "what": "ตรวจเวลาทำงานและ Work Center ที่ใช้จริง", "auto": "ระบบคำนวณ workorder cost จาก time tracking และ post labour entry ตอน MO done"},
        {"step": 4, "doc_type": "Stock Valuation", "name": mo.name, "state": "done", "menu": "Inventory > Reporting > Valuation", "what": "ดูต้นทุนรับเข้า, ตัด raw, รับ FG", "auto": "ระบบแยก SVL ตาม stock move แต่ละตัว"},
        {"step": 5, "doc_type": "Journal Entries", "name": ", ".join(account_moves.mapped("name")), "state": "posted", "menu": "Accounting > Journal Entries", "what": "ดู debit/credit ของ receipt, raw to WIP, FG from WIP, labour", "auto": "JE ทั้งหมดในชุดนี้ถูก post อัตโนมัติจาก inventory valuation และ MO labour posting"},
    ]

    tb_map = defaultdict(lambda: {"debit": Decimal("0"), "credit": Decimal("0"), "balance": Decimal("0")})
    for line in aml:
        key = (line.account_id.code or "", line.account_id.name or "", line.account_id.account_type or "")
        debit = Decimal(str(line.debit or 0))
        credit = Decimal(str(line.credit or 0))
        tb_map[key]["debit"] += debit
        tb_map[key]["credit"] += credit
        tb_map[key]["balance"] += debit - credit

    tb_rows = []
    for (code, name, account_type), values in sorted(tb_map.items()):
        tb_rows.append({"code": code, "name": name, "account_type": account_type, "debit": float(q2(values["debit"])), "credit": float(q2(values["credit"])), "balance": float(q2(values["balance"]))})

    trace_rows = []
    for move in stock_moves:
        related_am = move.account_move_ids.sorted(lambda rec: rec.id)
        related_svl = move.stock_valuation_layer_ids.sorted(lambda rec: rec.id)
        if related_am:
            for am in related_am:
                for line in am.line_ids.filtered(lambda l: l.account_id).sorted(lambda l: l.id):
                    trace_rows.append({
                        "step": 1 if move.picking_id else (2 if move.raw_material_production_id else 3),
                        "doc_type": _stock_doc_type(move),
                        "doc_no": _stock_doc_number(move),
                        "stock_move": move.reference or move.name,
                        "svl_ids": ", ".join(str(svl.id) for svl in related_svl),
                        "svl_value": float(sum(related_svl.mapped("value")) or 0.0),
                        "journal_entry": am.name,
                        "account_code": line.account_id.code,
                        "account_name": line.account_id.name,
                        "debit": float(line.debit or 0.0),
                        "credit": float(line.credit or 0.0),
                        "explain": "Receipt เปิด raw material" if move.picking_id else ("MO consume raw -> เข้า WIP" if move.raw_material_production_id else "MO รับ FG -> ปิด WIP เข้า FG"),
                    })
        else:
            trace_rows.append({
                "step": 0,
                "doc_type": _stock_doc_type(move),
                "doc_no": _stock_doc_number(move),
                "stock_move": move.reference or move.name,
                "svl_ids": ", ".join(str(svl.id) for svl in related_svl),
                "svl_value": float(sum(related_svl.mapped("value")) or 0.0),
                "journal_entry": "",
                "account_code": "",
                "account_name": "",
                "debit": 0.0,
                "credit": 0.0,
                "explain": "ไม่มี JE ผูกกับ stock move นี้",
            })
    for am in labour_moves:
        for line in am.line_ids.filtered(lambda l: l.account_id).sorted(lambda l: l.id):
            trace_rows.append({
                "step": 4,
                "doc_type": "MO Labour",
                "doc_no": mo.name,
                "stock_move": "",
                "svl_ids": "",
                "svl_value": 0.0,
                "journal_entry": am.name,
                "account_code": line.account_id.code,
                "account_name": line.account_id.name,
                "debit": float(line.debit or 0.0),
                "credit": float(line.credit or 0.0),
                "explain": "Labour posting จาก workorder cost ของ MO",
            })

    return {
        "tag": tag,
        "setup": setup,
        "receipt": receipt,
        "mo": mo,
        "finished_move": finished_move,
        "fg_lot": fg_lot,
        "stock_moves": stock_moves,
        "svls": svls,
        "account_moves": account_moves,
        "labour_moves": labour_moves,
        "aml": aml,
        "workorder_rows": workorder_rows,
        "component_rows": component_rows,
        "docs": docs,
        "tb_rows": tb_rows,
        "trace_rows": trace_rows,
        "receipt_cost": float(receipt_cost),
        "raw_cost": float(raw_cost),
        "workorder_cost": float(workorder_cost),
        "extra_cost_total": float(extra_cost_total),
        "fg_value": float(fg_value),
        "fg_unit_cost": float(fg_unit_cost),
        "modules": _installed_modules(env),
    }


def _account_move_source(trace, move):
    stock_moves = trace["stock_moves"].filtered(lambda m: move in m.account_move_ids)
    if stock_moves:
        first_move = stock_moves[0]
        return _stock_doc_type(first_move), _stock_doc_number(first_move)
    if move.id in trace["labour_moves"].ids:
        return "MO Labour", trace["mo"].name
    return "Other", move.ref or ""

def write_workbook(env, trace):
    path = _report_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    ws.append(["หัวข้อ", "ค่า"])
    for row in [
        ("Database", env.cr.dbname),
        ("Section", "MFG-30 Manufacturing Cost Accounting > 6.1 Overview Flow"),
        ("Run Tag", trace["tag"]),
        ("Product", trace["setup"]["product"].display_name),
        ("BOM", f"{trace['setup']['bom'].id} | {trace['setup']['bom'].code or '-'}"),
        ("Produce Qty", PRODUCE_QTY),
        ("Receipt", trace["receipt"].name),
        ("MO", trace["mo"].name),
        ("FG Lot", trace["fg_lot"].name),
        ("Raw Material Receipt Value", trace["receipt_cost"]),
        ("Raw Consumption Value", trace["raw_cost"]),
        ("Workorder Cost", trace["workorder_cost"]),
        ("Extra Cost Total", trace["extra_cost_total"]),
        ("Finished Goods Value", trace["fg_value"]),
        ("FG Unit Cost", trace["fg_unit_cost"]),
    ]:
        ws.append(list(row))
    style_header(ws)
    style_grid(ws, amount_cols=[2])
    auto_width(ws, 110)

    ws = wb.create_sheet("เลขเอกสาร")
    ws.append(["ลำดับ", "ประเภท", "เลขเอกสาร", "สถานะ", "เมนูที่เปิดดู", "คำอธิบาย"])
    for row in trace["docs"]:
        ws.append([row["step"], row["doc_type"], row["name"], row["state"], row["menu"], row["what"]])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 90)

    ws = wb.create_sheet("CUSTOM_CONTEXT")
    ws.append(["Module", "State", "หมายเหตุ"])
    notes = {
        "mrp_account": "คำนวณ valuation และ labour posting ของ MO",
        "mrp_workorder_hr_account": "ผูก workorder cost กับ analytic และ employee costing",
        "cost_sheet": "มี report เพิ่มในฐาน view แต่ scenario นี้เน้น trace จาก stock, SVL และ JE",
        "custom_landed_cost_account": "มีผลกับ landed cost ไม่ใช่ step หลักของ MO รอบนี้",
        "mrp_parallel_console": "ติดตั้งอยู่ในฐาน view แต่ scenario รอบนี้ปิด MO ด้วย flow ปกติของ MRP script",
    }
    for mod in trace["modules"]:
        ws.append([mod.name, mod.state, notes.get(mod.name, "")])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 90)

    ws = wb.create_sheet("SCREEN_FLOW")
    ws.append(["Step", "หน้าจอ Odoo", "เลขเอกสาร", "สิ่งที่ต้องดู", "สิ่งที่ต้องแคป", "ระบบทำอัตโนมัติอะไร", "เลขอ้างอิงเพิ่มเติม"])
    extra_refs = {
        1: ", ".join(sorted(set(trace["receipt"].move_ids.mapped("account_move_ids.name")))),
        2: ", ".join(sorted(set(trace["mo"].move_raw_ids.mapped("account_move_ids.name")))),
        3: ", ".join(sorted(set(trace["mo"].move_finished_ids.mapped("account_move_ids.name")) | set(trace["labour_moves"].mapped("name")))),
        4: ", ".join(str(svl.id) for svl in trace["svls"]),
        5: ", ".join(trace["account_moves"].mapped("name")),
    }
    capture_map = {
        1: "Receipt header + Operations + Detailed Operations + lots",
        2: "MO form header + Components tab",
        3: "Work Orders tab + duration + work center + FG lot",
        4: "Valuation list filter ตาม Receipt และ MO",
        5: "Journal Entry form + Journal Items",
    }
    verify_map = {
        1: "component, qty, lot, destination = คลังลอย",
        2: "product, qty, source/destination, state = Done",
        3: "duration 60 นาที, work center cost, labour JE",
        4: "receipt value, raw issue value, FG receipt value",
        5: "debit/credit ของ raw, WIP, FG, labour",
    }
    for row in trace["docs"]:
        ws.append([row["step"], row["menu"], row["name"], verify_map.get(row["step"], row["what"]), capture_map.get(row["step"], ""), row["auto"], extra_refs.get(row["step"], "")])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 100)

    ws = wb.create_sheet("MO_COST_SUMMARY")
    ws.append(["องค์ประกอบต้นทุน", "จำนวนเงิน", "มาจากไหน", "ดูจากหน้าจอไหน", "คำอธิบาย"])
    summary_rows = [
        ("Raw Material Receipt Value", trace["receipt_cost"], trace["receipt"].name, "Inventory > Operations > Receipts / Inventory > Reporting > Valuation", "มูลค่ารับเข้าวัตถุดิบตาม receipt จริง"),
        ("Raw Consumption to WIP", trace["raw_cost"], trace["mo"].name, "MO > Components / Inventory > Reporting > Valuation", "มูลค่าวัตถุดิบที่ถูก consume เข้า MO"),
        ("Workorder / Labour Cost", trace["workorder_cost"], trace["mo"].name, "MO > Work Orders / Accounting > Journal Entries", "เวลาทำงานจริง x Cost per Hour ของ Work Center"),
        ("Extra Cost", trace["extra_cost_total"], trace["mo"].name, "MO > Extra Unit Cost", "รอบนี้ไม่ได้ใส่ extra cost เพิ่ม"),
        ("Finished Goods Value", trace["fg_value"], trace["mo"].name, "Inventory > Reporting > Valuation", "ต้นทุนสินค้าสำเร็จรูปหลังปิด MO"),
        ("FG Unit Cost", trace["fg_unit_cost"], trace["mo"].name, "Inventory > Reporting > Valuation", "FG value / จำนวนที่ผลิต"),
    ]
    for row in summary_rows:
        ws.append(list(row))
    style_header(ws)
    style_grid(ws, amount_cols=[2])
    auto_width(ws, 100)

    ws = wb.create_sheet("WORKORDER_COST")
    ws.append(["Work Order", "Operation", "Work Center", "Duration (min)", "Cost/Hour", "Calculated Cost", "Expense Account", "Analytic"])
    for row in trace["workorder_rows"]:
        ws.append([row["wo_name"], row["operation"], row["workcenter"], row["duration_min"], row["costs_hour"], row["cal_cost"], row["expense_account"], row["analytic"]])
    style_header(ws)
    style_grid(ws, amount_cols=[4, 5, 6])
    auto_width(ws, 72)

    ws = wb.create_sheet("COMPONENT_TRACE")
    ws.append(["Component", "Planned Qty", "Done Qty", "UoM", "Receipt Lot", "Consumed Lot", "Consumption Value"])
    for row in trace["component_rows"]:
        ws.append([row["product"], row["planned_qty"], row["done_qty"], row["uom"], row["receipt_lot"], row["lot"], row["consumption_value"]])
    style_header(ws)
    style_grid(ws, amount_cols=[2, 3, 7])
    auto_width(ws, 76)
    ws = wb.create_sheet("STOCK_MOVES")
    ws.append(["Move ID", "Doc Type", "Doc No", "Product", "Planned Qty", "Done Qty", "Source", "Destination", "State", "Account Moves"])
    for move in trace["stock_moves"]:
        ws.append([move.id, _stock_doc_type(move), _stock_doc_number(move), move.product_id.display_name, float(move.product_uom_qty or 0.0), _move_qty_done(move), move.location_id.complete_name, move.location_dest_id.complete_name, move.state, ", ".join(move.account_move_ids.mapped("name"))])
    style_header(ws)
    style_grid(ws, amount_cols=[5, 6])
    auto_width(ws, 66)

    ws = wb.create_sheet("SVL")
    ws.append(["SVL ID", "Stock Move", "Doc Type", "Doc No", "Qty", "Unit Cost", "Value", "Journal Entry"])
    for svl in trace["svls"]:
        move = svl.stock_move_id
        ws.append([svl.id, move.reference or move.name, _stock_doc_type(move), _stock_doc_number(move), float(svl.quantity or 0.0), float(svl.unit_cost or 0.0), float(svl.value or 0.0), svl.account_move_id.name or ""])
    style_header(ws)
    style_grid(ws, amount_cols=[5, 6, 7])
    auto_width(ws, 52)

    ws = wb.create_sheet("ACCOUNT_MOVES")
    ws.append(["JE", "Date", "Journal", "Source Type", "Source Doc", "Ref", "State"])
    for move in trace["account_moves"]:
        source_type, source_doc = _account_move_source(trace, move)
        ws.append([move.name, str(move.date), move.journal_id.code, source_type, source_doc, move.ref or "", move.state])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 52)

    ws = wb.create_sheet("AML")
    ws.append(["JE", "Source Type", "Source Doc", "Account Code", "Account Name", "Debit", "Credit", "Balance", "Label"])
    for line in trace["aml"]:
        source_type, source_doc = _account_move_source(trace, line.move_id)
        ws.append([line.move_id.name, source_type, source_doc, line.account_id.code, line.account_id.name, float(line.debit or 0.0), float(line.credit or 0.0), float((line.debit or 0.0) - (line.credit or 0.0)), line.name or ""])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 7, 8])
    auto_width(ws, 58)

    ws = wb.create_sheet("TRACE_DETAIL")
    ws.append(["Step", "Doc Type", "Doc No", "Stock Move", "SVL IDs", "SVL Value", "JE", "Account Code", "Account Name", "Debit", "Credit", "อธิบาย"])
    for row in trace["trace_rows"]:
        ws.append([row["step"], row["doc_type"], row["doc_no"], row["stock_move"], row["svl_ids"], row["svl_value"], row["journal_entry"], row["account_code"], row["account_name"], row["debit"], row["credit"], row["explain"]])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 10, 11])
    auto_width(ws, 72)

    ws = wb.create_sheet("TB")
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit", "Balance"])
    for row in trace["tb_rows"]:
        ws.append([row["code"], row["name"], row["account_type"], row["debit"], row["credit"], row["balance"]])
    style_header(ws)
    style_grid(ws, amount_cols=[4, 5, 6])
    auto_width(ws, 52)

    wb.save(path)
    return path

def write_manual(env, trace):
    path = _manual_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    receipt_jes = ", ".join(sorted(set(trace["receipt"].move_ids.mapped("account_move_ids.name"))))
    raw_jes = ", ".join(sorted(set(trace["mo"].move_raw_ids.mapped("account_move_ids.name"))))
    fg_jes = ", ".join(sorted(set(trace["mo"].move_finished_ids.mapped("account_move_ids.name"))))
    labour_jes = ", ".join(trace["labour_moves"].mapped("name")) or "ไม่มี"
    module_lines = "\n".join(f"- `{mod.name}`: {mod.state}" for mod in trace["modules"])
    workorder_lines = "\n".join(
        f"- `{row['wo_name']}` | Work Center: `{row['workcenter']}` | Duration: `{row['duration_min']}` นาที | Cost/Hour: `{row['costs_hour']}` | Cost จริง: `{row['cal_cost']}`"
        for row in trace["workorder_rows"]
    ) or "- ไม่มี Work Order"
    component_lines = "\n".join(
        f"- `{row['product']}` | Planned `{row['planned_qty']}` {row['uom']} | Done `{row['done_qty']}` | Lot รับเข้า `{row['receipt_lot']}` | Lot ที่ consume `{row['lot']}` | มูลค่า `{row['consumption_value']}`"
        for row in trace["component_rows"]
    )

    content = f"""# MFG-30 Manufacturing Cost Accounting

## 6.1 ภาพรวมการไหลของข้อมูล (Overview Flow)

เอกสารนี้อ้างอิงการทดสอบจริงในฐาน `view` เมื่อ `{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}` โดยใช้สินค้าเดิมในระบบ ไม่ได้สร้าง product ใหม่ เพื่อให้สามารถเปิดเอกสารใน Odoo แล้วแปะรูปหน้าจอประกอบคู่มือได้ทันที

### 1. ขอบเขตการทดสอบ
- Database: `{env.cr.dbname}`
- Product: `{trace['setup']['product'].display_name}`
- BOM: `{trace['setup']['bom'].id} | {trace['setup']['bom'].code or '-'}`
- Qty ที่ผลิต: `{PRODUCE_QTY}`
- Receipt: `{trace['receipt'].name}`
- MO: `{trace['mo'].name}`
- FG Lot: `{trace['fg_lot'].name}`

### 2. Module ที่เกี่ยวข้องในฐาน view
{module_lines}

หมายเหตุ:
- `mrp_account` คือโมดูลหลักที่ทำให้เกิด `Stock Valuation` และ `Journal Entry` ของ MO
- `mrp_workorder_hr_account` ช่วยให้ cost ของ work order ไปผูกกับ analytic / employee costing
- `mrp_parallel_console` ติดตั้งอยู่ในฐาน `view` แต่รอบทดสอบนี้ใช้ flow ปกติของ MO เพื่อให้ trace ต้นทุนอ่านง่าย

### 3. ภาพรวมการไหลของข้อมูล
1. Validate `Receipt` -> ระบบสร้าง `stock.move` ของวัตถุดิบ -> สร้าง `SVL` -> post `JE` รับเข้าวัตถุดิบ
2. Confirm/Assign/Close `MO` -> ระบบ consume raw materials -> โอนมูลค่าเข้า `WIP`
3. ปิด `Work Order` -> ระบบคำนวณเวลา x `Cost per Hour` -> post `Labour Entry`
4. รับ `Finished Goods` จาก MO -> ระบบย้ายมูลค่าจาก `WIP` ไป `FG`
5. ผู้ใช้ตรวจสอบที่ `Valuation`, `Journal Entries`, และ `Journal Items` เพื่อดูว่าต้นทุนมาจากตรงไหนและวิ่งไปบัญชีใด

### 4. ต้นทุนจริงของรอบทดสอบ
- มูลค่ารับเข้าวัตถุดิบ: `{trace['receipt_cost']}`
- มูลค่าตัดวัตถุดิบเข้า MO: `{trace['raw_cost']}`
- ต้นทุน Work Order / Labour: `{trace['workorder_cost']}`
- Extra Cost: `{trace['extra_cost_total']}`
- มูลค่าสินค้าสำเร็จรูป: `{trace['fg_value']}`
- ต้นทุนต่อหน่วย FG: `{trace['fg_unit_cost']}`

สูตรที่ใช้ในระบบจาก `mrp_account` คือ:
- `FG Total Cost = Raw Consumption Cost + Work Center Cost + Extra Cost - By-product Share`
- รอบนี้ไม่มี by-product และไม่ได้ใส่ extra cost

### 5. รายละเอียด Work Order Cost
{workorder_lines}

### 6. รายละเอียด Component ที่ถูกใช้จริง
{component_lines}

### 7. Step-by-Step สำหรับแปะรูปหน้าจอ

#### Step 1: Receipt วัตถุดิบเข้า
- เมนู: `Inventory > Operations > Receipts`
- เอกสาร: `{trace['receipt'].name}`
- ให้แคป: Header, Operations, Detailed Operations, Lots/Serial
- ให้ตรวจ: product, qty, lot, source = Vendor, destination = คลังลอย
- หลังจากกด Validate ระบบจะสร้างเลข JE เหล่านี้อัตโนมัติ: `{receipt_jes}`
- จุดเชื่อมต่อ: Receipt นี้เป็นต้นทางของต้นทุน raw material ทั้งรอบ
- [แปะรูปหน้าจอ Receipt ที่นี่]

#### Step 2: Manufacturing Order
- เมนู: `Manufacturing > Operations > Manufacturing Orders`
- เอกสาร: `{trace['mo'].name}`
- ให้แคป: Header ของ MO, Components tab
- ให้ตรวจ: Product, Quantity, Source Location, Destination Location, FG Lot, State = Done
- Raw material ที่ consume จาก MO นี้จะสร้าง JE เหล่านี้อัตโนมัติ: `{raw_jes}`
- จุดเชื่อมต่อ: มูลค่าวัตถุดิบถูกย้ายจาก stock raw material ไป `WIP`
- [แปะรูปหน้าจอ MO Header/Components ที่นี่]

#### Step 3: Work Orders
- เมนู: `MO > Work Orders tab`
- เอกสารอ้างอิง: `{trace['mo'].name}`
- ให้แคป: Work Order line, Work Center, Duration
- ให้ตรวจ: Work Order เป็น `Done`, Duration = `{MANUAL_DURATION_MINUTES}` นาที, Work Center cost ถูกตั้งไว้
- Labour JE ที่เกิดอัตโนมัติ: `{labour_jes}`
- จุดเชื่อมต่อ: เวลาใน Work Order ถูกคูณด้วย Cost per Hour เพื่อสร้าง labour posting
- [แปะรูปหน้าจอ Work Orders ที่นี่]

#### Step 4: Stock Valuation
- เมนู: `Inventory > Reporting > Valuation`
- ให้ filter ด้วยเอกสาร `{trace['receipt'].name}` และ `{trace['mo'].name}`
- ให้แคป: บรรทัด valuation ของ receipt, raw consumption, finished goods
- ให้ตรวจ:
  - Receipt value = `{trace['receipt_cost']}`
  - Raw consumption value = `{trace['raw_cost']}`
  - FG value = `{trace['fg_value']}`
- จุดเชื่อมต่อ: หน้านี้เป็นตัวเชื่อมระหว่าง stock movement กับต้นทุนจริง
- [แปะรูปหน้าจอ Stock Valuation ที่นี่]

#### Step 5: Journal Entries
- เมนู: `Accounting > Journal Entries`
- ให้ filter ด้วย `{trace['receipt'].name}` และ `{trace['mo'].name}`
- JE ที่ต้องเห็น:
  - Receipt JE: `{receipt_jes}`
  - Raw Consumption JE: `{raw_jes}`
  - FG Receipt JE: `{fg_jes}`
  - Labour JE: `{labour_jes}`
- ให้แคป: Entry header และ Journal Items
- ให้ตรวจว่าบัญชี Debit/Credit วิ่งตาม flow นี้
  1. Receipt: `Dr Raw Material / Cr GRNI หรือบัญชีพักรับเข้า`
  2. Raw Consumption: `Dr WIP / Cr Raw Material`
  3. FG Receipt: `Dr Finished Goods / Cr WIP`
  4. Labour: `Dr WIP หรือ FG ตาม config / Cr Expense Absorption`
- [แปะรูปหน้าจอ Journal Entry ที่นี่]

### 8. จุดที่ต้องดูต่อใน Excel
- Workbook: `{_report_path()}`
- Sheet `SCREEN_FLOW`: ใช้เป็น check-list ว่าควรเปิดหน้าจอไหนก่อนหลัง
- Sheet `MO_COST_SUMMARY`: ใช้สรุปต้นทุนรวมของ MO รอบนี้
- Sheet `WORKORDER_COST`: ใช้ชี้ว่าต้นทุนเวลามาจาก work order ไหน
- Sheet `COMPONENT_TRACE`: ใช้จับคู่ receipt lot กับ consumed lot
- Sheet `TRACE_DETAIL`: ใช้ผูก `Document -> Stock Move -> SVL -> JE -> AML` ทีละบรรทัด

### 9. หมายเหตุทางบัญชี
- มูลค่าที่รับเข้า raw material (`{trace['receipt_cost']}`) ไม่จำเป็นต้องเท่ากับ FG cost เพราะรอบนี้มี labour cost เพิ่ม
- บัญชีที่เห็นจริงใน JE ให้ยึดจาก `Journal Items` ในระบบและ sheet `AML`/`TRACE_DETAIL` เป็นหลัก
- ถ้าต้องการเอาไปทำ manual ต่อในบทถัดไป แนะนำแตกเป็น 3 หัวข้อย่อย:
  1. `Receipt -> Raw Material Accounting`
  2. `MO Consume -> WIP`
  3. `FG Receipt + Labour Posting`
"""
    path.write_text(content, encoding="utf-8")
    return path

def main():
    local_env, close_cr = _get_env_and_cr()
    try:
        ensure_schema(local_env)
        setup = get_core_setup(local_env)
        tag = f"{TAG_PREFIX}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        receipt, lots_by_product = create_inbound_receipt(local_env, setup, tag)
        mo, finished_move, fg_lot = create_and_close_mo(local_env, setup, tag, lots_by_product)
        trace = build_trace(local_env, setup, receipt, mo, finished_move, fg_lot, tag)
        report_path = write_workbook(local_env, trace)
        manual_path = write_manual(local_env, trace)
        local_env.cr.commit()

        print("REPORT:", report_path)
        print("MANUAL:", manual_path)
        print("RECEIPT:", receipt.name)
        print("MO:", mo.name)
        print("FG LOT:", fg_lot.name)
        print("SVL COUNT:", len(trace["svls"]))
        print("JE:", ", ".join(trace["account_moves"].mapped("name")))
    finally:
        if close_cr is not None:
            close_cr.close()


if __name__ == "__main__":
    main()
