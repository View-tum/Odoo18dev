from __future__ import annotations

import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PRODUCT_CODE = "FG-PSS-TH-01001"
SALE_QTY = 120
SO_TYPE_ID = 1
PREFERRED_BANK_JOURNAL_CODE = "RKBK1"
REPORT_NAME = "poysian_odoo_actual_trace.xlsx"
CUSTOMER_NAME_PREFIX = "SIM DEMO POYSIAN ODOO VS EXCEL"
CLIENT_REF_PREFIX = "SIM-ODOO-EXCEL-POYSIAN"

EXPECTED_COMPARE = {
    "qty_sold": Decimal("120.00"),
    "invoice_total_gross": Decimal("21600.00"),
    "revenue_ex_vat": Decimal("20186.92"),
    "output_vat": Decimal("1413.08"),
    "cogs": Decimal("1672.31"),
}


def q2(value):
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _report_path() -> Path:
    return Path(__file__).resolve().parents[3] / "reports" / REPORT_NAME


def _header_style():
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _cell_border():
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def ensure_schema(env):
    env.cr.execute(
        "ALTER TABLE stock_picking ADD COLUMN IF NOT EXISTS manufacturing_type varchar"
    )
    env.cr.execute(
        "ALTER TABLE sale_sequence_type ADD COLUMN IF NOT EXISTS is_full_tax_invoice boolean DEFAULT true"
    )


def get_sale_defaults(env):
    recent = env["sale.order"].search(
        [("so_type_id", "=", SO_TYPE_ID), ("warehouse_id", "!=", False)],
        order="id desc",
        limit=1,
    )
    return {
        "warehouse_id": recent.warehouse_id.id or 1,
        "team_id": recent.team_id.id or 1,
        "pricelist_id": recent.pricelist_id.id or 1,
        "payment_term_id": recent.payment_term_id.id or False,
    }


def get_bank_setup(env):
    journal = env["account.journal"].search(
        [("type", "=", "bank"), ("code", "=", PREFERRED_BANK_JOURNAL_CODE)],
        limit=1,
    )
    if not journal:
        journal = env["account.journal"].search([("type", "=", "bank")], limit=1)
    method_line = env["account.payment.method.line"].search(
        [
            ("journal_id", "=", journal.id),
            ("payment_type", "=", "inbound"),
            ("name", "ilike", "Manual"),
        ],
        limit=1,
    )
    if not method_line:
        method_line = env["account.payment.method.line"].search(
            [("journal_id", "=", journal.id), ("payment_type", "=", "inbound")],
            limit=1,
        )
    return journal, method_line


def create_demo_partner(env, tag):
    return env["res.partner"].create(
        {
            "name": f"{CUSTOMER_NAME_PREFIX} {tag}",
            "customer_rank": 1,
            "approval_state": "approved",
        }
    )


def set_done_quantities(picking):
    for move_line in picking.move_line_ids:
        qty = move_line.quantity_product_uom or move_line.quantity
        move_line.qty_done = qty


def validate_picking(env, picking):
    set_done_quantities(picking)
    result = picking.button_validate()
    if isinstance(result, dict) and result.get("res_model") == "stock.backorder.confirmation":
        wizard = env[result["res_model"]].browse(result["res_id"])
        wizard.process()
    return result


def create_sale_flow(env):
    now = datetime.now()
    tag = now.strftime("%Y%m%d%H%M%S")
    product = env["product.product"].search([("default_code", "=", PRODUCT_CODE)], limit=1)
    if not product:
        raise ValueError(f"Product {PRODUCT_CODE} not found")

    sale_defaults = get_sale_defaults(env)
    bank_journal, payment_method_line = get_bank_setup(env)
    partner = create_demo_partner(env, tag)
    client_ref = f"{CLIENT_REF_PREFIX}-{tag}"

    order_vals = {
        "partner_id": partner.id,
        "date_order": now,
        "client_order_ref": client_ref,
        "so_type_id": SO_TYPE_ID,
        "warehouse_id": sale_defaults["warehouse_id"],
        "team_id": sale_defaults["team_id"],
        "pricelist_id": sale_defaults["pricelist_id"],
        "payment_term_id": sale_defaults["payment_term_id"],
        "order_line": [
            (0, 0, {"product_id": product.id, "product_uom_qty": SALE_QTY})
        ],
    }
    sale_order = env["sale.order"].create(order_vals)
    sale_order.action_confirm()

    pick = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "internal")[:1]
    validate_picking(env, pick)

    outgoing = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "outgoing")[:1]
    validate_picking(env, outgoing)

    invoice = sale_order._create_invoices()

    payment_register = env["account.payment.register"].with_context(
        active_model="account.move",
        active_ids=invoice.ids,
    ).create(
        {
            "journal_id": bank_journal.id,
            "payment_method_line_id": payment_method_line.id,
            "amount": invoice.amount_residual,
        }
    )
    payment_action = payment_register.action_create_payments()
    payment = env["account.payment"].browse(payment_action["res_id"])

    return {
        "tag": tag,
        "partner": partner,
        "product": product,
        "sale_order": sale_order,
        "pick": pick,
        "outgoing": outgoing,
        "invoice": invoice,
        "payment": payment,
        "bank_journal": bank_journal,
        "payment_method_line": payment_method_line,
    }


def _stock_valuation_layers(env, stock_moves):
    return env["stock.valuation.layer"].search(
        [("stock_move_id", "in", stock_moves.ids)], order="id"
    )


def _account_moves(invoice, payment, valuation_layers):
    stock_moves = valuation_layers.mapped("account_move_id")
    return (invoice | payment.move_id | stock_moves).sorted(lambda move: move.id)


def _gl_lines(account_moves):
    return account_moves.line_ids.filtered(lambda line: line.account_id).sorted(
        lambda line: (line.move_id.id, line.id)
    )


def build_trace(env, flow):
    sale_order = flow["sale_order"]
    pick = flow["pick"]
    outgoing = flow["outgoing"]
    invoice = flow["invoice"]
    payment = flow["payment"]
    product = flow["product"]

    pickings = sale_order.picking_ids.sorted(lambda picking: picking.id)
    stock_moves = pickings.move_ids.sorted(lambda move: move.id)
    valuation_layers = _stock_valuation_layers(env, stock_moves)
    account_moves = _account_moves(invoice, payment, valuation_layers)
    gl_lines = _gl_lines(account_moves)

    actual = {
        "qty_sold": q2(sum(outgoing.move_line_ids.mapped("quantity_product_uom"))),
        "invoice_total_gross": q2(invoice.amount_total),
        "revenue_ex_vat": q2(
            sum(
                -line.balance
                for line in gl_lines
                if line.account_id.account_type == "income"
            )
        ),
        "output_vat": q2(
            sum(
                -line.balance
                for line in gl_lines
                if line.account_id.id == 909 or "ภาษีขาย" in (line.account_id.name or "")
            )
        ),
        "cogs": q2(
            sum(
                line.balance
                for line in gl_lines
                if line.account_id.account_type == "expense_direct_cost"
            )
        ),
        "bank_increase": q2(
            sum(
                line.balance
                for line in gl_lines
                if line.account_id.account_type == "asset_cash"
            )
        ),
        "inventory_change": q2(
            sum(
                line.balance
                for line in gl_lines
                if line.account_id.id == 818
                or "สินค้าสำเร็จรูป - ยาดม" in (line.account_id.name or "")
            )
        ),
    }

    summary = {
        "tag": flow["tag"],
        "customer_name": flow["partner"].name,
        "product_code": product.default_code,
        "product_name": product.display_name,
        "sale_order": sale_order.name,
        "pick": pick.name,
        "outgoing": outgoing.name,
        "invoice": invoice.name,
        "payment": payment.name,
        "bank_journal": payment.journal_id.code,
        "invoice_journal": invoice.journal_id.code,
        "picked_lots": ", ".join(sorted(set(pick.move_line_ids.mapped("lot_id.name")))),
        "delivered_lots": ", ".join(sorted(set(outgoing.move_line_ids.mapped("lot_id.name")))),
        "qty_sold": actual["qty_sold"],
        "invoice_total_gross": actual["invoice_total_gross"],
        "revenue_ex_vat": actual["revenue_ex_vat"],
        "output_vat": actual["output_vat"],
        "cogs": actual["cogs"],
        "bank_increase": actual["bank_increase"],
        "inventory_change": actual["inventory_change"],
    }

    return {
        "summary": summary,
        "pickings": pickings,
        "stock_moves": stock_moves,
        "valuation_layers": valuation_layers,
        "account_moves": account_moves,
        "gl_lines": gl_lines,
        "actual": actual,
    }


def write_workbook(flow, trace):
    path = _report_path()
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_fill, header_font = _header_style()
    border = _cell_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    good_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")

    def style_header(ws, row=1):
        for cell in ws[row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

    def set_widths(ws, widths):
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width

    def style_grid(ws, start_row, end_row, amount_cols=None):
        amount_cols = amount_cols or []
        for row in range(start_row, end_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value is None:
                    continue
                cell.border = border
                cell.alignment = left
                if col in amount_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

    ws = wb.active
    ws.title = "README"
    ws.append(["Field", "Value"])
    for row in [
        ("Scenario", "Actual Odoo transaction trace for FG-PSS-TH-01001 using acctest"),
        ("Tag", trace["summary"]["tag"]),
        ("Customer", trace["summary"]["customer_name"]),
        ("Product", trace["summary"]["product_name"]),
        ("Sale Order", trace["summary"]["sale_order"]),
        ("Pick", trace["summary"]["pick"]),
        ("Delivery", trace["summary"]["outgoing"]),
        ("Invoice", trace["summary"]["invoice"]),
        ("Payment", trace["summary"]["payment"]),
        ("Scope", "Uses existing stock in acctest and creates a real SO -> Pick -> Delivery -> Invoice -> Payment flow"),
        ("Auto behavior", "Pick is created on SO confirm, Delivery is created on Pick validate, Invoice is auto-posted by custom module, stock/account entries are auto-generated by Odoo"),
        ("Schema patches", "Added stock_picking.manufacturing_type and sale_sequence_type.is_full_tax_invoice when missing"),
        ("Compare workbook", "Compare this file with reports/poysian_1x6_simulation_to_fs.xlsx"),
    ]:
        ws.append(list(row))
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row)
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True)
    set_widths(ws, {1: 22, 2: 120})

    ws = wb.create_sheet("DOC_FLOW")
    ws.append(
        [
            "Step",
            "Document",
            "Name",
            "State",
            "Manual or Auto",
            "What happened",
        ]
    )
    rows = [
        (1, "Sale Order", trace["summary"]["sale_order"], flow["sale_order"].state, "Manual", "Created by script"),
        (2, "Pick", trace["summary"]["pick"], flow["pick"].state, "Auto", "Created automatically when SO was confirmed"),
        (3, "Delivery", trace["summary"]["outgoing"], flow["outgoing"].state, "Auto", "Created automatically after Pick was validated"),
        (4, "Customer Invoice", trace["summary"]["invoice"], flow["invoice"].state, "Mixed", "Created by script and auto-posted by custom module sale_auto_confirm_invoice"),
        (5, "Customer Payment", trace["summary"]["payment"], flow["payment"].state, "Manual", "Created by payment register wizard"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row)
    set_widths(ws, {1: 8, 2: 18, 3: 22, 4: 12, 5: 16, 6: 60})

    ws = wb.create_sheet("PICKINGS")
    ws.append(
        [
            "Picking",
            "Type",
            "State",
            "Source",
            "Destination",
            "Product",
            "Lot",
            "Reserved Qty",
            "Done Qty",
        ]
    )
    for picking in trace["pickings"]:
        for move_line in picking.move_line_ids:
            ws.append(
                [
                    picking.name,
                    picking.picking_type_id.name,
                    picking.state,
                    picking.location_id.complete_name,
                    picking.location_dest_id.complete_name,
                    move_line.product_id.default_code,
                    move_line.lot_id.name,
                    float(move_line.quantity_product_uom or 0.0),
                    float(move_line.qty_done or 0.0),
                ]
            )
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row, amount_cols=[8, 9])
    set_widths(ws, {1: 18, 2: 14, 3: 10, 4: 24, 5: 24, 6: 20, 7: 14, 8: 12, 9: 12})

    ws = wb.create_sheet("SVL")
    ws.append(
        [
            "SVL ID",
            "Date",
            "Description",
            "Stock Move",
            "Qty",
            "Unit Cost",
            "Value",
            "Related Account Move",
        ]
    )
    for svl in trace["valuation_layers"]:
        ws.append(
            [
                svl.id,
                str(svl.create_date),
                svl.description,
                svl.stock_move_id.reference or svl.stock_move_id.picking_id.name,
                float(svl.quantity),
                float(svl.unit_cost),
                float(svl.value),
                svl.account_move_id.name or "",
            ]
        )
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row, amount_cols=[5, 6, 7])
    set_widths(ws, {1: 10, 2: 20, 3: 34, 4: 20, 5: 10, 6: 12, 7: 12, 8: 20})

    ws = wb.create_sheet("ACCOUNT_MOVES")
    ws.append(["Move", "Journal", "Date", "State", "Ref", "Partner"])
    for move in trace["account_moves"]:
        ws.append(
            [
                move.name,
                move.journal_id.code,
                str(move.date),
                move.state,
                move.ref or "",
                move.partner_id.name or "",
            ]
        )
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row)
    set_widths(ws, {1: 22, 2: 10, 3: 12, 4: 10, 5: 26, 6: 36})

    ws = wb.create_sheet("GL_LINES")
    ws.append(
        [
            "Move",
            "Journal",
            "Account Code",
            "Account Name",
            "Partner",
            "Debit",
            "Credit",
            "Balance",
            "Label",
        ]
    )
    for line in trace["gl_lines"]:
        ws.append(
            [
                line.move_id.name,
                line.move_id.journal_id.code,
                line.account_id.code or "",
                line.account_id.name,
                line.partner_id.name or "",
                float(line.debit),
                float(line.credit),
                float(line.balance),
                line.name or "",
            ]
        )
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row, amount_cols=[6, 7, 8])
    set_widths(ws, {1: 22, 2: 10, 3: 14, 4: 34, 5: 30, 6: 12, 7: 12, 8: 12, 9: 34})

    ws = wb.create_sheet("COMPARE_EXCEL")
    ws.append(["Metric", "Excel Expected", "Odoo Actual", "Diff"])
    compare_rows = []
    for key, expected_value in EXPECTED_COMPARE.items():
        actual_value = trace["actual"][key]
        diff = q2(actual_value - expected_value)
        compare_rows.append((key, float(expected_value), float(actual_value), float(diff)))
    for row in compare_rows:
        ws.append(list(row))
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row, amount_cols=[2, 3, 4])
    for row in range(2, ws.max_row + 1):
        diff_cell = ws.cell(row, 4)
        if abs(diff_cell.value) < 0.005:
            diff_cell.fill = good_fill
        else:
            diff_cell.fill = warn_fill
    set_widths(ws, {1: 24, 2: 16, 3: 16, 4: 14})

    ws = wb.create_sheet("AUTO_MAP")
    ws.append(["Area", "Manual", "Automatic"])
    for row in [
        ("Sales", "Create SO", "SO confirmation generates picking"),
        ("Warehouse", "Validate Pick / Validate Delivery", "Lot reservation and chained delivery document creation"),
        ("Accounting - Revenue", "Trigger invoice creation", "Custom module auto-posts invoice and Odoo generates tax/receivable lines"),
        ("Accounting - Inventory", "Validate delivery", "stock_account creates valuation layer and stock journal entry"),
        ("Payment", "Register payment", "Odoo creates payment journal entry and reconciles invoice"),
    ]:
        ws.append(list(row))
    style_header(ws, 1)
    style_grid(ws, 2, ws.max_row)
    set_widths(ws, {1: 18, 2: 28, 3: 72})

    wb.save(path)
    return path


def export_existing(env, sale_order_name):
    ensure_schema(env)
    sale_order = env["sale.order"].search([("name", "=", sale_order_name)], limit=1)
    if not sale_order:
        raise ValueError(f"Sale order {sale_order_name} not found")

    invoice = env["account.move"].search(
        [("invoice_origin", "=", sale_order.name), ("move_type", "=", "out_invoice")],
        order="id desc",
        limit=1,
    )
    payment = env["account.payment"].search(
        [
            ("partner_id", "=", sale_order.partner_id.id),
            ("amount", "=", invoice.amount_total),
            ("state", "=", "paid"),
        ],
        order="id desc",
        limit=1,
    )
    pick = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "internal")[:1]
    outgoing = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "outgoing")[:1]
    flow = {
        "tag": sale_order.client_order_ref or sale_order.name,
        "partner": sale_order.partner_id,
        "product": sale_order.order_line[:1].product_id,
        "sale_order": sale_order,
        "pick": pick,
        "outgoing": outgoing,
        "invoice": invoice,
        "payment": payment,
        "bank_journal": payment.journal_id,
        "payment_method_line": payment.payment_method_line_id,
    }
    trace = build_trace(env, flow)
    report_path = write_workbook(flow, trace)
    summary = {
        "sale_order": sale_order.name,
        "invoice": invoice.name,
        "payment": payment.name,
        "report_path": str(report_path),
    }
    print(json.dumps(summary, ensure_ascii=False, default=str))
    return summary


def run(env):
    ensure_schema(env)
    flow = create_sale_flow(env)
    trace = build_trace(env, flow)
    report_path = write_workbook(flow, trace)

    summary = {
        "tag": trace["summary"]["tag"],
        "sale_order": trace["summary"]["sale_order"],
        "pick": trace["summary"]["pick"],
        "outgoing": trace["summary"]["outgoing"],
        "invoice": trace["summary"]["invoice"],
        "payment": trace["summary"]["payment"],
        "qty_sold": float(trace["summary"]["qty_sold"]),
        "invoice_total_gross": float(trace["summary"]["invoice_total_gross"]),
        "revenue_ex_vat": float(trace["summary"]["revenue_ex_vat"]),
        "output_vat": float(trace["summary"]["output_vat"]),
        "cogs": float(trace["summary"]["cogs"]),
        "bank_increase": float(trace["summary"]["bank_increase"]),
        "report_path": str(report_path),
    }

    env.cr.commit()
    print(json.dumps(summary, ensure_ascii=False, default=str))
    return summary
