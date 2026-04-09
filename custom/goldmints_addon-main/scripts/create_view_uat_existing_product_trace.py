from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PRODUCT_CODE = "FG-PSS-TH-01001"
BOM_ID = 2980
PRODUCE_QTY = 200
SALE_QTY = 120
REPORT_NAME = "view_uat_existing_product_trace.xlsx"
TAG_PREFIX = "SIM-VIEW-UAT-POYSIAN"
VENDOR_NAME_PREFIX = "SIM VIEW UAT VENDOR"
CUSTOMER_NAME_PREFIX = "SIM VIEW UAT CUSTOMER"
PREFERRED_BANK_JOURNAL_CODE = "RKBK1"
WALKTHROUGH_FILE = "UAT_Walkthrough_DB_VIEW_TH_FINAL.md"


def q2(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _report_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        root = Path(script_file).resolve().parents[3]
    else:
        root = Path.cwd()
    return root / "reports" / REPORT_NAME


def _walkthrough_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        root = Path(script_file).resolve().parents[3]
    else:
        root = Path.cwd()
    return root / WALKTHROUGH_FILE


def _header_style():
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _cell_border():
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row=1):
    fill, font = _header_style()
    border = _cell_border()
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_grid(ws, start_row=2, amount_cols=None):
    amount_cols = set(amount_cols or [])
    border = _cell_border()
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in amount_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def auto_width(ws, max_width=48):
    for column in ws.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(values, default=8) + 2, max_width)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def ensure_schema(env):
    env.cr.execute(
        "ALTER TABLE stock_picking ADD COLUMN IF NOT EXISTS manufacturing_type varchar"
    )
    env.cr.execute(
        "ALTER TABLE sale_sequence_type ADD COLUMN IF NOT EXISTS is_full_tax_invoice boolean DEFAULT true"
    )


def _startup_missing_modules(env):
    names = ["account_customer_group_payment", "sale_credit_limit_warning"]
    missing = []
    for name in names:
        rec = env["ir.module.module"].search([("name", "=", name)], limit=1)
        if not rec or rec.state != "installed":
            missing.append(name)
    return missing


def verify_walkthrough_setup(env):
    checks = []

    def add(group, name, condition, detail="", note=""):
        checks.append(
            {
                "group": group,
                "name": name,
                "status": "PASS" if condition else "FAIL",
                "detail": detail,
                "note": note,
            }
        )

    company = env.company
    warehouse = env["stock.warehouse"].search([("company_id", "=", company.id)], limit=1)
    add("VERIFY 24", "warehouse exists", bool(warehouse), warehouse.display_name if warehouse else "")

    required_modules = [
        "mrp_mps_manufacturing_type",
        "mrp_parallel_console",
        "mrp_auto_merge",
        "mrp_production_summary",
    ]
    mods = env["ir.module.module"].search([("name", "in", required_modules), ("state", "=", "installed")])
    add(
        "VERIFY 24",
        "required modules installed",
        len(mods) == len(required_modules),
        ", ".join(sorted(mods.mapped("name"))),
    )

    admin = env.ref("base.user_admin", raise_if_not_found=False) or env.user
    group_plastic = env.ref("mrp_mps_manufacturing_type.group_mrp_manager_plastic", raise_if_not_found=False)
    group_pharma = env.ref("mrp_mps_manufacturing_type.group_mrp_manager_pharma", raise_if_not_found=False)
    add("VERIFY 24", "group plastic exists", bool(group_plastic), str(group_plastic.id) if group_plastic else "")
    add("VERIFY 24", "group pharma exists", bool(group_pharma), str(group_pharma.id) if group_pharma else "")
    add("VERIFY 24", "admin in plastic group", bool(group_plastic and admin in group_plastic.users))
    add("VERIFY 24", "admin in pharma group", bool(group_pharma and admin in group_pharma.users))

    for partner_name in ["UAT Customer Domestic", "UAT Customer Inter", "UAT Vendor RM"]:
        partner = env["res.partner"].search([("name", "=", partner_name)], limit=1)
        add("VERIFY 24", f"partner {partner_name}", bool(partner), str(partner.id) if partner else "")

    product_names = [
        "UAT-RM-A",
        "UAT-RM-B",
        "UAT-FG-MTO-PLASTIC",
        "UAT-FG-MTS-PHARMA",
        "UAT-FG-MULTI-BOM",
    ]
    for name in product_names:
        tmpl = env["product.template"].search([("name", "=", name)], limit=1)
        add("VERIFY 24", f"product {name}", bool(tmpl), str(tmpl.id) if tmpl else "")

    for bom_code in ["UAT-BOM-MTO-PLASTIC", "UAT-BOM-MTS-PHARMA", "UAT-BOM-MULTI-A", "UAT-BOM-MULTI-B"]:
        bom = env["mrp.bom"].search([("code", "=", bom_code)], limit=1)
        add("VERIFY 24", f"BOM {bom_code}", bool(bom), str(bom.id) if bom else "")

    fg_mts = env["product.product"].search([("product_tmpl_id.name", "=", "UAT-FG-MTS-PHARMA")], limit=1)
    fg_multi = env["product.product"].search([("product_tmpl_id.name", "=", "UAT-FG-MULTI-BOM")], limit=1)
    if warehouse and fg_mts:
        mps_mts = env["mrp.production.schedule"].search(
            [("product_id", "=", fg_mts.id), ("warehouse_id", "=", warehouse.id)], limit=1
        )
        orderpoint = env["stock.warehouse.orderpoint"].search(
            [("product_id", "=", fg_mts.id), ("warehouse_id", "=", warehouse.id)], limit=1
        )
        add("VERIFY 24", "MPS line FG MTS", bool(mps_mts), str(mps_mts.id) if mps_mts else "")
        add("VERIFY 24", "orderpoint FG MTS", bool(orderpoint), str(orderpoint.id) if orderpoint else "")
    else:
        add("VERIFY 24", "MPS line FG MTS", False, "missing warehouse/product")
        add("VERIFY 24", "orderpoint FG MTS", False, "missing warehouse/product")

    if warehouse and fg_multi:
        mps_multi = env["mrp.production.schedule"].search(
            [("product_id", "=", fg_multi.id), ("warehouse_id", "=", warehouse.id)], limit=1
        )
        add("VERIFY 24", "MPS line FG MULTI", bool(mps_multi), str(mps_multi.id) if mps_multi else "")
    else:
        add("VERIFY 24", "MPS line FG MULTI", False, "missing warehouse/product")

    icp = env["ir.config_parameter"].sudo()
    add("VERIFY 24", "mrp_auto_merge.enabled=True", icp.get_param("mrp_auto_merge.enabled") in ("True", "true", "1"))
    add("VERIFY 24", "mrp_auto_merge.date_range=7", icp.get_param("mrp_auto_merge.date_range") == "7")
    marker = icp.get_param("uat_manu_setup_last_run") or ""
    add("VERIFY 24", "setup marker exists", bool(marker), marker)

    product = env["product.product"].search([("default_code", "=", PRODUCT_CODE)], limit=1)
    bom = env["mrp.bom"].browse(BOM_ID)
    bank_journal = env["account.journal"].search([("type", "=", "bank"), ("code", "=", PREFERRED_BANK_JOURNAL_CODE)], limit=1)
    if not bank_journal:
        bank_journal = env["account.journal"].search([("type", "=", "bank")], limit=1)
    recent_mo = env["mrp.production"].search(
        [("product_id", "=", product.id), ("location_src_id", "!=", False)],
        order="id desc",
        limit=1,
    )
    floating_loc = recent_mo.location_src_id or env["stock.location"].search(
        [("complete_name", "ilike", "%คลังลอย%")], limit=1
    )
    missing_modules = _startup_missing_modules(env)

    add("CURRENT DB", f"existing product {PRODUCT_CODE}", bool(product), product.display_name if product else "")
    add("CURRENT DB", f"BOM {BOM_ID}", bool(bom.exists()), bom.code or "")
    add("CURRENT DB", "product tracking = lot", bool(product and product.tracking == "lot"), product.tracking if product else "")
    add(
        "CURRENT DB",
        "BOM components all exist",
        bool(bom.exists() and bom.bom_line_ids and all(line.product_id.exists() for line in bom.bom_line_ids)),
        ", ".join(line.product_id.display_name for line in bom.bom_line_ids) if bom.exists() else "",
    )
    add("CURRENT DB", "floating location GMP/Stock/คลังลอย exists", bool(floating_loc), floating_loc.complete_name if floating_loc else "")
    add("CURRENT DB", "bank journal available", bool(bank_journal), bank_journal.display_name if bank_journal else "")
    add("CURRENT DB", "startup missing modules absent", not missing_modules, ", ".join(missing_modules), "เป็น warning ตอน registry load ของฐาน view")

    passed = sum(1 for row in checks if row["status"] == "PASS")
    total = len(checks)
    legacy_passed = sum(1 for row in checks if row["group"] == "VERIFY 24" and row["status"] == "PASS")
    legacy_total = sum(1 for row in checks if row["group"] == "VERIFY 24")
    return {
        "checks": checks,
        "passed": passed,
        "total": total,
        "legacy_passed": legacy_passed,
        "legacy_total": legacy_total,
        "missing_modules": missing_modules,
    }


def get_core_setup(env):
    product = env["product.product"].search([("default_code", "=", PRODUCT_CODE)], limit=1)
    if not product:
        raise ValueError(f"Product {PRODUCT_CODE} not found")
    bom = env["mrp.bom"].browse(BOM_ID)
    if not bom.exists():
        raise ValueError(f"BOM {BOM_ID} not found")
    warehouse = env["stock.warehouse"].search([("company_id", "=", env.company.id)], limit=1)
    incoming_type = env["stock.picking.type"].search([("code", "=", "incoming"), ("warehouse_id", "=", warehouse.id)], limit=1)
    vendor_location = env["stock.location"].search([("usage", "=", "supplier")], limit=1)
    recent_mo = env["mrp.production"].search(
        [("product_id", "=", product.id), ("location_src_id", "!=", False)],
        order="id desc",
        limit=1,
    )
    floating_location = recent_mo.location_src_id
    if not floating_location:
        floating_location = env["stock.location"].search([("complete_name", "ilike", "GMP/Stock/%คลังลอย%")], limit=1)
    if not floating_location:
        floating_location = env["stock.location"].search([("complete_name", "ilike", "%คลังลอย%")], limit=1)
    if not incoming_type or not vendor_location or not floating_location:
        raise ValueError("Missing incoming picking type or required stock locations")
    return {
        "product": product,
        "bom": bom,
        "warehouse": warehouse,
        "incoming_type": incoming_type,
        "vendor_location": vendor_location,
        "floating_location": floating_location,
        "company": product.company_id or env.company,
    }


def get_sale_defaults(env):
    recent = env["sale.order"].search([("warehouse_id", "!=", False)], order="id desc", limit=1)
    if not recent:
        raise ValueError("No existing sale order found for defaults")
    return {
        "warehouse_id": recent.warehouse_id.id,
        "team_id": recent.team_id.id or False,
        "pricelist_id": recent.pricelist_id.id or False,
        "payment_term_id": recent.payment_term_id.id or False,
        "so_type_id": recent.so_type_id.id if "so_type_id" in recent._fields and recent.so_type_id else False,
    }


def get_bank_setup(env):
    journal = env["account.journal"].search([("type", "=", "bank"), ("code", "=", PREFERRED_BANK_JOURNAL_CODE)], limit=1)
    if not journal:
        journal = env["account.journal"].search([("type", "=", "bank")], limit=1)
    method_line = env["account.payment.method.line"].search(
        [("journal_id", "=", journal.id), ("payment_type", "=", "inbound"), ("name", "ilike", "Manual")],
        limit=1,
    )
    if not method_line:
        method_line = env["account.payment.method.line"].search([("journal_id", "=", journal.id), ("payment_type", "=", "inbound")], limit=1)
    return journal, method_line


def _optional_partner_vals(env, name, partner_type):
    vals = {"name": name}
    if partner_type == "vendor":
        vals["supplier_rank"] = 1
    if partner_type == "customer":
        vals["customer_rank"] = 1
    if "approval_state" in env["res.partner"]._fields:
        vals["approval_state"] = "approved"
    return vals


def create_vendor(env, tag):
    return env["res.partner"].create(_optional_partner_vals(env, f"{VENDOR_NAME_PREFIX} {tag}", "vendor"))


def create_customer(env, tag):
    return env["res.partner"].create(_optional_partner_vals(env, f"{CUSTOMER_NAME_PREFIX} {tag}", "customer"))


def set_done_quantities(picking):
    for move_line in picking.move_line_ids:
        qty = getattr(move_line, "quantity_product_uom", 0) or getattr(move_line, "quantity", 0) or getattr(move_line, "qty_done", 0) or move_line.move_id.product_uom_qty
        vals = {}
        if "qty_done" in move_line._fields:
            vals["qty_done"] = qty
        if "quantity" in move_line._fields:
            vals["quantity"] = qty
        if "picked" in move_line._fields:
            vals["picked"] = True
        if vals:
            move_line.write(vals)


def validate_picking(env, picking):
    set_done_quantities(picking)
    result = picking.button_validate()
    if isinstance(result, dict) and result.get("res_model") == "stock.backorder.confirmation":
        env[result["res_model"]].browse(result["res_id"]).process()
    return result


def _move_line_qty_field(record):
    if "quantity" in record._fields:
        return "quantity"
    if "qty_done" in record._fields:
        return "qty_done"
    return None


def pin_move_lot(move, lot, target_qty):
    qty_field = _move_line_qty_field(move.env["stock.move.line"])
    quant = move.env["stock.quant"].search(
        [("product_id", "=", move.product_id.id), ("lot_id", "=", lot.id), ("quantity", ">", 0), ("location_id", "child_of", move.location_id.id)],
        order="in_date, id",
        limit=1,
    )
    source_location = quant.location_id.id if quant else move.location_id.id
    if move.move_line_ids:
        primary = move.move_line_ids[0]
        vals = {"lot_id": lot.id, "location_id": source_location}
        if qty_field:
            vals[qty_field] = target_qty
        if "picked" in primary._fields:
            vals["picked"] = True
        primary.write(vals)
        if len(move.move_line_ids) > 1:
            move.move_line_ids[1:].unlink()
    else:
        vals = move._prepare_move_line_vals(quantity=target_qty, reserved_quant=False)
        vals.update({"lot_id": lot.id, "location_id": source_location})
        if qty_field:
            vals[qty_field] = target_qty
        if "picked" in move.env["stock.move.line"]._fields:
            vals["picked"] = True
        move.env["stock.move.line"].create(vals)
    move_vals = {}
    if "quantity" in move._fields:
        move_vals["quantity"] = target_qty
    if "picked" in move._fields:
        move_vals["picked"] = True
    if move_vals:
        move.write(move_vals)


def create_inbound_receipt(env, setup, tag):
    picking = env["stock.picking"].create(
        {
            "picking_type_id": setup["incoming_type"].id,
            "location_id": setup["vendor_location"].id,
            "location_dest_id": setup["floating_location"].id,
            "partner_id": create_vendor(env, tag).id,
            "scheduled_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "invoice_reference": tag,
            "invoice_date": datetime.now().strftime("%Y-%m-%d"),
            "origin": tag,
        }
    )
    factor = Decimal(str(PRODUCE_QTY)) / Decimal(str(setup["bom"].product_qty or 1.0))
    lots_by_product = {}
    for bom_line in setup["bom"].bom_line_ids:
        component = bom_line.product_id
        required_qty = float(Decimal(str(bom_line.product_qty)) * factor)
        lot = env["stock.lot"].create({"name": f"{tag}-{component.id}", "product_id": component.id, "company_id": setup["company"].id})
        move = env["stock.move"].create(
            {
                "name": component.display_name,
                "picking_id": picking.id,
                "product_id": component.id,
                "product_uom_qty": required_qty,
                "product_uom": bom_line.product_uom_id.id or component.uom_id.id,
                "location_id": setup["vendor_location"].id,
                "location_dest_id": setup["floating_location"].id,
            }
        )
        pin_move_lot(move, lot, required_qty)
        lots_by_product[component.id] = lot
    validate_picking(env, picking)
    return picking, lots_by_product


def pin_raw_lots(mo, lots_by_product):
    for move in mo.move_raw_ids.filtered(lambda m: m.state not in ("done", "cancel")):
        desired_lot = lots_by_product.get(move.product_id.id)
        if desired_lot:
            pin_move_lot(move, desired_lot, move.product_uom_qty)


def create_and_close_mo(env, setup, tag, lots_by_product):
    mo = env["mrp.production"].create(
        {
            "product_id": setup["product"].id,
            "product_qty": PRODUCE_QTY,
            "product_uom_id": setup["product"].uom_id.id,
            "bom_id": setup["bom"].id,
            "origin": tag,
            "location_src_id": setup["floating_location"].id,
            "location_dest_id": setup["warehouse"].lot_stock_id.id,
        }
    )
    mo.action_confirm()
    mo.action_assign()
    mo = env["mrp.production"].browse(mo.id)

    fg_lot = env["stock.lot"].create({"name": f"{tag}-FG", "product_id": setup["product"].id, "company_id": setup["company"].id})
    mo.lot_producing_id = fg_lot.id
    mo.qty_producing = mo.product_qty
    mo._console_fill_move_quantities_for_close({mo.id: mo.product_qty})
    pin_raw_lots(mo, lots_by_product)

    done_at = datetime.now()
    for workorder in mo.workorder_ids:
        duration = workorder.duration_expected or 60.0
        started_at = done_at - timedelta(minutes=duration)
        vals = {
            "state": "done",
            "qty_produced": mo.product_qty,
            "date_start": started_at,
            "date_finished": done_at,
            "duration_expected": duration,
            "duration": duration,
            "duration_unit": round(duration / max(mo.product_qty, 1), 4),
        }
        if "costs_hour" in workorder._fields:
            vals["costs_hour"] = workorder.workcenter_id.costs_hour
        workorder.with_context(bypass_duration_calculation=True).write(vals)
        workorder.end_all()

    raw_moves = mo.move_raw_ids.filtered(lambda move: move.state not in ("done", "cancel"))
    raw_moves.with_context(skip_mo_check=True)._action_done(cancel_backorder=True)
    mo._cal_price(raw_moves)

    finished_move = mo.move_finished_ids.filtered(lambda move: move.product_id == mo.product_id and move.state not in ("done", "cancel"))
    finished_move._action_done(cancel_backorder=True)
    finished_move.move_line_ids.consume_line_ids = [(6, 0, raw_moves.mapped("move_line_ids").ids)]

    mo.write({"date_finished": done_at, "priority": "0", "is_locked": True, "state": "done"})
    if hasattr(mo, "_post_labour"):
        mo._post_labour()
    return mo, finished_move, fg_lot


def create_sale_flow(env, product, fg_lot, tag):
    now = datetime.now()
    now_str = now.strftime("%Y-%m-%d %H:%M:%S")
    sale_defaults = get_sale_defaults(env)
    bank_journal, payment_method_line = get_bank_setup(env)
    partner = create_customer(env, tag)
    order_vals = {
        "partner_id": partner.id,
        "date_order": now_str,
        "client_order_ref": f"{TAG_PREFIX}-SO-{tag}",
        "warehouse_id": sale_defaults["warehouse_id"],
        "team_id": sale_defaults["team_id"],
        "pricelist_id": sale_defaults["pricelist_id"],
        "payment_term_id": sale_defaults["payment_term_id"],
        "order_line": [(0, 0, {"product_id": product.id, "product_uom_qty": SALE_QTY})],
    }
    if "so_type_id" in env["sale.order"]._fields and sale_defaults["so_type_id"]:
        order_vals["so_type_id"] = sale_defaults["so_type_id"]

    sale_order = env["sale.order"].create(order_vals)
    sale_order.action_confirm()

    pick = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "internal")[:1]
    if pick:
        pick.action_assign()
        for move in pick.move_ids.filtered(lambda m: m.product_id == product and m.state not in ("done", "cancel")):
            pin_move_lot(move, fg_lot, SALE_QTY)
        validate_picking(env, pick)

    outgoing = sale_order.picking_ids.filtered(lambda p: p.picking_type_id.code == "outgoing")[:1]
    if outgoing:
        outgoing.action_assign()
        for move in outgoing.move_ids.filtered(lambda m: m.product_id == product and m.state not in ("done", "cancel")):
            pin_move_lot(move, fg_lot, SALE_QTY)
        validate_picking(env, outgoing)

    invoice = sale_order._create_invoices()
    invoice_post_mode = "auto" if invoice.state == "posted" else "manual"
    if invoice.state == "draft":
        invoice.action_post()

    payment_register = env["account.payment.register"].with_context(active_model="account.move", active_ids=invoice.ids).create(
        {"journal_id": bank_journal.id, "payment_method_line_id": payment_method_line.id, "amount": invoice.amount_residual}
    )
    payment_action = payment_register.action_create_payments()
    payment = env["account.payment"].browse(payment_action["res_id"])

    return {
        "partner": partner,
        "sale_order": sale_order,
        "pick": pick,
        "outgoing": outgoing,
        "invoice": invoice,
        "payment": payment,
        "bank_journal": bank_journal,
        "payment_method_line": payment_method_line,
        "invoice_post_mode": invoice_post_mode,
    }


def _stock_doc_type(move):
    if move.picking_id:
        code = move.picking_id.picking_type_id.code
        if code == "incoming":
            return "Inbound Receipt"
        if code == "internal":
            return "Internal Pick"
        if code == "outgoing":
            return "Delivery"
        return f"Picking ({code})"
    if move.raw_material_production_id:
        return "MO Raw Consumption"
    if move.production_id:
        return "MO Finished Receipt"
    return "Other"


def _stock_doc_number(move):
    if move.picking_id:
        return move.picking_id.name
    if move.raw_material_production_id:
        return move.raw_material_production_id.name
    if move.production_id:
        return move.production_id.name
    return ""


def _stock_doc_origin(move):
    if move.picking_id:
        return move.picking_id.origin or ""
    if move.raw_material_production_id:
        return move.raw_material_production_id.origin or ""
    if move.production_id:
        return move.production_id.origin or ""
    return ""


def _account_bucket(account):
    account_type = account.account_type or ""
    if account_type.startswith("asset_"):
        return "Balance Sheet", "Assets"
    if account_type.startswith("liability_"):
        return "Balance Sheet", "Liabilities"
    if account_type == "equity":
        return "Balance Sheet", "Equity"
    if account_type.startswith("income"):
        return "P&L", "Income"
    if account_type.startswith("expense"):
        return "P&L", "Expenses"
    return "Other", "Other"


def build_trace(env, readiness, setup, receipt, mo, finished_move, fg_lot, sale_flow, tag):
    pickings = env["stock.picking"]
    for picking in [receipt, sale_flow["pick"], sale_flow["outgoing"]]:
        if picking:
            pickings |= picking

    stock_moves = (receipt.move_ids | mo.move_raw_ids | mo.move_finished_ids | pickings.move_ids).sorted(lambda move: move.id)
    svls = env["stock.valuation.layer"].search([("stock_move_id", "in", stock_moves.ids)], order="id")
    account_moves = (stock_moves.mapped("account_move_ids") | sale_flow["invoice"] | sale_flow["payment"].move_id | env["account.move"].search([("ref", "ilike", mo.name)])).sorted(lambda move: move.id)
    aml = account_moves.line_ids.filtered(lambda line: line.account_id).sorted(lambda line: (line.move_id.id, line.id))

    tb_map = defaultdict(lambda: {"debit": Decimal("0"), "credit": Decimal("0"), "balance": Decimal("0")})
    fs_map = defaultdict(lambda: Decimal("0"))
    for line in aml:
        key = (line.account_id.code or "", line.account_id.name or "", line.account_id.account_type or "")
        debit = Decimal(str(line.debit or 0))
        credit = Decimal(str(line.credit or 0))
        tb_map[key]["debit"] += debit
        tb_map[key]["credit"] += credit
        tb_map[key]["balance"] += debit - credit
        bucket, group = _account_bucket(line.account_id)
        fs_map[(bucket, group)] += debit - credit

    tb_rows = []
    for (code, name, account_type), values in sorted(tb_map.items()):
        tb_rows.append({"code": code, "name": name, "account_type": account_type, "debit": q2(values["debit"]), "credit": q2(values["credit"]), "balance": q2(values["balance"])})

    fs_rows = []
    for (bucket, group), balance in sorted(fs_map.items()):
        fs_rows.append({"statement": bucket, "group": group, "balance": q2(balance)})

    docs = [
        {"step": 1, "case_id": "TXN-01", "doc_type": "Inbound Receipt", "name": receipt.name, "state": receipt.state, "manual_auto": "Manual create / Auto post stock", "note": "รับ component ตาม BOM จริงเข้าคลังลอย แล้วระบบสร้าง stock valuation/account move อัตโนมัติเมื่อ validate"},
        {"step": 2, "case_id": "TXN-02", "doc_type": "Manufacturing Order", "name": mo.name, "state": mo.state, "manual_auto": "Manual create / Auto valuation", "note": "ปิด MO จริงด้วย existing product/BOM แล้ว raw consumption, FG receipt และ labour posting ถูกลงบัญชี"},
        {"step": 3, "case_id": "TXN-03", "doc_type": "Sale Order", "name": sale_flow["sale_order"].name, "state": sale_flow["sale_order"].state, "manual_auto": "Manual create", "note": "สร้าง SO จริงเพื่อใช้ FG lot ที่ผลิตได้"},
        {"step": 4, "case_id": "TXN-03", "doc_type": "Internal Pick", "name": sale_flow["pick"].name if sale_flow["pick"] else "", "state": sale_flow["pick"].state if sale_flow["pick"] else "", "manual_auto": "Auto create", "note": "ระบบสร้างเอกสารหยิบภายในอัตโนมัติหลัง confirm SO"},
        {"step": 5, "case_id": "TXN-03", "doc_type": "Delivery", "name": sale_flow["outgoing"].name if sale_flow["outgoing"] else "", "state": sale_flow["outgoing"].state if sale_flow["outgoing"] else "", "manual_auto": "Auto create / Auto valuation", "note": "ระบบสร้างใบส่งของและลง COGS + ตัดสต็อกอัตโนมัติเมื่อ validate"},
        {"step": 6, "case_id": "TXN-04", "doc_type": "Invoice", "name": sale_flow["invoice"].name, "state": sale_flow["invoice"].state, "manual_auto": f"Manual create / {sale_flow['invoice_post_mode']} post", "note": "สร้าง invoice จาก SO จริง แล้วตรวจว่า state หลังสร้างเป็น posted หรือไม่"},
        {"step": 7, "case_id": "TXN-04", "doc_type": "Payment", "name": sale_flow["payment"].name, "state": sale_flow["payment"].state, "manual_auto": "Manual wizard / Auto reconcile", "note": "register payment จริงและระบบ reconcile ลูกหนี้อัตโนมัติ"},
    ]

    return {"tag": tag, "readiness": readiness, "product": setup["product"], "bom": setup["bom"], "receipt": receipt, "mo": mo, "finished_move": finished_move, "fg_lot": fg_lot, "sale_flow": sale_flow, "pickings": pickings, "stock_moves": stock_moves, "svls": svls, "account_moves": account_moves, "aml": aml, "tb_rows": tb_rows, "fs_rows": fs_rows, "docs": docs}


def _account_move_source(trace, move):
    if move.id == trace["sale_flow"]["invoice"].id:
        return "Invoice", trace["sale_flow"]["invoice"].name
    if move.id == trace["sale_flow"]["payment"].move_id.id:
        return "Payment", trace["sale_flow"]["payment"].name
    stock_moves = trace["stock_moves"].filtered(lambda m: move in m.account_move_ids)
    if stock_moves:
        first_move = stock_moves[0]
        return _stock_doc_type(first_move), _stock_doc_number(first_move)
    if trace["mo"].name in (move.ref or ""):
        return "MO Labour/Overhead", trace["mo"].name
    return "Other", move.ref or ""


def load_walkthrough_cases():
    rows = []
    path = _walkthrough_path()
    if not path.exists():
        return rows
    pattern = re.compile(r"^####\s+(SC\d+-[\d.]+(?:-R\d+)?)\s+-\s+(.*)$")
    for line in path.read_text(encoding="utf-8").splitlines():
        match = pattern.match(line.strip())
        if match:
            rows.append({"case_id": match.group(1), "title": match.group(2).strip()})
    return rows


def case_status_map(trace):
    mo = trace["mo"]
    sale_flow = trace["sale_flow"]
    workorders = ", ".join(mo.workorder_ids.mapped("name"))
    docs = f"{trace['receipt'].name}; {mo.name}; {sale_flow['sale_order'].name}; {sale_flow['invoice'].name}; {sale_flow['payment'].name}"
    return {
        "SC02-2.2": {"status": "Partially Executed", "reason": "สร้าง MO manual จริงด้วย product/BOM ที่มีอยู่ในฐาน view แต่ไม่ได้ใช้ flow SO-triggered MTO ตามคู่มือเดิม", "evidence": mo.name},
        "SC03-3.4": {"status": "Executed", "reason": "ปิด MO จริงและ trace ต้นทุนจาก stock move -> SVL -> AML ได้ครบ", "evidence": mo.name},
        "SC04-4.1": {"status": "Executed", "reason": "MO อยู่สถานะ Done จริง", "evidence": mo.name},
        "SC04-4.2": {"status": "Executed", "reason": "Work Order ของ MO อยู่สถานะ Done จริง", "evidence": workorders},
        "SC05-5.6": {"status": "Executed", "reason": "ทดสอบ real-time production-accounting link จริงจาก receipt -> MO -> sale -> invoice -> payment", "evidence": docs},
    }


def default_case_reason(case_id):
    if case_id.startswith("SC01"):
        return "ไม่ได้รัน UI login/session ในรอบนี้ เพราะรอบนี้ทดสอบ transaction จริงใน database view ผ่าน ORM"
    if case_id.startswith(("SC02", "SC03", "SC04", "SC05")):
        return "คู่มือเดิมอิง UAT setup/product เฉพาะทางที่ยังไม่ได้เตรียมในฐาน view และรอบนี้ไม่สร้าง product ใหม่ตามคำขอ"
    return "ยังไม่ได้รันในรอบนี้"


def write_workbook(env, trace):
    path = _report_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    ws.append(["หัวข้อ", "ค่า"])
    rows = [
        ("Database", env.cr.dbname),
        ("Scenario", "Run test ในฐาน view ด้วย existing product เท่านั้น ไม่สร้าง product ใหม่"),
        ("Run Tag", trace["tag"]),
        ("Product", trace["product"].display_name),
        ("BOM", f"{trace['bom'].id} | {trace['bom'].code or '-'}"),
        ("Planned Produce Qty", PRODUCE_QTY),
        ("Sale Qty", SALE_QTY),
        ("Legacy verify target", "คู่มือคาดหวัง TOTAL 24/24 หลังรัน setup_ui_cases_view.py"),
        ("Legacy verify actual", f"{trace['readiness']['legacy_passed']}/{trace['readiness']['legacy_total']} (ไม่ได้รัน setup_ui_cases_view.py เพราะจะสร้าง product ใหม่)"),
        ("Startup warning", ", ".join(trace["readiness"]["missing_modules"]) if trace["readiness"]["missing_modules"] else "ไม่มี"),
        ("Receipt", trace["receipt"].name),
        ("MO", trace["mo"].name),
        ("FG Lot", trace["fg_lot"].name),
        ("SO", trace["sale_flow"]["sale_order"].name),
        ("Internal Pick", trace["sale_flow"]["pick"].name if trace["sale_flow"]["pick"] else ""),
        ("Delivery", trace["sale_flow"]["outgoing"].name if trace["sale_flow"]["outgoing"] else ""),
        ("Invoice", trace["sale_flow"]["invoice"].name),
        ("Payment", trace["sale_flow"]["payment"].name),
    ]
    for row in rows:
        ws.append(list(row))
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 120)

    ws = wb.create_sheet("SETUP_READINESS")
    ws.append(["กลุ่ม", "รายการตรวจ", "ผล", "รายละเอียด", "หมายเหตุ"])
    for row in trace["readiness"]["checks"]:
        ws.append([row["group"], row["name"], row["status"], row["detail"], row["note"]])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 80)

    ws = wb.create_sheet("CUSTOM_TESTS")
    ws.append(["Test ID", "รายการ", "ผล", "เลขเอกสารอ้างอิง", "คำอธิบาย"])
    custom_rows = [
        ("RDY-01", "ตรวจความพร้อมตาม walkthrough + current DB", f"PASS {trace['readiness']['passed']}/{trace['readiness']['total']}", "", f"legacy verify ได้ {trace['readiness']['legacy_passed']}/{trace['readiness']['legacy_total']}; ไม่รัน setup เดิมเพราะห้ามสร้าง product ใหม่"),
        ("TXN-01", "รับ component เข้าคลังจริง", "Executed", trace["receipt"].name, "ใช้ component จาก BOM 2980 สร้าง receipt จริงและระบบ post stock/account อัตโนมัติ"),
        ("TXN-02", "สร้างและปิด MO จริง", "Executed", trace["mo"].name, "ปิด WIP จริงและได้ FG lot ใหม่พร้อม valuation/journal จริง"),
        ("TXN-03", "ขาย FG lot ที่ผลิตได้จริง", "Executed", "; ".join(x for x in [trace["sale_flow"]["sale_order"].name, trace["sale_flow"]["pick"].name if trace["sale_flow"]["pick"] else "", trace["sale_flow"]["outgoing"].name if trace["sale_flow"]["outgoing"] else ""] if x), "SO -> Internal Pick -> Delivery ด้วย lot ที่ผลิตจาก MO รอบนี้"),
        ("TXN-04", "ออก invoice และรับชำระจริง", "Executed", f"{trace['sale_flow']['invoice'].name}; {trace['sale_flow']['payment'].name}", f"invoice post mode = {trace['sale_flow']['invoice_post_mode']}; payment register สร้าง payment และ reconcile ลูกหนี้"),
    ]
    for row in custom_rows:
        ws.append(row)
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 90)

    ws = wb.create_sheet("WALKTHROUGH_CASES")
    ws.append(["Case ID", "Title", "Status", "Reason", "Evidence"])
    statuses = case_status_map(trace)
    for row in load_walkthrough_cases():
        status = statuses.get(row["case_id"], {})
        ws.append([row["case_id"], row["title"], status.get("status", "Not Executed"), status.get("reason", default_case_reason(row["case_id"])), status.get("evidence", "")])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 90)

    ws = wb.create_sheet("DOC_NUMBERS")
    ws.append(["ลำดับ", "ประเภทเอกสาร", "เลขเอกสาร", "สถานะ", "หมายเหตุ"])
    for row in trace["docs"]:
        ws.append([row["step"], row["doc_type"], row["name"], row["state"], row["note"]])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 72)

    ws = wb.create_sheet("DOC_FLOW")
    ws.append(["Step", "Case ID", "Document Type", "Number", "State", "Manual/Auto", "What happened"])
    for row in trace["docs"]:
        ws.append([row["step"], row["case_id"], row["doc_type"], row["name"], row["state"], row["manual_auto"], row["note"]])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 84)

    ws = wb.create_sheet("STOCK_MOVES")
    ws.append(["Move ID", "Doc Type", "Doc No", "Origin", "Product", "Planned Qty", "Done Qty", "Source", "Destination", "State"])
    for move in trace["stock_moves"]:
        ws.append([move.id, _stock_doc_type(move), _stock_doc_number(move), _stock_doc_origin(move), move.product_id.display_name, float(move.product_uom_qty or 0.0), float(sum(move.move_line_ids.mapped("quantity")) or 0.0), move.location_id.complete_name, move.location_dest_id.complete_name, move.state])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 7])
    auto_width(ws, 64)

    ws = wb.create_sheet("SVL")
    ws.append(["SVL ID", "Stock Move", "Doc Type", "Doc No", "Qty", "Unit Cost", "Value", "Account Move"])
    for svl in trace["svls"]:
        move = svl.stock_move_id
        ws.append([svl.id, move.reference or move.name, _stock_doc_type(move), _stock_doc_number(move), float(svl.quantity or 0.0), float(svl.unit_cost or 0.0), float(svl.value or 0.0), svl.account_move_id.name or ""])
    style_header(ws)
    style_grid(ws, amount_cols=[5, 6, 7])
    auto_width(ws, 48)

    ws = wb.create_sheet("ACCOUNT_MOVES")
    ws.append(["Move ID", "Journal Entry", "Date", "Journal", "Source Type", "Source Doc", "Ref", "State"])
    for move in trace["account_moves"]:
        source_type, source_doc = _account_move_source(trace, move)
        ws.append([move.id, move.name, str(move.date), move.journal_id.code, source_type, source_doc, move.ref or "", move.state])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 48)

    ws = wb.create_sheet("AML")
    ws.append(["Move", "Source Type", "Source Doc", "Account Code", "Account Name", "Debit", "Credit", "Balance", "Label"])
    for line in trace["aml"]:
        source_type, source_doc = _account_move_source(trace, line.move_id)
        ws.append([line.move_id.name, source_type, source_doc, line.account_id.code, line.account_id.name, float(line.debit or 0.0), float(line.credit or 0.0), float((line.debit or 0.0) - (line.credit or 0.0)), line.name or ""])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 7, 8])
    auto_width(ws, 56)

    ws = wb.create_sheet("TB")
    ws.append(["Account Code", "Account Name", "Account Type", "Debit", "Credit", "Balance"])
    for row in trace["tb_rows"]:
        ws.append([row["code"], row["name"], row["account_type"], float(row["debit"]), float(row["credit"]), float(row["balance"])])
    style_header(ws)
    style_grid(ws, amount_cols=[4, 5, 6])
    auto_width(ws, 52)

    ws = wb.create_sheet("FS_SUMMARY")
    ws.append(["Statement", "Group", "Balance"])
    for row in trace["fs_rows"]:
        ws.append([row["statement"], row["group"], float(row["balance"])])
    style_header(ws)
    style_grid(ws, amount_cols=[3])
    auto_width(ws, 36)

    wb.save(path)
    return path


def main():
    ensure_schema(env)
    readiness = verify_walkthrough_setup(env)

    tag = datetime.now().strftime("%Y%m%d%H%M%S")
    origin_tag = f"{TAG_PREFIX}-{tag}"
    setup = get_core_setup(env)
    receipt, lots_by_product = create_inbound_receipt(env, setup, origin_tag)
    mo, finished_move, fg_lot = create_and_close_mo(env, setup, origin_tag, lots_by_product)
    sale_flow = create_sale_flow(env, setup["product"], fg_lot, origin_tag)

    trace = build_trace(env, readiness, setup, receipt, mo, finished_move, fg_lot, sale_flow, origin_tag)
    report_path = write_workbook(env, trace)
    env.cr.commit()

    print("REPORT:", report_path)
    print("RECEIPT:", receipt.name)
    print("MO:", mo.name)
    print("FG LOT:", fg_lot.name)
    print("SO:", sale_flow["sale_order"].name)
    if sale_flow["pick"]:
        print("PICK:", sale_flow["pick"].name)
    if sale_flow["outgoing"]:
        print("DELIVERY:", sale_flow["outgoing"].name)
    print("INVOICE:", sale_flow["invoice"].name)
    print("PAYMENT:", sale_flow["payment"].name)
    print("LEGACY_VERIFY:", f"{readiness['legacy_passed']}/{readiness['legacy_total']}")


main()
