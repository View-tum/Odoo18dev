from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PRODUCT_CODE = "FG-PSS-TH-01001"
BOM_ID = 2980
PRODUCE_QTY = 200
REPORT_NAME = "poysian_inbound_mo_detailed_trace.xlsx"
TAG_PREFIX = "SIM-INB-MO-POYSIAN"
VENDOR_NAME_PREFIX = "SIM VENDOR POYSIAN INBOUND MO"


def q2(value):
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _report_path() -> Path:
    return Path(__file__).resolve().parents[3] / "reports" / REPORT_NAME


def _header_style():
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _cell_border():
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_sheet(ws):
    fill, font = _header_style()
    border = _cell_border()
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    for column in ws.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(values, default=10) + 2, 42)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width
    ws.freeze_panes = "A2"


def _append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_sheet(ws)


def _optional_partner_vals(env, name):
    vals = {"name": name, "supplier_rank": 1}
    if "approval_state" in env["res.partner"]._fields:
        vals["approval_state"] = "approved"
    return vals


def _account_bucket(account):
    account_type = account.account_type or ""
    if account_type.startswith("asset_"):
        return "Balance Sheet", "Assets"
    if account_type.startswith("liability_"):
        return "Balance Sheet", "Liabilities"
    if account_type == "equity":
        return "Balance Sheet", "Equity"
    if account_type.startswith("income"):
        return "P&L", "Income"
    if account_type.startswith("expense"):
        return "P&L", "Expenses"
    if account_type == "off_balance":
        return "Off Balance", "Off Balance"
    return "Other", "Other"


def _doc_type(move):
    if move.picking_id:
        return "Inbound Receipt"
    if move.raw_material_production_id:
        return "MO Raw Consumption"
    if move.production_id:
        return "MO Finished Receipt"
    return "Other"


def _doc_number(move):
    if move.picking_id:
        return move.picking_id.name
    if move.raw_material_production_id:
        return move.raw_material_production_id.name
    if move.production_id:
        return move.production_id.name
    return ""


def _doc_origin(move):
    if move.picking_id:
        return move.picking_id.origin or ""
    if move.raw_material_production_id:
        return move.raw_material_production_id.origin or ""
    if move.production_id:
        return move.production_id.origin or ""
    return ""


def _doc_date(move):
    if move.picking_id:
        return move.picking_id.date_done or move.picking_id.scheduled_date
    if move.raw_material_production_id:
        return move.raw_material_production_id.date_finished or move.raw_material_production_id.date_start
    if move.production_id:
        return move.production_id.date_finished or move.production_id.date_start
    return move.date


def ensure_schema(env):
    env.cr.execute(
        "ALTER TABLE stock_picking ADD COLUMN IF NOT EXISTS manufacturing_type varchar"
    )
    env.cr.execute(
        "ALTER TABLE sale_sequence_type ADD COLUMN IF NOT EXISTS is_full_tax_invoice boolean DEFAULT true"
    )


def get_core_setup(env):
    product = env["product.product"].search([("default_code", "=", PRODUCT_CODE)], limit=1)
    if not product:
        raise ValueError(f"Product {PRODUCT_CODE} not found")
    bom = env["mrp.bom"].browse(BOM_ID)
    if not bom.exists():
        raise ValueError(f"BOM {BOM_ID} not found")
    incoming_type = env["stock.picking.type"].browse(1)
    if not incoming_type.exists():
        raise ValueError("Incoming picking type 1 not found")
    vendor_location = env["stock.location"].search([("usage", "=", "supplier")], limit=1)
    preprod_location = env["stock.location"].search(
        [("complete_name", "=", "GMP/Pre-Production")], limit=1
    )
    if not vendor_location or not preprod_location:
        raise ValueError("Required stock locations not found")
    return {
        "product": product,
        "bom": bom,
        "incoming_type": incoming_type,
        "vendor_location": vendor_location,
        "preprod_location": preprod_location,
        "company": product.company_id or env.company,
    }


def create_vendor(env, tag):
    return env["res.partner"].create(
        _optional_partner_vals(env, f"{VENDOR_NAME_PREFIX} {tag}")
    )


def create_inbound_receipt(env, setup, tag):
    picking = env["stock.picking"].create(
        {
            "picking_type_id": setup["incoming_type"].id,
            "location_id": setup["vendor_location"].id,
            "location_dest_id": setup["preprod_location"].id,
            "partner_id": create_vendor(env, tag).id,
            "scheduled_date": datetime.now(),
            "invoice_reference": tag,
            "invoice_date": datetime.now().date(),
            "origin": tag,
        }
    )
    lots_by_product = {}
    for bom_line in setup["bom"].bom_line_ids:
        component = bom_line.product_id
        lot = env["stock.lot"].create(
            {
                "name": f"{tag}-{component.id}",
                "product_id": component.id,
                "company_id": setup["company"].id,
            }
        )
        move = env["stock.move"].create(
            {
                "name": component.display_name,
                "picking_id": picking.id,
                "product_id": component.id,
                "product_uom_qty": bom_line.product_qty,
                "product_uom": component.uom_id.id,
                "location_id": setup["vendor_location"].id,
                "location_dest_id": setup["preprod_location"].id,
            }
        )
        move.quantity = bom_line.product_qty
        for move_line in move.move_line_ids:
            move_line.write({"lot_id": lot.id, "quantity": bom_line.product_qty})
        lots_by_product[component.id] = lot
    result = picking.button_validate()
    if isinstance(result, dict) and result.get("res_model") == "stock.backorder.confirmation":
        env[result["res_model"]].browse(result["res_id"]).process()
    return picking, lots_by_product


def pin_raw_lots(mo, lots_by_product):
    for move in mo.move_raw_ids.filtered(lambda m: m.state not in ("done", "cancel")):
        desired_lot = lots_by_product.get(move.product_id.id)
        if not desired_lot:
            continue
        target_qty = move.product_uom_qty
        quant = move.env["stock.quant"].search(
            [
                ("product_id", "=", move.product_id.id),
                ("lot_id", "=", desired_lot.id),
                ("quantity", ">", 0),
                ("location_id", "child_of", move.location_id.id),
            ],
            order="in_date, id",
            limit=1,
        )
        source_location = quant.location_id.id if quant else move.location_id.id
        if move.move_line_ids:
            primary = move.move_line_ids[0]
            primary.write(
                {
                    "lot_id": desired_lot.id,
                    "quantity": target_qty,
                    "location_id": source_location,
                    "picked": True,
                }
            )
            if len(move.move_line_ids) > 1:
                move.move_line_ids[1:].unlink()
        else:
            vals = move._prepare_move_line_vals(quantity=target_qty, reserved_quant=False)
            vals.update(
                {
                    "lot_id": desired_lot.id,
                    "quantity": target_qty,
                    "location_id": source_location,
                    "picked": True,
                }
            )
            move.env["stock.move.line"].create(vals)
        move.write({"quantity": target_qty, "picked": True})


def create_and_close_mo(env, setup, tag, lots_by_product):
    mo = env["mrp.production"].create(
        {
            "product_id": setup["product"].id,
            "product_qty": PRODUCE_QTY,
            "product_uom_id": setup["product"].uom_id.id,
            "bom_id": setup["bom"].id,
            "origin": tag,
        }
    )
    mo.action_confirm()
    mo.action_assign()
    mo = env["mrp.production"].browse(mo.id)

    fg_lot = env["stock.lot"].create(
        {
            "name": f"{tag}-FG",
            "product_id": setup["product"].id,
            "company_id": setup["company"].id,
        }
    )
    mo.lot_producing_id = fg_lot.id
    mo.qty_producing = mo.product_qty
    mo._console_fill_move_quantities_for_close({mo.id: mo.product_qty})
    pin_raw_lots(mo, lots_by_product)

    done_at = datetime.now()
    for workorder in mo.workorder_ids:
        duration = workorder.duration_expected or 60.0
        started_at = done_at - timedelta(minutes=duration)
        workorder.with_context(bypass_duration_calculation=True).write(
            {
                "state": "done",
                "qty_produced": mo.product_qty,
                "date_start": started_at,
                "date_finished": done_at,
                "costs_hour": workorder.workcenter_id.costs_hour,
                "duration_expected": duration,
                "duration": duration,
                "duration_unit": round(duration / max(mo.product_qty, 1), 2),
            }
        )
        workorder.end_all()

    raw_moves = mo.move_raw_ids.filtered(lambda move: move.state not in ("done", "cancel"))
    raw_moves.with_context(skip_mo_check=True)._action_done(cancel_backorder=True)
    mo._cal_price(raw_moves)

    finished_move = mo.move_finished_ids.filtered(
        lambda move: move.product_id == mo.product_id and move.state not in ("done", "cancel")
    )
    finished_move._action_done(cancel_backorder=True)
    finished_move.move_line_ids.consume_line_ids = [(6, 0, raw_moves.mapped("move_line_ids").ids)]

    mo.write(
        {
            "date_finished": done_at,
            "priority": "0",
            "is_locked": True,
            "state": "done",
        }
    )
    mo._post_labour()

    return mo, finished_move, fg_lot


def build_trace(flow):
    receipt = flow["receipt"]
    mo = flow["mo"]
    product = flow["product"]
    stock_moves = (receipt.move_ids | mo.move_raw_ids | mo.move_finished_ids).sorted(
        lambda move: move.id
    )
    move_lines = stock_moves.move_line_ids.sorted(lambda line: line.id)
    svls = stock_moves.stock_valuation_layer_ids.sorted(lambda svl: svl.id)
    account_moves = (
        stock_moves.mapped("account_move_ids")
        | mo.env["account.move"].search([("ref", "ilike", mo.name)])
    ).sorted(lambda move: move.id)
    aml = account_moves.line_ids.filtered(lambda line: line.account_id).sorted(
        lambda line: (line.move_id.id, line.id)
    )

    receipt_component_total = q2(sum(receipt.move_ids.mapped("stock_valuation_layer_ids.value")))
    finished_move = flow["finished_move"]
    finished_svl_total = q2(sum(finished_move.stock_valuation_layer_ids.mapped("value")))
    workcenter_cost = q2(sum(flow["mo"].workorder_ids.mapped(lambda wo: wo._cal_cost())))

    tb_map = defaultdict(lambda: {"debit": Decimal("0"), "credit": Decimal("0"), "balance": Decimal("0")})
    for line in aml:
        key = (
            line.account_id.code or "",
            line.account_id.name or "",
            line.account_id.account_type or "",
        )
        tb_map[key]["debit"] += Decimal(str(line.debit or 0))
        tb_map[key]["credit"] += Decimal(str(line.credit or 0))
        tb_map[key]["balance"] += Decimal(str(line.debit or 0)) - Decimal(str(line.credit or 0))

    tb_rows = []
    for (code, name, account_type), values in sorted(tb_map.items()):
        section, bucket = _account_bucket(type("Account", (), {"account_type": account_type})())
        tb_rows.append(
            {
                "account_code": code,
                "account_name": name,
                "account_type": account_type,
                "section": section,
                "bucket": bucket,
                "debit": q2(values["debit"]),
                "credit": q2(values["credit"]),
                "balance": q2(values["balance"]),
            }
        )

    fin_rows = []
    for row in tb_rows:
        amount = row["balance"]
        if row["section"] == "Balance Sheet" and row["bucket"] in ("Liabilities", "Equity"):
            amount = -amount
        if row["section"] == "P&L" and row["bucket"] == "Income":
            amount = -amount
        fin_rows.append(
            {
                "section": row["section"],
                "bucket": row["bucket"],
                "account_code": row["account_code"],
                "account_name": row["account_name"],
                "amount": q2(amount),
            }
        )

    trace_rows = []
    for move in stock_moves:
        svl_list = move.stock_valuation_layer_ids.sorted(lambda svl: svl.id)
        move_lines_text = ", ".join(
            f"{line.product_id.default_code or ''}:{line.quantity_product_uom or line.quantity}:{line.lot_id.name or ''}"
            for line in move.move_line_ids.sorted(lambda line: line.id)
        )
        if svl_list:
            for svl in svl_list:
                aml_list = svl.account_move_id.line_ids.filtered(lambda line: line.account_id).sorted(
                    lambda line: line.id
                )
                for line in aml_list:
                    section, bucket = _account_bucket(line.account_id)
                    trace_rows.append(
                        [
                            _doc_type(move),
                            _doc_number(move),
                            _doc_origin(move),
                            _doc_date(move),
                            move.id,
                            move.reference or move.name,
                            move.product_id.default_code or "",
                            move.product_id.display_name,
                            move.state,
                            move.quantity,
                            move.product_uom_qty,
                            move.location_id.complete_name,
                            move.location_dest_id.complete_name,
                            move_lines_text,
                            svl.id,
                            svl.quantity,
                            svl.unit_cost,
                            svl.value,
                            svl.account_move_id.name if svl.account_move_id else "",
                            line.id,
                            line.account_id.code or "",
                            line.account_id.name or "",
                            line.account_id.account_type or "",
                            line.debit,
                            line.credit,
                            line.balance,
                            section,
                            bucket,
                        ]
                    )
        else:
            trace_rows.append(
                [
                    _doc_type(move),
                    _doc_number(move),
                    _doc_origin(move),
                    _doc_date(move),
                    move.id,
                    move.reference or move.name,
                    move.product_id.default_code or "",
                    move.product_id.display_name,
                    move.state,
                    move.quantity,
                    move.product_uom_qty,
                    move.location_id.complete_name,
                    move.location_dest_id.complete_name,
                    move_lines_text,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    traced_account_move_ids = {move.id for move in stock_moves.mapped("account_move_ids")}
    extra_moves = account_moves.filtered(lambda move: move.id not in traced_account_move_ids)
    for move in extra_moves:
        for line in move.line_ids.filtered(lambda aml_line: aml_line.account_id).sorted(lambda aml_line: aml_line.id):
            section, bucket = _account_bucket(line.account_id)
            trace_rows.append(
                [
                    "MO Overhead",
                    mo.name,
                    mo.origin or "",
                    move.date,
                    "",
                    move.ref or move.name,
                    product.default_code or "",
                    product.display_name,
                    mo.state,
                    "",
                    mo.product_qty,
                    mo.location_src_id.complete_name,
                    mo.location_dest_id.complete_name,
                    "",
                    "",
                    "",
                    "",
                    "",
                    move.name,
                    line.id,
                    line.account_id.code or "",
                    line.account_id.name or "",
                    line.account_id.account_type or "",
                    line.debit,
                    line.credit,
                    line.balance,
                    section,
                    bucket,
                ]
            )

    lot_trace_rows = []
    for fg_line in finished_move.move_line_ids.sorted(lambda line: line.id):
        for consume_line in fg_line.consume_line_ids.sorted(lambda line: line.id):
            inbound_line = receipt.move_line_ids.filtered(
                lambda line, c=consume_line: line.product_id == c.product_id and line.lot_id == c.lot_id
            )[:1]
            lot_trace_rows.append(
                [
                    mo.name,
                    fg_line.lot_id.name if fg_line.lot_id else "",
                    fg_line.quantity_product_uom,
                    consume_line.product_id.default_code or "",
                    consume_line.product_id.display_name,
                    consume_line.lot_id.name if consume_line.lot_id else "",
                    consume_line.quantity_product_uom,
                    consume_line.move_id.id,
                    consume_line.move_id.name,
                    inbound_line.picking_id.name if inbound_line else "",
                    inbound_line.id if inbound_line else "",
                ]
            )

    summary = {
        "tag": flow["tag"],
        "product_code": product.default_code,
        "product_name": product.display_name,
        "receipt_name": receipt.name,
        "receipt_origin": receipt.origin,
        "mo_name": mo.name,
        "fg_lot": flow["fg_lot"].name,
        "produce_qty": mo.product_qty,
        "receipt_component_value": receipt_component_total,
        "finished_value": finished_svl_total,
        "workcenter_cost": workcenter_cost,
        "wip_balance": q2(
            sum(
                line.debit - line.credit
                for line in aml
                if (line.account_id.code or "") == "116021"
            )
        ),
        "account_move_names": ", ".join(account_moves.mapped("name")),
    }

    return {
        "summary": summary,
        "receipt": receipt,
        "mo": mo,
        "product": product,
        "stock_moves": stock_moves,
        "move_lines": move_lines,
        "svls": svls,
        "account_moves": account_moves,
        "aml": aml,
        "tb_rows": tb_rows,
        "fin_rows": fin_rows,
        "trace_rows": trace_rows,
        "lot_trace_rows": lot_trace_rows,
    }


def write_report(trace):
    report_path = _report_path()
    report_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    summary = trace["summary"]
    readme_rows = [
        ["Item", "Value"],
        ["Tag", summary["tag"]],
        ["Product", f'{summary["product_code"]} | {summary["product_name"]}'],
        ["Inbound Receipt", summary["receipt_name"]],
        ["MO", summary["mo_name"]],
        ["FG Lot", summary["fg_lot"]],
        ["Produce Qty", float(summary["produce_qty"])],
        ["Receipt Component Value", float(summary["receipt_component_value"])],
        ["Finished Goods Value", float(summary["finished_value"])],
        ["Workcenter Cost Absorbed", float(summary["workcenter_cost"])],
        ["WIP Ending Balance", float(summary["wip_balance"])],
        ["Stock Journal Entries", summary["account_move_names"]],
        [
            "Observation",
            "Current custom _post_inventory/button_mark_done path cancels the finished move in this database; "
            "the script used a lower-level ORM close sequence that still creates real stock/accounting entries.",
        ],
    ]
    for row in readme_rows:
        ws.append(row)
    _style_sheet(ws)

    doc_rows = [
        [
            "Inbound Receipt",
            trace["receipt"].name,
            trace["receipt"].state,
            trace["receipt"].origin or "",
            trace["receipt"].partner_id.display_name,
            trace["receipt"].location_id.complete_name,
            trace["receipt"].location_dest_id.complete_name,
            trace["receipt"].date_done,
            "Manual create/validate; Odoo auto-created stock moves, SVL, and stock journals.",
        ],
        [
            "Manufacturing Order",
            trace["mo"].name,
            trace["mo"].state,
            trace["mo"].origin or "",
            "",
            trace["mo"].location_src_id.complete_name,
            trace["mo"].location_dest_id.complete_name,
            trace["mo"].date_finished,
            "Manual create/confirm/assign; Odoo auto-created workorder, raw moves, finished move, SVL, and stock journals.",
        ],
    ]
    ws = wb.create_sheet("DOC_FLOW")
    _append_rows(
        ws,
        ["doc_type", "doc_no", "state", "origin", "partner", "source", "destination", "date_done", "automation_note"],
        doc_rows,
    )

    document_line_rows = []
    for move in trace["receipt"].move_ids.sorted(lambda move: move.id):
        document_line_rows.append(
            [
                "Inbound Receipt",
                trace["receipt"].name,
                move.id,
                move.product_id.default_code or "",
                move.product_id.display_name,
                move.product_uom_qty,
                move.quantity,
                ", ".join(move.move_line_ids.mapped("lot_id.name")),
            ]
        )
    for move in trace["mo"].move_raw_ids.sorted(lambda move: move.id):
        document_line_rows.append(
            [
                "MO Raw",
                trace["mo"].name,
                move.id,
                move.product_id.default_code or "",
                move.product_id.display_name,
                move.product_uom_qty,
                move.quantity,
                ", ".join(move.move_line_ids.mapped("lot_id.name")),
            ]
        )
    for move in trace["mo"].move_finished_ids.sorted(lambda move: move.id):
        document_line_rows.append(
            [
                "MO Finished",
                trace["mo"].name,
                move.id,
                move.product_id.default_code or "",
                move.product_id.display_name,
                move.product_uom_qty,
                move.quantity,
                ", ".join(move.move_line_ids.mapped("lot_id.name")),
            ]
        )
    ws = wb.create_sheet("DOCUMENT_LINES")
    _append_rows(
        ws,
        ["doc_type", "doc_no", "stock_move_id", "product_code", "product_name", "demand_qty", "done_qty", "lots"],
        document_line_rows,
    )

    move_rows = []
    for move in trace["stock_moves"]:
        move_rows.append(
            [
                _doc_type(move),
                _doc_number(move),
                move.id,
                move.reference or move.name,
                move.product_id.default_code or "",
                move.product_id.display_name,
                move.state,
                move.product_uom_qty,
                move.quantity,
                move.location_id.complete_name,
                move.location_dest_id.complete_name,
                move.price_unit,
                ", ".join(move.account_move_ids.mapped("name")),
            ]
        )
    ws = wb.create_sheet("STOCK_MOVES")
    _append_rows(
        ws,
        [
            "doc_type",
            "doc_no",
            "move_id",
            "move_ref",
            "product_code",
            "product_name",
            "state",
            "demand_qty",
            "done_qty",
            "source",
            "destination",
            "price_unit",
            "account_moves",
        ],
        move_rows,
    )

    move_line_rows = []
    for line in trace["move_lines"]:
        move_line_rows.append(
            [
                _doc_type(line.move_id),
                _doc_number(line.move_id),
                line.id,
                line.move_id.id,
                line.product_id.default_code or "",
                line.product_id.display_name,
                line.lot_id.name if line.lot_id else "",
                line.quantity,
                line.quantity_product_uom,
                line.picked,
                line.state,
                line.location_id.complete_name,
                line.location_dest_id.complete_name,
            ]
        )
    ws = wb.create_sheet("MOVE_LINES")
    _append_rows(
        ws,
        [
            "doc_type",
            "doc_no",
            "move_line_id",
            "move_id",
            "product_code",
            "product_name",
            "lot",
            "quantity",
            "quantity_product_uom",
            "picked",
            "state",
            "source",
            "destination",
        ],
        move_line_rows,
    )

    ws = wb.create_sheet("LOT_CONSUMPTION")
    _append_rows(
        ws,
        [
            "mo_name",
            "fg_lot",
            "fg_qty",
            "consumed_product_code",
            "consumed_product_name",
            "consumed_lot",
            "consumed_qty",
            "raw_move_id",
            "raw_move_name",
            "source_receipt",
            "source_move_line_id",
        ],
        trace["lot_trace_rows"],
    )

    svl_rows = []
    for svl in trace["svls"]:
        svl_rows.append(
            [
                svl.id,
                svl.stock_move_id.id if svl.stock_move_id else "",
                _doc_type(svl.stock_move_id),
                _doc_number(svl.stock_move_id),
                svl.stock_move_id.product_id.default_code or "",
                svl.stock_move_id.product_id.display_name if svl.stock_move_id else "",
                svl.quantity,
                svl.unit_cost,
                svl.value,
                svl.remaining_qty,
                svl.remaining_value,
                svl.lot_id.name if svl.lot_id else "",
                svl.account_move_id.name if svl.account_move_id else "",
            ]
        )
    ws = wb.create_sheet("SVL")
    _append_rows(
        ws,
        [
            "svl_id",
            "stock_move_id",
            "doc_type",
            "doc_no",
            "product_code",
            "product_name",
            "quantity",
            "unit_cost",
            "value",
            "remaining_qty",
            "remaining_value",
            "lot",
            "account_move",
        ],
        svl_rows,
    )

    account_move_rows = []
    for move in trace["account_moves"]:
        account_move_rows.append(
            [
                move.id,
                move.name,
                move.ref,
                move.date,
                move.state,
                move.journal_id.code,
                float(q2(sum(move.line_ids.mapped("debit")))),
                float(q2(sum(move.line_ids.mapped("credit")))),
            ]
        )
    ws = wb.create_sheet("ACCOUNT_MOVES")
    _append_rows(
        ws,
        ["move_id", "move_name", "ref", "date", "state", "journal", "total_debit", "total_credit"],
        account_move_rows,
    )

    aml_rows = []
    for line in trace["aml"]:
        section, bucket = _account_bucket(line.account_id)
        aml_rows.append(
            [
                line.id,
                line.move_id.name,
                line.move_id.ref,
                line.date,
                line.account_id.code or "",
                line.account_id.name or "",
                line.account_id.account_type or "",
                line.name,
                line.debit,
                line.credit,
                line.balance,
                section,
                bucket,
            ]
        )
    ws = wb.create_sheet("AML")
    _append_rows(
        ws,
        [
            "aml_id",
            "account_move",
            "ref",
            "date",
            "account_code",
            "account_name",
            "account_type",
            "line_name",
            "debit",
            "credit",
            "balance",
            "section",
            "bucket",
        ],
        aml_rows,
    )

    ws = wb.create_sheet("TRACE_DETAIL")
    _append_rows(
        ws,
        [
            "doc_type",
            "doc_no",
            "origin",
            "doc_date",
            "stock_move_id",
            "move_ref",
            "product_code",
            "product_name",
            "move_state",
            "move_done_qty",
            "move_demand_qty",
            "source",
            "destination",
            "move_lines",
            "svl_id",
            "svl_qty",
            "svl_unit_cost",
            "svl_value",
            "account_move",
            "aml_id",
            "account_code",
            "account_name",
            "account_type",
            "debit",
            "credit",
            "balance",
            "fs_section",
            "fs_bucket",
        ],
        trace["trace_rows"],
    )

    tb_sheet_rows = [
        [
            row["account_code"],
            row["account_name"],
            row["account_type"],
            row["section"],
            row["bucket"],
            float(row["debit"]),
            float(row["credit"]),
            float(row["balance"]),
        ]
        for row in trace["tb_rows"]
    ]
    ws = wb.create_sheet("TB")
    _append_rows(
        ws,
        ["account_code", "account_name", "account_type", "section", "bucket", "debit", "credit", "balance"],
        tb_sheet_rows,
    )

    fin_sheet_rows = [
        [
            row["section"],
            row["bucket"],
            row["account_code"],
            row["account_name"],
            float(row["amount"]),
        ]
        for row in trace["fin_rows"]
    ]
    ws = wb.create_sheet("FIN_STMT")
    _append_rows(
        ws,
        ["section", "bucket", "account_code", "account_name", "presented_amount"],
        fin_sheet_rows,
    )

    wb.save(report_path)
    return report_path


def run(env):
    ensure_schema(env)
    setup = get_core_setup(env)
    tag = f"{TAG_PREFIX}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    try:
        receipt, lots_by_product = create_inbound_receipt(env, setup, tag)
        mo, finished_move, fg_lot = create_and_close_mo(env, setup, tag, lots_by_product)
        flow = {
            "tag": tag,
            "product": setup["product"],
            "receipt": receipt,
            "mo": mo,
            "finished_move": finished_move,
            "fg_lot": fg_lot,
        }
        trace = build_trace(flow)
        report_path = write_report(trace)
        env.cr.commit()
        return {
            "tag": tag,
            "receipt_name": receipt.name,
            "mo_name": mo.name,
            "fg_lot": fg_lot.name,
            "report_path": str(report_path),
            "wip_balance": str(trace["summary"]["wip_balance"]),
        }
    except Exception:
        env.cr.rollback()
        raise
