from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_REPORT = "view_uat_account_knock_detail_th.xlsx"
OUTPUT_REPORT = "view_uat_account_settle_checklist_th.xlsx"


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _reports_dir() -> Path:
    return _root_path() / "reports"


def _to_decimal(value) -> Decimal:
    return Decimal(str(round(float(value or 0.0), 2)))


def _sheet_dict_rows(ws):
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in row):
            continue
        rows.append({headers[idx]: row[idx] for idx in range(len(headers))})
    return rows


def _style_border() -> Border:
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _phase_from_entry(entry: dict, docs_by_type: dict[str, str]) -> str:
    je = str(entry["JE"] or "")
    ref = str(entry["Ref"] or "")
    if je == docs_by_type.get("Invoice"):
        return "Invoice"
    if je == docs_by_type.get("Payment"):
        return "Payment"
    if docs_by_type.get("Delivery") and docs_by_type["Delivery"] in ref:
        return "Delivery"
    if docs_by_type.get("MO") and docs_by_type["MO"] in ref:
        return "MO"
    if docs_by_type.get("Receipt") and docs_by_type["Receipt"] in ref:
        return "Receipt"
    return ""


def _screen_for_phase(phase: str) -> str:
    return {
        "Receipt": "Inventory > Receipts",
        "MO": "Manufacturing > Manufacturing Orders",
        "Delivery": "Inventory > Delivery Orders",
        "Invoice": "Accounting > Customers > Invoices",
        "Payment": "Accounting > Customers > Payments",
    }.get(phase, "")


def _side_for_phase(phase: str) -> str:
    return {
        "Receipt": "ขาเข้า",
        "MO": "ผลิต",
        "Delivery": "ขาออก",
        "Invoice": "ขาออก",
        "Payment": "ขาออก",
    }.get(phase, "")


def _doc_for_phase(phase: str, docs_by_type: dict[str, str]) -> str:
    return docs_by_type.get(phase, "")


def _status_fill(status: str) -> PatternFill:
    return {
        "ค้าง / ยังไม่ settle": PatternFill("solid", fgColor="F4B084"),
        "settle แล้ว": PatternFill("solid", fgColor="C6E0B4"),
        "settle แบบกลุ่ม": PatternFill("solid", fgColor="D9EAD3"),
        "ปลายทางงบ / ไม่ต้อง settle": PatternFill("solid", fgColor="D9E2F3"),
        "คงเหลือ stock": PatternFill("solid", fgColor="D9E2F3"),
    }.get(status, PatternFill())


def _exact_pair(entries: list[dict], code: str):
    debit_entries = [entry for entry in entries if entry["Account Code"] == code and _to_decimal(entry["Debit"]) > 0]
    credit_entries = [entry for entry in entries if entry["Account Code"] == code and _to_decimal(entry["Credit"]) > 0]

    for debit_entry in debit_entries:
        if debit_entry.get("matched"):
            continue
        debit_amount = _to_decimal(debit_entry["Debit"])
        for credit_entry in credit_entries:
            if credit_entry.get("matched"):
                continue
            if _to_decimal(credit_entry["Credit"]) != debit_amount:
                continue
            debit_entry["matched"] = True
            credit_entry["matched"] = True
            debit_entry["settle_ref"] = f"{credit_entry['JE']} | {credit_entry['Ref']}"
            credit_entry["settle_ref"] = f"{debit_entry['JE']} | {debit_entry['Ref']}"
            debit_entry["settle_amount"] = float(debit_amount)
            credit_entry["settle_amount"] = float(debit_amount)
            debit_entry["remaining_amount"] = 0.0
            credit_entry["remaining_amount"] = 0.0
            debit_entry["settle_status"] = "settle แล้ว"
            credit_entry["settle_status"] = "settle แล้ว"


def _apply_custom_settle(entries: list[dict], docs_by_type: dict[str, str], reconcile_rows: list[dict]):
    receipt = docs_by_type.get("Receipt", "")
    mo = docs_by_type.get("MO", "")
    delivery = docs_by_type.get("Delivery", "")
    invoice = docs_by_type.get("Invoice", "")
    payment = docs_by_type.get("Payment", "")

    _exact_pair(entries, "116003")
    _exact_pair(entries, "116902")
    _exact_pair(entries, "113001")

    for entry in entries:
        code = entry["Account Code"]
        debit = _to_decimal(entry["Debit"])
        credit = _to_decimal(entry["Credit"])
        amount = float(debit or credit)
        entry.setdefault("settle_amount", 0.0)
        entry.setdefault("remaining_amount", amount)
        entry.setdefault("settle_ref", "")
        entry.setdefault("settle_status", "")

        if code == "116901":
            entry["settle_status"] = "ค้าง / ยังไม่ settle"
            entry["settle_ref"] = "รอ Vendor Bill / AP"
            entry["settle_amount"] = 0.0
            entry["remaining_amount"] = amount
            continue

        if code == "116021":
            entry["settle_status"] = "settle แบบกลุ่ม"
            entry["settle_ref"] = (
                f"กลุ่ม WIP {mo}: STJ/26/03/04662 + STJ/26/03/04663 + STJ/26/03/04664 + "
                f"STJ/26/03/04665 + STJ/26/03/04666"
            )
            entry["settle_amount"] = amount
            entry["remaining_amount"] = 0.0
            continue

        if code == "116031" and entry["JE"] == "STJ/26/03/04659":
            entry["settle_status"] = "settle แล้ว"
            entry["settle_ref"] = "STJ/26/03/04662 | GMP/MO/02763 - ยาดมโป๊ยเซียนกล่อง 5 โหล"
            entry["settle_amount"] = amount
            entry["remaining_amount"] = 0.0
            continue

        if code == "116031" and entry["JE"] == "STJ/26/03/04662":
            entry["settle_status"] = "settle แล้ว"
            entry["settle_ref"] = f"STJ/26/03/04659 | {receipt} - ยาดมโป๊ยเซียนกล่อง 5 โหล"
            entry["settle_amount"] = amount
            entry["remaining_amount"] = 0.0
            continue

        if code == "116031" and entry["JE"] == "STJ/26/03/04665":
            entry["settle_status"] = "คงเหลือ stock"
            entry["settle_ref"] = f"STJ/26/03/04668 | {delivery} - partial ตัดออก 17,938.69"
            entry["settle_amount"] = 17938.69
            entry["remaining_amount"] = round(amount - 17938.69, 2)
            continue

        if code == "116031" and entry["JE"] == "STJ/26/03/04668":
            entry["settle_status"] = "settle แล้ว"
            entry["settle_ref"] = "STJ/26/03/04665 | GMP/MO/02763 - FG รับเข้าคลัง"
            entry["settle_amount"] = amount
            entry["remaining_amount"] = 0.0
            continue

        if code == "113001" and reconcile_rows:
            reconcile = reconcile_rows[0]
            entry["settle_ref"] = (
                f"{reconcile['Invoice']} <-> {reconcile['Payment']} | partial reconcile id {reconcile['Partial Reconcile ID']}"
            )

        if entry["settle_status"]:
            continue

        if code in ("410001", "410032", "215001", "510001", "111302", "590001"):
            entry["settle_status"] = "ปลายทางงบ / ไม่ต้อง settle"
            entry["settle_ref"] = "ปลายทางงบหรือบัญชีปลายทางของรายการนี้"
            entry["settle_amount"] = amount
            entry["remaining_amount"] = 0.0
            continue

        if code == "116003" and entry["settle_status"] == "":
            entry["settle_status"] = "settle แล้ว"
            entry["remaining_amount"] = 0.0
            continue

        if code == "116902" and entry["settle_status"] == "":
            entry["settle_status"] = "settle แล้ว"
            entry["remaining_amount"] = 0.0
            continue

        if code == "113001" and entry["settle_status"] == "":
            entry["settle_status"] = "settle แล้ว"
            entry["remaining_amount"] = 0.0
            continue


def _origin_note(entry: dict) -> str:
    code = entry["Account Code"]
    phase = entry["phase"]
    if code == "116901":
        return "มาจาก Receipt ตอนรับสินค้าเข้าระบบ"
    if code == "116003" and _to_decimal(entry["Debit"]) > 0:
        return "มาจาก Receipt ที่รับวัตถุดิบ/บรรจุภัณฑ์เข้า stock"
    if code == "116003" and _to_decimal(entry["Credit"]) > 0:
        return "มาจาก MO ที่ consume วัตถุดิบเข้า WIP"
    if code == "116021":
        return "มาจาก MO ฝั่งงานระหว่างทำ"
    if code == "116902":
        return "มาจาก Delivery ตอนส่งของ"
    if code == "113001":
        return "มาจาก Invoice/Payment ฝั่งลูกหนี้"
    if phase == "Invoice":
        return "มาจาก Invoice ที่ระบบ post อัตโนมัติ"
    if phase == "Payment":
        return "มาจาก Payment ที่ระบบ post และ reconcile อัตโนมัติ"
    return "มาจาก JE ของ transaction นี้"


def _detail_note(entry: dict) -> str:
    code = entry["Account Code"]
    status = entry["settle_status"]
    debit = _to_decimal(entry["Debit"])
    credit = _to_decimal(entry["Credit"])

    if code == "116901":
        return "บัญชีพักสินค้าขาเข้า เปิดเครดิตตอนรับเข้า แต่ใน scenario นี้ยังไม่มี Vendor Bill/AP มาปิด จึงให้บัญชีติ๊กติดตามค้างไว้"
    if code == "116003":
        if debit > 0:
            return "บรรทัดนี้เป็น stock เข้า ฝั่งวัตถุดิบ/บรรจุภัณฑ์ และถูก MO consume ภายหลัง ให้ติ๊ก settle เมื่อเทียบกับ JE consume ฝั่งเครดิตครบแล้ว"
        return "บรรทัดนี้เป็นการตัดวัตถุดิบจาก stock เข้า WIP จับคู่กับ receipt ฝั่งเดบิตจำนวนเท่ากัน"
    if code == "116021":
        return "WIP ต้องดูแบบกลุ่ม ไม่ควรดูบรรทัดเดี่ยว เพราะ raw consume, FG receipt และ overhead ร่วมกันทำให้ยอดสุทธิเป็นศูนย์"
    if code == "116031":
        if entry["JE"] == "STJ/26/03/04665":
            return "FG รับเข้าคลังจาก MO แล้วถูก delivery ตัดออกบางส่วน เหลือ stock ปลายงวดจึงไม่ใช่ open issue ทาง settle"
        return "บัญชี stock ฝั่งสินค้าสำเร็จรูป/ชิ้นส่วน ให้ดูว่า movement ถัดไปตัดออกครบหรือยัง"
    if code == "116902":
        return "บัญชีสินค้าขาออกเปิดตอน Delivery และถูก Invoice knock ไป COGS ให้ติ๊ก settle เมื่อเทียบสอง JE นี้ตรงกัน"
    if code == "113001":
        return "ลูกหนี้ invoice ถูก payment ปิดด้วย partial reconcile จริงในระบบ ให้บัญชีติ๊ก settle หลังตรวจ ref และยอด"
    if status == "ปลายทางงบ / ไม่ต้อง settle":
        return "บรรทัดนี้เป็นปลายทางของรายได้ ต้นทุน ภาษี ธนาคาร หรือ resource absorption ไม่ใช่บัญชีที่ต้องตาม settle ต่อ"
    return "ตรวจ ref และยอดเดบิต/เครดิตประกอบกับ JE คู่ settle"


def _load_source():
    source_path = _reports_dir() / SOURCE_REPORT
    workbook = load_workbook(source_path, data_only=True)

    docs_rows = _sheet_dict_rows(workbook["เลขเอกสาร"])
    docs_by_type = {row["ประเภทเอกสาร"]: row["เลขที่"] for row in docs_rows}
    knock_rows = _sheet_dict_rows(workbook["ไทม์ไลน์_KNOCK"])
    knock_by_je = {row["JE"]: row for row in knock_rows}
    reconcile_rows = _sheet_dict_rows(workbook["RECONCILE_AR"])
    detail_rows = _sheet_dict_rows(workbook["JE_DETAIL"])

    entries = []
    for index, row in enumerate(detail_rows, start=1):
        entry = dict(row)
        entry["row_no"] = index
        entry["phase"] = _phase_from_entry(entry, docs_by_type)
        entry["side"] = _side_for_phase(entry["phase"])
        entry["screen"] = _screen_for_phase(entry["phase"])
        entry["doc_no"] = _doc_for_phase(entry["phase"], docs_by_type)
        knock_row = knock_by_je.get(entry["JE"], {})
        entry["je_origin"] = knock_row.get("เปิดบัญชีอะไร", "")
        entry["je_knock"] = knock_row.get("เคลียร์/knock อะไร", "")
        entry["je_result"] = knock_row.get("ผลหลังรายการ", "")
        entries.append(entry)

    _apply_custom_settle(entries, docs_by_type, reconcile_rows)

    return docs_rows, entries


def _write_report():
    docs_rows, entries = _load_source()
    output_path = _reports_dir() / OUTPUT_REPORT

    wb = Workbook()
    ws = wb.active
    ws.title = "เช็กลิสต์_settle"

    header_fill = _header_fill("1F4E78")
    section_fill = _header_fill("4F81BD")
    yellow_fill = _header_fill("FFF2CC")
    border = _style_border()

    ws.merge_cells("A1:T1")
    ws["A1"] = "เช็กลิสต์ไล่บรรทัดสำหรับบัญชีตรวจและติ๊ก settle"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = border

    intro_rows = [
        ("วิธีใช้", "ไล่จากบนลงล่างทีละบรรทัด ตรวจ JE, Ref, Account, Debit/Credit, คู่ settle, ยอดคงค้าง แล้วค่อยติ๊กคอลัมน์ตรวจ/settle"),
        ("หมายเหตุ 1", "บรรทัดที่สถานะเป็น 'ค้าง / ยังไม่ settle' คือยังมีงานบัญชีค้าง เช่น 116901 รอ Vendor Bill/AP"),
        ("หมายเหตุ 2", "บรรทัดที่สถานะเป็น 'settle แบบกลุ่ม' เช่น 116021 WIP ต้องดูรวมทั้งกลุ่ม JE ไม่ควรดูเป็นคู่บรรทัดเดียว"),
        ("หมายเหตุ 3", "บรรทัดที่สถานะเป็น 'ปลายทางงบ / ไม่ต้อง settle' ให้ตรวจความถูกต้องของ posting แต่ไม่ต้องตามคู่ settle ต่อ"),
    ]
    for row_index, (label, value) in enumerate(intro_rows, start=2):
        ws.cell(row_index, 1, label)
        ws.cell(row_index, 2, value)
        ws.cell(row_index, 1).font = Font(bold=True)

    row = 7
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=20)
    ws.cell(row, 1, "เลขเอกสารอ้างอิง")
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = section_fill
    ws.cell(row, 1).border = border
    row += 1

    ws.append(["ประเภทเอกสาร", "เลขที่", "วันที่", "สถานะ", "JE ที่เกี่ยวข้อง"])
    header_row = row
    row += 1
    for doc in docs_rows:
        ws.append([doc["ประเภทเอกสาร"], doc["เลขที่"], doc["วันที่"], doc["สถานะ"], doc["JE ที่เกี่ยวข้อง"]])
        row += 1

    detail_start = row + 1
    ws.merge_cells(start_row=detail_start, start_column=1, end_row=detail_start, end_column=20)
    ws.cell(detail_start, 1, "ไล่ทีละบรรทัดเพื่อให้บัญชีติ๊ก settle ตามได้")
    ws.cell(detail_start, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(detail_start, 1).fill = section_fill
    ws.cell(detail_start, 1).border = border

    headers = [
        "ลำดับ",
        "ติ๊กตรวจ",
        "ติ๊ก settle",
        "วันที่",
        "ฝั่ง",
        "หน้าจอ",
        "เลขเอกสาร",
        "JE",
        "Ref",
        "Account Code",
        "Account Name",
        "Debit",
        "Credit",
        "ยอดที่ใช้ settle",
        "ยอดคงค้างบรรทัดนี้",
        "สถานะ settle",
        "คู่ settle / ref",
        "มาจากตรงไหน",
        "อธิบายละเอียด",
        "หมายเหตุบัญชี",
    ]
    header_line = detail_start + 1
    for index, header in enumerate(headers, start=1):
        ws.cell(header_line, index, header)

    for index, entry in enumerate(entries, start=1):
        row_idx = header_line + index
        ws.cell(row_idx, 1, index)
        ws.cell(row_idx, 2, "")
        ws.cell(row_idx, 3, "")
        ws.cell(row_idx, 4, entry["วันที่"])
        ws.cell(row_idx, 5, entry["side"])
        ws.cell(row_idx, 6, entry["screen"])
        ws.cell(row_idx, 7, entry["doc_no"])
        ws.cell(row_idx, 8, entry["JE"])
        ws.cell(row_idx, 9, entry["Ref"])
        ws.cell(row_idx, 10, entry["Account Code"])
        ws.cell(row_idx, 11, entry["Account Name"])
        ws.cell(row_idx, 12, float(entry["Debit"] or 0.0))
        ws.cell(row_idx, 13, float(entry["Credit"] or 0.0))
        ws.cell(row_idx, 14, float(entry.get("settle_amount", 0.0)))
        ws.cell(row_idx, 15, float(entry.get("remaining_amount", 0.0)))
        ws.cell(row_idx, 16, entry.get("settle_status", ""))
        ws.cell(row_idx, 17, entry.get("settle_ref", ""))
        ws.cell(row_idx, 18, _origin_note(entry))
        ws.cell(row_idx, 19, _detail_note(entry))
        ws.cell(row_idx, 20, "")

    for current_row in range(1, ws.max_row + 1):
        for current_col in range(1, 21):
            cell = ws.cell(current_row, current_col)
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, 21):
        cell = ws.cell(header_row, col)
        if col <= 5 and cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, 21):
        cell = ws.cell(header_line, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(header_line + 1, ws.max_row + 1):
        ws.cell(row_idx, 2).fill = yellow_fill
        ws.cell(row_idx, 3).fill = yellow_fill
        status_cell = ws.cell(row_idx, 16)
        status_cell.fill = _status_fill(str(status_cell.value or ""))
        for amount_col in (12, 13, 14, 15):
            ws.cell(row_idx, amount_col).number_format = "#,##0.00"

    ws.freeze_panes = "A14"

    widths = {
        1: 8,
        2: 10,
        3: 11,
        4: 12,
        5: 10,
        6: 28,
        7: 18,
        8: 20,
        9: 44,
        10: 12,
        11: 28,
        12: 12,
        13: 12,
        14: 14,
        15: 15,
        16: 18,
        17: 42,
        18: 28,
        19: 50,
        20: 24,
    }
    for column_index, width in widths.items():
        ws.column_dimensions[get_column_letter(column_index)].width = width

    wb.save(output_path)
    return output_path


report = _write_report()
print(report)
