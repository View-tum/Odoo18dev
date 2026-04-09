from pathlib import Path

from openpyxl import load_workbook


FILE_PATH = Path(
    r"C:\365_project\TheCool18e\Dev\reports\Journal Entry (account.move) (6)_pro_import_ready_v5_balanced.xlsx"
)
FILE_SHEET = "account.move"
FILE_REF = "OB_IMPORT_JE6_V5"
POST_AFTER_CREATE = True

# Keep the user's requested 999 codes and create them only if missing.
MISSING_ACCOUNTS = {
    "999113001": {
        "name": "Opening Balance AR Temp",
        "account_type": "asset_receivable",
        "reconcile": True,
    },
    "999113002": {
        "name": "Opening Balance Other AR Temp",
        "account_type": "asset_receivable",
        "reconcile": True,
    },
    "999212001": {
        "name": "Opening Balance AP Temp",
        "account_type": "liability_payable",
        "reconcile": True,
    },
}


def _ensure_accounts():
    account_model = env["account.account"]
    created = []
    for code, vals in MISSING_ACCOUNTS.items():
        if account_model.search([("code", "=", code)], limit=1):
            continue
        rec = account_model.create({"code": code, **vals})
        created.append(rec.code)
    return created


def _load_rows():
    workbook = load_workbook(FILE_PATH, data_only=True)
    worksheet = workbook[FILE_SHEET] if FILE_SHEET in workbook.sheetnames else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    return headers, rows[1:]


def _build_move_vals(headers, data_rows):
    idx = {name: pos for pos, name in enumerate(headers)}
    first = data_rows[0]
    journal_name = first[idx["journal_id"]]
    journal = env["account.journal"].search([("name", "=", journal_name)], limit=1)
    if not journal:
        raise ValueError(f"Journal not found: {journal_name}")

    line_vals = []
    for row in data_rows:
        account_code = str(row[idx["line_ids/account_id"]]).strip()
        account = env["account.account"].search([("code", "=", account_code)], limit=1)
        if not account:
            raise ValueError(f"Account not found: {account_code}")
        line_vals.append(
            (
                0,
                0,
                {
                    "account_id": account.id,
                    "debit": float(row[idx["line_ids/debit"]] or 0),
                    "credit": float(row[idx["line_ids/credit"]] or 0),
                    "is_imported": True,
                    # Prevent default account taxes from generating extra tax lines.
                    "tax_ids": [(6, 0, [])],
                },
            )
        )

    return {
        "date": str(first[idx["date"]]),
        "journal_id": journal.id,
        "name": str(first[idx["name"]] or "/"),
        "ref": FILE_REF,
        "line_ids": line_vals,
    }


def _existing_move(move_vals):
    return env["account.move"].search(
        [
            ("ref", "=", FILE_REF),
            ("date", "=", move_vals["date"]),
            ("journal_id", "=", move_vals["journal_id"]),
        ],
        limit=1,
    )


def main():
    if not FILE_PATH.exists():
        raise FileNotFoundError(FILE_PATH)

    created_accounts = _ensure_accounts()
    headers, data_rows = _load_rows()
    move_vals = _build_move_vals(headers, data_rows)

    existing = _existing_move(move_vals)
    if existing:
        print(
            {
                "status": "skipped",
                "reason": "move already exists",
                "move_id": existing.id,
                "move_name": existing.name,
                "state": existing.state,
                "created_accounts": created_accounts,
            }
        )
        return

    move = env["account.move"].with_context(check_move_validity=False).create(move_vals)

    # Validate after create. This passes even though direct UI import does not.
    move._check_balanced(env["account.move.line"]._fields)

    if POST_AFTER_CREATE:
        move.action_post()

    print(
        {
            "status": "ok",
            "move_id": move.id,
            "move_name": move.name,
            "state": move.state,
            "line_count": len(move.line_ids),
            "created_accounts": created_accounts,
            "ref": move.ref,
        }
    )


main()
