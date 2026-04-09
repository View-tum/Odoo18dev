from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TAG_PREFIX = "SIM-VIEW-UAT-POYSIAN-"
REPORT_NAME = "view_uat_odoo_vs_excel_step_detail.xlsx"
WALKTHROUGH_FILE = "UAT_Walkthrough_DB_VIEW_TH_FINAL.md"


def _root_path() -> Path:
    script_file = globals().get("__file__")
    if script_file:
        return Path(script_file).resolve().parents[3]
    return Path.cwd()


def _report_path() -> Path:
    return _root_path() / "reports" / REPORT_NAME


def _walkthrough_path() -> Path:
    return _root_path() / WALKTHROUGH_FILE


def _header_style():
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    return fill, font


def _cell_border():
    thin = Side(border_style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row=1):
    fill, font = _header_style()
    border = _cell_border()
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_grid(ws, start_row=2, amount_cols=None):
    amount_cols = set(amount_cols or [])
    border = _cell_border()
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in amount_cols and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def auto_width(ws, max_width=56):
    for column in ws.columns:
        values = [len(str(cell.value)) for cell in column if cell.value is not None]
        width = min(max(values, default=8) + 2, max_width)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def _find_latest_bundle(env):
    receipt = env["stock.picking"].search(
        [("origin", "ilike", TAG_PREFIX + "%"), ("picking_type_id.code", "=", "incoming")],
        order="id desc",
        limit=1,
    )
    if not receipt:
        raise ValueError("No receipt found for latest SIM-VIEW-UAT bundle")

    origin = receipt.origin
    mo = env["mrp.production"].search([("origin", "=", origin)], order="id desc", limit=1)
    if not mo:
        raise ValueError("No MO found for origin %s" % origin)

    so = env["sale.order"].search(
        [("client_order_ref", "=", f"{TAG_PREFIX}SO-{origin}")],
        order="id desc",
        limit=1,
    )
    if not so:
        so = env["sale.order"].search(
            [
                ("partner_id.name", "ilike", "SIM VIEW UAT CUSTOMER%"),
                ("create_date", ">=", receipt.create_date),
            ],
            order="id desc",
            limit=1,
        )
    if not so:
        raise ValueError("No sale order found for origin %s" % origin)

    pick = so.picking_ids.filtered(lambda p: p.picking_type_id.code == "internal")[:1]
    delivery = so.picking_ids.filtered(lambda p: p.picking_type_id.code == "outgoing")[:1]
    invoice = env["account.move"].search([("invoice_origin", "=", so.name)], order="id desc", limit=1)
    if not invoice:
        raise ValueError("No invoice found for SO %s" % so.name)

    receivable_lines = invoice.line_ids.filtered(lambda l: l.account_id.account_type == "asset_receivable")
    partials = receivable_lines.matched_debit_ids | receivable_lines.matched_credit_ids
    payment_move = (partials.debit_move_id.move_id | partials.credit_move_id.move_id).filtered(lambda m: m != invoice)[:1]
    payment = env["account.payment"].search([("move_id", "=", payment_move.id)], limit=1) if payment_move else env["account.payment"]

    return {
        "origin": origin,
        "receipt": receipt,
        "mo": mo,
        "so": so,
        "pick": pick,
        "delivery": delivery,
        "invoice": invoice,
        "payment": payment,
    }


def _walkthrough_rows():
    path = _walkthrough_path()
    if not path.exists():
        return []
    pattern = re.compile(r"^####\s+(SC\d+-[\d.]+(?:-R\d+)?)\s+-\s+(.*)$")
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        match = pattern.match(line.strip())
        if match:
            rows.append((match.group(1), match.group(2).strip()))
    return rows


def _bundle_coverage():
    executed = {
        "SC02-2.2": ("Partially Executed", "ใช้ existing product/BOM จริง แต่เป็น MO manual ไม่ใช่ MTO ตาม walkthrough เดิม"),
        "SC04-4.1": ("Executed", "MO ปิดจริงในฐาน view"),
        "SC04-4.2": ("Executed", "Work Order อยู่สถานะ done จริง"),
        "SC05-5.6": ("Executed", "trace จริงจาก Receipt -> MO -> Sale -> Invoice -> Payment"),
    }
    rows = []
    for case_id, title in _walkthrough_rows():
        status, reason = executed.get(
            case_id,
            (
                "Not Executed",
                "รอบนี้ไม่ได้รันทุก case ใน walkthrough; รันเฉพาะ readiness + transaction จริงด้วย existing product",
            ),
        )
        rows.append((case_id, title, status, reason))
    return rows


def _doc_moves(doc):
    if not doc:
        return doc.env["stock.move"]
    return doc.move_ids


def _stock_value_summary(env, moves):
    svls = env["stock.valuation.layer"].search([("stock_move_id", "in", moves.ids)], order="id") if moves else env["stock.valuation.layer"]
    account_moves = (moves.mapped("account_move_ids") | svls.mapped("account_move_id")).sorted(lambda m: m.id) if moves else env["account.move"]
    return svls, account_moves


def _aml_rows(account_moves):
    return account_moves.line_ids.filtered(lambda l: l.account_id).sorted(lambda l: (l.move_id.id, l.id))


def _fmt_doc_date(doc):
    date_val = getattr(doc, "date", False) or getattr(doc, "scheduled_date", False) or getattr(doc, "date_deadline", False) or getattr(doc, "date_order", False) or getattr(doc, "invoice_date", False) or getattr(doc, "payment_date", False)
    return str(date_val) if date_val else ""


def _fmt_receipt_screen(receipt):
    parts = []
    for move in receipt.move_ids:
        qty_done = sum(move.move_line_ids.mapped("quantity")) or move.quantity or move.product_uom_qty
        parts.append(f"{move.product_id.display_name}: รับ {qty_done} {move.product_uom.name}")
    return "\n".join(parts)


def _fmt_mo_screen(mo):
    raw_lines = []
    for move in mo.move_raw_ids.filtered(lambda m: m.state == "done"):
        qty_done = sum(move.move_line_ids.mapped("quantity")) or move.quantity or move.product_uom_qty
        raw_lines.append(f"ใช้ {move.product_id.display_name} = {qty_done} {move.product_uom.name}")
    return "\n".join(
        [
            f"สินค้า: {mo.product_id.display_name}",
            f"ผลิตจริง: {mo.product_qty} {mo.product_uom_id.name}",
            f"FG Lot: {mo.lot_producing_id.name or ''}",
            *raw_lines,
        ]
    )


def _fmt_picking_screen(picking):
    if not picking:
        return ""
    parts = []
    for move in picking.move_ids:
        qty_done = sum(move.move_line_ids.mapped("quantity")) or move.quantity or move.product_uom_qty
        lots = ", ".join(move.move_line_ids.mapped("lot_id.name"))
        label = f"{move.product_id.display_name}: {qty_done} {move.product_uom.name}"
        if lots:
            label += f" | lot: {lots}"
        parts.append(label)
    return "\n".join(parts)


def _fmt_invoice_screen(invoice):
    lines = []
    for line in invoice.invoice_line_ids:
        lines.append(f"{line.product_id.display_name}: qty {line.quantity} | subtotal {line.price_subtotal} | total {line.price_total}")
    return "\n".join(
        [
            f"คู่ค้า: {invoice.partner_id.display_name}",
            f"Untaxed: {invoice.amount_untaxed}",
            f"VAT: {invoice.amount_tax}",
            f"Total: {invoice.amount_total}",
            *lines,
        ]
    )


def _fmt_payment_screen(payment, invoice):
    if not payment:
        return "ไม่พบ payment ที่ reconcile"
    ref = getattr(payment, "ref", False) or getattr(payment.move_id, "ref", False) or payment.name or invoice.name
    return "\n".join(
        [
            f"Journal: {payment.journal_id.display_name}",
            f"Amount: {payment.amount}",
            f"Ref: {ref}",
            f"State: {payment.state}",
        ]
    )


def _statement_hint(aml):
    accounts = aml.mapped("account_id")
    names = []
    if accounts.filtered(lambda a: a.account_type.startswith("income")):
        names.append("P&L รายได้")
    if accounts.filtered(lambda a: a.account_type.startswith("expense")):
        names.append("P&L ค่าใช้จ่าย")
    if accounts.filtered(lambda a: a.account_type.startswith("asset_")):
        names.append("งบดุล สินทรัพย์")
    if accounts.filtered(lambda a: a.account_type.startswith("liability_")):
        names.append("งบดุล หนี้สิน")
    return ", ".join(names)


def _doc_je_names(account_moves):
    return ", ".join(account_moves.mapped("name"))


def _doc_numbers(bundle):
    return [
        ("Receipt", bundle["receipt"].name),
        ("MO", bundle["mo"].name),
        ("SO", bundle["so"].name),
        ("Internal Pick", bundle["pick"].name if bundle["pick"] else ""),
        ("Delivery", bundle["delivery"].name if bundle["delivery"] else ""),
        ("Invoice", bundle["invoice"].name),
        ("Payment", bundle["payment"].name if bundle["payment"] else ""),
    ]


def _build_steps(env, bundle):
    receipt_svls, receipt_am = _stock_value_summary(env, bundle["receipt"].move_ids)
    mo_stock_moves = (bundle["mo"].move_raw_ids | bundle["mo"].move_finished_ids).sorted(lambda m: m.id)
    mo_svls, mo_am = _stock_value_summary(env, mo_stock_moves)
    mo_am |= env["account.move"].search([("ref", "ilike", bundle["mo"].name)], order="id")
    pick_svls, pick_am = _stock_value_summary(env, bundle["pick"].move_ids if bundle["pick"] else env["stock.move"])
    out_svls, out_am = _stock_value_summary(env, bundle["delivery"].move_ids if bundle["delivery"] else env["stock.move"])
    invoice_am = bundle["invoice"]
    payment_am = bundle["payment"].move_id if bundle["payment"] else env["account.move"]

    return [
        {
            "step": 1,
            "screen": "Inventory > Receipts",
            "doc_type": "Receipt",
            "doc": bundle["receipt"],
            "user_action": "สร้างใบรับเข้าและกด Validate",
            "system_auto": "ระบบสร้าง stock move done, SVL และ JE รับวัตถุดิบอัตโนมัติ",
            "screen_values": _fmt_receipt_screen(bundle["receipt"]),
            "svls": receipt_svls,
            "account_moves": receipt_am,
        },
        {
            "step": 2,
            "screen": "Manufacturing > Operations > Manufacturing Orders",
            "doc_type": "MO",
            "doc": bundle["mo"],
            "user_action": "Confirm MO, ใส่ FG lot, ปิด WO และปิด MO",
            "system_auto": "ระบบ consume วัตถุดิบ, รับ FG, คำนวณ valuation และ post labour/WIP",
            "screen_values": _fmt_mo_screen(bundle["mo"]),
            "svls": mo_svls.sorted(lambda s: s.id),
            "account_moves": mo_am.sorted(lambda m: m.id),
        },
        {
            "step": 3,
            "screen": "Sales > Orders > Quotations / Orders",
            "doc_type": "SO",
            "doc": bundle["so"],
            "user_action": "สร้าง SO และกด Confirm",
            "system_auto": "ระบบสร้าง Internal Pick และ Delivery ให้อัตโนมัติ",
            "screen_values": f"{bundle['so'].partner_id.display_name}\n{bundle['so'].order_line[0].product_id.display_name}: qty {bundle['so'].order_line[0].product_uom_qty}",
            "svls": env["stock.valuation.layer"],
            "account_moves": env["account.move"],
        },
        {
            "step": 4,
            "screen": "Inventory > Operations > Internal Transfers",
            "doc_type": "Internal Pick",
            "doc": bundle["pick"],
            "user_action": "Assign lot และ Validate internal pick",
            "system_auto": "ระบบขยับ stock ภายใน; ปกติไม่ลงบัญชี",
            "screen_values": _fmt_picking_screen(bundle["pick"]),
            "svls": pick_svls,
            "account_moves": pick_am,
        },
        {
            "step": 5,
            "screen": "Inventory > Operations > Delivery Orders",
            "doc_type": "Delivery",
            "doc": bundle["delivery"],
            "user_action": "Assign lot และ Validate delivery",
            "system_auto": "ระบบตัด stock, สร้าง SVL และ post COGS/Stock Output อัตโนมัติ",
            "screen_values": _fmt_picking_screen(bundle["delivery"]),
            "svls": out_svls,
            "account_moves": out_am,
        },
        {
            "step": 6,
            "screen": "Accounting > Customers > Invoices",
            "doc_type": "Invoice",
            "doc": bundle["invoice"],
            "user_action": "Create Invoice จาก SO",
            "system_auto": "custom flow ในฐานนี้ post invoice อัตโนมัติหลัง create",
            "screen_values": _fmt_invoice_screen(bundle["invoice"]),
            "svls": env["stock.valuation.layer"],
            "account_moves": invoice_am,
        },
        {
            "step": 7,
            "screen": "Accounting > Customers > Payments",
            "doc_type": "Payment",
            "doc": bundle["payment"],
            "user_action": "Register Payment",
            "system_auto": "ระบบสร้าง payment entry และ reconcile AR อัตโนมัติ",
            "screen_values": _fmt_payment_screen(bundle["payment"], bundle["invoice"]),
            "svls": env["stock.valuation.layer"],
            "account_moves": payment_am,
        },
    ]


def write_report(env, bundle):
    path = _report_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    steps = _build_steps(env, bundle)

    ws = wb.active
    ws.title = "README"
    ws.append(["หัวข้อ", "ค่า"])
    for label, value in [
        ("Database", env.cr.dbname),
        ("Scenario", "เทียบหน้าจอ Odoo กับตัวเลขใน Excel จาก transaction จริงชุดล่าสุดในฐาน view"),
        ("Origin Tag", bundle["origin"]),
        ("Product", bundle["mo"].product_id.display_name),
        ("Coverage", "ไม่ใช่การรัน walkthrough ทั้งหมด; เป็น readiness + real transaction chain ของ existing product"),
        * _doc_numbers(bundle),
    ]:
        ws.append([label, value])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 96)

    ws = wb.create_sheet("STEP_COMPARE")
    ws.append([
        "Step",
        "หน้าจอ Odoo",
        "ประเภทเอกสาร",
        "เลขเอกสาร",
        "วันที่",
        "สถานะ",
        "ผู้ใช้ทำอะไร",
        "ระบบทำอะไรอัตโนมัติ",
        "ค่าที่เห็นบนหน้าจอ",
        "JE",
        "Debit รวม",
        "Credit รวม",
        "Balance",
        "กระทบงบ",
    ])
    for step in steps:
        account_moves = step["account_moves"]
        aml = _aml_rows(account_moves)
        debit = sum(aml.mapped("debit"))
        credit = sum(aml.mapped("credit"))
        balance = debit - credit
        ws.append([
            step["step"],
            step["screen"],
            step["doc_type"],
            step["doc"].name if step["doc"] else "",
            _fmt_doc_date(step["doc"]) if step["doc"] else "",
            step["doc"].state if step["doc"] else "",
            step["user_action"],
            step["system_auto"],
            step["screen_values"],
            _doc_je_names(account_moves),
            float(debit),
            float(credit),
            float(balance),
            _statement_hint(aml),
        ])
    style_header(ws)
    style_grid(ws, amount_cols=[11, 12, 13])
    auto_width(ws, 84)

    ws = wb.create_sheet("ODOO_SCREEN_FIELDS")
    ws.append(["Step", "Doc Type", "Doc No", "Field Group", "Value"])
    for step in steps:
        doc = step["doc"]
        if not doc:
            continue
        rows = [
            ("Header", f"ชื่อเอกสาร: {doc.display_name if hasattr(doc, 'display_name') else doc.name}"),
            ("Header", f"State: {doc.state}"),
            ("Header", f"Date: {_fmt_doc_date(doc)}"),
        ]
        if step["doc_type"] == "Receipt":
            rows.append(("Lines", _fmt_receipt_screen(doc)))
        elif step["doc_type"] == "MO":
            rows.append(("Header", f"BOM: {bundle['mo'].bom_id.code or bundle['mo'].bom_id.id}"))
            rows.append(("Header", f"FG Lot: {bundle['mo'].lot_producing_id.name or ''}"))
            rows.append(("Lines", _fmt_mo_screen(doc)))
        elif step["doc_type"] in ("Internal Pick", "Delivery"):
            rows.append(("Lines", _fmt_picking_screen(doc)))
        elif step["doc_type"] == "SO":
            rows.append(("Header", f"Customer: {doc.partner_id.display_name}"))
            rows.append(("Lines", f"{doc.order_line[0].product_id.display_name} qty {doc.order_line[0].product_uom_qty}"))
        elif step["doc_type"] == "Invoice":
            rows.append(("Header", f"Customer: {doc.partner_id.display_name}"))
            rows.append(("Header", f"Untaxed: {doc.amount_untaxed}"))
            rows.append(("Header", f"Tax: {doc.amount_tax}"))
            rows.append(("Header", f"Total: {doc.amount_total}"))
            rows.append(("Lines", _fmt_invoice_screen(doc)))
        elif step["doc_type"] == "Payment":
            ref = getattr(doc, "ref", False) or getattr(doc.move_id, "ref", False) or doc.name or ""
            rows.append(("Header", f"Journal: {doc.journal_id.display_name}"))
            rows.append(("Header", f"Amount: {doc.amount}"))
            rows.append(("Header", f"Ref: {ref}"))

        for group, value in rows:
            ws.append([step["step"], step["doc_type"], doc.name, group, value])
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 96)

    ws = wb.create_sheet("EXCEL_NUMBERS")
    ws.append(["Step", "Doc Type", "Doc No", "Category", "Reference", "Qty", "Unit Cost", "Value", "หมายเหตุ"])
    for step in steps:
        for svl in step["svls"]:
            ws.append([
                step["step"],
                step["doc_type"],
                step["doc"].name if step["doc"] else "",
                "SVL",
                svl.account_move_id.name or svl.stock_move_id.reference or "",
                float(svl.quantity or 0.0),
                float(svl.unit_cost or 0.0),
                float(svl.value or 0.0),
                svl.stock_move_id.product_id.display_name,
            ])
        for move in step["account_moves"]:
            aml = _aml_rows(move)
            ws.append([
                step["step"],
                step["doc_type"],
                step["doc"].name if step["doc"] else "",
                "JE Total",
                move.name,
                "",
                "",
                float(sum(aml.mapped("debit")) - sum(aml.mapped("credit"))),
                move.ref or "",
            ])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 7, 8])
    auto_width(ws, 56)

    ws = wb.create_sheet("DOC_TO_STOCK_SVL")
    ws.append(["Step", "Doc Type", "Doc No", "Stock Move ID", "Product", "Move Qty", "From", "To", "SVL ID", "SVL Qty", "SVL Value", "JE"])
    for step in steps:
        move_ids = []
        if step["doc_type"] == "MO":
            move_ids = (bundle["mo"].move_raw_ids | bundle["mo"].move_finished_ids).sorted(lambda m: m.id)
        elif hasattr(step["doc"], "move_ids"):
            move_ids = step["doc"].move_ids.sorted(lambda m: m.id)
        else:
            move_ids = env["stock.move"]
        svl_map = {}
        for svl in step["svls"]:
            svl_map.setdefault(svl.stock_move_id.id, [])
            svl_map[svl.stock_move_id.id].append(svl)
        for move in move_ids:
            svls = svl_map.get(move.id) or [None]
            for svl in svls:
                ws.append([
                    step["step"],
                    step["doc_type"],
                    step["doc"].name if step["doc"] else "",
                    move.id,
                    move.product_id.display_name,
                    float(move.quantity or move.product_uom_qty or 0.0),
                    move.location_id.complete_name,
                    move.location_dest_id.complete_name,
                    svl.id if svl else "",
                    float(svl.quantity or 0.0) if svl else "",
                    float(svl.value or 0.0) if svl else "",
                    svl.account_move_id.name if svl and svl.account_move_id else "",
                ])
    style_header(ws)
    style_grid(ws, amount_cols=[6, 10, 11])
    auto_width(ws, 52)

    ws = wb.create_sheet("DEBIT_CREDIT_DETAIL")
    ws.append(["Step", "Doc Type", "Doc No", "JE", "Account Code", "Account Name", "Debit", "Credit", "Balance", "Label"])
    for step in steps:
        for move in step["account_moves"]:
            for line in _aml_rows(move):
                ws.append([
                    step["step"],
                    step["doc_type"],
                    step["doc"].name if step["doc"] else "",
                    move.name,
                    line.account_id.code,
                    line.account_id.name,
                    float(line.debit or 0.0),
                    float(line.credit or 0.0),
                    float((line.debit or 0.0) - (line.credit or 0.0)),
                    line.name or "",
                ])
    style_header(ws)
    style_grid(ws, amount_cols=[7, 8, 9])
    auto_width(ws, 52)

    ws = wb.create_sheet("TEST_COVERAGE")
    ws.append(["Case ID", "Title", "Status", "Reason"])
    for row in _bundle_coverage():
        ws.append(list(row))
    style_header(ws)
    style_grid(ws)
    auto_width(ws, 96)

    wb.save(path)
    return path


def main():
    bundle = _find_latest_bundle(env)
    path = write_report(env, bundle)
    print("REPORT:", path)
    for label, value in _doc_numbers(bundle):
        print(f"{label.upper()}:", value)


main()
