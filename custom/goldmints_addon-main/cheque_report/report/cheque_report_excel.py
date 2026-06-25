# -*- coding: utf-8 -*-
import io
from datetime import date
from odoo import http
from odoo.http import request
import openpyxl
from openpyxl.styles import Font, Alignment


class ChequeReportExcelController(http.Controller):

    @http.route("/web/cheque_report/export_excel", type="http", auth="user")
    def export_excel(self, wizard_id, **kwargs):
        wizard = request.env["cheque.report"].browse(int(wizard_id))
        if not wizard.exists():
            return request.not_found()

        # ดึงข้อมูลการแบ่งหน้าจากโมเดลหลัก
        pages_data = wizard.get_report_data()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        report_font = Font(name="AngsanaUPC", size=16)

        column_widths = {
            "A": 5.57,
            "B": 18.57,
            "C": 13.71,
            "D": 12.57,
            "E": 10.57,
            "F": 16.57,
            "G": 14.14,
            "H": 13.14,
        }

        row_heights = {
            1: 16.5,
            2: 16.5,
            3: 20.25,
            4: 23.25,
            5: 18,
            6: 23.25,
            7: 18,
            8: 23.25,
            9: 23.25,
            10: 15.75,
            11: 16.5,
            12: 23.25,
            13: 19.5,
            14: 17.25,
            15: 16.5,
            16: 15,
            17: 18.75,
            18: 18,
            19: 16.5,
            20: 17.25,
            21: 19.5,
            22: 15,
            23: 23.25,
            24: 17.25,
            25: 23.25,
            26: 15,
            27: 23.25,
            28: 18,
            29: 23.25,
            30: 15.75,
            31: 19.5,
            32: 17.25,
            33: 23.25,
            34: 15.75,
            35: 23.25,
            36: 15,
            37: 23.25,
            38: 18.75,
            39: 23.25,
            40: 15,
            41: 23.25,
            42: 15.75,
            43: 19.5,
            44: 16.5,
            45: 17.25,
            46: 22.5,
            47: 21.75,
            48: 15.75,
            49: 23.25,
            50: 21,
            51: 21,
            52: 21,
        }

        # วนลูปสร้างแต่ละแผ่นงานตามเลขหน้า
        for page in pages_data:
            sheet_name = f"Sheet {page['page_no']}"
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            # กำหนดขนาดคอลัมน์และแถวประจำหน้า
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for row, height in row_heights.items():
                ws.row_dimensions[row].height = height

            # ฟังก์ชันช่วยเขียนข้อมูลลงเซลล์และกำหนดฟอนต์ Angsana 16
            def write_cell(cell_coord, value):
                ws[cell_coord] = value
                ws[cell_coord].font = report_font
                ws[cell_coord].alignment = Alignment(
                    horizontal="left", vertical="center"
                )

            # --- [ ข้อมูล HEADER ] ---
            header = page["header"]
            write_cell("B4", header["date"])
            write_cell("D4", header["cheque_bank_branch"])
            write_cell("H4", header["cheque_bank_branch"])
            write_cell("B6", header["bank_number_full"])
            write_cell("E6", header["company_bank_name"])

            # --- [ ข้อมูล SUB HEADER / FOOTER ] ---
            footer = page["footer"]
            write_cell("B12", footer["page_total_bahttext"])

            write_cell("F12", footer["page_total_amount"])
            ws["F12"].number_format = "#,##0.00"
            ws["F12"].alignment = Alignment(horizontal="right", vertical="center")

            write_cell("H12", footer["page_items_count"])
            ws["H12"].alignment = Alignment(horizontal="left", vertical="center")

            # แผนผังพิกัดแถวเช็ครายการที่ 1 - 5
            line_mapping = {
                1: {
                    "cheque_number": "B21",
                    "bank_name": "C21",
                    "branch_name": "D21",
                    "cheque_date": "F21",
                    "cheque_amount": "H21",
                    "partner_name": "B23",
                    "partner_code": "E23",
                    "invoice_name": "B25",
                    "sale_region": "E25",
                    "saleperson_name": "G25",
                },
                2: {
                    "cheque_number": "B27",
                    "bank_name": "C27",
                    "branch_name": "D27",
                    "cheque_date": "F27",
                    "cheque_amount": "H27",
                    "partner_name": "B29",
                    "partner_code": "E29",
                    "invoice_name": "B31",
                    "sale_region": "E31",
                    "saleperson_name": "G31",
                },
                3: {
                    "cheque_number": "B33",
                    "bank_name": "C33",
                    "branch_name": "D33",
                    "cheque_date": "F33",
                    "cheque_amount": "H33",
                    "partner_name": "B35",
                    "partner_code": "E35",
                    "invoice_name": "B37",
                    "sale_region": "E37",
                    "saleperson_name": "G37",
                },
                4: {
                    "cheque_number": "B39",
                    "bank_name": "C39",
                    "branch_name": "D39",
                    "cheque_date": "F39",
                    "cheque_amount": "H39",
                    "partner_name": "B41",
                    "partner_code": "E41",
                    "invoice_name": "B43",
                    "sale_region": "E43",
                    "saleperson_name": "G43",
                },
                5: {
                    "cheque_number": "B45",
                    "bank_name": "C45",
                    "branch_name": "D45",
                    "cheque_date": "F45",
                    "cheque_amount": "H45",
                    "partner_name": "B47",
                    "partner_code": "E47",
                    "invoice_name": "B49",
                    "sale_region": "E49",
                    "saleperson_name": "G49",
                },
            }

            for idx, line in enumerate(page["lines"], start=1):
                coords = line_mapping[idx]

                write_cell(coords["cheque_number"], line["cheque_number"])

                write_cell(coords["bank_name"], line["bank_name"])
                ws[coords["bank_name"]].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

                write_cell(coords["branch_name"], line["branch_name"])
                write_cell(coords["cheque_date"], line["cheque_date"])

                write_cell(coords["cheque_amount"], line["cheque_amount"])
                ws[coords["cheque_amount"]].number_format = "#,##0.00"
                ws[coords["cheque_amount"]].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

                write_cell(coords["partner_name"], line["partner_name"])
                write_cell(coords["partner_code"], line["partner_code"])
                write_cell(coords["invoice_name"], line["invoice_name"])

                write_cell(coords["sale_region"], line["sale_region"])
                ws[coords["sale_region"]].alignment = Alignment(
                    horizontal="left", vertical="center"
                )

                write_cell(coords["saleperson_name"], line["saleperson_name"])
                ws[coords["saleperson_name"]].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        excel_file = fp.read()
        fp.close()

        filename = f"Cheque_Deposit_Report_{date.today().strftime('%d-%m-%Y')}.xlsx"

        return request.make_response(
            excel_file,
            headers=[
                (
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                ("Content-Disposition", f"attachment; filename={filename};"),
            ],
        )
