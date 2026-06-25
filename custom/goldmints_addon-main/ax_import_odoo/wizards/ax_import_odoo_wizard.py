# -*- coding: utf-8 -*-
import base64
import io
import re
import zipfile
from datetime import date, datetime, timedelta
from html import escape

from openpyxl import load_workbook

from odoo import Command, api, fields, models, _
from odoo.exceptions import UserError


class AXImportOdooWizard(models.TransientModel):
    _name = "ax.import.odoo.wizard"
    _description = "AX Import Odoo"

    voucher_file = fields.Binary(string="AX Voucher XLSX File", required=True)
    voucher_filename = fields.Char()

    default_journal_id = fields.Many2one(
        "account.journal",
        string="Default Journal",
        required=True,
        domain="[('company_id', '=', company_id), ('type', 'in', ['general', 'bank', 'cash'])]",
        check_company=True,
        default=lambda self: self._default_journal(),
        help="Journal used for the created Odoo journal entries.",
    )
    company_id = fields.Many2one(
        "res.company",
        required=True,
        default=lambda self: self.env.company,
    )

    post_entries = fields.Boolean(
        string="Post Journal Entries",
        help="If enabled, created entries are posted automatically. If posting fails, they remain in draft.",
    )
    skip_existing_moves = fields.Boolean(
        string="Skip Existing AX Vouchers",
        default=True,
        help="Skip a sheet when another journal entry with the same AX voucher number already exists.",
    )
    account_code_mode = fields.Selection(
        [
            ("main", "Use Main Account Before Dash"),
            ("full", "Use Full AX Account Number"),
        ],
        string="Account Code Matching",
        default="main",
        required=True,
        help="AX account numbers often include a dash suffix for dimensions. Main mode maps 700018-7 to account 700018.",
    )
    create_missing_accounts = fields.Boolean(
        string="Create Missing Accounts",
        default=True,
        help="Create an account when the AX account number cannot be matched.",
    )
    create_missing_partners = fields.Boolean(
        string="Create Missing Partners",
        default=True,
        help="Create vendor/customer partners from the AX voucher header when no partner is found.",
    )
    strict_partner_matching = fields.Boolean(
        string="Strict Partner Matching",
        help="Fail the import if a partner from the voucher header cannot be found or created.",
    )
    partner_policy = fields.Selection(
        [
            ("receivable_payable", "Receivable/Payable Lines Only"),
            ("all", "All Lines"),
            ("none", "Do Not Assign"),
        ],
        string="Partner Assignment",
        default="receivable_payable",
        required=True,
    )

    created_move_ids = fields.Many2many("account.move", string="Created Entries", readonly=True)
    created_move_count = fields.Integer(compute="_compute_created_move_count", readonly=True)
    skipped_move_count = fields.Integer(string="Skipped Entries", readonly=True)
    failed_move_count = fields.Integer(string="Failed Entries", readonly=True)
    log = fields.Text(readonly=True)

    @api.model
    def _default_journal(self):
        return self.env["account.journal"].search(
            [
                ("company_id", "=", self.env.company.id),
                ("type", "=", "general"),
            ],
            order="sequence, id",
            limit=1,
        )

    @api.depends("created_move_ids")
    def _compute_created_move_count(self):
        for wizard in self:
            wizard.created_move_count = len(wizard.created_move_ids)

    def _reload(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    def _log_line(self, message):
        self.ensure_one()
        now = fields.Datetime.to_string(fields.Datetime.now())
        self.log = (self.log or "") + "[%s] %s\n" % (now, message)

    def action_import_ax_import_odoo(self):
        self.ensure_one()
        if not self.voucher_file:
            raise UserError(_("Please upload an AX voucher XLSX file first."))

        content = base64.b64decode(self.voucher_file)
        if not self._is_xlsx_file(self.voucher_filename, content):
            raise UserError(_("Only XLSX workbooks are supported."))

        self.created_move_ids = [Command.clear()]
        self.skipped_move_count = 0
        self.failed_move_count = 0
        self.log = False

        vouchers = self._parse_workbook(content, self.voucher_filename or "ax_import_odoo.xlsx")
        if not vouchers:
            raise UserError(_("No AX voucher sheets were found in the uploaded workbook."))

        created_moves = self.env["account.move"]
        skipped_count = 0
        failed = []
        cache = {
            "accounts": {},
            "partners": {},
            "missing_partners": set(),
        }

        self._log_line(_("Starting AX voucher import with %s sheet(s).") % len(vouchers))
        for voucher in vouchers:
            try:
                with self.env.cr.savepoint():
                    move, skipped = self._create_move_from_voucher(voucher, cache)
                    if skipped:
                        skipped_count += 1
                    elif move:
                        created_moves |= move
            except Exception as exc:
                label = voucher.get("voucher_number") or voucher.get("sheet_name") or "-"
                message = _("Failed to import AX voucher %s: %s") % (label, str(exc))
                failed.append(message)
                self._log_line(message)

        if not created_moves and not skipped_count and failed:
            raise UserError("\n".join(failed[:10]))

        if self.post_entries and created_moves:
            try:
                with self.env.cr.savepoint():
                    created_moves.action_post()
                self._log_line(_("Posted entries: %s") % len(created_moves))
            except Exception as exc:
                self._log_line(_("Posting failed. Entries kept in draft: %s") % str(exc))

        self.created_move_ids = [Command.set(created_moves.ids)]
        self.skipped_move_count = skipped_count
        self.failed_move_count = len(failed)
        self._log_line(
            _("Import completed. Created=%s, Skipped=%s, Failed=%s")
            % (len(created_moves), skipped_count, len(failed))
        )
        return self._reload()

    def _create_move_from_voucher(self, voucher, cache):
        voucher_number = self._to_str(voucher.get("voucher_number"))
        if self.skip_existing_moves and voucher_number:
            existing = self.env["account.move"].search(
                [
                    ("company_id", "=", self.company_id.id),
                    ("ax_voucher_number", "=", voucher_number),
                ],
                limit=1,
            )
            if existing:
                self._log_line(_("Skipped existing AX voucher %s (%s).") % (voucher_number, existing.display_name))
                return False, True

        move_date = voucher.get("transaction_date") or voucher.get("document_date") or voucher.get("due_date")
        if not move_date:
            raise UserError(_("Transaction date is required."))

        line_vals = []
        total_debit = 0.0
        total_credit = 0.0

        for line in voucher.get("lines", []):
            row_number = line.get("row_number")
            account = self._resolve_account(
                line.get("account_code"),
                line.get("account_name"),
                cache,
                row_number,
            )
            debit = self._to_float(line.get("debit"), row_number, "debit")
            credit = self._to_float(line.get("credit"), row_number, "credit")
            if debit < 0 or credit < 0:
                raise UserError(_("Row %s: debit/credit cannot be negative.") % row_number)
            if debit and credit:
                raise UserError(_("Row %s: line cannot contain both debit and credit.") % row_number)
            if not debit and not credit:
                continue

            total_debit += debit
            total_credit += credit

            vals = {
                "name": self._to_str(line.get("description")) or self._to_str(line.get("account_name")) or voucher_number,
                "account_id": account.id,
                "debit": debit,
                "credit": credit,
                "tax_ids": [Command.clear()],
                "tax_tag_ids": [Command.clear()],
                "ax_account_code": self._to_str(line.get("account_code")) or False,
            }

            partner = self._resolve_partner_for_line(voucher, account, cache, row_number)
            if partner:
                vals["partner_id"] = partner.id

            line_vals.append(Command.create(vals))

        if not line_vals:
            raise UserError(_("No importable journal lines found."))

        company_currency = self.company_id.currency_id
        if company_currency.compare_amounts(total_debit, total_credit) != 0:
            raise UserError(
                _("Voucher %s is not balanced (debit=%s, credit=%s).")
                % (voucher_number or "-", total_debit, total_credit)
            )

        move_vals = {
            "move_type": "entry",
            "date": move_date,
            "journal_id": self.default_journal_id.id,
            "company_id": self.company_id.id,
            "ref": self._prepare_move_ref(voucher),
            "line_ids": line_vals,
            "narration": self._prepare_narration(voucher),
            "ax_voucher_number": voucher_number or False,
            "ax_voucher_type": voucher.get("voucher_type") or False,
            "ax_document_number": voucher.get("invoice_number") or False,
            "ax_journal_number": voucher.get("journal_number") or False,
            "ax_journal_description": voucher.get("journal_description") or False,
            "ax_source_filename": voucher.get("source_filename") or False,
        }

        move = self.env["account.move"].with_context(skip_computed_taxes=True).create(move_vals)
        self._log_line(_("Created journal entry %s for AX voucher %s.") % (move.display_name, voucher_number or "-"))
        return move, False

    def _parse_workbook(self, content, filename):
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        except Exception as exc:
            raise UserError(_("Unable to read XLSX workbook: %s") % str(exc))

        vouchers = []
        for sheet in workbook.worksheets:
            voucher = self._parse_sheet(sheet, filename)
            if voucher:
                vouchers.append(voucher)
        return vouchers

    def _parse_sheet(self, sheet, filename):
        header = self._find_line_table_header(sheet)
        if not header:
            return False

        lines = self._parse_line_table(sheet, header)
        if not lines:
            return False

        vendor_raw = self._value_after_label(sheet, ("vendor",))
        customer_raw = self._value_after_label(sheet, ("customer",))
        partner_type = "vendor" if vendor_raw else "customer" if customer_raw else False
        partner_code, partner_name = self._split_code_name(vendor_raw or customer_raw)

        journal_raw = self._value_after_label(sheet, ("journalno",))
        journal_number, journal_description = self._split_code_name(journal_raw)

        voucher = {
            "source_filename": filename,
            "sheet_name": sheet.title,
            "voucher_number": self._to_str(self._value_after_label(sheet, ("voucherno",))),
            "voucher_type": self._find_voucher_type(sheet),
            "transaction_date": self._to_date_or_false(self._value_after_label(sheet, ("transactiondate",))),
            "document_date": self._to_date_or_false(self._value_after_label(sheet, ("documentdate",))),
            "due_date": self._to_date_or_false(self._value_after_label(sheet, ("duedate",))),
            "invoice_number": self._to_str(self._value_after_label(sheet, ("invoicetaxinvoiceno",))),
            "withholding_tax_amount": self._to_float(
                self._value_after_label(sheet, ("withholdingtax",)),
                field_name="withholding tax",
            ),
            "bank": self._to_str(self._value_after_label(sheet, ("bank",))),
            "partner_raw": self._to_str(vendor_raw or customer_raw),
            "partner_type": partner_type,
            "partner_code": partner_code,
            "partner_name": partner_name,
            "journal_raw": self._to_str(journal_raw),
            "journal_number": journal_number,
            "journal_description": journal_description,
            "settlements": self._parse_settlements(sheet),
            "lines": lines,
        }
        return voucher

    def _find_line_table_header(self, sheet):
        for row_index in range(1, sheet.max_row + 1):
            columns = {}
            for column_index in range(1, sheet.max_column + 1):
                normalized = self._normalize_token(sheet.cell(row_index, column_index).value)
                if normalized == "accountno":
                    columns["account_code"] = column_index
                elif normalized == "accountdescription":
                    columns["account_name"] = column_index
                elif normalized == "description":
                    columns["description"] = column_index
                elif normalized == "debit":
                    columns["debit"] = column_index
                elif normalized == "credit":
                    columns["credit"] = column_index

            required = {"account_code", "account_name", "description", "debit", "credit"}
            if required.issubset(columns):
                return {"row": row_index, "columns": columns}
        return False

    def _parse_line_table(self, sheet, header):
        columns = header["columns"]
        rows = []
        started = False
        for row_index in range(header["row"] + 1, sheet.max_row + 1):
            account_code = self._to_str(sheet.cell(row_index, columns["account_code"]).value)
            if self._normalize_token(account_code) == "total":
                break

            account_name = self._to_str(sheet.cell(row_index, columns["account_name"]).value)
            description = self._to_str(sheet.cell(row_index, columns["description"]).value)
            debit = sheet.cell(row_index, columns["debit"]).value
            credit = sheet.cell(row_index, columns["credit"]).value

            if not any((account_code, account_name, description, debit, credit)):
                if started:
                    continue
                continue

            started = True
            if not account_code:
                raise UserError(_("Sheet %s row %s: Account No. is required.") % (sheet.title, row_index))

            rows.append(
                {
                    "row_number": row_index,
                    "account_code": account_code,
                    "account_name": account_name,
                    "description": description,
                    "debit": debit,
                    "credit": credit,
                }
            )
        return rows

    def _parse_settlements(self, sheet):
        start_row = False
        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                if "settlement" in self._normalize_token(sheet.cell(row_index, column_index).value):
                    start_row = row_index
                    break
            if start_row:
                break
        if not start_row:
            return []

        header_row = False
        columns = {}
        for row_index in range(start_row, min(start_row + 5, sheet.max_row) + 1):
            found = {}
            for column_index in range(1, sheet.max_column + 1):
                normalized = self._normalize_token(sheet.cell(row_index, column_index).value)
                if normalized == "journalno":
                    found["journal_no"] = column_index
                elif normalized == "voucherno":
                    found["voucher_no"] = column_index
                elif normalized == "invoiceno":
                    found["invoice_no"] = column_index
                elif normalized == "date":
                    found["date"] = column_index
                elif normalized == "originalamount":
                    found["original_amount"] = column_index
                elif normalized == "whtamount":
                    found["wht_amount"] = column_index
                elif normalized == "settleamount":
                    found["settle_amount"] = column_index
            if {"voucher_no", "invoice_no", "original_amount", "settle_amount"}.issubset(found):
                header_row = row_index
                columns = found
                break

        if not header_row:
            return []

        settlements = []
        first_data_row = header_row + 1
        main_col = columns.get("journal_no") or columns.get("voucher_no")
        for row_index in range(first_data_row, sheet.max_row + 1):
            first_value = self._to_str(sheet.cell(row_index, main_col).value) if main_col else ""
            if self._normalize_token(first_value) == "total":
                break

            row = {}
            for key, column_index in columns.items():
                row[key] = sheet.cell(row_index, column_index).value

            if not any(value not in (None, "", False) for value in row.values()):
                if settlements:
                    break
                continue

            settlements.append(
                {
                    "journal_no": self._to_str(row.get("journal_no")),
                    "voucher_no": self._to_str(row.get("voucher_no")),
                    "invoice_no": self._to_str(row.get("invoice_no")),
                    "date": self._to_date_or_false(row.get("date")),
                    "original_amount": self._to_float(row.get("original_amount")),
                    "wht_amount": self._to_float(row.get("wht_amount")),
                    "settle_amount": self._to_float(row.get("settle_amount")),
                }
            )
        return settlements

    def _resolve_account(self, account_code, account_name, cache, row_number):
        raw_code = self._to_str(account_code)
        raw_name = self._to_str(account_name)
        if not raw_code:
            raise UserError(_("Row %s: Account No. is required.") % row_number)

        cache_key = (self.account_code_mode, raw_code, raw_name)
        if cache_key in cache["accounts"]:
            return cache["accounts"][cache_key]

        account_model = self.env["account.account"].with_context(active_test=False).with_company(self.company_id)
        for candidate in self._account_code_candidates(raw_code):
            account = account_model.search(
                [
                    ("company_ids", "in", self.company_id.id),
                    ("code", "=ilike", candidate),
                ],
                limit=1,
            )
            if account:
                cache["accounts"][cache_key] = account
                return account

        if raw_name:
            account = account_model.search(
                [
                    ("company_ids", "in", self.company_id.id),
                    ("name", "=ilike", raw_name),
                ],
                limit=1,
            )
            if account:
                cache["accounts"][cache_key] = account
                return account

        if not self.create_missing_accounts:
            raise UserError(_("Row %s: account not found (%s).") % (row_number, raw_code))

        account = self._create_missing_account(raw_code, raw_name, row_number)
        cache["accounts"][cache_key] = account
        return account

    def _create_missing_account(self, account_code, account_name, row_number):
        code = self._account_code_for_create(account_code)
        if not code:
            raise UserError(_("Row %s: account code is invalid (%s).") % (row_number, account_code))

        account_type = self._guess_account_type(code, account_name)
        vals = {
            "code": code,
            "name": account_name or _("AX Account %s") % code,
            "account_type": account_type,
            "company_ids": [Command.set([self.company_id.id])],
        }
        if account_type in ("asset_receivable", "liability_payable"):
            vals["reconcile"] = True

        account_model = self.env["account.account"].with_context(active_test=False).with_company(self.company_id)
        try:
            account = account_model.create(vals)
        except Exception:
            sanitized = self._sanitize_account_code(code)
            if not sanitized or sanitized == code:
                raise
            existing = account_model.search(
                [
                    ("company_ids", "in", self.company_id.id),
                    ("code", "=ilike", sanitized),
                ],
                limit=1,
            )
            if existing:
                return existing
            vals["code"] = sanitized
            account = account_model.create(vals)

        self._log_line(_("Row %s: account '%s' was missing and auto-created.") % (row_number, account.code))
        return account

    def _resolve_partner_for_line(self, voucher, account, cache, row_number):
        if self.partner_policy == "none":
            return False

        account_type = getattr(account, "account_type", "") or ""
        if self.partner_policy == "receivable_payable" and account_type not in (
            "asset_receivable",
            "liability_payable",
        ):
            return False

        partner_name = self._to_str(voucher.get("partner_name"))
        partner_code = self._to_str(voucher.get("partner_code"))
        if not partner_name and not partner_code:
            return False

        cache_key = (partner_code, partner_name, voucher.get("partner_type"))
        if cache_key in cache["partners"]:
            return cache["partners"][cache_key]

        partner = self._find_partner(partner_code, partner_name)
        if not partner and self.create_missing_partners:
            partner = self._create_missing_partner(partner_code, partner_name, voucher.get("partner_type"))

        if not partner:
            label = partner_name or partner_code
            if self.strict_partner_matching:
                raise UserError(_("Row %s: partner not found (%s).") % (row_number, label))
            if label not in cache["missing_partners"]:
                cache["missing_partners"].add(label)
                self._log_line(_("Partner not found on row %s, left empty: %s") % (row_number, label))
            return False

        cache["partners"][cache_key] = partner
        return partner

    def _find_partner(self, partner_code, partner_name):
        Partner = self.env["res.partner"].with_context(active_test=False)
        code = self._to_str(partner_code)
        name = self._to_str(partner_name)
        if code and name:
            return Partner.search(
                [
                    "|",
                    ("ref", "=ilike", code),
                    ("name", "=ilike", name),
                ],
                limit=1,
            )
        if code:
            return Partner.search([("ref", "=ilike", code)], limit=1)
        if name:
            return Partner.search([("name", "=ilike", name)], limit=1)
        return Partner.browse()

    def _create_missing_partner(self, partner_code, partner_name, partner_type):
        vals = {
            "name": partner_name or partner_code,
            "ref": partner_code or False,
            "company_type": "company",
        }
        if partner_type == "vendor":
            vals["supplier_rank"] = 1
        elif partner_type == "customer":
            vals["customer_rank"] = 1
        partner = self.env["res.partner"].create(vals)
        self._log_line(_("Created partner from AX header: %s") % partner.display_name)
        return partner

    def _prepare_move_ref(self, voucher):
        parts = []
        for key in ("voucher_number", "invoice_number", "journal_number"):
            value = self._to_str(voucher.get(key))
            if value and value not in parts:
                parts.append(value)
        journal_description = self._to_str(voucher.get("journal_description"))
        if journal_description:
            parts.append(journal_description)
        return " | ".join(parts) or False

    def _prepare_narration(self, voucher):
        lines = [
            _("Imported from AX voucher XLSX"),
            _("Source file: %s") % (voucher.get("source_filename") or "-"),
            _("Sheet: %s") % (voucher.get("sheet_name") or "-"),
        ]
        for label, key in (
            (_("Voucher"), "voucher_number"),
            (_("Voucher Type"), "voucher_type"),
            (_("Partner"), "partner_raw"),
            (_("Invoice/Tax Invoice"), "invoice_number"),
            (_("Journal"), "journal_raw"),
            (_("Bank"), "bank"),
        ):
            value = self._to_str(voucher.get(key))
            if value:
                lines.append("%s: %s" % (label, value))

        wht_amount = voucher.get("withholding_tax_amount")
        if wht_amount:
            lines.append(_("Withholding Tax: %s") % wht_amount)

        settlements = voucher.get("settlements") or []
        for settlement in settlements:
            pieces = []
            for label, key in (
                (_("Voucher"), "voucher_no"),
                (_("Invoice"), "invoice_no"),
                (_("Original"), "original_amount"),
                (_("WHT"), "wht_amount"),
                (_("Settled"), "settle_amount"),
            ):
                value = settlement.get(key)
                if value not in (None, "", False):
                    pieces.append("%s %s" % (label, value))
            if pieces:
                lines.append(_("Settlement: %s") % ", ".join(pieces))

        return "<br/>".join(escape(line) for line in lines)

    def _find_voucher_type(self, sheet):
        for row in sheet.iter_rows(values_only=True):
            for value in row:
                text = self._to_str(value).upper()
                if "PAYMENT VOUCHER" in text:
                    return "payment"
                if "RECEIPT VOUCHER" in text:
                    return "receipt"
                if "PAYABLE VOUCHER" in text:
                    return "payable"
                if "JOURNAL VOUCHER" in text:
                    return "journal"
        return False

    def _value_after_label(self, sheet, normalized_labels):
        labels = set(normalized_labels)
        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row_index, column_index).value
                normalized = self._normalize_token(cell_value)
                if normalized not in labels:
                    continue

                for offset in range(1, min(35, sheet.max_column - column_index) + 1):
                    value = sheet.cell(row_index, column_index + offset).value
                    if value not in (None, "", False):
                        return value
        return False

    def _split_code_name(self, value):
        text = self._to_str(value)
        if not text:
            return "", ""
        if ":" in text:
            code, name = text.split(":", 1)
            return code.strip(), name.strip()
        return "", text

    def _account_code_candidates(self, account_code):
        raw = self._to_str(account_code).replace(" ", "")
        main = self._main_account_code(raw)
        sanitized = self._sanitize_account_code(raw)
        if self.account_code_mode == "full":
            candidates = [raw, sanitized, main]
        else:
            candidates = [main, raw, sanitized]

        result = []
        for candidate in candidates:
            if candidate and candidate not in result:
                result.append(candidate)
        return result

    def _account_code_for_create(self, account_code):
        raw = self._to_str(account_code).replace(" ", "")
        if self.account_code_mode == "full":
            return raw
        return self._main_account_code(raw) or raw

    def _main_account_code(self, account_code):
        raw = self._to_str(account_code).replace(" ", "")
        if "-" in raw:
            return raw.split("-", 1)[0]
        return raw

    def _sanitize_account_code(self, account_code):
        return re.sub(r"[^A-Za-z0-9.]", "", self._to_str(account_code))

    def _guess_account_type(self, account_code, account_name):
        code = self._to_str(account_code)
        name = self._to_str(account_name).lower()

        if "receivable" in name or "ลูกหนี้" in name or code.startswith("113"):
            return "asset_receivable"
        if "payable" in name or "เจ้าหนี้" in name:
            return "liability_payable"
        if "bank" in name or "cash" in name or "เงินสด" in name or code.startswith("111"):
            return "asset_cash"

        first_digit = ""
        for char in code:
            if char.isdigit():
                first_digit = char
                break

        if first_digit == "1":
            return "asset_current"
        if first_digit == "2":
            return "liability_current"
        if first_digit == "3":
            return "equity"
        if first_digit == "4":
            return "income"
        if first_digit == "5":
            return "expense_direct_cost"
        return "expense"

    def _normalize_token(self, value):
        text = self._to_str(value).lower()
        if not text:
            return ""
        return re.sub(r"[^a-z0-9]", "", text)

    def _to_str(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _to_float(self, value, row_number=None, field_name="value"):
        if value in (None, "", False):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)

        text = self._to_str(value)
        if not text:
            return 0.0

        is_negative = text.startswith("(") and text.endswith(")")
        if is_negative:
            text = text[1:-1]

        text = text.replace(",", "")
        text = text.replace(" ", "")
        try:
            amount = float(text)
        except Exception:
            if row_number:
                raise UserError(_("Row %s: invalid %s value '%s'.") % (row_number, field_name, value))
            raise UserError(_("Invalid number value: %s") % value)

        if is_negative:
            amount *= -1
        return amount

    def _to_date_or_false(self, value):
        if value in (None, "", False):
            return False
        return self._to_date(value)

    def _to_date(self, value):
        if isinstance(value, datetime):
            return self._normalize_year(value.date())
        if isinstance(value, date):
            return self._normalize_year(value)
        if isinstance(value, (int, float)):
            return self._excel_float_to_date(float(value))

        text = self._to_str(value)
        if not text:
            return False

        short_thai = self._parse_short_thai_date(text)
        if short_thai:
            return short_thai

        if re.fullmatch(r"\d{8}", text):
            for fmt in ("%Y%m%d",):
                try:
                    return self._normalize_year(datetime.strptime(text, fmt).date())
                except Exception:
                    pass

        if re.fullmatch(r"-?\d+(\.\d+)?", text):
            return self._excel_float_to_date(float(text))

        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%d-%m-%Y",
            "%d.%m.%Y",
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M:%S",
        ]
        for fmt in formats:
            try:
                return self._normalize_year(datetime.strptime(text, fmt).date())
            except Exception:
                continue

        try:
            parsed = fields.Date.to_date(text)
            if parsed:
                return self._normalize_year(parsed)
        except Exception:
            pass

        raise UserError(_("Invalid date value: %s") % value)

    def _parse_short_thai_date(self, text):
        match = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2})", self._to_str(text))
        if not match:
            return False
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))
        if year >= 43:
            year = 2500 + year - 543
        else:
            year = 2000 + year
        return date(year, month, day)

    def _normalize_year(self, value):
        if value and value.year > 2400:
            return date(value.year - 543, value.month, value.day)
        return value

    def _excel_float_to_date(self, serial):
        if serial <= 0:
            raise UserError(_("Invalid Excel date serial: %s") % serial)
        base = date(1899, 12, 30)
        return base + timedelta(days=int(serial))

    def _is_xlsx_file(self, filename, content):
        if filename and filename.lower().endswith((".xlsx", ".xlsm")):
            return True
        try:
            return zipfile.is_zipfile(io.BytesIO(content))
        except Exception:
            return False
