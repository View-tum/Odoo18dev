import base64
import io
import logging
import math
import re
from datetime import date, datetime, time

from openpyxl import Workbook, load_workbook

from odoo import Command, _, fields, models
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)


class SaleOrderImportWizard(models.TransientModel):
    _name = "sale.order.import.wizard"
    _description = "Import Sale Orders from XLSX"

    data_file = fields.Binary(string="XLSX File", required=True)
    filename = fields.Char(string="Filename")
    update_existing = fields.Boolean(
        string="Update Existing Orders",
        help="If enabled, draft orders matched by name or document reference will "
        "be overwritten with the lines from the file.",
        default=False,
    )
    apply_discount_to_price = fields.Boolean(
        string="Embed Discount in Unit Price",
        help="If enabled, the discount percent from the file will be applied to the unit price "
        "and the discount field will be set to 0 to avoid pricelist discount restrictions.",
        default=False,
    )
    log_message = fields.Text(string="Log", readonly=True)
    success_count = fields.Integer(string="Successful", readonly=True)
    fail_count = fields.Integer(string="Failed", readonly=True)
    failure_file = fields.Binary(string="Failed Records File", readonly=True)
    failure_filename = fields.Char(string="Failed File Name", readonly=True)

    HEADER_ALIASES = {
        "document_ref": "document_ref",
        "partner_id/id": "partner_id/id",
        "partner_id/name": "partner_id/name",
        "partner_invoice_id/id": "partner_invoice_id/id",
        "partner_invoice_id/name": "partner_invoice_id/name",
        "partner_shipping_id/id": "partner_shipping_id/id",
        "partner_shipping_id/name": "partner_shipping_id/name",
        "date_order": "date_order",
        "commitment_date": "commitment_date",
        "payment_term_id": "payment_term_id",
        "sale_note": "sale_note",
        "order_line/product_id": "order_line/product_id",
        "order_line/product_id/name": "order_line/product_id/name",
        "order_line/product_uom_qty": "order_line/product_uom_qty",
        "order_line/price_unit": "order_line/price_unit",
        "order_line/discount": "order_line/discount",
        "so_type_id": "so_type_id",
        "warehouse_id": "warehouse_id",
        "user_id": "user_id",
        "website_id": "website_id",
        "name": "name",
        "ax_sale_order": "ax_sale_order",
        "ax sale order": "ax_sale_order",
        "id": "id",
    }

    ORDER_FIELDS = [
        "id",
        "name",
        "document_ref",
        "ax_sale_order",
        "partner_id/id",
        "partner_id/name",
        "partner_invoice_id/id",
        "partner_invoice_id/name",
        "partner_shipping_id/id",
        "partner_shipping_id/name",
        "date_order",
        "commitment_date",
        "payment_term_id",
        "sale_note",
        "so_type_id",
        "warehouse_id",
        "user_id",
        "website_id",
    ]

    LINE_FIELDS = [
        "order_line/product_id",
        "order_line/product_id/name",
        "order_line/product_uom_qty",
        "order_line/price_unit",
        "order_line/discount",
    ]

    def action_import(self):
        self.ensure_one()
        if not self.data_file:
            raise UserError(_("Please upload an XLSX file to import."))

        orders_payload = self._read_workbook()
        created_count = 0
        updated_count = 0
        fail_count = 0
        failures = []

        for order_key, payload in orders_payload.items():
            try:
                with self.env.cr.savepoint():
                    created, message = self._create_or_update_order(payload)
                    if created is True:
                        created_count += 1
                    elif created is False:
                        updated_count += 1
                    else:
                        fail_count += 1
                        failures.append(
                            {
                                "order_key": order_key,
                                "error": message or _("Skipped"),
                                "order": payload.get("order", {}),
                                "lines": payload.get("lines", []),
                            }
                        )
                        continue
            except ValidationError as exc:
                fail_count += 1
                failures.append(
                    {
                        "order_key": order_key,
                        "error": self._error_message(exc),
                        "order": payload.get("order", {}),
                        "lines": payload.get("lines", []),
                    }
                )
            except UserError as exc:
                fail_count += 1
                failures.append(
                    {
                        "order_key": order_key,
                        "error": self._error_message(exc),
                        "order": payload.get("order", {}),
                        "lines": payload.get("lines", []),
                    }
                )
            except Exception as exc:  # noqa: BLE001
                fail_count += 1
                _logger.exception("Unexpected error while importing sale order %s", order_key)
                failures.append(
                    {
                        "order_key": order_key,
                        "error": str(exc),
                        "order": payload.get("order", {}),
                        "lines": payload.get("lines", []),
                    }
                )

        success_total = created_count + updated_count
        vals = {
            "success_count": success_total,
            "fail_count": fail_count,
            "log_message": _("Successful: %s, Failed: %s") % (success_total, fail_count),
            "failure_file": False,
            "failure_filename": False,
        }
        if failures:
            data, fname = self._export_failures_xlsx(failures)
            vals.update(
                {
                    "failure_file": base64.b64encode(data),
                    "failure_filename": fname,
                }
            )
        # sudo to avoid write hooks failing on transient fields
        self.sudo().write(vals)

        return {
            "name": _("Sale Order Import"),
            "type": "ir.actions.act_window",
            "res_model": "sale.order.import.wizard",
            "view_mode": "form",
            "target": "new",
            "res_id": self.id,
        }

    # -- Workbook parsing -------------------------------------------------
    def _read_workbook(self):
        decoded = base64.b64decode(self.data_file)
        orders = {}

        # Prefer polars (fast) > pandas > streaming openpyxl.
        processed_sheets = self._read_with_polars(decoded, orders)
        if not processed_sheets:
            processed_sheets = self._read_with_pandas(decoded, orders)
        if not processed_sheets:
            processed_sheets = self._read_with_openpyxl(decoded, orders)

        if not processed_sheets:
            raise UserError(_("The provided file is empty or has no readable sheets."))

        return orders

    def _read_with_polars(self, decoded, orders):
        pl = self._ensure_polars()
        if not pl:
            return 0

        # Use openpyxl once to list sheets without loading all data.
        try:
            wb = load_workbook(io.BytesIO(decoded), read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as exc:  # noqa: BLE001
            _logger.debug("Openpyxl failed to read sheet names for polars: %s", exc)
            return 0

        processed = 0
        for sheet_name in sheet_names:
            try:
                df = pl.read_excel(io.BytesIO(decoded), sheet_name=sheet_name)
            except Exception as exc:  # noqa: BLE001
                _logger.debug("Polars failed reading sheet %s: %s", sheet_name, exc)
                continue

            if df.height == 0:
                continue

            header_index, headers_found = self._find_header_row_pl(df)
            if header_index is None:
                continue

            headers = self._normalize_headers(headers_found)
            current_key = None
            ax_format = self._is_ax_format(headers)

            for row in df.rows()[header_index + 1 :]:
                row_values = self._row_to_dict(headers, row)
                if ax_format:
                    row_values = self._convert_ax_row(row_values)
                current_key = self._consume_row(row_values, orders, current_key)

            processed += 1

        return processed

    def _read_with_pandas(self, decoded, orders):
        try:
            import pandas as pd
        except Exception:
            return 0

        try:
            excel = pd.ExcelFile(io.BytesIO(decoded), engine="openpyxl")
        except Exception as exc:  # noqa: BLE001
            _logger.debug("Pandas Excel load failed: %s", exc)
            return 0

        processed = 0
        for sheet_name in excel.sheet_names:
            try:
                df = excel.parse(sheet_name=sheet_name, header=None)
            except Exception as exc:  # noqa: BLE001
                _logger.debug("Failed reading sheet %s via pandas: %s", sheet_name, exc)
                continue

            if df.empty:
                continue

            header_index, headers_found = self._find_header_row_df(df)
            if header_index is None:
                continue

            headers = self._normalize_headers(headers_found)
            current_key = None
            ax_format = self._is_ax_format(headers)

            for row in df.iloc[header_index + 1 :].itertuples(index=False, name=None):
                row_values = self._row_to_dict(headers, row)
                if ax_format:
                    row_values = self._convert_ax_row(row_values)
                current_key = self._consume_row(row_values, orders, current_key)

            processed += 1

        return processed

    def _find_header_row_df(self, df):
        for idx, row in df.iterrows():
            headers = self._match_header_row(list(row))
            if headers:
                return idx, headers
        return None, None

    def _find_header_row_pl(self, df):
        for idx, row in enumerate(df.rows()):
            headers = self._match_header_row(list(row))
            if headers:
                return idx, headers
        return None, None

    def _read_with_openpyxl(self, decoded, orders):
        try:
            workbook = load_workbook(io.BytesIO(decoded), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise UserError(_("Could not read the XLSX file: %s") % exc) from exc

        processed_sheets = 0

        for sheet in workbook.worksheets:
            header = None
            current_key = None
            ax_format = False

            for row in sheet.iter_rows(values_only=True):
                if header is None:
                    headers_found = self._match_header_row(list(row))
                    if not headers_found:
                        continue
                    header = self._normalize_headers(headers_found)
                    ax_format = self._is_ax_format(header)
                    processed_sheets += 1
                    continue

                row_values = self._row_to_dict(header, row)
                if ax_format:
                    row_values = self._convert_ax_row(row_values)
                current_key = self._consume_row(row_values, orders, current_key)

        workbook.close()
        return processed_sheets

    def _row_to_dict(self, headers, row):
        result = {}
        for index, value in enumerate(row):
            if index >= len(headers):
                continue
            header_key = headers[index]
            if not header_key:
                continue
            value = self._clean_cell_value(value)
            if value in (None, ""):
                continue
            result[header_key] = value
        return result

    def _map_header(self, header):
        if self._is_empty_cell(header):
            return ""
        normalized = str(header).strip().lower()
        return self.HEADER_ALIASES.get(normalized, normalized)

    def _normalize_headers(self, headers):
        seen = {}
        normalized = []
        for h in headers:
            key = h or ""
            count = seen.get(key, 0)
            if key == "name":
                key = "partner_name_ax" if count == 0 else "line_name_ax"
            elif count:
                key = f"{key}_{count}"
            seen[h] = count + 1
            normalized.append(key)
        return normalized

    def _match_header_row(self, row):
        if not row or all(self._is_empty_cell(cell) for cell in row):
            return None
        mapped = [self._map_header(cell) for cell in row]
        non_empty = [h for h in mapped if h]
        if len(non_empty) < 2:
            return None
        # Recognize either AX format or Odoo export header
        if "sales id" in non_empty and "item id" in non_empty:
            return mapped
        expected = set(self.ORDER_FIELDS + self.LINE_FIELDS)
        if any(h in expected for h in non_empty):
            return mapped
        return None

    def _clean_cell_value(self, value):
        if self._is_empty_cell(value):
            return None
        if isinstance(value, str):
            return value.strip()
        return value

    def _is_ax_format(self, headers):
        lower = set([h for h in headers if h])
        return "sales id" in lower and "item id" in lower

    def _convert_ax_row(self, row):
        # Map AX export columns to Odoo import structure
        name_customer = row.get("partner_name_ax") or row.get("name")
        line_name = row.get("line_name_ax") or row.get("name")
        sale_price = row.get("sale price") or row.get("price unit")
        discount = row.get("line percent") or row.get("line disc.") or row.get("line discount")
        result = {
            "document_ref": row.get("sales id"),
            "ax_sale_order": row.get("sales id"),
            "partner_id/name": name_customer,
            "partner_invoice_id/name": row.get("invoice account") or name_customer,
            "partner_shipping_id/name": row.get("delivery name") or name_customer,
            "date_order": row.get("delivery date"),
            "commitment_date": row.get("delivery date"),
            "payment_term_id": row.get("payment"),
            "sale_note": row.get("delivery address") or row.get("sale status"),
            "so_type_id": row.get("taxgroup") or row.get("sales line status"),
            "order_line/product_id": row.get("item id"),
            "order_line/product_id/name": line_name,
            "order_line/product_uom_qty": row.get("order qty") or row.get("quantity") or row.get("qty"),
            "order_line/price_unit": sale_price,
            "order_line/discount": discount,
        }
        return result

    # -- Record creation helpers -----------------------------------------
    def _create_or_update_order(self, payload):
        order_vals = payload.get("order", {})
        line_vals = payload.get("lines", [])

        partner = self._find_partner(order_vals.get("partner_id/id"), order_vals.get("partner_id/name"))
        if not partner:
            raise ValidationError(_("Customer not found."))

        partner_invoice = self._find_partner(
            order_vals.get("partner_invoice_id/id"),
            order_vals.get("partner_invoice_id/name"),
            parent=partner,
            child_type="invoice",
        ) or self._ensure_child_partner(order_vals.get("partner_invoice_id/name"), partner, "invoice") or partner
        partner_shipping = self._find_partner(
            order_vals.get("partner_shipping_id/id"),
            order_vals.get("partner_shipping_id/name"),
            parent=partner,
            child_type="delivery",
        ) or self._ensure_child_partner(order_vals.get("partner_shipping_id/name"), partner, "delivery") or partner
        partner_invoice = partner_invoice or partner
        partner_shipping = partner_shipping or partner

        # Auto-approve partner if the approval module is installed.
        if partner and "approval_state" in partner._fields and partner.approval_state != "approved":
            partner.sudo().with_context(no_vat_validation=True).write({"approval_state": "approved"})
        # Also approve child contacts we are about to use.
        for child in {partner_invoice, partner_shipping}:
            if child and "approval_state" in child._fields and child.approval_state != "approved":
                child.sudo().with_context(no_vat_validation=True).write({"approval_state": "approved"})

        so_type = self._find_so_type(order_vals.get("so_type_id"))
        if not so_type:
            raise ValidationError(_("SO Type not found."))

        payment_term = self._find_payment_term(order_vals.get("payment_term_id"))
        warehouse = self._find_warehouse(order_vals.get("warehouse_id"))
        user = self._find_user(order_vals.get("user_id")) or self.env.user
        website = self._find_website(order_vals.get("website_id"))
        SaleOrder = self.env["sale.order"]
        has_ax_sale_order = "ax_sale_order" in SaleOrder._fields

        ax_sale_order = order_vals.get("ax_sale_order")
        document_ref = order_vals.get("document_ref") or ax_sale_order

        values = {
            "partner_id": partner.id,
            "partner_invoice_id": partner_invoice.id,
            "partner_shipping_id": partner_shipping.id,
            "so_type_id": so_type.id,
            "payment_term_id": payment_term.id if payment_term else False,
            "warehouse_id": warehouse.id if warehouse else False,
            "user_id": user.id,
            "website_id": website.id if website else False,
            "document_ref": document_ref or False,
            "sale_note": order_vals.get("sale_note") or False,
        }

        if has_ax_sale_order and ax_sale_order:
            values["ax_sale_order"] = ax_sale_order

        if order_vals.get("name"):
            values["name"] = order_vals["name"]
        if order_vals.get("date_order"):
            values["date_order"] = self._parse_datetime(order_vals.get("date_order"))
        if order_vals.get("commitment_date"):
            values["commitment_date"] = self._parse_datetime(order_vals.get("commitment_date"))

        lines_commands = self._prepare_lines(line_vals)

        # Allow bypassing discount caps that come from pricelist rules during imports.
        ctx = dict(self.env.context, ignore_pricelist_discount_limit=True)
        SaleOrderCtx = self.env["sale.order"].with_context(ctx)

        order = self._match_order(values)
        if order:
            if not self.update_existing:
                return None, _("Skipped (already exists).")
            if order.state != "draft":
                return None, _("Skipped (order not in draft).")
            order.with_context(ctx).write(values)
            order.with_context(ctx).write({"order_line": [Command.clear()] + lines_commands})
            return False, _("Updated.")

        values["order_line"] = lines_commands
        order = SaleOrderCtx.with_context(no_vat_validation=True).create(values)
        return True, _("Created (%s)") % order.name

    def _prepare_lines(self, line_vals):
        commands = []
        for line in line_vals:
            product = self._find_product(
                line.get("order_line/product_id"),
                line.get("order_line/product_id/name"),
            )
            if not product:
                raise ValidationError(_("Product not found for line %s") % (line.get("order_line/product_id") or ""))

            qty = self._to_float(line.get("order_line/product_uom_qty"), default=1.0)
            price = self._to_float(line.get("order_line/price_unit"), default=0.0)
            discount = self._to_float(line.get("order_line/discount"), default=0.0)
            if self.apply_discount_to_price and discount:
                price = price * (1 - (discount / 100.0))
                discount = 0.0
            name = (
                line.get("order_line/product_id/name")
                or line.get("order_line/product_id")
                or product.get_product_multiline_description_sale()
            )
            line_values = {
                "product_id": product.id,
                "name": name,
                "product_uom_qty": qty,
                "price_unit": price,
                "discount": discount,
                "product_uom": product.uom_id.id,
            }
            commands.append(Command.create(line_values))
        return commands

    def _create_placeholder_product(self, line):
        code = self._extract_product_code(line.get("order_line/product_id") or "")
        name = line.get("order_line/product_id/name") or line.get("order_line/product_id") or _("Placeholder Product")
        vals = {
            "name": name,
            "default_code": code or False,
            "type": "service",
            "sale_ok": True,
            "purchase_ok": False,
        }
        return self.env["product.product"].sudo().create(vals)

    def _error_message(self, exc):
        # ValidationError/UserError sometimes lack .name
        if hasattr(exc, "name") and exc.name:
            return exc.name
        if exc.args:
            return str(exc.args[0])
        return str(exc)

    def _export_failures_xlsx(self, failures):
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed Records"

        headers = ["order_key", "error"] + self.ORDER_FIELDS + ["lines"]
        ws.append(headers)

        for failure in failures:
            order_data = failure.get("order", {}) or {}
            lines_data = failure.get("lines", []) or []
            line_strings = []
            for idx, line in enumerate(lines_data, start=1):
                parts = []
                product_display = line.get("order_line/product_id/name") or line.get("order_line/product_id")
                if product_display:
                    parts.append(f"product={product_display}")
                qty = line.get("order_line/product_uom_qty")
                if qty not in (None, ""):
                    parts.append(f"qty={qty}")
                price = line.get("order_line/price_unit")
                if price not in (None, ""):
                    parts.append(f"price={price}")
                discount = line.get("order_line/discount")
                if discount not in (None, ""):
                    parts.append(f"discount={discount}")
                line_strings.append(f"#{idx} " + ", ".join(parts))

            row = [
                failure.get("order_key"),
                failure.get("error"),
            ] + [order_data.get(field) for field in self.ORDER_FIELDS] + ["\n".join(line_strings) or None]
            ws.append(row)

        stream = io.BytesIO()
        wb.save(stream)
        filename = f"sale_order_import_failed_{fields.Date.today()}.xlsx"
        return stream.getvalue(), filename

    # -- Finders ---------------------------------------------------------
    def _find_partner(self, xml_id=None, name=None, parent=None, child_type=None):
        partner = None
        if xml_id:
            # Allow both xmlids and raw integer IDs.
            if isinstance(xml_id, (int, float)) or (isinstance(xml_id, str) and xml_id.isdigit()):
                partner = self.env["res.partner"].browse(int(xml_id))
                partner = partner if partner.exists() else None
            if not partner and isinstance(xml_id, str) and "." in xml_id:
                try:
                    partner = self.env.ref(xml_id, raise_if_not_found=False)
                except ValueError:
                    partner = None

        domain = []
        if parent:
            domain.append(("parent_id", "=", parent.id))
        if child_type:
            domain.append(("type", "=", child_type))
        if name:
            name_domain = [("name", "ilike", name)]
            domain.extend(name_domain)
            partner = partner or self.env["res.partner"].search(domain, limit=1)

        if partner:
            return partner
        if name:
            return self.env["res.partner"].with_context(no_vat_validation=True).create(
                {
                    "name": name,
                    "parent_id": parent.id if parent else False,
                    "type": child_type or "contact",
                }
            )
        return partner

    def _ensure_child_partner(self, name, parent, child_type):
        if not parent:
            return None
        if not name:
            return parent
        return self.env["res.partner"].with_context(no_vat_validation=True).create(
            {
                "name": name,
                "parent_id": parent.id,
                "type": child_type or "contact",
            }
        )

    def _find_payment_term(self, name_or_xmlid):
        """Find or create a payment term using the value from the import."""
        if not name_or_xmlid:
            return None

        PaymentTerm = self.env["account.payment.term"]
        term = None

        if isinstance(name_or_xmlid, str) and "." in name_or_xmlid:
            term = self.env.ref(name_or_xmlid, raise_if_not_found=False)

        if not term:
            term = PaymentTerm.search([("name", "ilike", name_or_xmlid)], limit=1)

        if term:
            return term

        return PaymentTerm.create(
            {
                "name": name_or_xmlid,
                "company_id": self.env.company.id,
                # Simple immediate payment term when a new one is needed.
                "line_ids": [
                    (
                        0,
                        0,
                        {
                            "value": "percent",
                            "delay_type": "days_after",
                            "nb_days": 0,
                        },
                    )
                ],
            }
        )

    def _find_so_type(self, name_or_xmlid):
        if not name_or_xmlid:
            return self.env["sale.sequence.type"].search([], limit=1)
        if isinstance(name_or_xmlid, str) and "." in name_or_xmlid:
            so_type = self.env.ref(name_or_xmlid, raise_if_not_found=False)
            if so_type:
                return so_type
        so_type = self.env["sale.sequence.type"].search([("name", "ilike", name_or_xmlid)], limit=1)
        if so_type:
            return so_type
        return self.env["sale.sequence.type"].create(
            {
                "name": name_or_xmlid,
                "company_id": self.env.company.id,
            }
        )

    def _find_warehouse(self, name):
        if not name:
            return self.env["stock.warehouse"].search([], limit=1)
        Warehouse = self.env["stock.warehouse"]
        domain = ["|", ("code", "=", name), ("name", "ilike", name)]
        warehouse = Warehouse.search(domain, limit=1)
        if warehouse:
            return warehouse

        base_code = "".join(ch for ch in (name or "").upper() if ch.isalnum())[:5] or "WH"
        code = base_code
        suffix = 1
        while Warehouse.search([("code", "=", code)], limit=1):
            code = f"{base_code[:4]}{suffix}"
            suffix += 1

        return Warehouse.create(
            {
                "name": name,
                "code": code,
                "partner_id": self.env.company.partner_id.id,
                "company_id": self.env.company.id,
            }
        )

    def _find_user(self, name_or_login):
        if not name_or_login:
            return None
        User = self.env["res.users"]
        user = User.search(
            ["|", ("login", "=", name_or_login), ("name", "ilike", name_or_login)], limit=1
        )
        if user:
            return user

        login_base = "".join(ch for ch in name_or_login.lower() if ch.isalnum() or ch in {".", "_"}) or "user"
        login = login_base
        idx = 1
        while User.search([("login", "=", login)], limit=1):
            login = f"{login_base}_{idx}"
            idx += 1

        group_user = self.env.ref("base.group_user")
        return User.create(
            {
                "name": name_or_login,
                "login": login,
                "groups_id": [(6, 0, [group_user.id])] if group_user else False,
                "company_id": self.env.company.id,
            }
        )

    def _find_product(self, code_or_xmlid=None, name=None):
        Product = self.env["product.product"]
        product = None
        if isinstance(code_or_xmlid, str) and "." in code_or_xmlid and "[" not in code_or_xmlid:
            try:
                product = self.env.ref(code_or_xmlid, raise_if_not_found=False)
            except ValueError:
                product = None

        code = self._extract_product_code(code_or_xmlid or name) or code_or_xmlid
        if code:
            # Try legacy code then default_code.
            product = product or Product.search([("old_default_code", "=", code)], limit=1)
            product = product or Product.search([("default_code", "=", code)], limit=1)

        if not product and name:
            product = Product.search(["|", ("old_default_code", "=", name), ("name", "ilike", name)], limit=1)
        return product

    def _find_website(self, name_or_xmlid):
        if "website" not in self.env:
            return False
        Website = self.env["website"]
        if not name_or_xmlid:
            return Website.search([], limit=1) or Website.create({"name": _("Website")})
        if isinstance(name_or_xmlid, str) and "." in name_or_xmlid:
            website = self.env.ref(name_or_xmlid, raise_if_not_found=False)
            if website:
                return website
        website = Website.search(
            ["|", ("name", "ilike", name_or_xmlid), ("code", "=", name_or_xmlid)],
            limit=1,
        )
        if website:
            return website
        return Website.create({"name": name_or_xmlid})

    def _extract_product_code(self, value):
        if not value or not isinstance(value, str):
            return None
        match = re.search(r"\[(.*?)\]", value)
        if match:
            return match.group(1).strip()
        return None

    def _match_order(self, values):
        SaleOrder = self.env["sale.order"]
        has_ax_sale_order = "ax_sale_order" in SaleOrder._fields
        name = values.get("name")
        document_ref = values.get("document_ref")
        ax_sale_order = values.get("ax_sale_order") if has_ax_sale_order else False
        if not (name or document_ref or ax_sale_order):
            return None
        conditions = []
        if name:
            conditions.append(("name", "=", name))
        if document_ref:
            conditions.append(("document_ref", "=", document_ref))
        if ax_sale_order:
            conditions.append(("ax_sale_order", "=", ax_sale_order))

        domain = []
        for condition in conditions:
            domain = ["|"] + domain + [condition] if domain else [condition]
        return SaleOrder.search(domain, limit=1)

    # -- Parsers ---------------------------------------------------------
    def _consume_row(self, row_values, orders, current_key):
        if not row_values and not current_key:
            return current_key

        order_key = (
            row_values.get("id")
            or row_values.get("name")
            or row_values.get("ax_sale_order")
            or row_values.get("document_ref")
            or current_key
        )
        if not order_key:
            return current_key

        current_key = order_key
        data = orders.setdefault(order_key, {"order": {}, "lines": []})
        for field_name in self.ORDER_FIELDS:
            if field_name in row_values:
                data["order"].setdefault(field_name, row_values[field_name])

        if any(row_values.get(key) not in (None, "") for key in self.LINE_FIELDS):
            line_values = {key: row_values.get(key) for key in self.LINE_FIELDS if key in row_values}
            data["lines"].append(line_values)
        return current_key

    def _is_empty_cell(self, value):
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        if isinstance(value, float) and math.isnan(value):
            return True
        try:
            import pandas as pd
            if pd.isna(value):
                return True
        except Exception:
            pass
        return False

    def _ensure_polars(self):
        try:
            import polars as pl  # type: ignore
            return pl
        except Exception:
            _logger.info("polars not installed; falling back to pandas/openpyxl")
            return None

    def _parse_datetime(self, value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        if isinstance(value, str):
            cleaned = value.strip()
            for fmt in (
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y-%m-%d",
                "%d/%m/%Y %H:%M:%S",
                "%d/%m/%Y %H:%M",
                "%d/%m/%Y",
            ):
                try:
                    return datetime.strptime(cleaned, fmt)
                except ValueError:
                    continue
            try:
                return fields.Datetime.to_datetime(cleaned)
            except Exception:  # noqa: BLE001
                raise ValidationError(_("Invalid datetime format: %s") % value)
        raise ValidationError(_("Invalid datetime value: %s") % value)

    def _to_float(self, value, default=0.0):
        if value in (None, ""):
            return default
        if isinstance(value, str):
            # Remove commas and spaces commonly found in exports
            value = value.replace(",", "").strip()
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(_("Invalid number: %s") % value) from exc
