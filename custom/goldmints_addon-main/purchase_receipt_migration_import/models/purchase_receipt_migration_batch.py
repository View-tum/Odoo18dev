import base64
import io
import re
from collections import defaultdict
from datetime import date, datetime, time

from openpyxl import load_workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools.float_utils import float_compare, float_is_zero


class PurchaseReceiptMigrationBatch(models.Model):
    _name = "purchase.receipt.migration.batch"
    _description = "Purchase + Receipt Migration Batch"
    _order = "id desc"

    name = fields.Char(default="New", copy=False, readonly=True)
    company_id = fields.Many2one("res.company", required=True, default=lambda self: self.env.company)
    filename = fields.Char()
    data_file = fields.Binary(required=True)
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("parsed", "Parsed"),
            ("po_created", "PO Created"),
            ("done", "Done"),
        ],
        default="draft",
        readonly=True,
    )
    default_picking_type_id = fields.Many2one(
        "stock.picking.type",
        required=True,
        domain="[('code', '=', 'incoming'), '|', ('company_id', '=', False), ('company_id', '=', company_id)]",
        default=lambda self: self._default_picking_type(),
    )
    note = fields.Text(readonly=True)

    po_line_ids = fields.One2many("purchase.receipt.migration.po.line", "batch_id")
    receipt_line_ids = fields.One2many("purchase.receipt.migration.receipt.line", "batch_id")

    po_line_count = fields.Integer(compute="_compute_counts")
    po_mapped_count = fields.Integer(compute="_compute_counts")
    po_error_count = fields.Integer(compute="_compute_counts")
    receipt_line_count = fields.Integer(compute="_compute_counts")
    receipt_mapped_count = fields.Integer(compute="_compute_counts")
    receipt_error_count = fields.Integer(compute="_compute_counts")

    @api.model
    def _default_picking_type(self):
        return self.env["stock.picking.type"].search(
            [
                ("code", "=", "incoming"),
                "|",
                ("company_id", "=", False),
                ("company_id", "=", self.env.company.id),
            ],
            order="warehouse_id, id",
            limit=1,
        )

    @api.model_create_multi
    def create(self, vals_list):
        sequence = self.env["ir.sequence"]
        for vals in vals_list:
            if vals.get("name", "New") == "New":
                vals["name"] = sequence.next_by_code("purchase.receipt.migration.batch") or "New"
        return super().create(vals_list)

    @api.depends("po_line_ids.state", "receipt_line_ids.state")
    def _compute_counts(self):
        for rec in self:
            rec.po_line_count = len(rec.po_line_ids)
            rec.po_mapped_count = len(rec.po_line_ids.filtered(lambda line: line.state == "mapped"))
            rec.po_error_count = len(rec.po_line_ids.filtered(lambda line: line.state == "error"))
            rec.receipt_line_count = len(rec.receipt_line_ids)
            rec.receipt_mapped_count = len(rec.receipt_line_ids.filtered(lambda line: line.state == "mapped"))
            rec.receipt_error_count = len(rec.receipt_line_ids.filtered(lambda line: line.state == "error"))

    def action_reset_lines(self):
        for rec in self:
            rec.po_line_ids.unlink()
            rec.receipt_line_ids.unlink()
            rec.write({"state": "draft", "note": False})
        return True

    def action_parse_file(self):
        self.ensure_one()
        if not self.data_file:
            raise UserError(_("Please upload an XLSX file first."))

        self.action_reset_lines()
        workbook = load_workbook(io.BytesIO(base64.b64decode(self.data_file)), data_only=True)

        if "PO_LINES" not in workbook.sheetnames:
            raise UserError(_("Workbook must contain sheet 'PO_LINES'."))
        if "RECEIPT_LINES" not in workbook.sheetnames:
            raise UserError(_("Workbook must contain sheet 'RECEIPT_LINES'."))

        po_count = self._parse_po_sheet(workbook["PO_LINES"])
        receipt_count = self._parse_receipt_sheet(workbook["RECEIPT_LINES"])
        self.action_recompute_mapping()
        self.write(
            {
                "state": "parsed",
                "note": _("Parsed %s PO rows and %s receipt rows.") % (po_count, receipt_count),
            }
        )
        return True

    def action_recompute_mapping(self):
        self.ensure_one()
        for line in self.po_line_ids:
            line.write(self._map_po_line(line))
        for line in self.receipt_line_ids:
            line.write(self._map_receipt_line(line))
        return True

    def action_create_purchase_orders(self):
        self.ensure_one()
        lines = self.po_line_ids.filtered(lambda line: line.state == "mapped" and not line.purchase_order_id)
        if not lines:
            raise UserError(_("No mapped PO lines ready to create."))

        grouped_lines = defaultdict(lambda: self.env["purchase.receipt.migration.po.line"])
        for line in lines:
            grouped_lines[line.po_number] |= line

        created_orders = self.env["purchase.order"]
        for po_number, po_lines in grouped_lines.items():
            self._validate_po_group(po_number, po_lines)
            existing_order = self.env["purchase.order"].search(
                [("name", "=", po_number), ("company_id", "=", self.company_id.id)],
                limit=1,
            )
            if existing_order:
                raise UserError(_("PO %s already exists in Odoo. Import stops to avoid duplicates.") % po_number)

            head = po_lines[:1]
            order_vals = {
                "name": po_number,
                "partner_id": head.partner_id.id,
                "company_id": self.company_id.id,
                "currency_id": head.currency_id.id,
                "picking_type_id": head.picking_type_id.id,
                "date_order": self._combine_date(head.order_date),
                "origin": _("Migration Import"),
            }
            if head.payment_term_id and "payment_term_id" in self.env["purchase.order"]._fields:
                order_vals["payment_term_id"] = head.payment_term_id.id
            purchase_order = self.env["purchase.order"].create(order_vals)
            created_orders |= purchase_order

            for line in po_lines:
                purchase_line = self.env["purchase.order.line"].create(
                    {
                        "order_id": purchase_order.id,
                        "product_id": line.product_id.id,
                        "name": line.line_note or line.product_name or line.product_id.display_name,
                        "product_qty": line.order_qty,
                        "product_uom": line.uom_id.id,
                        "price_unit": line.unit_price,
                        "date_planned": self._combine_date(line.planned_date or line.order_date),
                    }
                )
                line.write(
                    {
                        "purchase_order_id": purchase_order.id,
                        "purchase_line_id": purchase_line.id,
                        "state": "created",
                        "error_message": False,
                    }
                )

        self.write(
            {
                "state": "po_created",
                "note": _("Created %s purchase orders.") % len(created_orders),
            }
        )
        return True

    def action_create_receipts(self):
        self.ensure_one()
        lines = self.receipt_line_ids.filtered(lambda line: line.state == "mapped" and not line.move_line_id)
        if not lines:
            raise UserError(_("No mapped receipt lines ready to create."))

        grouped_lines = defaultdict(lambda: self.env["purchase.receipt.migration.receipt.line"])
        for line in lines.sorted(key=lambda rec: (rec.po_number or "", rec.receipt_date or fields.Datetime.now())):
            grouped_lines[(line.po_number, line.receipt_date)] |= line

        processed_pickings = self.env["stock.picking"]
        for (po_number, receipt_date), receipt_lines in grouped_lines.items():
            purchase_order = receipt_lines[:1].purchase_order_id or self.env["purchase.order"].search(
                [("name", "=", po_number), ("company_id", "=", self.company_id.id)],
                limit=1,
            )
            if not purchase_order:
                raise UserError(_("PO %s does not exist yet. Create the PO first.") % po_number)
            self._ensure_purchase_order_confirmed(purchase_order)

            picking = self._get_open_incoming_picking(purchase_order)
            if not picking:
                raise UserError(_("PO %s has no open incoming picking left for receipt import.") % po_number)

            write_vals = {"scheduled_date": receipt_date}
            if "date_of_transfer" in picking._fields:
                write_vals["date_of_transfer"] = receipt_date
            invoice_refs = [value for value in receipt_lines.mapped("invoice_reference") if value]
            invoice_dates = [value for value in receipt_lines.mapped("invoice_date") if value]
            if "invoice_reference" in picking._fields and invoice_refs:
                write_vals["invoice_reference"] = invoice_refs[0]
            if "invoice_date" in picking._fields and invoice_dates:
                write_vals["invoice_date"] = invoice_dates[0]
            picking.with_context(skip_update_effective_date=True).write(write_vals)

            prepared_moves = self.env["stock.move"]
            for line in receipt_lines:
                move = self._find_target_move(picking, line)
                if not move:
                    raise UserError(_("Picking %s has no open move for product %s.") % (picking.name, line.product_code))

                if move not in prepared_moves:
                    self._prepare_move_for_receipt_import(move)
                    prepared_moves |= move

                remaining_qty = move.product_uom_qty - sum(move.move_line_ids.mapped("quantity"))
                if float_compare(line.qty_done, remaining_qty, precision_rounding=move.product_uom.rounding) > 0:
                    raise ValidationError(
                        _("Receipt qty %.2f is greater than remaining qty %.2f for PO %s product %s.")
                        % (line.qty_done, remaining_qty, po_number, line.product_code)
                    )

                move_line_vals = {
                    "picking_id": picking.id,
                    "move_id": move.id,
                    "company_id": self.company_id.id,
                    "product_id": line.product_id.id,
                    "product_uom_id": move.product_uom.id,
                    "quantity": line.qty_done,
                    "location_id": picking.location_id.id,
                    "location_dest_id": line.dest_location_id.id,
                }
                existing_lot = False
                if line.product_id.tracking in ("lot", "serial"):
                    if not line.lot_name:
                        raise ValidationError(_("Lot/Serial is required for tracked product %s.") % line.product_code)
                    if line.product_id.tracking == "serial" and not float_is_zero(line.qty_done - 1.0, precision_rounding=1e-6):
                        raise ValidationError(_("Serial-tracked product %s must have qty_done = 1 per row.") % line.product_code)
                    existing_lot = self.env["stock.lot"].search(
                        [
                            ("name", "=", line.lot_name),
                            ("product_id", "=", line.product_id.id),
                            ("company_id", "=", self.company_id.id),
                        ],
                        limit=1,
                    )
                    if existing_lot:
                        move_line_vals["lot_id"] = existing_lot.id
                    else:
                        move_line_vals["lot_name"] = line.lot_name
                if line.manufacturing_date and "manufacturing_date" in self.env["stock.move.line"]._fields:
                    move_line_vals["manufacturing_date"] = line.manufacturing_date

                move_line = self.env["stock.move.line"].create(move_line_vals)
                line.write(
                    {
                        "purchase_order_id": purchase_order.id,
                        "purchase_line_id": move.purchase_line_id.id if move.purchase_line_id else False,
                        "picking_id": picking.id,
                        "move_line_id": move_line.id,
                        "lot_id": existing_lot.id if existing_lot else False,
                    }
                )

            picking.with_context(skip_backorder=True, skip_update_effective_date=True).button_validate()
            processed_pickings |= picking
            for line in receipt_lines:
                if line.move_line_id and line.move_line_id.lot_id and not line.lot_id:
                    line.lot_id = line.move_line_id.lot_id.id
                line.write({"state": "done", "error_message": False})

        self.write(
            {
                "state": "done",
                "note": _("Created and validated %s receipts.") % len(processed_pickings),
            }
        )
        return True

    def _ensure_purchase_order_confirmed(self, purchase_order):
        """Support companies using double validation on purchase orders."""
        if purchase_order.state in ("draft", "sent"):
            purchase_order.button_confirm()
        if purchase_order.state == "to approve":
            purchase_order.button_approve(force=True)

    def _prepare_move_for_receipt_import(self, move):
        """Remove system-generated placeholder move lines before writing imported lots.

        In Odoo 18 incoming pickings may already contain draft move lines with the full
        planned quantity on each move. Those lines have no lot/serial and are not the
        migrated receipt lines we want to preserve. If we keep them, remaining quantity
        becomes zero and the import cannot create the real receipt rows.
        """
        placeholder_lines = move.move_line_ids.filtered(
            lambda ml: not ml.lot_id
            and not getattr(ml, "lot_name", False)
            and not ml.package_id
            and not ml.result_package_id
            and not ml.owner_id
            and ml.location_id == move.picking_id.location_id
            and float_compare(ml.quantity, 0.0, precision_rounding=move.product_uom.rounding) > 0
        )
        if placeholder_lines:
            placeholder_lines.unlink()

    def _validate_po_group(self, po_number, po_lines):
        if len(po_lines.mapped("partner_id")) != 1:
            raise ValidationError(_("PO %s mixes multiple vendors.") % po_number)
        if len(po_lines.mapped("currency_id")) != 1:
            raise ValidationError(_("PO %s mixes multiple currencies.") % po_number)
        if len(po_lines.mapped("picking_type_id")) != 1:
            raise ValidationError(_("PO %s mixes multiple receipt operation types.") % po_number)

    def _get_open_incoming_picking(self, purchase_order):
        pickings = purchase_order.picking_ids.filtered(
            lambda picking: picking.state not in ("done", "cancel") and picking.picking_type_id.code == "incoming"
        )
        return pickings.sorted(key=lambda picking: (picking.scheduled_date or fields.Datetime.now(), picking.id))[:1]

    def _find_target_move(self, picking, line):
        moves = picking.move_ids_without_package.filtered(lambda move: move.state not in ("done", "cancel"))
        if line.purchase_line_id:
            moves = moves.filtered(lambda move: move.purchase_line_id == line.purchase_line_id)
        else:
            moves = moves.filtered(lambda move: move.product_id == line.product_id)
        return moves.sorted(key=lambda move: move.id)[:1]

    def _map_po_line(self, line):
        errors = []
        partner, partner_error = self._resolve_partner(line.vendor_code, line.vendor_name)
        if partner_error:
            errors.append(partner_error)
        product, product_error = self._resolve_product(line.product_code)
        if product_error:
            errors.append(product_error)
        currency = self._match_currency(line.currency_code)
        if not currency:
            errors.append(_("Currency not found"))
        uom = self._match_uom(line.uom_name) if line.uom_name else False
        if not uom and product:
            uom = product.uom_po_id or product.uom_id
        if not uom:
            errors.append(_("UoM not found"))
        payment_term = self._match_payment_term(line.payment_term_code) if line.payment_term_code else False
        picking_type = self._match_receipt_operation(line.receipt_operation_name) or self.default_picking_type_id
        if not picking_type:
            errors.append(_("Receipt operation type not found"))
        if float_compare(line.order_qty, 0.0, precision_rounding=(uom.rounding if uom else 0.01)) <= 0:
            errors.append(_("Order quantity must be greater than zero"))

        return {
            "partner_id": partner.id if partner else False,
            "product_id": product.id if product else False,
            "uom_id": uom.id if uom else False,
            "currency_id": currency.id if currency else False,
            "payment_term_id": payment_term.id if payment_term else False,
            "picking_type_id": picking_type.id if picking_type else False,
            "state": "error" if errors else "mapped",
            "error_message": "; ".join(errors) if errors else False,
        }

    def _map_receipt_line(self, line):
        errors = []
        product, product_error = self._resolve_product(line.product_code)
        if product_error:
            errors.append(product_error)
        purchase_order = self.env["purchase.order"].search(
            [("name", "=", line.po_number), ("company_id", "=", self.company_id.id)],
            limit=1,
        )
        if not purchase_order:
            po_line = self.po_line_ids.filtered(
                lambda row: row.po_number == line.po_number and row.product_code == line.product_code
            )[:1]
            purchase_order = po_line.purchase_order_id if po_line else False
        partner = purchase_order.partner_id if purchase_order else False
        if not partner and line.vendor_code:
            related_po_lines = self.po_line_ids.filtered(
                lambda row: row.po_number == line.po_number and row.partner_id and row.vendor_code == line.vendor_code
            )
            unique_partners = related_po_lines.mapped("partner_id")
            if len(unique_partners) == 1:
                partner = unique_partners[:1]
        if not partner and line.vendor_code:
            partner, partner_error = self._resolve_partner(line.vendor_code, line.vendor_name)
            if partner_error:
                errors.append(partner_error)
        dest_location = self._match_dest_location(line.dest_location_complete_name)
        if not dest_location:
            errors.append(_("Destination location not found or not internal"))
        if float_compare(line.qty_done, 0.0, precision_rounding=(product.uom_id.rounding if product else 0.01)) <= 0:
            errors.append(_("Receipt quantity must be greater than zero"))
        if product and product.tracking in ("lot", "serial") and not line.lot_name:
            errors.append(_("Lot/Serial is required for tracked product"))
        if product and product.tracking == "serial" and not float_is_zero(line.qty_done - 1.0, precision_rounding=1e-6):
            errors.append(_("Serial-tracked receipt rows must have qty_done = 1"))

        picking_type = (purchase_order and purchase_order.picking_type_id) or self.default_picking_type_id
        if picking_type and getattr(picking_type, "require_invoice_info", False):
            if not line.invoice_reference or not line.invoice_date:
                errors.append(_("Invoice Reference and Invoice Date are required for this receipt operation type"))

        purchase_line = False
        if purchase_order and product:
            if line.po_line_ref:
                source_po_line = self.po_line_ids.filtered(
                    lambda row: row.po_number == line.po_number and row.po_line_ref == line.po_line_ref
                )[:1]
                if source_po_line and source_po_line.product_id != product:
                    errors.append(_("PO line ref does not match product"))
                purchase_line = source_po_line.purchase_line_id if source_po_line else False
                if not purchase_line:
                    errors.append(_("PO line ref not found or purchase order not created yet"))
            elif line.po_source_row_no:
                source_po_line = self.po_line_ids.filtered(
                    lambda row: row.po_number == line.po_number and row.source_row_no == line.po_source_row_no
                )[:1]
                if source_po_line and source_po_line.product_id != product:
                    errors.append(_("PO source row does not match product"))
                purchase_line = source_po_line.purchase_line_id if source_po_line else False
                if not purchase_line:
                    errors.append(_("PO source row not found or purchase order not created yet"))
            else:
                matched_lines = purchase_order.order_line.filtered(lambda row: row.product_id == product)
                if len(matched_lines) > 1:
                    errors.append(_("Multiple PO lines found for this product; fill PO Source Row No to disambiguate"))
                purchase_line = matched_lines[:1]
        return {
            "partner_id": partner.id if partner else False,
            "product_id": product.id if product else False,
            "purchase_order_id": purchase_order.id if purchase_order else False,
            "purchase_line_id": purchase_line.id if purchase_line else False,
            "dest_location_id": dest_location.id if dest_location else False,
            "state": "error" if errors else "mapped",
            "error_message": "; ".join(errors) if errors else False,
        }

    def _normalize_name(self, value):
        if not value:
            return False
        return re.sub(r"\s+", " ", str(value).strip()).lower()

    def _resolve_partner(self, vendor_code, vendor_name=False):
        if not vendor_code:
            return False, _("Vendor code is required")

        partners = self.env["res.partner"].search(
            [
                ("ref", "=", vendor_code),
                "|",
                ("company_id", "=", False),
                ("company_id", "=", self.company_id.id),
            ]
        )
        if not partners:
            return False, _("Vendor code %s not found") % vendor_code
        if len(partners) == 1:
            return partners[:1], False

        normalized_name = self._normalize_name(vendor_name)
        if normalized_name:
            exact_name_partners = partners.filtered(lambda partner: self._normalize_name(partner.name) == normalized_name)
            if len(exact_name_partners) == 1:
                return exact_name_partners[:1], False

        return False, _("Vendor code %s matches multiple vendors in Odoo") % vendor_code

    def _resolve_product(self, product_code):
        if not product_code:
            return False, _("Product code is required")

        products = self.env["product.product"].search([("default_code", "=", product_code)])
        if len(products) == 1:
            return products[:1], False
        if len(products) > 1:
            return False, _("Product code %s matches multiple products in Odoo") % product_code

        if "old_default_code" in self.env["product.template"]._fields:
            products = self.env["product.product"].search([("product_tmpl_id.old_default_code", "=", product_code)])
            if len(products) == 1:
                return products[:1], False
            if len(products) > 1:
                return False, _("Old product code %s matches multiple products in Odoo") % product_code

        return False, _("Product code %s not found") % product_code

    def _match_currency(self, currency_code):
        return self.env["res.currency"].search([("name", "=", currency_code)], limit=1)

    def _match_uom(self, uom_name):
        return self.env["uom.uom"].search([("name", "=", uom_name)], limit=1)

    def _match_payment_term(self, code):
        return self.env["account.payment.term"].search([("name", "=", code)], limit=1)

    def _match_receipt_operation(self, op_name):
        if not op_name:
            return False
        return self.env["stock.picking.type"].search(
            [
                ("name", "=", op_name),
                ("code", "=", "incoming"),
                "|",
                ("company_id", "=", False),
                ("company_id", "=", self.company_id.id),
            ],
            limit=1,
        )

    def _match_dest_location(self, complete_name):
        return self.env["stock.location"].search(
            [("complete_name", "=", complete_name), ("usage", "=", "internal")],
            limit=1,
        )

    def _parse_po_sheet(self, worksheet):
        headers = self._extract_headers(worksheet)
        required = {"po_number", "vendor_code", "product_code", "currency_code", "order_qty", "unit_price"}
        missing = required - set(headers)
        if missing:
            raise UserError(_("PO_LINES is missing required columns: %s") % ", ".join(sorted(missing)))
        count = 0
        for row_no, row in self._iter_sheet_rows(worksheet, headers, 2):
            if not any(value not in (None, "") for value in row.values()):
                continue
            self.env["purchase.receipt.migration.po.line"].create(
                {
                    "batch_id": self.id,
                    "source_row_no": row_no,
                    "po_line_ref": self._text(row.get("po_line_ref")),
                    "po_number": self._text(row.get("po_number")),
                    "vendor_code": self._text(row.get("vendor_code")),
                    "vendor_name": self._text(row.get("vendor_name") or row.get("vendor_match_name") or row.get("vendor_name_source")),
                    "order_date": self._date(row.get("order_date")),
                    "planned_date": self._date(row.get("planned_date")),
                    "currency_code": self._text(row.get("currency_code")),
                    "payment_term_code": self._text(row.get("payment_term_code")),
                    "product_code": self._text(row.get("product_code")),
                    "product_name": self._text(row.get("product_name")),
                    "uom_name": self._text(row.get("uom_name")),
                    "order_qty": self._float(row.get("order_qty")),
                    "unit_price": self._float(row.get("unit_price")),
                    "line_note": self._text(row.get("line_note")),
                    "receipt_operation_name": self._text(row.get("receipt_operation_name")),
                }
            )
            count += 1
        return count

    def _parse_receipt_sheet(self, worksheet):
        headers = self._extract_headers(worksheet)
        required = {"po_number", "product_code", "receipt_date", "qty_done", "dest_location_complete_name"}
        missing = required - set(headers)
        if missing:
            raise UserError(_("RECEIPT_LINES is missing required columns: %s") % ", ".join(sorted(missing)))
        count = 0
        for row_no, row in self._iter_sheet_rows(worksheet, headers, 2):
            if not any(value not in (None, "") for value in row.values()):
                continue
            self.env["purchase.receipt.migration.receipt.line"].create(
                {
                    "batch_id": self.id,
                    "source_row_no": row_no,
                    "po_line_ref": self._text(row.get("po_line_ref")),
                    "po_number": self._text(row.get("po_number")),
                    "po_source_row_no": int(self._float(row.get("po_source_row_no"))) if row.get("po_source_row_no") not in (None, "", False) else False,
                    "vendor_code": self._text(row.get("vendor_code")),
                    "vendor_name": self._text(row.get("vendor_name") or row.get("vendor_match_name") or row.get("vendor_name_source")),
                    "product_code": self._text(row.get("product_code")),
                    "qty_done": self._float(row.get("qty_done")),
                    "receipt_date": self._datetime(row.get("receipt_date")),
                    "lot_name": self._text(row.get("lot_name")),
                    "dest_location_complete_name": self._text(row.get("dest_location_complete_name")),
                    "invoice_reference": self._text(row.get("invoice_reference")),
                    "invoice_date": self._date(row.get("invoice_date")),
                    "manufacturing_date": self._datetime(row.get("manufacturing_date")),
                    "line_note": self._text(row.get("line_note")),
                }
            )
            count += 1
        return count

    @staticmethod
    def _extract_headers(worksheet):
        headers = {}
        for idx, cell in enumerate(worksheet[1], start=1):
            key = PurchaseReceiptMigrationBatch._normalize_header(cell.value)
            if key:
                headers[key] = idx
        return headers

    @staticmethod
    def _iter_sheet_rows(worksheet, headers, start_row):
        for row_no in range(start_row, worksheet.max_row + 1):
            row_values = {}
            for key, column_no in headers.items():
                row_values[key] = worksheet.cell(row=row_no, column=column_no).value
            yield row_no, row_values

    @staticmethod
    def _normalize_header(value):
        return str(value or "").strip().lower().replace(" ", "_")

    @staticmethod
    def _text(value):
        if value in (None, False, ""):
            return False
        text = str(value).strip()
        if text.endswith(".0"):
            try:
                float(text)
                text = text[:-2]
            except Exception:
                pass
        return text or False

    @staticmethod
    def _float(value):
        if value in (None, False, ""):
            return 0.0
        return float(value)

    @staticmethod
    def _date(value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return fields.Date.to_date(value)

    @staticmethod
    def _datetime(value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        return fields.Datetime.to_datetime(value)

    @staticmethod
    def _combine_date(value):
        if not value:
            return fields.Datetime.now()
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, time.min)
        return fields.Datetime.to_datetime(value)
