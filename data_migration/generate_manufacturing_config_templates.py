import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent


def load_json(name):
    return json.loads((BASE / name).read_text(encoding="utf-8"))


SEL = load_json("model_import_selections_view.json")
REF = load_json("pro_reference_dump.json")
RR = load_json("pro_routes_rules_dump.json")
PRD = load_json("pro_workcenter_product_mt_dump.json")
WC_FULL = load_json("pro_workcenter_dump_full.json") if (BASE / "pro_workcenter_dump_full.json").exists() else []

header_fill = PatternFill("solid", fgColor="1F4E78")
ref_fill = PatternFill("solid", fgColor="D9EAF7")
inst_fill = PatternFill("solid", fgColor="FFF2CC")
white_font = Font(color="FFFFFF", bold=True)
bold_font = Font(bold=True)
wrap = Alignment(wrap_text=True, vertical="top")


def add_sheet(wb, name, headers, rows, widths=None, header_type="main", freeze="A2"):
    ws = wb.create_sheet(name)
    ws.append(headers)
    fill = header_fill if header_type == "main" else ref_fill
    for cell in ws[1]:
        cell.fill = fill
        cell.font = white_font if header_type == "main" else bold_font
        cell.alignment = wrap
    for row in rows:
        ws.append(row)
    ws.freeze_panes = freeze
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    if widths:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for idx, header in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(16, min(42, len(str(header)) + 4))
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    readme_lines = [
        ("Purpose", "Template สำหรับ import เข้า Odoo: Product/Workcenter Manufacturing Type, Operation Type, Route, Rule"),
        ("Import Order", "1) IMP_Product_MfgType / IMP_Workcenter_MfgType  2) IMP_OperationType  3) IMP_Route  4) IMP_Rule"),
        ("Important", "Route และ Rule ควร import ใน Developer Mode / Technical Import และควรทดสอบใน VIEW ก่อน PRO"),
        ("Matching Policy", "Warehouse ใช้ชื่อ exact, Location ใช้ Complete Name exact, Operation Type/Route ใน Rule ให้ใช้ External ID จากไฟล์นี้"),
        ("Manufacturing Type Values", "plastic, pharma, packaging"),
        ("Rule Action Values", "pull, push, pull_push, manufacture, buy"),
        ("Rule Supply Method", "make_to_stock, make_to_order, mts_else_mto"),
        ("Rule Group Propagation", "none, propagate, fixed"),
        ("Rule Automatic Move", "manual, transparent"),
        ("Picking Type Code", "incoming, outgoing, internal, mrp_operation, repair_operation"),
        ("Reservation Method", "at_confirm, manual, by_date"),
        ("Create Backorder", "ask, always, never"),
        ("External ID Pattern", "ใช้ prefix เช่น mig_optype_..., mig_route_..., mig_rule_..."),
        ("Reference DB", "อ้างอิงค่าจาก database pro ณ วันที่สร้างไฟล์นี้"),
    ]
    ws.append(["Key", "Value"])
    for cell in ws[1]:
        cell.fill = inst_fill
        cell.font = bold_font
    for line in readme_lines:
        ws.append(line)
    for c in ["A", "B"]:
        ws.column_dimensions[c].width = 42 if c == "A" else 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    add_sheet(
        wb,
        "IMP_Product_MfgType",
        ["Internal Reference", "Name", "Manufacturing Type"],
        [],
        widths=[24, 46, 22],
    )
    add_sheet(
        wb,
        "IMP_Workcenter_MfgType",
        ["Work Center", "Manufacturing Type"],
        [],
        widths=[40, 22],
    )
    add_sheet(
        wb,
        "IMP_OperationType",
        [
            "External ID",
            "Operation Type",
            "Type of Operation",
            "Warehouse",
            "Sequence Prefix",
            "Source Location",
            "Destination Location",
            "Create New Lots/Serial Numbers",
            "Use Existing Lots/Serial Numbers",
            "Reservation Method",
            "Create Backorder",
            "Show Detailed Operations",
            "Require Invoice Reference/Date",
            "Propagate Invoice Reference/Date",
            "Show Delivery Address",
        ],
        [],
        widths=[28, 34, 20, 34, 20, 38, 38, 18, 18, 20, 18, 18, 18, 18, 18],
    )
    add_sheet(
        wb,
        "IMP_Route",
        [
            "External ID",
            "Route",
            "Company",
            "Sequence",
            "Applicable on Product",
            "Applicable on Product Category",
            "Applicable on Packaging",
            "Selectable on Sales Order Line",
            "Supplied Warehouse",
            "Active",
        ],
        [],
        widths=[32, 40, 34, 12, 18, 24, 22, 24, 34, 12],
    )
    add_sheet(
        wb,
        "IMP_Rule",
        [
            "External ID",
            "Name",
            "Route/External ID",
            "Company",
            "Warehouse",
            "Action",
            "Supply Method",
            "Source Location",
            "Destination Location",
            "Operation Type/External ID",
            "Propagation of Procurement Group",
            "Automatic Move",
            "Lead Time",
            "Cancel Next Move",
            "Sequence",
        ],
        [],
        widths=[32, 40, 32, 34, 34, 16, 24, 38, 38, 32, 26, 20, 12, 16, 12],
    )

    sel_rows = []
    for item in SEL:
        for key, label in item["selection"]:
            sel_rows.append([item["model"], item["field"], key, label])
    add_sheet(wb, "REF_Selections", ["Model", "Field", "Technical Value", "Label"], sel_rows, widths=[24, 28, 24, 40], header_type="ref")

    warehouse_rows = [[r["id"], r["name"], r["company_id"][1] if r.get("company_id") else ""] for r in REF["stock.warehouse"]]
    add_sheet(wb, "REF_Warehouses_Pro", ["Database ID", "Warehouse", "Company"], warehouse_rows, widths=[12, 34, 34], header_type="ref")

    loc_rows = [[r["id"], r["complete_name"], r["usage"], r["company_id"][1] if r.get("company_id") else ""] for r in REF["stock.location"]]
    add_sheet(wb, "REF_Locations_Pro", ["Database ID", "Complete Name", "Usage", "Company"], loc_rows, widths=[12, 48, 18, 34], header_type="ref")

    pt_rows = []
    for r in REF["stock.picking.type"]:
        pt_rows.append([
            r["id"],
            r["name"],
            r["code"],
            r["sequence_code"],
            r["warehouse_id"][1] if r.get("warehouse_id") else "",
            r["default_location_src_id"][1] if r.get("default_location_src_id") else "",
            r["default_location_dest_id"][1] if r.get("default_location_dest_id") else "",
            r.get("require_invoice_info", False),
            r.get("propagate_invoice_info", False),
        ])
    add_sheet(
        wb,
        "REF_OperationTypes_Pro",
        ["Database ID", "Operation Type", "Type of Operation", "Sequence Prefix", "Warehouse", "Source Location", "Destination Location", "Require Invoice Info", "Propagate Invoice Info"],
        pt_rows,
        widths=[12, 34, 20, 20, 34, 40, 40, 18, 18],
        header_type="ref",
    )

    route_rows = []
    for idx, r in enumerate(RR["routes"], start=1):
        ext = f"mig_route_{idx:03d}"
        route_rows.append([
            ext,
            r["id"],
            r["name"],
            r["company_id"][1] if r.get("company_id") else "",
            r["sequence"],
            r.get("product_selectable", False),
            r.get("product_categ_selectable", False),
            r.get("packaging_selectable", False),
            r.get("sale_selectable", False),
            r["supplied_wh_id"][1] if r.get("supplied_wh_id") else "",
        ])
    add_sheet(
        wb,
        "REF_Routes_Pro",
        ["Suggested External ID", "Database ID", "Route", "Company", "Sequence", "Applicable on Product", "Applicable on Product Category", "Applicable on Packaging", "Selectable on Sales Order Line", "Supplied Warehouse"],
        route_rows,
        widths=[24, 12, 44, 34, 12, 18, 24, 22, 24, 34],
        header_type="ref",
    )

    route_ext_by_id = {r["id"]: f"mig_route_{i:03d}" for i, r in enumerate(RR["routes"], start=1)}
    op_ext_by_id = {}
    for r in REF["stock.picking.type"]:
        slug = "".join(ch.lower() if ch.isalnum() else "_" for ch in r["name"]).strip("_")
        while "__" in slug:
            slug = slug.replace("__", "_")
        op_ext_by_id[r["id"]] = f"mig_optype_{slug[:40]}"

    rule_rows = []
    for idx, r in enumerate(RR["rules"], start=1):
        rule_rows.append([
            f"mig_rule_{idx:03d}",
            r["id"],
            r["name"],
            route_ext_by_id.get(r["route_id"][0], ""),
            r["route_id"][1] if r.get("route_id") else "",
            r["warehouse_id"][1] if r.get("warehouse_id") else "",
            r["action"],
            r["procure_method"],
            r["location_src_id"][1] if r.get("location_src_id") else "",
            r["location_dest_id"][1] if r.get("location_dest_id") else "",
            op_ext_by_id.get(r["picking_type_id"][0], "") if r.get("picking_type_id") else "",
            r["picking_type_id"][1] if r.get("picking_type_id") else "",
            r["group_propagation_option"],
            r["auto"],
            r["delay"],
            r["propagate_cancel"],
            r["sequence"],
        ])
    add_sheet(
        wb,
        "REF_Rules_Pro",
        ["Suggested External ID", "Database ID", "Name", "Route/External ID", "Route", "Warehouse", "Action", "Supply Method", "Source Location", "Destination Location", "Operation Type/External ID", "Operation Type", "Propagation of Procurement Group", "Automatic Move", "Lead Time", "Cancel Next Move", "Sequence"],
        rule_rows,
        widths=[24, 12, 38, 24, 38, 28, 16, 24, 40, 40, 24, 34, 26, 20, 12, 16, 12],
        header_type="ref",
    )

    prod_rows = [[r["id"], r["default_code"], r["name"], r.get("manufacturing_type") or ""] for r in PRD["products"]]
    add_sheet(wb, "REF_ProductMT_Pro", ["Database ID", "Internal Reference", "Name", "Manufacturing Type"], prod_rows, widths=[12, 24, 46, 22], header_type="ref")

    wc_rows = [[r["id"], r["name"], r.get("manufacturing_type") or ""] for r in WC_FULL]
    add_sheet(wb, "REF_Workcenter_Pro", ["Database ID", "Work Center", "Manufacturing Type"], wc_rows, widths=[12, 42, 22], header_type="ref")

    example_ws = wb.create_sheet("EXAMPLE_Copy_From_Pro")
    example_ws.append(["Template Sheet", "How to use"])
    for cell in example_ws[1]:
        cell.fill = inst_fill
        cell.font = bold_font
    example_rows = [
        ("IMP_Product_MfgType", "คัดลอกเฉพาะคอลัมน์ Internal Reference, Name, Manufacturing Type จาก REF_ProductMT_Pro แล้วแก้ค่าที่ต้อง migrate"),
        ("IMP_Workcenter_MfgType", "คัดลอก Work Center กับ Manufacturing Type จาก REF_Workcenter_Pro ถ้าต้องอัปเดต workcenter"),
        ("IMP_OperationType", "ดูรูปแบบค่าที่ถูกต้องจาก REF_OperationTypes_Pro แล้วใส่ External ID ของรายการใหม่"),
        ("IMP_Route", "ดู route เดิมจาก REF_Routes_Pro แล้วสร้าง External ID ใหม่สำหรับรายการที่ import"),
        ("IMP_Rule", "ใช้ Route/External ID และ Operation Type/External ID เพื่อเชื่อม relation ให้แน่นอน"),
    ]
    for row in example_rows:
        example_ws.append(row)
    for c in ["A", "B"]:
        example_ws.column_dimensions[c].width = 28 if c == "A" else 110
    for row in example_ws.iter_rows():
        for cell in row:
            cell.alignment = wrap

    out = BASE / "Odoo_Manufacturing_Config_Import_Templates_PRO.xlsx"
    wb.save(out)

    csv_specs = {
        "IMP_Product_MfgType.csv": ["Internal Reference", "Name", "Manufacturing Type"],
        "IMP_Workcenter_MfgType.csv": ["Work Center", "Manufacturing Type"],
        "IMP_OperationType.csv": ["External ID", "Operation Type", "Type of Operation", "Warehouse", "Sequence Prefix", "Source Location", "Destination Location", "Create New Lots/Serial Numbers", "Use Existing Lots/Serial Numbers", "Reservation Method", "Create Backorder", "Show Detailed Operations", "Require Invoice Reference/Date", "Propagate Invoice Reference/Date", "Show Delivery Address"],
        "IMP_Route.csv": ["External ID", "Route", "Company", "Sequence", "Applicable on Product", "Applicable on Product Category", "Applicable on Packaging", "Selectable on Sales Order Line", "Supplied Warehouse", "Active"],
        "IMP_Rule.csv": ["External ID", "Name", "Route/External ID", "Company", "Warehouse", "Action", "Supply Method", "Source Location", "Destination Location", "Operation Type/External ID", "Propagation of Procurement Group", "Automatic Move", "Lead Time", "Cancel Next Move", "Sequence"],
    }
    for fn, headers in csv_specs.items():
        with (BASE / fn).open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

    (BASE / "Odoo_Manufacturing_Config_Import_Readme.md").write_text(
        "# Odoo Manufacturing Config Import Templates\n\n"
        "ไฟล์หลัก: `data_migration/Odoo_Manufacturing_Config_Import_Templates_PRO.xlsx`\n\n"
        "## ใช้สำหรับ\n"
        "1. Product Manufacturing Type\n"
        "2. Workcenter Manufacturing Type\n"
        "3. Operation Type\n"
        "4. Route\n"
        "5. Rule\n\n"
        "## ลำดับ import ที่แนะนำ\n"
        "1. `IMP_Product_MfgType`\n"
        "2. `IMP_Workcenter_MfgType`\n"
        "3. `IMP_OperationType`\n"
        "4. `IMP_Route`\n"
        "5. `IMP_Rule`\n\n"
        "## หมายเหตุสำคัญ\n"
        "- `Manufacturing Type` ใช้ค่า technical เท่านั้น: `plastic`, `pharma`, `packaging`\n"
        "- `IMP_Rule` ควรอ้าง `Route/External ID` และ `Operation Type/External ID` เพื่อกัน matching ผิด\n"
        "- Location ให้ใช้ค่า exact จาก `REF_Locations_Pro` โดยยึด `Complete Name`\n"
        "- Warehouse ให้ใช้ชื่อ exact จาก `REF_Warehouses_Pro`\n"
        "- Route / Rule เป็น config เชิงเทคนิค ควรทดสอบ import ใน `view` ก่อน `pro`\n",
        encoding="utf-8",
    )

    print(out)


if __name__ == "__main__":
    main()
