from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE = Path(r"C:\Users\tumsu\Downloads\ATHENS Price HIKVISION 2026 .xlsx")
OUT_DIR = ROOT / "output" / "spreadsheet"
IMAGE_DIR = OUT_DIR / "hikvision_product_images_inspect"
REPORT_PATH = OUT_DIR / "HIKVISION_product_inspection_report.xlsx"
CONTACT_PREFIX = OUT_DIR / "HIKVISION_product_image_contact_sheet"

SUMMARY_HEADERS = [
    "Sheet",
    "Product rows",
    "Images in sheet",
    "Products with exact-row image",
    "Products with mapped image",
    "Missing mapped image",
    "Exact duplicate code groups",
    "Section count",
]

PRODUCT_HEADERS = [
    "ลำดับ",
    "Source Sheet",
    "Source Row",
    "Brand",
    "Raw Section",
    "Proposed Category",
    "Item No.",
    "Product Code / Name",
    "Normalized SKU Key",
    "Line",
    "Details",
    "MSRP",
    "Dealer",
    "VIP/User",
    "Data Sheet",
    "Warranty",
    "Exact Row Image Count",
    "Mapped Image Count",
    "Primary Image",
    "Issue Status",
    "Issues",
]

DUPLICATE_HEADERS = [
    "Duplicate Type",
    "Normalized SKU Key",
    "Count",
    "Source Sheet",
    "Source Row",
    "Raw Section",
    "Product Code / Name",
    "Dealer",
    "VIP/User",
    "Issues",
]

CATEGORY_HEADERS = [
    "ลำดับ",
    "Brand",
    "Proposed Category",
    "Parent Category",
    "Product Count",
    "Raw Sections Included",
]

IMAGE_HEADERS = [
    "Source Sheet",
    "Source Row",
    "Image Index",
    "Anchor Row",
    "Anchor Col",
    "Width",
    "Height",
    "Area",
    "Mapped To Product Row",
    "Mapped Product Code",
    "Image File",
]


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if len(text) >= 2 and text[0] == text[-1] == '"':
        text = text[1:-1].strip()
    return text


def one_line(value: Any, limit: int | None = None) -> str:
    text = clean(value).replace("\n", " / ")
    text = re.sub(r"\s+", " ", text).strip()
    if limit and len(text) > limit:
        return text[: limit - 3].rstrip() + "..."
    return text


def is_item_number(value: Any) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.0)?", clean(value)))


def is_number_like(value: Any) -> bool:
    if value is None or clean(value) == "":
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    return bool(re.fullmatch(r"-?\d+(?:,\d{3})*(?:\.\d+)?", clean(value)))


def exact_key(value: str) -> str:
    text = value.upper().replace("รุ่น", " ")
    text = re.sub(r'"', "", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def proposed_category(section: str, line: str, code: str) -> str:
    text = f"{section} {line} {code}".upper()
    if "DVR" in text or "NVR" in text:
        return "เครื่องบันทึก DVR/NVR"
    if "ANPR" in text or "ENTRANCE" in text or "BARRIER" in text or "TMG" in text or "TMC" in text:
        return "ระบบอ่านป้ายทะเบียน ANPR และไม้กั้น"
    if "ACCESS" in text or "K1T" in text or "KAS" in text or "K4H" in text or "K7" in text:
        return "ระบบ Access Control"
    if "CABLE" in text or "POE" in text or "SWITCH" in text or "ROUTER" in text:
        return "อุปกรณ์เครือข่ายและสายสัญญาณ"
    if "CAMERA" in text or "2CD" in text or "2CE" in text:
        return "กล้องวงจรปิด HIKVISION"
    return "สินค้า HIKVISION อื่นๆ"


def sheet_rule(sheet_name: str) -> tuple[int, int, int]:
    if sheet_name == "UNIVIEW":
        return 5, 1, 1
    if sheet_name == "Tiandy":
        return 3, 1, 3
    if sheet_name == "EZVIZ":
        return 5, 2, 4
    if sheet_name == "HIKFIRE":
        return 3, 1, 2
    if sheet_name == "HIKVISION Access & Entrance":
        return 4, 1, 3
    if sheet_name == "HIKVISION CCTV":
        return 6, 1, 3
    return 1, 1, 1


def scan_rows(ws, sheet_name: str) -> tuple[list[dict[str, Any]], list[tuple[int, str]]]:
    rows: list[dict[str, Any]] = []
    sections: list[tuple[int, str]] = []
    current_section = ""
    start_row, item_col, code_col = sheet_rule(sheet_name)

    for row in range(1, ws.max_row + 1):
        if sheet_name == "HIKVISION CCTV":
            if clean(ws.cell(row, 2).value) and not clean(ws.cell(row, 1).value) and not clean(ws.cell(row, 3).value):
                current_section = one_line(ws.cell(row, 2).value)
                sections.append((row, current_section))
        elif sheet_name == "HIKVISION Access & Entrance":
            if row == 2 and clean(ws.cell(row, 1).value):
                current_section = one_line(ws.cell(row, 1).value)
                sections.append((row, current_section))
            elif clean(ws.cell(row, 1).value) and not is_item_number(ws.cell(row, 1).value) and not clean(ws.cell(row, 3).value):
                current_section = one_line(ws.cell(row, 1).value)
                sections.append((row, current_section))
        elif sheet_name == "UNIVIEW":
            if clean(ws.cell(row, 1).value) and not clean(ws.cell(row, 3).value) and row != 1:
                current_section = one_line(ws.cell(row, 1).value)
                sections.append((row, current_section))

        item_value = ws.cell(row, item_col).value
        code = clean(ws.cell(row, code_col).value)
        product = False
        if row >= start_row and code:
            if sheet_name in {"Tiandy", "HIKFIRE", "HIKVISION Access & Entrance", "HIKVISION CCTV"}:
                product = is_item_number(item_value)
            elif sheet_name == "EZVIZ":
                product = bool(clean(ws.cell(row, 2).value) and clean(ws.cell(row, 4).value))
            elif sheet_name == "UNIVIEW":
                product = bool(clean(ws.cell(row, 1).value) and (clean(ws.cell(row, 3).value) or clean(ws.cell(row, 14).value)))

        if not product:
            continue

        rows.append(
            {
                "source_sheet": sheet_name,
                "source_row": row,
                "item_no": one_line(item_value),
                "raw_section": current_section,
                "code": code,
            }
        )

    return rows, sections


def extract_images(ws, sheet_name: str) -> tuple[dict[int, list[dict[str, Any]]], list[dict[str, Any]]]:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    images_by_row: dict[int, list[dict[str, Any]]] = defaultdict(list)
    all_images: list[dict[str, Any]] = []
    prefix = "access" if sheet_name == "HIKVISION Access & Entrance" else "cctv"
    for index, image in enumerate(ws._images, start=1):
        anchor = image.anchor._from
        row = anchor.row + 1
        col = anchor.col + 1
        width = int(getattr(image, "width", 0) or 0)
        height = int(getattr(image, "height", 0) or 0)
        fmt = (getattr(image, "format", None) or "png").lower()
        if fmt == "jpeg":
            ext = "jpg"
        else:
            ext = fmt if re.fullmatch(r"[a-z0-9]+", fmt) else "png"
        out_path = IMAGE_DIR / f"hikvision_{prefix}_row_{row:04d}_image_{index:03d}.{ext}"
        out_path.write_bytes(image._data())
        record = {
            "sheet": sheet_name,
            "index": index,
            "anchor_row": row,
            "anchor_col": col,
            "width": width,
            "height": height,
            "area": width * height,
            "path": out_path,
        }
        images_by_row[row].append(record)
        all_images.append(record)
    return images_by_row, all_images


def largest_image(records: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not records:
        return None
    return max(records, key=lambda item: int(item["area"]))


def build_hikvision_products(wb) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    products: list[dict[str, Any]] = []
    image_records: list[dict[str, Any]] = []
    sheet_meta: dict[str, Any] = {}

    for sheet_name in ["HIKVISION Access & Entrance", "HIKVISION CCTV"]:
        ws = wb[sheet_name]
        source_rows, sections = scan_rows(ws, sheet_name)
        images_by_row, all_images = extract_images(ws, sheet_name)
        product_by_row = {row["source_row"]: row for row in source_rows}
        section_rows = {row for row, _ in sections}

        for row in source_rows:
            source_row = row["source_row"]
            exact_images = images_by_row.get(source_row, [])
            mapped_images = list(exact_images)
            mapped_from = ""
            if not mapped_images and source_row - 1 in section_rows and images_by_row.get(source_row - 1):
                mapped_images = list(images_by_row[source_row - 1])
                mapped_from = f"mapped from section row {source_row - 1}"
            primary = largest_image(mapped_images)

            if sheet_name == "HIKVISION Access & Entrance":
                line = row["raw_section"]
                details = clean(ws.cell(source_row, 4).value)
                msrp = ""
                dealer = ws.cell(source_row, 6).value
                vip = ws.cell(source_row, 7).value
                datasheet = ""
                warranty_raw = clean(ws.cell(source_row, 8).value)
                warranty = f"{warranty_raw} years" if warranty_raw else ""
            else:
                line = clean(ws.cell(source_row, 4).value)
                details = clean(ws.cell(source_row, 5).value)
                msrp = ws.cell(source_row, 6).value
                dealer = ws.cell(source_row, 7).value
                vip = ws.cell(source_row, 8).value
                datasheet = clean(ws.cell(source_row, 9).value)
                warranty = clean(ws.cell(source_row, 10).value)

            category = proposed_category(row["raw_section"], line, row["code"])
            issues: list[str] = []
            price_values = [msrp, dealer, vip]
            present_prices = [value for value in price_values if clean(value) != ""]
            invalid_prices = [clean(value) for value in present_prices if not is_number_like(value)]
            if not details:
                issues.append("missing details")
            if sheet_name == "HIKVISION CCTV" and not line:
                issues.append("missing line")
            if not present_prices:
                issues.append("missing prices")
            if invalid_prices:
                issues.append("non-numeric price: " + ", ".join(invalid_prices))
            if not warranty:
                issues.append("missing warranty")
            if not mapped_images:
                issues.append("missing image")
            elif mapped_from:
                issues.append(mapped_from)
            if len(mapped_images) > 1:
                issues.append(f"multiple images on row: {len(mapped_images)}")
            if "\n" in row["code"] or '"""' in row["code"] or row["code"].startswith('"'):
                issues.append("product code/name needs cleanup")

            product = {
                **row,
                "brand": "HIKVISION",
                "normalized_key": exact_key(row["code"]),
                "line": one_line(line),
                "details": details,
                "msrp": msrp,
                "dealer": dealer,
                "vip": vip,
                "datasheet": datasheet,
                "warranty": warranty,
                "proposed_category": category,
                "exact_image_count": len(exact_images),
                "mapped_image_count": len(mapped_images),
                "primary_image": str(primary["path"]) if primary else "",
                "issues": issues,
                "status": "OK" if not issues else "CHECK",
            }
            products.append(product)

        for record in all_images:
            mapped_row = ""
            mapped_code = ""
            if record["anchor_row"] in product_by_row:
                mapped_row = record["anchor_row"]
                mapped_code = product_by_row[record["anchor_row"]]["code"]
            elif record["anchor_row"] in section_rows and record["anchor_row"] + 1 in product_by_row:
                mapped_row = record["anchor_row"] + 1
                mapped_code = product_by_row[record["anchor_row"] + 1]["code"]
            image_records.append({**record, "mapped_row": mapped_row, "mapped_code": mapped_code})

        sheet_meta[sheet_name] = {
            "rows": source_rows,
            "sections": sections,
            "image_count": len(all_images),
            "exact_products_with_images": sum(1 for row in source_rows if images_by_row.get(row["source_row"])),
            "mapped_products_with_images": sum(
                1
                for row in source_rows
                if images_by_row.get(row["source_row"])
                or (row["source_row"] - 1 in section_rows and images_by_row.get(row["source_row"] - 1))
            ),
        }

    counts = Counter(product["normalized_key"] for product in products)
    sheets_by_key: dict[str, set[str]] = defaultdict(set)
    for product in products:
        sheets_by_key[product["normalized_key"]].add(product["source_sheet"])

    for product in products:
        if counts[product["normalized_key"]] > 1:
            product["issues"].append("exact duplicate product code")
            product["status"] = "CHECK"
        if len(sheets_by_key[product["normalized_key"]]) > 1:
            product["issues"].append("duplicate across HIKVISION sheets")
            product["status"] = "CHECK"

    duplicate_rows: list[dict[str, Any]] = []
    for key, count in sorted(counts.items()):
        if count <= 1:
            continue
        duplicate_type = "Cross-sheet" if len(sheets_by_key[key]) > 1 else "Same-sheet"
        for product in products:
            if product["normalized_key"] == key:
                duplicate_rows.append({**product, "duplicate_type": duplicate_type, "duplicate_count": count})

    return products, duplicate_rows, image_records, sheet_meta


def build_summary(wb, hikvision_products: list[dict[str, Any]], sheet_meta: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        product_rows, sections = scan_rows(ws, sheet_name)
        images_by_row = Counter()
        for image in ws._images:
            images_by_row[image.anchor._from.row + 1] += 1
        exact_with_images = sum(1 for row in product_rows if images_by_row[row["source_row"]])
        mapped_with_images = exact_with_images
        if sheet_name in sheet_meta:
            mapped_with_images = sheet_meta[sheet_name]["mapped_products_with_images"]
        duplicate_groups = sum(1 for _, count in Counter(exact_key(row["code"]) for row in product_rows).items() if count > 1)
        rows.append(
            [
                sheet_name,
                len(product_rows),
                len(ws._images),
                exact_with_images,
                mapped_with_images,
                len(product_rows) - mapped_with_images,
                duplicate_groups,
                len(sections),
            ]
        )

    total_hikvision = len(hikvision_products)
    unique_hikvision = len({product["normalized_key"] for product in hikvision_products})
    rows.append(["HIKVISION TOTAL", total_hikvision, "", "", "", total_hikvision - unique_hikvision, "", ""])
    rows.append(["HIKVISION UNIQUE SKU KEY", unique_hikvision, "", "", "", "", "", ""])
    return rows


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(one_line(cell.value, 80)) for cell in column if cell.value is not None) if column else 12
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 55)
    ws.auto_filter.ref = ws.dimensions


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)


def write_report(wb_source, products: list[dict[str, Any]], duplicates: list[dict[str, Any]], image_records: list[dict[str, Any]], sheet_meta: dict[str, Any]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_rows = build_summary(wb_source, products, sheet_meta)
    append_rows(ws_summary, SUMMARY_HEADERS, summary_rows)
    style_sheet(ws_summary)
    add_table(ws_summary, "HIKVISIONSummary")

    ws_products = wb.create_sheet("HIKVISION_Product_Audit")
    product_rows = []
    for index, product in enumerate(products, start=1):
        product_rows.append(
            [
                index,
                product["source_sheet"],
                product["source_row"],
                product["brand"],
                product["raw_section"],
                product["proposed_category"],
                product["item_no"],
                one_line(product["code"]),
                product["normalized_key"],
                product["line"],
                one_line(product["details"], 500),
                product["msrp"],
                product["dealer"],
                product["vip"],
                product["datasheet"],
                product["warranty"],
                product["exact_image_count"],
                product["mapped_image_count"],
                product["primary_image"],
                product["status"],
                "; ".join(product["issues"]),
            ]
        )
    append_rows(ws_products, PRODUCT_HEADERS, product_rows)
    style_sheet(ws_products)
    add_table(ws_products, "HIKVISIONProducts")

    ws_dup = wb.create_sheet("Duplicates")
    duplicate_rows = []
    for item in duplicates:
        duplicate_rows.append(
            [
                item["duplicate_type"],
                item["normalized_key"],
                item["duplicate_count"],
                item["source_sheet"],
                item["source_row"],
                item["raw_section"],
                one_line(item["code"]),
                item["dealer"],
                item["vip"],
                "; ".join(item["issues"]),
            ]
        )
    append_rows(ws_dup, DUPLICATE_HEADERS, duplicate_rows)
    style_sheet(ws_dup)
    add_table(ws_dup, "HIKVISIONDuplicates")

    ws_categories = wb.create_sheet("Proposed_Categories")
    category_rows = []
    category_counts = Counter(product["proposed_category"] for product in products)
    raw_by_category: dict[str, set[str]] = defaultdict(set)
    for product in products:
        raw_by_category[product["proposed_category"]].add(product["raw_section"])
    category_rows.append([1, "HIKVISION", "HIKVISION", "", len(products), "สร้างเป็นหมวดแบรนด์หลัก"])
    for index, (category, count) in enumerate(sorted(category_counts.items()), start=2):
        category_rows.append([index, "HIKVISION", category, "HIKVISION", count, ", ".join(sorted(raw_by_category[category]))])
    append_rows(ws_categories, CATEGORY_HEADERS, category_rows)
    style_sheet(ws_categories)
    add_table(ws_categories, "HIKVISIONCategories")

    ws_images = wb.create_sheet("Image_Audit")
    image_rows = []
    for record in image_records:
        image_rows.append(
            [
                record["sheet"],
                record.get("mapped_row", ""),
                record["index"],
                record["anchor_row"],
                record["anchor_col"],
                record["width"],
                record["height"],
                record["area"],
                record.get("mapped_row", ""),
                one_line(record.get("mapped_code", "")),
                str(record["path"]),
            ]
        )
    append_rows(ws_images, IMAGE_HEADERS, image_rows)
    style_sheet(ws_images)
    add_table(ws_images, "HIKVISIONImages")

    for worksheet in wb.worksheets:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

    wb.save(REPORT_PATH)


def load_font(size: int) -> ImageFont.ImageFont:
    font_candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\tahoma.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
    ]
    for font_path in font_candidates:
        if Path(font_path).exists():
            return ImageFont.truetype(font_path, size)
    return ImageFont.load_default()


def draw_wrapped(draw: ImageDraw.ImageDraw, xy: tuple[int, int], text: str, font: ImageFont.ImageFont, fill: str, width: int, line_height: int, max_lines: int) -> None:
    words = re.split(r"(\s+)", text)
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = current + word
        if draw.textbbox((0, 0), candidate, font=font)[2] <= width:
            current = candidate
        else:
            if current.strip():
                lines.append(current.strip())
            current = word.strip()
        if len(lines) >= max_lines:
            break
    if current.strip() and len(lines) < max_lines:
        lines.append(current.strip())
    for index, line in enumerate(lines[:max_lines]):
        draw.text((xy[0], xy[1] + index * line_height), line, font=font, fill=fill)


def make_contact_sheets(products: list[dict[str, Any]]) -> list[Path]:
    page_paths: list[Path] = []
    with_images = [product for product in products if product["primary_image"]]
    per_page = 30
    cols = 5
    card_w = 260
    card_h = 260
    margin = 28
    title_h = 58
    thumb_h = 150
    font_title = load_font(20)
    font = load_font(14)
    small = load_font(12)
    pages = [with_images[index : index + per_page] for index in range(0, len(with_images), per_page)]
    for page_no, page_products in enumerate(pages, start=1):
        rows = (len(page_products) + cols - 1) // cols
        canvas = Image.new("RGB", (margin * 2 + cols * card_w, margin * 2 + title_h + rows * card_h), "white")
        draw = ImageDraw.Draw(canvas)
        draw.text((margin, margin), f"HIKVISION Product Image Contact Sheet {page_no}/{len(pages)}", font=font_title, fill="#111827")
        for index, product in enumerate(page_products):
            col = index % cols
            row = index // cols
            x = margin + col * card_w
            y = margin + title_h + row * card_h
            draw.rectangle((x, y, x + card_w - 12, y + card_h - 12), outline="#D1D5DB", width=1)
            try:
                img = Image.open(product["primary_image"]).convert("RGB")
                img.thumbnail((card_w - 32, thumb_h), Image.Resampling.LANCZOS)
                ix = x + (card_w - 12 - img.width) // 2
                iy = y + 10 + (thumb_h - img.height) // 2
                canvas.paste(img, (ix, iy))
            except Exception:
                draw.text((x + 12, y + 56), "IMAGE ERROR", font=font, fill="#B91C1C")
            text_y = y + thumb_h + 18
            draw.text((x + 12, text_y), f"Row {product['source_row']} | {product['source_sheet'].replace('HIKVISION ', '')}", font=small, fill="#4B5563")
            draw_wrapped(draw, (x + 12, text_y + 20), one_line(product["code"]), font, "#111827", card_w - 36, 17, 3)
            if product["status"] != "OK":
                draw.text((x + 12, y + card_h - 34), "CHECK", font=small, fill="#B91C1C")
        out_path = Path(f"{CONTACT_PREFIX}_{page_no}.jpg")
        canvas.save(out_path, quality=92)
        page_paths.append(out_path)
    return page_paths


def main() -> int:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE
    if not source.exists():
        print(f"Source file not found: {source}")
        return 1
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(source, data_only=True)
    products, duplicates, image_records, sheet_meta = build_hikvision_products(wb)
    write_report(wb, products, duplicates, image_records, sheet_meta)
    contact_sheets = make_contact_sheets(products)
    total = len(products)
    unique = len({product["normalized_key"] for product in products})
    check_count = sum(1 for product in products if product["status"] != "OK")
    print(f"source: {source}")
    print(f"hikvision_product_rows: {total}")
    print(f"hikvision_unique_sku_keys: {unique}")
    print(f"rows_needing_check: {check_count}")
    print(f"duplicate_rows: {len(duplicates)}")
    print(f"report: {REPORT_PATH}")
    for path in contact_sheets:
        print(f"contact_sheet: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
