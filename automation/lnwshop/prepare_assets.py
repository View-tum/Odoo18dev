from __future__ import annotations

from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops, ImageEnhance


ROOT = Path(__file__).resolve().parents[2]
SOURCE_EXCEL = ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_AI_ready.xlsx"
WATERMARK = Path(__file__).resolve().parent / "watermark_current_transparent_hq.png"
WATERMARKED_DIR = ROOT / "output" / "spreadsheet" / "imou_product_images_watermarked"
WATERMARKED_EXCEL = ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_AI_ready_watermarked.xlsx"
CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_categories.xlsx"
DEFAULT_PRODUCT_TYPE_1 = "สินค้าอุตสาหกรรม"
DEFAULT_PRODUCT_TYPE_2 = "อุปกรณ์ใช้ในโรงงานอุตสาหกรรม"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def crop_watermark(mark: Image.Image) -> Image.Image:
    rgba = mark.convert("RGBA")
    white = Image.new("RGBA", rgba.size, (255, 255, 255, 0))
    pixels = rgba.load()
    data = []
    for y in range(rgba.height):
        for x in range(rgba.width):
            r, g, b, a = pixels[x, y]
            if a < 8 or (r >= 253 and g >= 253 and b >= 253):
                data.append((255, 255, 255, 0))
            else:
                data.append((r, g, b, a))
    candidate = Image.new("RGBA", rgba.size)
    candidate.putdata(data)
    bbox = ImageChops.difference(candidate, white).getbbox()
    if not bbox:
        return strengthen_watermark(rgba)
    return strengthen_watermark(candidate.crop(bbox))


def strengthen_watermark(mark: Image.Image) -> Image.Image:
    rgba = mark.convert("RGBA")
    pixels = rgba.load()
    data = []
    for y in range(rgba.height):
        for x in range(rgba.width):
            r, g, b, a = pixels[x, y]
            if a < 8:
                data.append((255, 255, 255, 0))
                continue
            boosted_alpha = 255 if a >= 120 else min(255, int(a * 1.65))
            data.append((r, g, b, boosted_alpha))
    boosted = Image.new("RGBA", rgba.size)
    boosted.putdata(data)
    return boosted


def apply_watermark(image_path: Path, out_path: Path, mark: Image.Image) -> None:
    with Image.open(image_path) as source:
        base = source.convert("RGBA")
        base_w, base_h = base.size
        if base_w <= 0 or base_h <= 0:
            return

        target_w = max(78, int(base_w * 0.65))
        ratio = target_w / mark.width
        target_h = max(1, int(mark.height * ratio))
        if target_h > base_h * 0.50:
            target_h = max(42, int(base_h * 0.50))
            ratio = target_h / mark.height
            target_w = max(1, int(mark.width * ratio))

        resized = mark.resize((target_w, target_h), Image.Resampling.LANCZOS)
        alpha = resized.getchannel("A")
        alpha = ImageEnhance.Brightness(alpha).enhance(1.35)
        resized.putalpha(alpha)

        pad = max(4, int(min(base_w, base_h) * 0.04))
        x = max(pad, base_w - target_w - pad)
        y = max(pad, base_h - target_h - pad)
        base.alpha_composite(resized, (x, y))

        out_path.parent.mkdir(parents=True, exist_ok=True)
        base.save(out_path, "PNG")


def header_index(ws) -> dict[str, int]:
    return {cell.value: cell.column for cell in ws[1]}


def set_cell(ws, row: int, header: str, value: Any, indexes: dict[str, int]) -> None:
    ws.cell(row, indexes[header]).value = value


def build_watermarked_workbook() -> tuple[int, dict[str, str]]:
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)
    if not WATERMARK.exists():
        raise FileNotFoundError(WATERMARK)

    WATERMARKED_DIR.mkdir(parents=True, exist_ok=True)
    mark = crop_watermark(Image.open(WATERMARK))

    wb = load_workbook(SOURCE_EXCEL)
    ws = wb["LnwShop_AI_Ready"]
    indexes = header_index(ws)
    image_col = indexes["รูปภาพหลัก/ไฟล์ภาพ"]
    brand_col = indexes["แบรนด์"]
    category_col = indexes["หมวดหมู่แนะนำใน LnwShop"]

    original_header = "รูปภาพต้นฉบับ"
    watermarked_header = "รูปภาพพร้อมลายน้ำ"
    original_category_header = "หมวดหมู่แนะนำเดิม"
    product_type_1_header = "Product Type ระดับ 1"
    product_type_2_header = "Product Type ระดับ 2"
    if original_header not in indexes:
        ws.cell(1, ws.max_column + 1).value = original_header
    indexes = header_index(ws)
    if watermarked_header not in indexes:
        ws.cell(1, ws.max_column + 1).value = watermarked_header
    indexes = header_index(ws)
    if original_category_header not in indexes:
        ws.cell(1, ws.max_column + 1).value = original_category_header
    indexes = header_index(ws)
    if product_type_1_header not in indexes:
        ws.cell(1, ws.max_column + 1).value = product_type_1_header
    indexes = header_index(ws)
    if product_type_2_header not in indexes:
        ws.cell(1, ws.max_column + 1).value = product_type_2_header
    indexes = header_index(ws)

    image_map: dict[str, str] = {}
    count = 0
    for row in range(2, ws.max_row + 1):
        current = clean(ws.cell(row, image_col).value)
        brand = clean(ws.cell(row, brand_col).value) or "IMOU"
        original_category = clean(ws.cell(row, category_col).value)
        if original_category:
            set_cell(ws, row, original_category_header, original_category, indexes)
            if not original_category.startswith(f"{brand} >"):
                ws.cell(row, category_col).value = f"{brand} > {original_category}"
        set_cell(ws, row, product_type_1_header, DEFAULT_PRODUCT_TYPE_1, indexes)
        set_cell(ws, row, product_type_2_header, DEFAULT_PRODUCT_TYPE_2, indexes)
        if not current or "ไม่มีรูป" in current:
            continue
        src_path = Path(current)
        if not src_path.is_absolute():
            src_path = SOURCE_EXCEL.parent / src_path
        if not src_path.exists():
            continue

        out_name = f"{src_path.stem}_watermarked.png"
        out_path = WATERMARKED_DIR / out_name
        apply_watermark(src_path, out_path, mark)
        rel_out = out_path.relative_to(SOURCE_EXCEL.parent)
        rel_src = src_path.relative_to(SOURCE_EXCEL.parent)

        set_cell(ws, row, original_header, str(rel_src), indexes)
        set_cell(ws, row, watermarked_header, str(rel_out), indexes)
        ws.cell(row, image_col).value = str(rel_out)
        image_map[current] = str(rel_out)
        count += 1

    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(indexes[original_header])].width = 42
    ws.column_dimensions[get_column_letter(indexes[watermarked_header])].width = 42
    ws.column_dimensions[get_column_letter(indexes[original_category_header])].width = 42
    ws.column_dimensions[get_column_letter(indexes[product_type_1_header])].width = 28
    ws.column_dimensions[get_column_letter(indexes[product_type_2_header])].width = 40
    wb.save(WATERMARKED_EXCEL)
    return count, image_map


def category_description(name: str, parent: str) -> str:
    if parent:
        return (
            f"{name} สำหรับร้าน One Tech Solution รวมสินค้า IMOU ที่คัดไว้สำหรับการใช้งานจริง "
            f"ในหมวด {parent} พร้อมข้อมูลสินค้าไทยและอังกฤษเพื่อช่วยค้นหาบนหน้าเว็บ"
        )
    return (
        f"{name} จาก One Tech Solution รวมสินค้า IMOU และอุปกรณ์ที่เกี่ยวข้อง "
        "จัดหมวดหมู่เพื่อให้ลูกค้าเลือกสินค้าได้ง่าย"
    )


def build_categories() -> int:
    wb_src = load_workbook(WATERMARKED_EXCEL if WATERMARKED_EXCEL.exists() else SOURCE_EXCEL, data_only=True)
    ws_src = wb_src["LnwShop_AI_Ready"]
    indexes = header_index(ws_src)
    cat_col = indexes["หมวดหมู่แนะนำใน LnwShop"]
    img_col = indexes["รูปภาพหลัก/ไฟล์ภาพ"]
    brand_col = indexes["แบรนด์"]

    categories: "OrderedDict[str, dict[str, str]]" = OrderedDict()

    def add_category(name: str, parent: str, image: str, source: str) -> None:
        key = f"{parent}>{name}" if parent else name
        if key not in categories:
            categories[key] = {
                "name": name,
                "parent": parent,
                "description": category_description(name, parent),
                "image": image,
                "status": "รอตรวจสอบ",
                "product_type_1": DEFAULT_PRODUCT_TYPE_1,
                "product_type_2": DEFAULT_PRODUCT_TYPE_2,
                "source": source,
            }
        elif not categories[key]["image"] and image:
            categories[key]["image"] = image

    for row in range(2, ws_src.max_row + 1):
        chain = clean(ws_src.cell(row, cat_col).value)
        image = clean(ws_src.cell(row, img_col).value)
        brand = clean(ws_src.cell(row, brand_col).value) or "IMOU"
        if not chain:
            continue
        parts = [part.strip() for part in chain.split(">") if part.strip()]
        if parts and parts[0] == brand:
            parts = parts[1:]
        add_category(brand, "", image, "สร้างเป็นหมวดแบรนด์หลัก")
        parent = brand
        for part in parts:
            source = "สร้างจากหมวดหมู่หลักในไฟล์สินค้า" if parent == brand else "สร้างจากหมวดหมู่ย่อยในไฟล์สินค้า"
            add_category(part, parent, image, source)
            parent = part

    wb = Workbook()
    ws = wb.active
    ws.title = "Categories_AI_Ready"
    headers = [
        "ลำดับ",
        "สถานะ",
        "ชื่อหมวดหมู่",
        "หมวดหมู่ย่อยของ",
        "รายละเอียดหมวดหมู่",
        "แสดงหมวดหมู่ที่หน้าร้าน",
        "Product Type ระดับ 1",
        "Product Type ระดับ 2",
        "รูปภาพหมวดหมู่",
        "SEO Title",
        "SEO Keywords",
        "SEO Meta Description",
        "แหล่งข้อมูล",
    ]
    ws.append(headers)
    for idx, item in enumerate(categories.values(), start=1):
        name = item["name"]
        parent = item["parent"]
        seo_title = f"{name} | One Tech Solution IMOU"
        keywords = ", ".join(
            token
            for token in [name, parent, "IMOU", "One Tech Solution", "กล้องวงจรปิด", "IP Camera", "CCTV"]
            if token
        )
        meta = f"{name} {('ในหมวด ' + parent) if parent else ''} สินค้า IMOU จาก One Tech Solution พร้อมรายละเอียดไทยและอังกฤษ"
        ws.append(
            [
                idx,
                item["status"],
                name,
                parent or "ไม่มีหมวดหมู่หลัก",
                item["description"],
                "เปิด",
                item["product_type_1"],
                item["product_type_2"],
                item["image"],
                seo_title,
                keywords,
                meta,
                item["source"],
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for sheet in [ws]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(bottom=thin_gray)
    widths = [8, 16, 34, 28, 70, 24, 28, 40, 50, 45, 60, 70, 35]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 64
    end_col = get_column_letter(ws.max_column)
    tab = Table(displayName="LnwShopCategories", ref=f"A1:{end_col}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    wb.save(CATEGORY_EXCEL)
    return len(categories)


def main() -> int:
    image_count, _ = build_watermarked_workbook()
    category_count = build_categories()
    print(f"watermarked images: {image_count}")
    print(f"watermarked excel: {WATERMARKED_EXCEL}")
    print(f"categories: {category_count}")
    print(f"category excel: {CATEGORY_EXCEL}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
