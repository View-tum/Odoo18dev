from __future__ import annotations

import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

from prepare_assets import WATERMARK, apply_watermark, crop_watermark
from prepare_hikfire_assets import CATEGORY_HEADERS, DEFAULT_PRODUCT_TYPE_1, DEFAULT_PRODUCT_TYPE_2, PRODUCT_HEADERS


ROOT = Path(__file__).resolve().parents[2]
SOURCE_EXCEL = Path(r"C:\Users\tumsu\Downloads\ATHENS Price HIKVISION 2026 .xlsx")
AUDIT_EXCEL = ROOT / "output" / "spreadsheet" / "HIKVISION_product_inspection_report.xlsx"
WATERMARKED_DIR = ROOT / "output" / "spreadsheet" / "hikvision_product_images_watermarked"
PRODUCT_EXCEL = ROOT / "output" / "spreadsheet" / "HIKVISION_LnwShop_AI_ready_watermarked.xlsx"
CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "HIKVISION_LnwShop_categories.xlsx"
REVIEW_EXCEL = ROOT / "output" / "spreadsheet" / "HIKVISION_LnwShop_review.xlsx"
BRAND = "HIKVISION"
EXISTING_ROOT_CATEGORY = "HikVision"
LNWSHOP_CATEGORY_ALIASES = {
    "CCTV กล้องวงจรปิด": "CCTV กล้อวงจรปิด",
}


EXISTING_CATEGORY_ORDER = [
    "CCTV กล้องวงจรปิด",
    "NVR",
    "Access Control",
    "Access Control Accessories Kit",
    "Entrance & Exit (Car Park)",
    "Switches, AP & Router",
    "Accessories",
]


CATEGORY_DESCRIPTIONS = {
    "CCTV กล้องวงจรปิด": "กล้องวงจรปิด HikVision ทั้งระบบ Analog, IP Camera, ColorVu, AcuSense และกล้องสำหรับงานรักษาความปลอดภัย",
    "NVR": "เครื่องบันทึกภาพ HikVision DVR และ NVR สำหรับระบบกล้องวงจรปิดในบ้าน อาคาร ร้านค้า และโรงงาน",
    "Access Control": "อุปกรณ์ควบคุมประตู HikVision เครื่องสแกนใบหน้า เครื่องอ่านบัตร และระบบ Access Control",
    "Access Control Accessories Kit": "ชุดอุปกรณ์และอะไหล่ Access Control HikVision เช่น ชุดล็อกประตู ปุ่มกดออก อุปกรณ์จ่ายไฟ และขาจับ",
    "Entrance & Exit (Car Park)": "อุปกรณ์ระบบอ่านป้ายทะเบียน HikVision ANPR ไม้กั้นรถยนต์ อุปกรณ์ Barrier Gate และอุปกรณ์ประกอบ",
    "Switches, AP & Router": "อุปกรณ์เครือข่าย HikVision PoE Switch, Ethernet Switch, Wireless Router และ Access Point สำหรับระบบกล้องวงจรปิด",
    "Accessories": "อุปกรณ์เสริม HikVision สายสัญญาณ อุปกรณ์ประกอบ และสินค้าเสริมสำหรับงานติดตั้ง",
    "กล้องวงจรปิด HIKVISION": "กล้องวงจรปิด HIKVISION ทั้งระบบ Analog, IP Camera, ColorVu, AcuSense และกล้องสำหรับงานรักษาความปลอดภัย",
    "เครื่องบันทึก DVR/NVR": "เครื่องบันทึกภาพ HIKVISION DVR และ NVR สำหรับระบบกล้องวงจรปิดในบ้าน อาคาร ร้านค้า และโรงงาน",
    "ระบบ Access Control": "อุปกรณ์ควบคุมประตู Access Control HIKVISION เครื่องสแกนใบหน้า เครื่องอ่านบัตร ปุ่มกดออก และอุปกรณ์ล็อกประตู",
    "ระบบอ่านป้ายทะเบียน ANPR และไม้กั้น": "อุปกรณ์ระบบอ่านป้ายทะเบียน HIKVISION ANPR ไม้กั้นรถยนต์ อุปกรณ์ Barrier Gate และอุปกรณ์ประกอบ",
    "อุปกรณ์เครือข่ายและสายสัญญาณ": "อุปกรณ์เครือข่าย HIKVISION สาย LAN, PoE Switch, Ethernet Switch และ Wireless Router สำหรับระบบกล้องวงจรปิด",
    "สินค้า HIKVISION อื่นๆ": "สินค้าและอุปกรณ์ HIKVISION สำหรับงานระบบรักษาความปลอดภัยและงานติดตั้ง",
}


THAI_PREFIX = {
    "CCTV กล้องวงจรปิด": "กล้องวงจรปิด HikVision",
    "NVR": "เครื่องบันทึกภาพ HikVision",
    "Access Control": "อุปกรณ์ควบคุมประตู Access Control HikVision",
    "Access Control Accessories Kit": "ชุดอุปกรณ์ Access Control HikVision",
    "Entrance & Exit (Car Park)": "อุปกรณ์ระบบอ่านป้ายทะเบียนและไม้กั้น HikVision",
    "Switches, AP & Router": "อุปกรณ์เครือข่าย HikVision",
    "Accessories": "อุปกรณ์เสริม HikVision",
    "กล้องวงจรปิด HIKVISION": "กล้องวงจรปิด HIKVISION",
    "เครื่องบันทึก DVR/NVR": "เครื่องบันทึกภาพ HIKVISION",
    "ระบบ Access Control": "อุปกรณ์ควบคุมประตู Access Control HIKVISION",
    "ระบบอ่านป้ายทะเบียน ANPR และไม้กั้น": "อุปกรณ์ระบบอ่านป้ายทะเบียนและไม้กั้น HIKVISION",
    "อุปกรณ์เครือข่ายและสายสัญญาณ": "อุปกรณ์เครือข่ายและสายสัญญาณ HIKVISION",
    "สินค้า HIKVISION อื่นๆ": "สินค้า HIKVISION",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def one_line(value: Any) -> str:
    text = clean(value).replace("\n", " / ")
    return re.sub(r"\s+", " ", text).strip()


def exact_key(value: str) -> str:
    text = value.upper().replace("รุ่น", " ")
    text = re.sub(r'"', "", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def model_from_code(value: str) -> str:
    text = clean(value)
    text = text.replace('"""', "").replace('"', "").strip()
    text = re.sub(r"^HIKVISION\s+รุ่น\s+", "", text, flags=re.IGNORECASE)
    text = re.split(r"\n| / ประกอบไปด้วย|ประกอบไปด้วย", text, maxsplit=1)[0].strip()
    text = re.sub(r"\s+", " ", text)
    return text or one_line(value)


def sku_for(key: str, model: str) -> str:
    sku_key = key or exact_key(model)
    if not sku_key:
        sku_key = re.sub(r"[^A-Z0-9]+", "", model.upper())[:40]
    return f"HIKVISION-{sku_key}"


def slug_for(key: str, model: str) -> str:
    base = key or exact_key(model) or model
    slug = re.sub(r"[^a-z0-9]+", "-", base.lower()).strip("-")
    return f"hikvision-{slug}" if slug else "hikvision-product"


def english_category(category: str) -> str:
    mapping = {
        "CCTV กล้องวงจรปิด": "CCTV Camera",
        "NVR": "DVR NVR Recorder",
        "Access Control": "Access Control",
        "Access Control Accessories Kit": "Access Control Accessories Kit",
        "Entrance & Exit (Car Park)": "ANPR and Barrier Gate",
        "Switches, AP & Router": "Switches AP Router",
        "Accessories": "Accessories",
        "กล้องวงจรปิด HIKVISION": "CCTV Camera",
        "เครื่องบันทึก DVR/NVR": "DVR NVR Recorder",
        "ระบบ Access Control": "Access Control",
        "ระบบอ่านป้ายทะเบียน ANPR และไม้กั้น": "ANPR and Barrier Gate",
        "อุปกรณ์เครือข่ายและสายสัญญาณ": "Network and Cable",
    }
    return mapping.get(category, "Security Product")


def lnwshop_category_name(category: str) -> str:
    return LNWSHOP_CATEGORY_ALIASES.get(category, category)


def existing_category_for(record: dict[str, Any], model: str, proposed_category: str) -> str:
    raw_section = clean(record.get("Raw Section")).upper()
    line = clean(record.get("Line")).upper()
    code = model.upper()
    text = f"{raw_section} {line} {code}"
    if "ANPR" in text or "ENTRANCE" in text or "BARRIER" in text or "TMC" in code or "TMG" in code:
        return "Entrance & Exit (Car Park)"
    if "SWITCH" in text or "ROUTER" in text or code.startswith("DS-3E") or code.startswith("DS-3WR"):
        return "Switches, AP & Router"
    if "CABLE" in text or code.startswith("DS-1LN"):
        return "Accessories"
    if "ACCESS" in text or code.startswith("DS-K") or code.startswith("HVS-K"):
        if any(token in code for token in ["KAS", "K4H", "K7", "K2M", "K2T", "AW24", "P07"]):
            return "Access Control Accessories Kit"
        return "Access Control"
    if "DVR" in text or "NVR" in text or re.match(r"^(I?DS-72|DS-77)", code):
        return "NVR"
    if "CAMERA" in text or code.startswith("DS-2C") or code.startswith("IDS-2C"):
        return "CCTV กล้องวงจรปิด"
    if proposed_category == "ระบบ Access Control":
        return "Access Control"
    if proposed_category == "ระบบอ่านป้ายทะเบียน ANPR และไม้กั้น":
        return "Entrance & Exit (Car Park)"
    if proposed_category == "อุปกรณ์เครือข่ายและสายสัญญาณ":
        return "Switches, AP & Router"
    if proposed_category == "เครื่องบันทึก DVR/NVR":
        return "NVR"
    if proposed_category == "กล้องวงจรปิด HIKVISION":
        return "CCTV กล้องวงจรปิด"
    return "Accessories"


def ensure_audit() -> None:
    if AUDIT_EXCEL.exists():
        return
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)
    subprocess.run([sys.executable, str(Path(__file__).resolve().parent / "inspect_hikvision_products.py"), str(SOURCE_EXCEL)], check=True)


def audit_records() -> list[dict[str, Any]]:
    ensure_audit()
    wb = load_workbook(AUDIT_EXCEL, data_only=True)
    ws = wb["HIKVISION_Product_Audit"]
    headers = [cell.value for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row_number in range(2, ws.max_row + 1):
        values = [cell.value for cell in ws[row_number]]
        row = dict(zip(headers, values))
        row["_audit_row"] = row_number
        rows.append(row)
    return rows


def rel_to_output(path_value: str) -> str:
    if not path_value:
        return ""
    path = Path(path_value)
    if not path.is_absolute():
        path = (ROOT / "output" / "spreadsheet" / path).resolve()
    try:
        return str(path.relative_to(ROOT / "output" / "spreadsheet"))
    except ValueError:
        return str(path)


def ensure_watermarked_images(records: list[dict[str, Any]]) -> None:
    WATERMARKED_DIR.mkdir(parents=True, exist_ok=True)
    if not WATERMARK.exists():
        raise FileNotFoundError(WATERMARK)
    mark = crop_watermark(Image.open(WATERMARK))
    wb = load_workbook(AUDIT_EXCEL)
    ws = wb["HIKVISION_Product_Audit"]
    headers = [cell.value for cell in ws[1]]
    if "Watermarked Image" not in headers:
        col = ws.max_column + 1
        ws.cell(1, col).value = "Watermarked Image"
        headers.append("Watermarked Image")
    else:
        col = headers.index("Watermarked Image") + 1
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    for row in records:
        audit_row = int(row["_audit_row"])
        primary = Path(clean(row.get("Primary Image")))
        if not primary.exists():
            continue
        model = model_from_code(clean(row.get("Product Code / Name")))
        source_row = int(row.get("Source Row") or 0)
        prefix = "access" if "Access" in clean(row.get("Source Sheet")) else "cctv"
        safe = re.sub(r"[^A-Za-z0-9]+", "-", model.upper()).strip("-")[:70] or f"row-{source_row}"
        out_path = WATERMARKED_DIR / f"hikvision_{prefix}_row_{source_row:04d}_{safe}_watermarked.png"
        apply_watermark(primary, out_path, mark)
        ws.cell(audit_row, col).value = str(out_path)
        row["Watermarked Image"] = str(out_path)
    ws.column_dimensions[get_column_letter(col)].width = 54
    wb.save(AUDIT_EXCEL)


def thai_name(model: str, category: str) -> str:
    prefix = THAI_PREFIX.get(category, "สินค้า HIKVISION")
    return f"{prefix} รุ่น {model}"


def english_name(model: str, category: str, line: str) -> str:
    suffix = line or english_category(category)
    return f"HIKVISION {model} {suffix}".strip()


def warranty_text(value: str) -> str:
    return clean(value) or "รอตรวจสอบ"


def short_description(name_th: str, model: str, category: str) -> str:
    return f"{name_th} สำหรับงานระบบรักษาความปลอดภัย กล้องวงจรปิด Access Control เครือข่าย และงานติดตั้ง HIKVISION"


def thai_detail(name_th: str, model: str, category: str, details: str, warranty: str) -> str:
    parts = [
        name_th,
        f"สินค้าแบรนด์ HIKVISION รุ่น {model} จัดอยู่ในหมวด {category} เหมาะสำหรับงานติดตั้งระบบรักษาความปลอดภัย อาคาร ร้านค้า โรงงาน และโครงการ",
    ]
    if details:
        parts.append(f"รายละเอียดจากไฟล์ต้นฉบับ: {one_line(details)}")
    if warranty:
        parts.append(f"การรับประกัน: {warranty}")
    parts.append("ราคาและจำนวน stock ในไฟล์นี้ตั้งเป็น 0 เพื่อให้ตรวจสอบก่อนลงร้านจริง")
    return "\n".join(parts)


def feature_text(details: str, warranty: str, category: str) -> str:
    items = [f"- หมวดสินค้า: {category}"]
    if details:
        snippets = [part.strip() for part in re.split(r"/|\n", details) if part.strip()]
        for snippet in snippets[:4]:
            items.append(f"- {snippet}")
    if warranty:
        items.append(f"- Warranty: {warranty}")
    return "\n".join(items)


def source_specs(row: dict[str, Any], model: str) -> str:
    fields = [
        ("Source Sheet", clean(row.get("Source Sheet"))),
        ("Source Row", clean(row.get("Source Row"))),
        ("Product Code", model),
        ("Raw Section", clean(row.get("Raw Section"))),
        ("Line", clean(row.get("Line"))),
        ("Details", clean(row.get("Details"))),
        ("Warranty", clean(row.get("Warranty"))),
        ("Original MSRP", clean(row.get("MSRP"))),
        ("Original Dealer", clean(row.get("Dealer"))),
        ("Original VIP/User", clean(row.get("VIP/User"))),
        ("Audit Issues", clean(row.get("Issues"))),
    ]
    return "\n".join(f"{name}: {value}" for name, value in fields if value)


def select_unique_records(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selected: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in records:
        key = clean(row.get("Normalized SKU Key")) or exact_key(clean(row.get("Product Code / Name")))
        if key in seen:
            skipped.append(row)
            continue
        seen.add(key)
        selected.append(row)
    return selected, skipped


def build_product_rows(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ensure_watermarked_images(records)
    selected, skipped = select_unique_records(records)
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(selected, start=1):
        raw_code = clean(record.get("Product Code / Name"))
        model = model_from_code(raw_code)
        key = clean(record.get("Normalized SKU Key")) or exact_key(raw_code)
        proposed_category = clean(record.get("Proposed Category")) or "สินค้า HIKVISION อื่นๆ"
        category = existing_category_for(record, model, proposed_category)
        category_chain = f"{EXISTING_ROOT_CATEGORY} > {lnwshop_category_name(category)}"
        line = clean(record.get("Line"))
        details = clean(record.get("Details"))
        warranty = warranty_text(clean(record.get("Warranty")))
        name_th = thai_name(model, category)
        name_en = english_name(model, category, line)
        watermarked = rel_to_output(clean(record.get("Watermarked Image")))
        original_image = rel_to_output(clean(record.get("Primary Image")))
        issue_text = clean(record.get("Issues"))
        note_parts = ["ตั้งราคาและ stock เป็น 0 เพื่อให้ตรวจสอบก่อนลงร้านจริง", f"เตรียมจาก audit row {record.get('_audit_row')}"]
        if issue_text:
            note_parts.append(f"Audit note: {issue_text}")
        seo_title_th = f"{name_th} | One Tech Solution"
        seo_title_en = f"{name_en} | One Tech Solution"
        keywords = ", ".join(
            token
            for token in [
                model,
                BRAND,
                "Hikvision",
                category,
                line,
                "CCTV",
                "Security",
                "Access Control",
                "ANPR",
                "One Tech Solution",
                "ซ่อมติดตั้งสั่งซื้ออะไหล่",
                "ขายส่ง",
                "คอนโดโรงเรียนโรงแรมหมู่บ้าน",
            ]
            if token
        )
        meta = f"{name_th} / {name_en} สินค้า HIKVISION สำหรับงานระบบรักษาความปลอดภัย ราคา 0 บาทในไฟล์ตรวจสอบก่อนลงร้าน"
        rows.append(
            {
                "ลำดับ": index,
                "สถานะลงสินค้า": "รอตรวจสอบ",
                "เผยแพร่บนร้าน": "ปิด",
                "แบรนด์": BRAND,
                "หมวดหมู่หลัก": category,
                "หมวดหมู่ย่อย": "",
                "หมวดหมู่แนะนำใน LnwShop": category_chain,
                "ชื่อสินค้า (ไทย)": name_th,
                "Product Name (EN)": name_en,
                "SKU / Article Code": sku_for(key, model),
                "Vendor Part No.": raw_code,
                "Model": model,
                "EAN/UPC": "",
                "ราคาปกติ (บาท)": 0,
                "ราคาขาย (บาท)": 0,
                "ราคาพิเศษ (บาท)": 0,
                "จำนวนสต๊อกตั้งต้น": 0,
                "สถานะสต๊อก": "รอตรวจสอบ",
                "น้ำหนักรวม (kg)": "",
                "กว้าง (cm)": "",
                "ยาว (cm)": "",
                "สูง (cm)": "",
                "ขนาดรวม": "",
                "การรับประกัน": warranty,
                "รายละเอียดสั้น (ไทย)": short_description(name_th, model, category),
                "รายละเอียดเต็ม (ไทย)": thai_detail(name_th, model, category, details, warranty),
                "จุดเด่นสินค้า (ไทย)": feature_text(details, warranty, category),
                "Product Detail (EN)": one_line(details),
                "SEO Title (TH)": seo_title_th,
                "SEO Title (EN)": seo_title_en,
                "SEO Title (TH+EN)": f"{seo_title_th} / {seo_title_en}",
                "SEO Keywords (TH+EN)": keywords,
                "SEO Meta Description (TH+EN)": meta,
                "URL Slug": slug_for(key, model),
                "Search Tags": ", ".join(token for token in [model, BRAND, category, line, "CCTV", "Security", "ซ่อมติดตั้งสั่งซื้ออะไหล่", "ขายส่ง", "คอนโดโรงเรียนโรงแรมหมู่บ้าน"] if token),
                "รูปภาพหลัก/ไฟล์ภาพ": watermarked,
                "หมายเหตุสำหรับลงสินค้า": "; ".join(note_parts),
                "ข้อมูลสเปกต้นฉบับ": source_specs(record, model),
                "แหล่งข้อมูลในไฟล์ต้นฉบับ": f"{record.get('Source Sheet')} row {record.get('Source Row')}",
                "รูปภาพต้นฉบับ": original_image,
                "รูปภาพพร้อมลายน้ำ": watermarked,
                "หมวดหมู่แนะนำเดิม": f"{clean(record.get('Raw Section'))}; proposed={proposed_category}",
                "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                "ราคาต้นฉบับ Dealer": 0,
                "ราคาต้นฉบับ SI": 0,
                "ราคาต้นฉบับ MSRP": 0,
            }
        )
    return rows, skipped


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(one_line(cell.value)) for cell in column if cell.value is not None) if column else 10
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 48)
    ws.auto_filter.ref = ws.dimensions


def add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    table = Table(displayName=name, ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    add_table(ws, table_name)
    wb.save(path)


def category_description(name: str) -> str:
    return CATEGORY_DESCRIPTIONS.get(name, f"{name} สินค้า HIKVISION สำหรับจัดกลุ่มสินค้าในร้าน One Tech Solution")


def build_category_rows(product_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: list[dict[str, Any]] = []
    by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in product_rows:
        by_category[clean(row.get("หมวดหมู่หลัก"))].append(row)
    first_image = product_rows[0].get("รูปภาพหลัก/ไฟล์ภาพ", "") if product_rows else ""
    categories.append(
        {
            "ลำดับ": 1,
            "สถานะ": "ใช้งาน",
            "ชื่อหมวดหมู่": EXISTING_ROOT_CATEGORY,
            "หมวดหมู่ย่อยของ": "",
            "รายละเอียดหมวดหมู่": "รวมสินค้า HikVision กล้องวงจรปิด เครื่องบันทึกภาพ Access Control ระบบอ่านป้ายทะเบียน อุปกรณ์เครือข่าย และอุปกรณ์รักษาความปลอดภัย",
            "แสดงหมวดหมู่ที่หน้าร้าน": "เปิด",
            "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
            "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
            "รูปภาพหมวดหมู่": first_image,
            "SEO Title": "HikVision | One Tech Solution",
            "SEO Keywords": "HikVision, HIKVISION, กล้องวงจรปิด, CCTV, Access Control, ANPR, One Tech Solution",
            "SEO Meta Description": "สินค้า HikVision สำหรับระบบรักษาความปลอดภัย กล้องวงจรปิด Access Control และงานติดตั้งจาก One Tech Solution",
            "แหล่งข้อมูล": "หมวดนี้มีอยู่แล้วใน LnwShop",
        }
    )
    index = 2
    for name in EXISTING_CATEGORY_ORDER:
        rows = by_category.get(name, [])
        if not rows:
            continue
        lnwshop_name = lnwshop_category_name(name)
        image = rows[0].get("รูปภาพหลัก/ไฟล์ภาพ", "")
        categories.append(
            {
                "ลำดับ": index,
                "สถานะ": "ใช้งาน",
                "ชื่อหมวดหมู่": lnwshop_name,
                "หมวดหมู่ย่อยของ": EXISTING_ROOT_CATEGORY,
                "รายละเอียดหมวดหมู่": category_description(name),
                "แสดงหมวดหมู่ที่หน้าร้าน": "เปิด",
                "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                "รูปภาพหมวดหมู่": image,
                "SEO Title": f"{name} | One Tech Solution",
                "SEO Keywords": ", ".join([name, "HikVision", BRAND, "CCTV", "Security", "One Tech Solution"]),
                "SEO Meta Description": f"{name} สินค้า HikVision สำหรับงานระบบรักษาความปลอดภัยและงานติดตั้งจาก One Tech Solution",
                "แหล่งข้อมูล": f"หมวดนี้มีอยู่แล้วใน LnwShop; มีสินค้าในไฟล์ {len(rows)} รายการ",
            }
        )
        index += 1
    return categories


def write_review(product_rows: list[dict[str, Any]], category_rows: list[dict[str, Any]], skipped: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = [
        ["Metric", "Value"],
        ["Input audit rows", len(product_rows) + len(skipped)],
        ["Prepared unique product rows", len(product_rows)],
        ["Skipped duplicate rows", len(skipped)],
        ["Category rows", len(category_rows)],
        ["Price rule", "All price fields are 0"],
        ["Stock rule", "All stock fields are 0"],
        ["Watermark", str(WATERMARK)],
        ["Product Excel", str(PRODUCT_EXCEL)],
        ["Category Excel", str(CATEGORY_EXCEL)],
    ]
    for row in summary:
        ws.append(row)
    style_sheet(ws)
    ws_products = wb.create_sheet("Products")
    ws_products.append(PRODUCT_HEADERS)
    for row in product_rows:
        ws_products.append([row.get(header, "") for header in PRODUCT_HEADERS])
    style_sheet(ws_products)
    add_table(ws_products, "HIKVISIONReviewProducts")
    ws_categories = wb.create_sheet("Categories")
    ws_categories.append(CATEGORY_HEADERS)
    for row in category_rows:
        ws_categories.append([row.get(header, "") for header in CATEGORY_HEADERS])
    style_sheet(ws_categories)
    add_table(ws_categories, "HIKVISIONReviewCategories")
    ws_skipped = wb.create_sheet("Skipped_Duplicates")
    skipped_headers = ["Source Sheet", "Source Row", "Product Code / Name", "Normalized SKU Key", "Proposed Category", "Issues"]
    ws_skipped.append(skipped_headers)
    for row in skipped:
        ws_skipped.append([row.get(header, "") for header in skipped_headers])
    style_sheet(ws_skipped)
    add_table(ws_skipped, "HIKVISIONSkippedDuplicates")
    wb.save(REVIEW_EXCEL)


def validate(product_rows: list[dict[str, Any]], category_rows: list[dict[str, Any]]) -> list[str]:
    errors: list[str] = []
    sku_counts = Counter(clean(row.get("SKU / Article Code")) for row in product_rows)
    duplicates = [sku for sku, count in sku_counts.items() if count > 1]
    if duplicates:
        errors.append(f"Duplicate SKU: {duplicates[:10]}")
    for index, row in enumerate(product_rows, start=2):
        for header in ["รูปภาพหลัก/ไฟล์ภาพ", "ชื่อสินค้า (ไทย)", "SKU / Article Code", "หมวดหมู่แนะนำใน LnwShop"]:
            if not clean(row.get(header)):
                errors.append(f"Missing {header} at prepared row {index}")
        for header in ["ราคาปกติ (บาท)", "ราคาขาย (บาท)", "ราคาพิเศษ (บาท)", "จำนวนสต๊อกตั้งต้น"]:
            if row.get(header) != 0:
                errors.append(f"{header} is not 0 at prepared row {index}")
        image_path = ROOT / "output" / "spreadsheet" / clean(row.get("รูปภาพหลัก/ไฟล์ภาพ"))
        if not image_path.exists():
            errors.append(f"Missing image at prepared row {index}: {image_path}")
    if not category_rows:
        errors.append("No categories")
    category_names = {clean(row.get("ชื่อหมวดหมู่")) for row in category_rows}
    for row in product_rows:
        category = clean(row.get("หมวดหมู่หลัก"))
        if category and lnwshop_category_name(category) not in category_names:
            errors.append(f"Product category not found in category workbook: {category}")
    return errors


def main() -> int:
    records = audit_records()
    product_rows, skipped = build_product_rows(records)
    category_rows = build_category_rows(product_rows)
    errors = validate(product_rows, category_rows)
    if errors:
        for error in errors:
            print(f"ERROR: {error}", file=sys.stderr)
        return 1
    write_workbook(PRODUCT_EXCEL, "LnwShop_AI_Ready", PRODUCT_HEADERS, product_rows, "HIKVISIONProducts")
    write_workbook(CATEGORY_EXCEL, "Categories_AI_Ready", CATEGORY_HEADERS, category_rows, "HIKVISIONCategories")
    write_review(product_rows, category_rows, skipped)
    print(f"source audit rows: {len(records)}")
    print(f"prepared products: {len(product_rows)}")
    print(f"skipped duplicates: {len(skipped)}")
    print(f"categories: {len(category_rows)}")
    print(f"product excel: {PRODUCT_EXCEL}")
    print(f"category excel: {CATEGORY_EXCEL}")
    print(f"review excel: {REVIEW_EXCEL}")
    print(f"watermarked images: {WATERMARKED_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
