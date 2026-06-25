from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


ROOT = Path(__file__).resolve().parents[2]
WATERMARKED_EXCEL = ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_AI_ready_watermarked.xlsx"
DEFAULT_EXCEL = WATERMARKED_EXCEL if WATERMARKED_EXCEL.exists() else ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_AI_ready.xlsx"
DEFAULT_CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "IMOU_LnwShop_categories.xlsx"
DEFAULT_PROFILE = Path(__file__).resolve().parent / "edge-profile"
DEFAULT_INSPECT = Path(__file__).resolve().parent / "last_inspect.json"
DEFAULT_SCREENSHOT_DIR = ROOT / "output" / "playwright"
DEFAULT_START_URL = "https://a.lnwstore.com/onetechsolution/inventory/"
DEFAULT_ADD_PRODUCT_URL = "https://a.lnwstore.com/onetechsolution/inventory/product/new/"
DEFAULT_ADD_CATEGORY_URL = "https://a.lnwstore.com/onetechsolution/inventory/category/new/"


FIELD_MAP = {
    "brand": "แบรนด์",
    "name_th": "ชื่อสินค้า (ไทย)",
    "name_en": "Product Name (EN)",
    "sku": "SKU / Article Code",
    "vendor": "Vendor Part No.",
    "model": "Model",
    "barcode": "EAN/UPC",
    "normal_price": "ราคาปกติ (บาท)",
    "sale_price": "ราคาขาย (บาท)",
    "special_price": "ราคาพิเศษ (บาท)",
    "stock": "จำนวนสต๊อกตั้งต้น",
    "weight": "น้ำหนักรวม (kg)",
    "width": "กว้าง (cm)",
    "length": "ยาว (cm)",
    "height": "สูง (cm)",
    "short_description": "รายละเอียดสั้น (ไทย)",
    "full_description": "รายละเอียดเต็ม (ไทย)",
    "seo_title": "SEO Title (TH+EN)",
    "seo_keywords": "SEO Keywords (TH+EN)",
    "seo_description": "SEO Meta Description (TH+EN)",
    "tags": "Search Tags",
    "image": "รูปภาพหลัก/ไฟล์ภาพ",
    "shop_category": "หมวดหมู่แนะนำใน LnwShop",
    "product_type_1": "Product Type ระดับ 1",
    "product_type_2": "Product Type ระดับ 2",
}


LABELS = {
    "brand": ["แบรนด์", "ยี่ห้อ", "Brand"],
    "name_th": ["ชื่อสินค้า", "ชื่อสินค้า *", "Product Name", "Name"],
    "sku": ["รหัสสินค้า", "SKU", "Article Code", "รหัส"],
    "normal_price": ["ราคาปกติ", "ราคา", "Price"],
    "sale_price": ["ราคาขาย", "ราคาสินค้า", "Sale Price"],
    "special_price": ["ราคาพิเศษ", "Special Price"],
    "stock": ["จำนวนสินค้า", "สต๊อก", "Stock", "Inventory"],
    "weight": ["น้ำหนัก", "Weight"],
    "width": ["กว้าง", "Width"],
    "length": ["ยาว", "Length"],
    "height": ["สูง", "Height"],
    "full_description": ["รายละเอียดสินค้า", "รายละเอียด", "Description"],
    "seo_title": ["SEO Title", "Meta Title", "หัวข้อ SEO", "Title"],
    "seo_keywords": ["SEO Keywords", "Meta Keywords", "Keywords", "คำค้นหา"],
    "seo_description": ["SEO Description", "Meta Description", "Description"],
    "tags": ["Tags", "Tag", "ป้ายสินค้า"],
    "barcode": ["Barcode", "EAN", "UPC", "บาร์โค้ด"],
    "category_name": ["ชื่อหมวดหมู่", "Category Name", "Name"],
    "category_parent": ["หมวดหมู่ย่อยของ", "หมวดหมู่หลัก", "Parent Category", "Parent"],
    "category_description": ["รายละเอียดหมวดหมู่", "รายละเอียด", "Description"],
    "product_type": ["ประเภทสินค้า", "Product Type"],
}


CATEGORY_FIELD_MAP = {
    "status": "สถานะ",
    "name": "ชื่อหมวดหมู่",
    "parent": "หมวดหมู่ย่อยของ",
    "description": "รายละเอียดหมวดหมู่",
    "display": "แสดงหมวดหมู่ที่หน้าร้าน",
    "product_type_1": "Product Type ระดับ 1",
    "product_type_2": "Product Type ระดับ 2",
    "image": "รูปภาพหมวดหมู่",
    "seo_title": "SEO Title",
    "seo_keywords": "SEO Keywords",
    "seo_description": "SEO Meta Description",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_product(excel_path: Path, row_number: int) -> dict[str, str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["LnwShop_AI_Ready"]
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[row_number]]
    data_by_header = dict(zip(headers, row))

    product: dict[str, str] = {}
    for key, header in FIELD_MAP.items():
        product[key] = clean(data_by_header.get(header))

    image_value = product.get("image", "")
    if image_value and "ไม่มีรูป" not in image_value:
        image_path = Path(image_value)
        if not image_path.is_absolute():
            image_path = excel_path.parent / image_path
        product["image_abs"] = str(image_path.resolve())
    else:
        product["image_abs"] = ""

    product["excel_row"] = str(row_number)
    product["display_name"] = product.get("name_th") or product.get("name_en") or product.get("sku")
    return product


def read_category(excel_path: Path, row_number: int) -> dict[str, str]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Categories_AI_Ready"]
    headers = [cell.value for cell in ws[1]]
    row = [cell.value for cell in ws[row_number]]
    data_by_header = dict(zip(headers, row))

    category: dict[str, str] = {}
    for key, header in CATEGORY_FIELD_MAP.items():
        category[key] = clean(data_by_header.get(header))

    image_value = category.get("image", "")
    if image_value:
        image_path = Path(image_value)
        if not image_path.is_absolute():
            image_path = excel_path.parent / image_path
        category["image_abs"] = str(image_path.resolve())
    else:
        category["image_abs"] = ""

    category["excel_row"] = str(row_number)
    category["display_name"] = category.get("name", "")
    return category


def print_product(product: dict[str, str]) -> None:
    keys = [
        "excel_row",
        "display_name",
        "brand",
        "sku",
        "model",
        "shop_category",
        "product_type_1",
        "product_type_2",
        "normal_price",
        "sale_price",
        "special_price",
        "stock",
        "image_abs",
    ]
    for key in keys:
        print(f"{key}: {product.get(key, '')}")


def print_category(category: dict[str, str]) -> None:
    keys = ["excel_row", "display_name", "parent", "product_type_1", "product_type_2", "display", "image_abs"]
    for key in keys:
        print(f"{key}: {category.get(key, '')}")


def visible_field_info(page) -> list[dict[str, str]]:
    return page.evaluate(
        """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          const labelText = (el) => {
            const labels = el.labels ? Array.from(el.labels).map(l => l.innerText.trim()).join(' | ') : '';
            const aria = el.getAttribute('aria-label') || '';
            const placeholder = el.getAttribute('placeholder') || '';
            const name = el.getAttribute('name') || '';
            const id = el.getAttribute('id') || '';
            const parent = el.closest('label, .form-group, .row, tr, div');
            const context = parent ? parent.innerText.trim().replace(/\\s+/g, ' ').slice(0, 240) : '';
            return { labels, aria, placeholder, name, id, context };
          };
          return Array.from(document.querySelectorAll('input, textarea, select, [contenteditable="true"]'))
            .filter(visible)
            .map((el, index) => ({
              index,
              tag: el.tagName.toLowerCase(),
              type: el.getAttribute('type') || '',
              value: (el.value || el.textContent || '').slice(0, 80),
              ...labelText(el)
            }));
        }
        """
    )


def visible_buttons(page) -> list[dict[str, str]]:
    return page.evaluate(
        """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          return Array.from(document.querySelectorAll('button, input[type=submit], a'))
            .filter(visible)
            .map((el, index) => ({
              index,
              tag: el.tagName.toLowerCase(),
              type: el.getAttribute('type') || '',
              text: (el.innerText || el.value || el.getAttribute('aria-label') || '').trim().replace(/\\s+/g, ' ').slice(0, 120),
              href: el.getAttribute('href') || ''
            }))
            .filter(item => item.text || item.href);
        }
        """
    )


def fill_by_context(page, labels: list[str], value: str, prefer_multiline: bool = False) -> dict[str, Any]:
    labels = [label.lower() for label in labels]
    return page.evaluate(
        """
        ({ labels, value, preferMultiline }) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          const fieldText = (el) => {
            const parts = [];
            if (el.labels) parts.push(...Array.from(el.labels).map(l => l.innerText || ''));
            for (const attr of ['aria-label', 'placeholder', 'name', 'id', 'class']) {
              parts.push(el.getAttribute(attr) || '');
            }
            const parent = el.closest('label, .form-group, .row, tr, div');
            if (parent) parts.push((parent.innerText || '').slice(0, 260));
            return parts.join(' ').replace(/\\s+/g, ' ').toLowerCase();
          };
          const candidates = Array.from(document.querySelectorAll('input:not([type=hidden]):not([type=file]), textarea, [contenteditable="true"]'))
            .filter(visible)
            .filter(el => !['checkbox', 'radio', 'submit', 'button'].includes((el.getAttribute('type') || '').toLowerCase()));
          const sorted = candidates.sort((a, b) => {
            const aMulti = a.tagName.toLowerCase() === 'textarea' || a.isContentEditable;
            const bMulti = b.tagName.toLowerCase() === 'textarea' || b.isContentEditable;
            if (preferMultiline && aMulti !== bMulti) return aMulti ? -1 : 1;
            return 0;
          });
          for (const el of sorted) {
            const haystack = fieldText(el);
            if (labels.some(label => haystack.includes(label))) {
              if (el.isContentEditable) {
                el.focus();
                el.innerText = value;
              } else {
                el.focus();
                el.value = value;
              }
              el.dispatchEvent(new Event('input', { bubbles: true }));
              el.dispatchEvent(new Event('change', { bubbles: true }));
              return { ok: true, matched: labels.find(label => haystack.includes(label)), tag: el.tagName.toLowerCase(), context: haystack.slice(0, 220) };
            }
          }
          return { ok: false };
        }
        """,
        {"labels": labels, "value": value, "preferMultiline": prefer_multiline},
    )


def select_by_context(page, labels: list[str], option_text: str) -> dict[str, Any]:
    if not option_text or option_text == "ไม่มีหมวดหมู่หลัก":
        return {"ok": True, "skipped": True, "reason": "no parent category"}
    labels = [label.lower() for label in labels]
    return page.evaluate(
        """
        ({ labels, optionText }) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          const fieldText = (el) => {
            const parts = [];
            if (el.labels) parts.push(...Array.from(el.labels).map(l => l.innerText || ''));
            for (const attr of ['aria-label', 'placeholder', 'name', 'id', 'class']) {
              parts.push(el.getAttribute(attr) || '');
            }
            const parent = el.closest('label, .form-group, .row, tr, div');
            if (parent) parts.push((parent.innerText || '').slice(0, 260));
            return parts.join(' ').replace(/\\s+/g, ' ').toLowerCase();
          };
          for (const el of Array.from(document.querySelectorAll('select')).filter(visible)) {
            const haystack = fieldText(el);
            if (!labels.some(label => haystack.includes(label))) continue;
            const options = Array.from(el.options);
            let option = options.find(opt => (opt.textContent || '').trim() === optionText);
            if (!option) option = options.find(opt => (opt.textContent || '').includes(optionText));
            if (!option) return { ok: false, reason: 'option not found', available: options.map(opt => (opt.textContent || '').trim()).slice(0, 60) };
            el.value = option.value;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            return { ok: true, selected: (option.textContent || '').trim() };
          }
          return { ok: false, reason: 'select not found' };
        }
        """,
        {"labels": labels, "optionText": option_text},
    )


def select_option_by_text_anywhere(page, option_text: str, label_hint: str = "") -> dict[str, Any]:
    if not option_text:
        return {"ok": False, "reason": "empty option text"}
    result = page.evaluate(
        """
        ({ optionText, labelHint }) => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          const fieldText = (el) => {
            const parts = [];
            if (el.labels) parts.push(...Array.from(el.labels).map(l => l.innerText || ''));
            for (const attr of ['aria-label', 'placeholder', 'name', 'id', 'class']) {
              parts.push(el.getAttribute(attr) || '');
            }
            const parent = el.closest('label, .form-group, .row, tr, div');
            if (parent) parts.push((parent.innerText || '').slice(0, 320));
            return parts.join(' ').replace(/\\s+/g, ' ');
          };
          const selects = Array.from(document.querySelectorAll('select')).filter(visible);
          const withOption = [];
          for (const el of selects) {
            const options = Array.from(el.options || []);
            let option = options.find(opt => (opt.textContent || '').trim() === optionText);
            if (!option) option = options.find(opt => (opt.textContent || '').includes(optionText));
            if (option) withOption.push({ el, option, context: fieldText(el) });
          }
          if (withOption.length < 1) return { ok: false, reason: 'option not found in visible selects' };
          let picked = withOption[0];
          if (labelHint) {
            const hinted = withOption.find(item => item.context.toLowerCase().includes(labelHint.toLowerCase()));
            if (hinted) picked = hinted;
          }
          picked.el.value = picked.option.value;
          picked.el.dispatchEvent(new Event('input', { bubbles: true }));
          picked.el.dispatchEvent(new Event('change', { bubbles: true }));
          return { ok: true, selected: (picked.option.textContent || '').trim(), context: picked.context.slice(0, 220) };
        }
        """,
        {"optionText": option_text, "labelHint": label_hint},
    )
    if result.get("ok"):
        return result

    # Fallback for custom dropdowns.
    try:
        page.get_by_text(re.compile(re.escape(option_text), re.I)).first.click(timeout=1200)
        page.wait_for_timeout(500)
        return {"ok": True, "selected": option_text, "fallback": "text click"}
    except Exception:
        return result


def select_product_type(page, first_level: str, second_level: str) -> dict[str, Any]:
    return page.evaluate(
        """
        async ({ firstLevel, secondLevel }) => {
          const sleep = (ms) => new Promise(resolve => setTimeout(resolve, ms));
          const setByText = (select, text) => {
            if (!select || !text) return { ok: false, reason: 'missing select/text' };
            const options = Array.from(select.options || []);
            let option = options.find(opt => (opt.textContent || '').trim() === text);
            if (!option) option = options.find(opt => (opt.textContent || '').includes(text));
            if (!option) return { ok: false, reason: 'option not found', text, available: options.map(opt => (opt.textContent || '').trim()).slice(0, 80) };
            select.value = option.value;
            select.dispatchEvent(new Event('input', { bubbles: true }));
            select.dispatchEvent(new Event('change', { bubbles: true }));
            return { ok: true, selected: (option.textContent || '').trim(), value: option.value };
          };
          let selects = Array.from(document.querySelectorAll('#lnw_category_id select.drop_down'));
          if (selects.length < 1) return { ok: false, skipped: true, reason: 'product type selects not found on this page' };
          const level1 = setByText(selects[0], firstLevel);
          await sleep(1200);
          selects = Array.from(document.querySelectorAll('#lnw_category_id select.drop_down'));
          const level2 = setByText(selects[1], secondLevel);
          return { level_1: level1, level_2: level2 };
        }
        """,
        {"firstLevel": first_level, "secondLevel": second_level},
    )


def select_shop_category(page, category_chain: str) -> dict[str, Any]:
    if not category_chain:
        return {"ok": False, "reason": "empty category"}
    return page.evaluate(
        """
        ({ categoryChain }) => {
          const normalize = (text) => (text || '').replace(/\\s+/g, ' ').replace(/›/g, '>').replace(/â€º/g, '>').trim();
          const leafOf = (text) => normalize(text).split('>').map(x => x.trim()).filter(Boolean).pop();
          const loose = (text) => normalize(text).toLowerCase().replace(/[\\s\\u200B-\\u200D\\uFEFF]/g, '').replace(/\\u0e07/g, '');
          const leaf = leafOf(categoryChain);
          const select = document.querySelector('.shopCat select, div.shopCat select');
          if (!select) return { ok: false, reason: 'shop category select not found' };
          const options = Array.from(select.options || []);
          const normalizedChain = normalize(categoryChain);
          let option = options.find(opt => normalize(opt.textContent) === normalizedChain);
          if (!option) option = options.find(opt => normalize(opt.textContent).endsWith(normalizedChain));
          if (!option) {
            const looseChain = loose(normalizedChain);
            option = options.find(opt => loose(opt.textContent) === looseChain || loose(opt.textContent).endsWith(looseChain));
          }
          if (!option && leaf) option = options.find(opt => leafOf(opt.textContent) === leaf);
          if (!option && leaf) {
            const looseLeaf = loose(leaf);
            option = options.find(opt => loose(leafOf(opt.textContent)) === looseLeaf || loose(opt.textContent).includes(looseLeaf));
          }
          if (!option && leaf) option = options.find(opt => normalize(opt.textContent).includes(leaf));
          if (!option) return { ok: false, reason: 'category option not found', categoryChain, available: options.map(opt => normalize(opt.textContent)).slice(0, 120) };
          select.value = option.value;
          select.dispatchEvent(new Event('input', { bubbles: true }));
          select.dispatchEvent(new Event('change', { bubbles: true }));
          if (window.jQuery) {
            window.jQuery(select).trigger('chosen:updated').trigger('change');
          }
          return { ok: true, selected: normalize(option.textContent), value: option.value };
        }
        """,
        {"categoryChain": category_chain},
    )


def fill_rich_text_anywhere(page, value: str) -> dict[str, Any]:
    for frame in page.frames:
        try:
            result = frame.evaluate(
                """
                (value) => {
                  const visible = (el) => {
                    const rect = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
                  };
                  const el = Array.from(document.querySelectorAll('body[contenteditable="true"], [contenteditable="true"], textarea'))
                    .find(visible);
                  if (!el) return { ok: false };
                  el.focus();
                  if (el.tagName.toLowerCase() === 'textarea') {
                    el.value = value;
                  } else {
                    el.innerHTML = value.replace(/\\n/g, '<br>');
                  }
                  el.dispatchEvent(new Event('input', { bubbles: true }));
                  el.dispatchEvent(new Event('change', { bubbles: true }));
                  return { ok: true, frameUrl: window.location.href, tag: el.tagName.toLowerCase() };
                }
                """,
                value,
            )
            if result.get("ok"):
                return result
        except Exception:
            continue
    return {"ok": False, "reason": "no rich text editor found"}


def click_optional_tab(page, names: list[str]) -> bool:
    for name in names:
        try:
            page.get_by_text(re.compile(name, re.I)).first.click(timeout=1200)
            page.wait_for_timeout(500)
            return True
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue
    return False


def safe_goto(page, url: str, wait_ms: int = 1000) -> dict[str, Any]:
    last_error = ""
    for attempt in range(1, 4):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=45000)
            page.wait_for_timeout(wait_ms)
            return {"ok": True, "method": "direct-url", "url": page.url, "attempt": attempt}
        except PlaywrightError as exc:
            last_error = str(exc)
            if url in page.url or "interrupted by another navigation" in last_error:
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                except Exception:
                    page.wait_for_timeout(1200)
                if url in page.url:
                    page.wait_for_timeout(wait_ms)
                    return {
                        "ok": True,
                        "method": "direct-url-recovered",
                        "url": page.url,
                        "attempt": attempt,
                        "warning": last_error[:240],
                    }
            page.wait_for_timeout(1200)
    return {"ok": False, "method": "direct-url", "url": page.url, "reason": last_error[:500]}


def click_add_product(page) -> dict[str, Any]:
    if DEFAULT_ADD_PRODUCT_URL not in page.url:
        return safe_goto(page, DEFAULT_ADD_PRODUCT_URL)
    patterns = [
        r"^\s*\+\s*สินค้า\s*$",
        r"^\s*สินค้า\s*$",
        r"เพิ่มสินค้า",
        r"Add\s*Product",
        r"New\s*Product",
    ]
    for pattern in patterns:
        try:
            page.get_by_text(re.compile(pattern, re.I)).first.click(timeout=2000)
            page.wait_for_load_state("domcontentloaded", timeout=5000)
            page.wait_for_timeout(1000)
            return {"ok": True, "pattern": pattern, "url": page.url}
        except PlaywrightTimeoutError:
            continue
        except Exception as exc:
            return {"ok": False, "pattern": pattern, "reason": str(exc), "url": page.url}
    return {"ok": False, "reason": "add product button not found", "url": page.url}


def click_add_category(page) -> dict[str, Any]:
    if DEFAULT_ADD_CATEGORY_URL not in page.url:
        return safe_goto(page, DEFAULT_ADD_CATEGORY_URL)
    patterns = [
        r"^\s*\+\s*หมวดหมู่\s*$",
        r"^\s*หมวดหมู่\s*$",
        r"เพิ่มหมวดหมู่",
        r"Add\s*Category",
        r"New\s*Category",
    ]
    for pattern in patterns:
        try:
            page.get_by_text(re.compile(pattern, re.I)).first.click(timeout=2000)
            page.wait_for_load_state("domcontentloaded", timeout=5000)
            page.wait_for_timeout(1000)
            return {"ok": True, "pattern": pattern, "url": page.url}
        except PlaywrightTimeoutError:
            continue
        except Exception as exc:
            return {"ok": False, "pattern": pattern, "reason": str(exc), "url": page.url}
    return {"ok": False, "reason": "add category button not found", "url": page.url}


def click_save(page) -> dict[str, Any]:
    try:
        page.locator('input[type="submit"][value="บันทึก"], input.button.primary[value="บันทึก"]').first.click(timeout=2500)
        page.wait_for_load_state("domcontentloaded", timeout=8000)
        page.wait_for_timeout(1200)
        return {"ok": True, "method": "submit-input", "url": page.url}
    except Exception:
        pass

    patterns = [
        r"^\s*บันทึก\s*$",
        r"บันทึกข้อมูล",
        r"บันทึกสินค้า",
        r"บันทึกหมวดหมู่",
        r"^\s*Save\s*$",
        r"Save\s*Changes",
    ]
    for pattern in patterns:
        try:
            page.get_by_text(re.compile(pattern, re.I)).first.click(timeout=2500)
            page.wait_for_load_state("domcontentloaded", timeout=8000)
            page.wait_for_timeout(1200)
            return {"ok": True, "pattern": pattern, "url": page.url}
        except PlaywrightTimeoutError:
            continue
        except Exception as exc:
            return {"ok": False, "pattern": pattern, "reason": str(exc), "url": page.url}
    return {"ok": False, "reason": "save button not found", "url": page.url}


def kg_to_grams(value: str) -> str:
    if not value:
        return ""
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError:
        return value
    grams = int(round(number * 1000))
    return str(grams)


def fill_first_selector(page, selectors: list[str], value: str) -> dict[str, Any]:
    if value is None:
        value = ""
    return page.evaluate(
        """
        ({ selectors, value }) => {
          for (const selector of selectors) {
            const el = document.querySelector(selector);
            if (!el) continue;
            if (el.disabled || el.readOnly) continue;
            if (el.isContentEditable) {
              el.focus();
              el.innerText = value;
            } else {
              el.focus && el.focus();
              el.value = value;
            }
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            return { ok: true, selector, tag: el.tagName.toLowerCase() };
          }
          return { ok: false, reason: 'selector not found', selectors };
        }
        """,
        {"selectors": selectors, "value": value},
    )


def set_rich_text_by_id(page, editor_id: str, value: str) -> dict[str, Any]:
    return page.evaluate(
        """
        ({ editorId, value }) => {
          const methods = [];
          const escapeHtml = (text) => String(text)
            .replace(/&/g, '&amp;')
            .replace(/</g, '&lt;')
            .replace(/>/g, '&gt;')
            .replace(/"/g, '&quot;')
            .replace(/'/g, '&#039;');
          const html = String(value || '').split(/\\r?\\n/).map(escapeHtml).join('<br>');

          if (window.tinymce && window.tinymce.get(editorId)) {
            const editor = window.tinymce.get(editorId);
            editor.setContent(html);
            editor.fire('change');
            methods.push('tinymce');
          }

          const textarea = document.getElementById(editorId);
          if (textarea) {
            textarea.value = value || '';
            textarea.dispatchEvent(new Event('input', { bubbles: true }));
            textarea.dispatchEvent(new Event('change', { bubbles: true }));
            methods.push('textarea');
          }

          const iframe = document.getElementById(`${editorId}_ifr`);
          const doc = iframe && (iframe.contentDocument || iframe.contentWindow?.document);
          const body = doc && doc.body;
          if (body) {
            body.innerHTML = html;
            body.dispatchEvent(new Event('input', { bubbles: true }));
            body.dispatchEvent(new Event('change', { bubbles: true }));
            methods.push('iframe');
          }

          return methods.length ? { ok: true, editorId, methods } : { ok: false, editorId, reason: 'editor not found' };
        }
        """,
        {"editorId": editor_id, "value": value or ""},
    )


def ensure_product_pro_mode(page) -> dict[str, Any]:
    return page.evaluate(
        """
        async () => {
          const sleep = (ms) => new Promise(resolve => setTimeout(resolve, ms));
          const hasProFields = () => Array.from(document.querySelectorAll('input, textarea, select')).some((el) => {
            const parent = el.closest('label, .edit_field, .gadget, .field, .row, div');
            const text = ((parent && parent.innerText) || '').replace(/\\s+/g, ' ');
            return text.includes('รหัสสินค้า') || text.includes('น้ำหนัก') || text.includes('บาร์โค้ด');
          });
          if (hasProFields()) return { ok: true, mode: 'pro-fields-visible' };

          const switchEl = document.querySelector('.edit_mode_gadget .switch, .edit_mode_gadget .switchery, .edit_mode_gadget [role="switch"]');
          if (!switchEl) return { ok: false, reason: 'mode switch not found' };
          switchEl.click();
          await sleep(1200);
          return { ok: hasProFields(), mode: hasProFields() ? 'switched-to-pro' : 'switch-clicked-no-pro-fields' };
        }
        """
    )


def select_category_parent(page, parent_name: str) -> dict[str, Any]:
    if not parent_name or parent_name == "ไม่มีหมวดหมู่หลัก":
        return {"ok": True, "skipped": True, "reason": "no parent category"}
    return page.evaluate(
        """
        ({ parentName }) => {
          const normalize = (text) => (text || '')
            .replace(/\\u00a0/g, ' ')
            .replace(/^[\\s\\-]+/, '')
            .replace(/\\s+/g, ' ')
            .trim();
          const select = document.querySelector('select[varname="parent_id"]');
          if (!select) return { ok: false, reason: 'parent select not found' };
          const options = Array.from(select.options || []);
          let option = options.find(opt => normalize(opt.textContent) === parentName);
          if (!option) option = options.find(opt => normalize(opt.textContent).endsWith(parentName));
          if (!option) option = options.find(opt => normalize(opt.textContent).includes(parentName));
          if (!option) return { ok: false, reason: 'parent option not found', parentName, available: options.map(opt => normalize(opt.textContent)).slice(0, 120) };
          select.value = option.value;
          select.dispatchEvent(new Event('input', { bubbles: true }));
          select.dispatchEvent(new Event('change', { bubbles: true }));
          return { ok: true, selected: normalize(option.textContent), value: option.value };
        }
        """,
        {"parentName": parent_name},
    )


def fill_tags(page, tags: str) -> dict[str, Any]:
    if not tags:
        return {"ok": False, "reason": "empty tags"}
    return page.evaluate(
        """
        ({ tags }) => {
          const input = document.querySelector('.tagContainer input, input[placeholder*="Keywords"]');
          if (!input) return { ok: false, reason: 'tag input not found' };
          input.focus();
          input.value = tags;
          input.dispatchEvent(new Event('input', { bubbles: true }));
          input.dispatchEvent(new Event('change', { bubbles: true }));
          input.dispatchEvent(new KeyboardEvent('keydown', { bubbles: true, key: 'Enter', code: 'Enter' }));
          input.dispatchEvent(new KeyboardEvent('keyup', { bubbles: true, key: 'Enter', code: 'Enter' }));
          return { ok: true };
        }
        """,
        {"tags": tags},
    )


def fill_product_seo(page, seo_title: str, seo_description: str, seo_keywords: str) -> dict[str, Any]:
    return page.evaluate(
        """
        ({ seoTitle, seoDescription, seoKeywords }) => {
          const setValue = (el, value) => {
            if (!el || value === undefined || value === null || value === '') return false;
            el.focus && el.focus();
            el.value = value;
            el.dispatchEvent(new Event('input', { bubbles: true }));
            el.dispatchEvent(new Event('change', { bubbles: true }));
            return true;
          };
          const result = {};
          result.title = setValue(document.querySelector('#g-seo .seo-title input, .seoContent .seo-title input'), seoTitle);
          result.description = setValue(document.querySelector('#g-seo .seo-description input, .seoContent .seo-description input'), seoDescription);
          result.keywords = setValue(document.querySelector('#g-seo .seo-keyword input, .seoContent .seo-keyword input'), seoKeywords);

          if (!result.title || !result.description || !result.keywords) {
            const inputs = Array.from(document.querySelectorAll('#g-seo input[type="text"], .seoContent input[type="text"]'));
            if (!result.title) result.title = setValue(inputs[0], seoTitle);
            if (!result.description) result.description = setValue(inputs[1], seoDescription);
            if (!result.keywords) result.keywords = setValue(inputs[2], seoKeywords);
          }
          return { ok: Boolean(result.title || result.description || result.keywords), ...result };
        }
        """,
        {"seoTitle": seo_title, "seoDescription": seo_description, "seoKeywords": seo_keywords},
    )


def upload_dialog_state(page) -> dict[str, Any]:
    return page.evaluate(
        """
        () => {
          const visible = (el) => {
            const rect = el.getBoundingClientRect();
            const style = window.getComputedStyle(el);
            return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
          };
          const textOf = (el) => `${el.innerText || ''} ${el.value || ''} ${el.getAttribute('aria-label') || ''}`
            .replace(/\\s+/g, ' ')
            .trim();
          const visibleText = Array.from(document.querySelectorAll('body *'))
            .filter(visible)
            .map(textOf)
            .filter(Boolean)
            .join(' ');
          const lowerText = visibleText.toLowerCase();
          const busy = lowerText.includes('กำลังอัพโหลด') || lowerText.includes('uploading') || lowerText.includes('please wait');
          const confirmButtons = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a, .button'))
            .filter(visible)
            .map((el) => ({ text: textOf(el), className: String(el.className || '') }))
            .filter((item) => item.className.includes('btn-confirm-selected') || /^เพิ่ม\\s*\\(\\d+\\)$/.test(item.text) || /^Add\\s*\\(\\d+\\)$/i.test(item.text));
          return {
            ok: !busy && confirmButtons.length > 0,
            busy,
            confirmCount: confirmButtons.length,
            confirmText: confirmButtons[0]?.text || '',
            visibleText: visibleText.slice(0, 300),
          };
        }
        """
    )


def wait_upload_ready(page, timeout_ms: int = 90000) -> dict[str, Any]:
    attempts = max(1, timeout_ms // 500)
    last_state: dict[str, Any] = {}
    for _ in range(attempts):
        last_state = upload_dialog_state(page)
        if last_state.get("ok"):
            return last_state
        page.wait_for_timeout(500)
    return {"ok": False, "reason": "upload dialog not ready before timeout", "state": last_state}


def wait_upload_dialog_closed(page, timeout_ms: int = 20000) -> dict[str, Any]:
    attempts = max(1, timeout_ms // 500)
    last_state: dict[str, Any] = {}
    for _ in range(attempts):
        last_state = upload_dialog_state(page)
        if not last_state.get("busy") and int(last_state.get("confirmCount") or 0) < 1:
            return {"ok": True}
        page.wait_for_timeout(500)
    return {"ok": False, "reason": "upload dialog did not close after confirm", "state": last_state}


def click_upload_confirm(page) -> dict[str, Any]:
    ready = wait_upload_ready(page)
    if not ready.get("ok"):
        return ready
    for _ in range(24):
        result = page.evaluate(
            """
            () => {
              const visible = (el) => {
                const rect = el.getBoundingClientRect();
                const style = window.getComputedStyle(el);
                return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
              };
              const textOf = (el) => `${el.innerText || ''} ${el.value || ''} ${el.getAttribute('aria-label') || ''}`
                .replace(/\\s+/g, ' ')
                .trim();
              const candidates = Array.from(document.querySelectorAll('button, input[type="button"], input[type="submit"], a, .button'))
                .filter(visible)
                .map((el) => ({ el, text: textOf(el), rect: el.getBoundingClientRect(), className: String(el.className || '') }))
                .filter((item) => item.className.includes('btn-confirm-selected') || /^เพิ่ม\\s*\\(\\d+\\)$/.test(item.text) || /^Add\\s*\\(\\d+\\)$/i.test(item.text));
              if (candidates.length < 1) return { ok: false, reason: 'confirm button not found' };
              candidates.sort((a, b) => {
                const aScore = (a.text.includes('(') ? 10 : 0) + (a.className.includes('primary') ? 3 : 0) + a.rect.left / 1000 + a.rect.top / 1000;
                const bScore = (b.text.includes('(') ? 10 : 0) + (b.className.includes('primary') ? 3 : 0) + b.rect.left / 1000 + b.rect.top / 1000;
                return bScore - aScore;
              });
              candidates[0].el.click();
              return { ok: true, text: candidates[0].text, className: candidates[0].className };
            }
            """
        )
        if result.get("ok"):
            result["ready"] = ready
            result["closed"] = wait_upload_dialog_closed(page)
            return result
        page.wait_for_timeout(500)
    return {"ok": False, "reason": "upload confirm button not found after waiting"}


def upload_image(page, image_path: str) -> dict[str, Any]:
    if not image_path:
        return {"ok": False, "reason": "no image path"}
    path = Path(image_path)
    if not path.exists():
        return {"ok": False, "reason": f"missing file: {path}"}
    try:
        inputs = page.locator('input[type="file"]')
        count = inputs.count()
        if count < 1:
            for selector in [
                ".add_image_area",
                "button.add_image_area",
                "#thumbnailImages button",
                "#thumbnailImages a",
                "#thumbnailImages .button",
            ]:
                trigger = page.locator(selector)
                try:
                    if trigger.count() < 1:
                        continue
                    with page.expect_file_chooser(timeout=2200) as chooser_info:
                        trigger.first.click(timeout=2200)
                    chooser_info.value.set_files(str(path))
                    confirm = click_upload_confirm(page)
                    return {
                        "ok": bool(confirm.get("ok") and confirm.get("closed", {}).get("ok")),
                        "file": str(path),
                        "method": "file-chooser",
                        "selector": selector,
                        "confirm": confirm,
                    }
                except PlaywrightTimeoutError:
                    page.wait_for_timeout(500)
                    break
                except Exception:
                    continue

            click_result = page.evaluate(
                """
                () => {
                  const visible = (el) => {
                    const rect = el.getBoundingClientRect();
                    const style = window.getComputedStyle(el);
                    return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none';
                  };
                  const direct = document.querySelector('.add_image_area, #thumbnailImages .button, #thumbnailImages a, #thumbnailImages button');
                  if (direct) {
                    direct.click();
                    return { ok: true, selector: 'direct image control' };
                  }
                  const labels = ['เลือกภาพ', 'Upload ภาพใหม่', 'เพิ่มรูป', 'รูปภาพ', 'Image'];
                  const candidates = Array.from(document.querySelectorAll('button, input[type="button"], a, label, div, span')).filter(visible);
                  for (const el of candidates) {
                    const text = `${el.innerText || ''} ${el.value || ''}`.replace(/\\s+/g, ' ').trim();
                    if (labels.some(label => text.includes(label))) {
                      el.click();
                      return { ok: true, text };
                    }
                  }
                  return { ok: false, reason: 'image upload trigger not found' };
                }
                """
            )
            page.wait_for_timeout(800)
            inputs = page.locator('input[type="file"]')
            count = inputs.count()
            if count < 1:
                return {"ok": False, "reason": "no file input found", "click": click_result}
        inputs.nth(count - 1).set_input_files(str(path))
        confirm = click_upload_confirm(page)
        return {"ok": bool(confirm.get("ok") and confirm.get("closed", {}).get("ok")), "file": str(path), "confirm": confirm}
    except Exception as exc:
        return {"ok": False, "reason": str(exc)}


def fill_product(page, product: dict[str, str]) -> dict[str, Any]:
    results: dict[str, Any] = {}
    results["pro_mode"] = ensure_product_pro_mode(page)

    results["name_th"] = fill_first_selector(page, [".a-name input", "input.input-text.not_disabled"], product.get("name_th", ""))
    if not results["name_th"].get("ok"):
        results["name_th_fallback"] = fill_by_context(page, LABELS.get("name_th", ["name_th"]), product.get("name_th", ""))

    results["retail_price"] = fill_first_selector(page, [".real-price input", "input.has-unit-right"], "0")
    results["discount_price"] = fill_first_selector(page, [".discount_true input[type='number']", ".sale input[type='number']"], "0")

    short_desc = product.get("short_description", "")
    if short_desc:
        results["short_description"] = set_rich_text_by_id(page, "tinymce_short_description", short_desc)

    full_desc = product.get("full_description", "")
    if full_desc:
        results["full_description"] = set_rich_text_by_id(page, "tinymce_description", full_desc)
        if not results["full_description"].get("ok"):
            results["full_description_rich_text_fallback"] = fill_rich_text_anywhere(page, full_desc)

    if product.get("tags", ""):
        results["tags"] = fill_tags(page, product.get("tags", ""))

    fields = [
        ("sku", product.get("sku", ""), False),
        ("barcode", product.get("barcode", ""), False),
        ("brand", product.get("brand", ""), False),
        ("stock", product.get("stock", "0") or "0", False),
        ("weight", kg_to_grams(product.get("weight", "")), False),
        ("width", product.get("width", ""), False),
        ("length", product.get("length", ""), False),
        ("height", product.get("height", ""), False),
    ]
    for key, value, multiline in fields:
        if not value:
            continue
        results[key] = fill_by_context(page, LABELS.get(key, [key]), value, multiline)

    results["shop_category"] = select_shop_category(page, product.get("shop_category", ""))
    results["product_type"] = select_product_type(
        page,
        product.get("product_type_1", "สินค้าอุตสาหกรรม"),
        product.get("product_type_2", "อุปกรณ์ใช้ในโรงงานอุตสาหกรรม"),
    )

    results["seo"] = fill_product_seo(
        page,
        product.get("seo_title", ""),
        product.get("seo_description", ""),
        product.get("seo_keywords", ""),
    )

    clicked_image = click_optional_tab(page, ["รูปภาพ", "ภาพสินค้า", "Images", "Media"])
    results["image_tab_clicked"] = clicked_image
    results["image"] = upload_image(page, product.get("image_abs", ""))
    return results


def fill_category(page, category: dict[str, str]) -> dict[str, Any]:
    results: dict[str, Any] = {}
    results["category_name"] = fill_first_selector(page, ['input[varname="name"]'], category.get("name", ""))
    if not results["category_name"].get("ok"):
        results["category_name_fallback"] = fill_by_context(page, LABELS["category_name"], category.get("name", ""))
    results["category_parent"] = select_category_parent(page, category.get("parent", ""))
    desc = category.get("description", "")
    results["category_description"] = set_rich_text_by_id(page, "tinymce_desc", desc)
    if not results["category_description"].get("ok"):
        results["category_description_fallback"] = fill_by_context(page, LABELS["category_description"], desc, prefer_multiline=True)
    results["seo_title"] = fill_first_selector(page, ['input[varname="seo-title"]'], category.get("seo_title", ""))
    results["seo_description"] = fill_first_selector(page, ['input[varname="seo-description"]'], category.get("seo_description", ""))
    results["seo_keywords"] = fill_first_selector(page, ['input[varname="seo-keyword"]'], category.get("seo_keywords", ""))

    clicked_image = click_optional_tab(page, ["ภาพหมวดหมู่", "รูปภาพ", "Images", "Media"])
    results["image_tab_clicked"] = clicked_image
    results["image"] = upload_image(page, category.get("image_abs", ""))

    return results


def product_auto_save_blockers(results: dict[str, Any]) -> list[str]:
    blockers: list[str] = []
    if not results.get("shop_category", {}).get("ok"):
        blockers.append("shop_category")
    product_type = results.get("product_type", {})
    if not product_type.get("level_1", {}).get("ok") or not product_type.get("level_2", {}).get("ok"):
        blockers.append("product_type")
    if not results.get("image", {}).get("ok"):
        blockers.append("image")
    return blockers


def run_browser(args, action) -> None:
    DEFAULT_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    args.profile.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(args.profile),
            channel="msedge",
            headless=False,
            viewport={"width": 1440, "height": 950},
            args=["--start-maximized"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        start_result = safe_goto(page, args.url or DEFAULT_START_URL)
        if not start_result.get("ok"):
            raise RuntimeError(f"Could not open start URL: {start_result}")
        action(page)
        context.close()


def command_inspect(args) -> None:
    def action(page) -> None:
        print("Edge automation window is open.")
        print("Login to LnwShop in that window if needed, then return here and press Enter.")
        input()
        add_result = {"skipped": True}
        if args.click_add_product:
            add_result = click_add_product(page)
            print(f"add product click: {add_result}")
            if not add_result.get("ok"):
                print("If the form is not open yet, click '+ สินค้า' in Edge, then press Enter.")
                input()
        page.wait_for_timeout(1000)
        data = {
            "url": page.url,
            "title": page.title(),
            "add_product_click": add_result,
            "fields": visible_field_info(page),
            "buttons": visible_buttons(page),
        }
        DEFAULT_INSPECT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        screenshot = DEFAULT_SCREENSHOT_DIR / "lnwshop_inspect.png"
        page.screenshot(path=str(screenshot), full_page=True)
        print(f"saved inspect: {DEFAULT_INSPECT}")
        print(f"saved screenshot: {screenshot}")
        print(f"visible fields: {len(data['fields'])}")
        print(f"visible buttons/links: {len(data['buttons'])}")

    run_browser(args, action)


def command_fill(args) -> None:
    product = read_product(args.excel, args.row)
    print_product(product)

    def action(page) -> None:
        if args.wait_before_fill:
            print("Make sure the inventory page is ready, then press Enter to continue.")
            input()
        if args.click_add_product:
            add_result = click_add_product(page)
            print(f"add product click: {add_result}")
            if not add_result.get("ok"):
                print("If the product form is not open yet, click '+ สินค้า' in Edge, then press Enter to fill.")
                input()
        results = fill_product(page, product)
        if args.auto_save:
            blockers = product_auto_save_blockers(results)
            if blockers:
                results["save"] = {"ok": False, "skipped": True, "reason": "auto-save blocked", "blockers": blockers}
            else:
                results["save"] = click_save(page)
        result_path = Path(__file__).resolve().parent / "last_fill_result.json"
        result_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        screenshot = DEFAULT_SCREENSHOT_DIR / f"lnwshop_fill_row_{args.row}.png"
        page.screenshot(path=str(screenshot), full_page=True)
        print(f"saved fill result: {result_path}")
        print(f"saved screenshot: {screenshot}")
        if args.auto_save:
            if results.get("save", {}).get("skipped"):
                print(f"Auto-save skipped because: {', '.join(results['save']['blockers'])}")
            else:
                print("Auto-save mode completed. Review the page in Edge, then press Enter here when done.")
                input()
        else:
            print("Manual-save mode: review the page in Edge. Press Enter here when done.")
            input()

    run_browser(args, action)


def command_inspect_category(args) -> None:
    def action(page) -> None:
        print("Edge automation window is open.")
        print("Login to LnwShop in that window if needed, then return here and press Enter.")
        input()
        add_result = {"skipped": True}
        if args.click_add_category:
            add_result = click_add_category(page)
            print(f"add category click: {add_result}")
            if not add_result.get("ok"):
                print("If the category form is not open yet, click '+ หมวดหมู่' in Edge, then press Enter.")
                input()
        page.wait_for_timeout(1000)
        data = {
            "url": page.url,
            "title": page.title(),
            "add_category_click": add_result,
            "fields": visible_field_info(page),
            "buttons": visible_buttons(page),
        }
        inspect_path = Path(__file__).resolve().parent / "last_category_inspect.json"
        inspect_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        screenshot = DEFAULT_SCREENSHOT_DIR / "lnwshop_category_inspect.png"
        page.screenshot(path=str(screenshot), full_page=True)
        print(f"saved inspect: {inspect_path}")
        print(f"saved screenshot: {screenshot}")
        print(f"visible fields: {len(data['fields'])}")
        print(f"visible buttons/links: {len(data['buttons'])}")

    run_browser(args, action)


def command_fill_category(args) -> None:
    category = read_category(args.category_excel, args.row)
    print_category(category)

    def action(page) -> None:
        if args.wait_before_fill:
            print("Make sure the inventory page is ready, then press Enter to continue.")
            input()
        if args.click_add_category:
            add_result = click_add_category(page)
            print(f"add category click: {add_result}")
            if not add_result.get("ok"):
                print("If the category form is not open yet, click '+ หมวดหมู่' in Edge, then press Enter to fill.")
                input()
        results = fill_category(page, category)
        if args.auto_save:
            results["save"] = click_save(page)
        result_path = Path(__file__).resolve().parent / "last_category_fill_result.json"
        result_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        screenshot = DEFAULT_SCREENSHOT_DIR / f"lnwshop_category_fill_row_{args.row}.png"
        page.screenshot(path=str(screenshot), full_page=True)
        print(f"saved fill result: {result_path}")
        print(f"saved screenshot: {screenshot}")
        if args.auto_save:
            print("Auto-save mode completed. Review the page in Edge, then press Enter here when done.")
            input()
        else:
            print("Manual-save mode: review the page in Edge. Press Enter here when done.")
            input()

    run_browser(args, action)


def sheet_last_row(excel_path: Path, sheet_name: str) -> int:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    return ws.max_row


def command_fill_categories(args) -> None:
    end_row = args.end_row or sheet_last_row(args.category_excel, "Categories_AI_Ready")
    print(f"Category batch: rows {args.start_row}-{end_row}")

    def action(page) -> None:
        print("Make sure the inventory page is ready, then press Enter to start category batch.")
        input()
        for row_number in range(args.start_row, end_row + 1):
            category = read_category(args.category_excel, row_number)
            print("\n" + "=" * 72)
            print_category(category)
            add_result = click_add_category(page)
            print(f"add category click: {add_result}")
            if not add_result.get("ok"):
                print("Open '+ หมวดหมู่' manually in Edge, then press Enter to fill this row.")
                input()
            results = fill_category(page, category)
            if args.auto_save:
                results["save"] = click_save(page)
            result_path = Path(__file__).resolve().parent / f"last_category_fill_row_{row_number}.json"
            result_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
            screenshot = DEFAULT_SCREENSHOT_DIR / f"lnwshop_category_fill_row_{row_number}.png"
            page.screenshot(path=str(screenshot), full_page=True)
            print(f"saved fill result: {result_path}")
            print(f"saved screenshot: {screenshot}")
            if args.auto_save:
                if args.no_pause:
                    continue
                print("Auto-save mode: press Enter for the next row, or Ctrl+C to stop.")
            else:
                print("Manual-save mode: review/save in Edge, then press Enter for the next row, or Ctrl+C to stop.")
            input()

    run_browser(args, action)


def command_fill_products(args) -> None:
    end_row = args.end_row or sheet_last_row(args.excel, "LnwShop_AI_Ready")
    print(f"Product batch: rows {args.start_row}-{end_row}")

    def action(page) -> None:
        print("Make sure the inventory page is ready, then press Enter to start product batch.")
        input()
        for row_number in range(args.start_row, end_row + 1):
            product = read_product(args.excel, row_number)
            print("\n" + "=" * 72)
            print_product(product)
            add_result = click_add_product(page)
            print(f"add product click: {add_result}")
            if not add_result.get("ok"):
                print("Open '+ สินค้า' manually in Edge, then press Enter to fill this row.")
                input()
            results = fill_product(page, product)
            if args.auto_save:
                blockers = product_auto_save_blockers(results)
                if blockers:
                    results["save"] = {"ok": False, "skipped": True, "reason": "auto-save blocked", "blockers": blockers}
                else:
                    results["save"] = click_save(page)
            result_path = Path(__file__).resolve().parent / f"last_product_fill_row_{row_number}.json"
            result_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
            screenshot = DEFAULT_SCREENSHOT_DIR / f"lnwshop_product_fill_row_{row_number}.png"
            page.screenshot(path=str(screenshot), full_page=True)
            print(f"saved fill result: {result_path}")
            print(f"saved screenshot: {screenshot}")
            if args.auto_save and results.get("save", {}).get("skipped"):
                raise RuntimeError(f"Stopped before save at row {row_number}: {', '.join(results['save']['blockers'])}")
            if args.auto_save:
                if args.no_pause:
                    continue
                print("Auto-save mode: press Enter for the next row, or Ctrl+C to stop.")
            else:
                print("Manual-save mode: review/save in Edge, then press Enter for the next row, or Ctrl+C to stop.")
            input()

    run_browser(args, action)


def command_dry_run(args) -> None:
    product = read_product(args.excel, args.row)
    print_product(product)


def command_dry_run_category(args) -> None:
    category = read_category(args.category_excel, args.row)
    print_category(category)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fill LnwShop product form from prepared Excel.")
    sub = parser.add_subparsers(dest="command", required=True)

    def add_common(p: argparse.ArgumentParser) -> None:
        p.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
        p.add_argument("--category-excel", type=Path, default=DEFAULT_CATEGORY_EXCEL)
        p.add_argument("--row", type=int, default=2, help="Excel row number in LnwShop_AI_Ready sheet.")
        p.add_argument("--profile", type=Path, default=DEFAULT_PROFILE)
        p.add_argument("--url", default=os.environ.get("LNWSHOP_URL", DEFAULT_START_URL))

    inspect = sub.add_parser("inspect", help="Open Edge and dump visible form fields/buttons.")
    add_common(inspect)
    inspect.add_argument("--no-click-add-product", action="store_false", dest="click_add_product")
    inspect.set_defaults(click_add_product=True)
    inspect.set_defaults(func=command_inspect)

    fill = sub.add_parser("fill", help="Fill one product row and stop before saving.")
    add_common(fill)
    fill.add_argument("--wait-before-fill", action="store_true", default=True)
    fill.add_argument("--no-click-add-product", action="store_false", dest="click_add_product")
    fill.add_argument("--auto-save", action="store_true", help="Click the save button after filling. Use only after manual tests pass.")
    fill.set_defaults(click_add_product=True)
    fill.set_defaults(func=command_fill)

    dry = sub.add_parser("dry-run", help="Read one product row without opening browser.")
    add_common(dry)
    dry.set_defaults(func=command_dry_run)

    inspect_category = sub.add_parser("inspect-category", help="Open Edge and dump visible category form fields/buttons.")
    add_common(inspect_category)
    inspect_category.add_argument("--no-click-add-category", action="store_false", dest="click_add_category")
    inspect_category.set_defaults(click_add_category=True)
    inspect_category.set_defaults(func=command_inspect_category)

    fill_category_parser = sub.add_parser("fill-category", help="Fill one category row and stop before saving.")
    add_common(fill_category_parser)
    fill_category_parser.add_argument("--wait-before-fill", action="store_true", default=True)
    fill_category_parser.add_argument("--no-click-add-category", action="store_false", dest="click_add_category")
    fill_category_parser.add_argument("--auto-save", action="store_true", help="Click the save button after filling. Use only after manual tests pass.")
    fill_category_parser.set_defaults(click_add_category=True)
    fill_category_parser.set_defaults(func=command_fill_category)

    dry_category = sub.add_parser("dry-run-category", help="Read one category row without opening browser.")
    add_common(dry_category)
    dry_category.set_defaults(func=command_dry_run_category)

    fill_categories = sub.add_parser("fill-categories", help="Fill a range of category rows in one browser session.")
    add_common(fill_categories)
    fill_categories.add_argument("--start-row", type=int, default=2)
    fill_categories.add_argument("--end-row", type=int)
    fill_categories.add_argument("--auto-save", action="store_true", help="Click save after each row. Use only after manual tests pass.")
    fill_categories.add_argument("--no-pause", action="store_true", help="Do not pause between rows when auto-save is enabled.")
    fill_categories.set_defaults(func=command_fill_categories)

    fill_products = sub.add_parser("fill-products", help="Fill a range of product rows in one browser session.")
    add_common(fill_products)
    fill_products.add_argument("--start-row", type=int, default=2)
    fill_products.add_argument("--end-row", type=int)
    fill_products.add_argument("--auto-save", action="store_true", help="Click save after each row. Use only after manual tests pass.")
    fill_products.add_argument("--no-pause", action="store_true", help="Do not pause between rows when auto-save is enabled.")
    fill_products.set_defaults(func=command_fill_products)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if hasattr(args, "category_excel") and args.command in {"inspect-category", "fill-category", "dry-run-category", "fill-categories"} and not args.category_excel.exists():
        print(f"Category Excel file not found: {args.category_excel}", file=sys.stderr)
        return 2
    if not args.excel.exists():
        print(f"Excel file not found: {args.excel}", file=sys.stderr)
        return 2
    args.func(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
