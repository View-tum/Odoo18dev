from __future__ import annotations

import re
from copy import copy
from collections import OrderedDict
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

from prepare_assets import apply_watermark, crop_watermark


ROOT = Path(__file__).resolve().parents[2]
SOURCE_EXCEL = ROOT / "automation" / "ATHENS Price HIKFIRE2026 .xlsx"
WATERMARK = Path(__file__).resolve().parent / "สำเนาของ watermark (15).png"
IMAGE_DIR = ROOT / "output" / "spreadsheet" / "hikfire_product_images"
WATERMARKED_DIR = ROOT / "output" / "spreadsheet" / "hikfire_product_images_watermarked"
PRODUCT_EXCEL = ROOT / "output" / "spreadsheet" / "HIKFIRE_LnwShop_AI_ready_watermarked.xlsx"
CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "HIKFIRE_LnwShop_categories.xlsx"
REVIEW_EXCEL = ROOT / "output" / "spreadsheet" / "HIKFIRE_LnwShop_review.xlsx"
BRAND = "HIKFIRE"
DEFAULT_PRODUCT_TYPE_1 = "สินค้าอุตสาหกรรม"
DEFAULT_PRODUCT_TYPE_2 = "อุปกรณ์ใช้ในโรงงานอุตสาหกรรม"


PRODUCT_HEADERS = [
    "ลำดับ",
    "สถานะลงสินค้า",
    "เผยแพร่บนร้าน",
    "แบรนด์",
    "หมวดหมู่หลัก",
    "หมวดหมู่ย่อย",
    "หมวดหมู่แนะนำใน LnwShop",
    "ชื่อสินค้า (ไทย)",
    "Product Name (EN)",
    "SKU / Article Code",
    "Vendor Part No.",
    "Model",
    "EAN/UPC",
    "ราคาปกติ (บาท)",
    "ราคาขาย (บาท)",
    "ราคาพิเศษ (บาท)",
    "จำนวนสต๊อกตั้งต้น",
    "สถานะสต๊อก",
    "น้ำหนักรวม (kg)",
    "กว้าง (cm)",
    "ยาว (cm)",
    "สูง (cm)",
    "ขนาดรวม",
    "การรับประกัน",
    "รายละเอียดสั้น (ไทย)",
    "รายละเอียดเต็ม (ไทย)",
    "จุดเด่นสินค้า (ไทย)",
    "Product Detail (EN)",
    "SEO Title (TH)",
    "SEO Title (EN)",
    "SEO Title (TH+EN)",
    "SEO Keywords (TH+EN)",
    "SEO Meta Description (TH+EN)",
    "URL Slug",
    "Search Tags",
    "รูปภาพหลัก/ไฟล์ภาพ",
    "หมายเหตุสำหรับลงสินค้า",
    "ข้อมูลสเปกต้นฉบับ",
    "แหล่งข้อมูลในไฟล์ต้นฉบับ",
    "รูปภาพต้นฉบับ",
    "รูปภาพพร้อมลายน้ำ",
    "หมวดหมู่แนะนำเดิม",
    "Product Type ระดับ 1",
    "Product Type ระดับ 2",
    "ราคาต้นฉบับ Dealer",
    "ราคาต้นฉบับ SI",
    "ราคาต้นฉบับ MSRP",
]


CATEGORY_HEADERS = [
    "ลำดับ",
    "สถานะ",
    "ชื่อหมวดหมู่",
    "หมวดหมู่ย่อยของ",
    "รายละเอียดหมวดหมู่",
    "แสดงหมวดหมู่ที่หน้าร้าน",
    "Product Type ระดับ 1",
    "Product Type ระดับ 2",
    "รูปภาพหมวดหมู่",
    "SEO Title",
    "SEO Keywords",
    "SEO Meta Description",
    "แหล่งข้อมูล",
]


IMAGE_BY_ROW = {
    3: 19,
    4: 18,
    5: 4,
    6: 2,
    7: 3,
    8: 1,
    9: 5,
    10: 6,
    11: 7,
    12: 8,
    13: 9,
    14: 10,
    15: 11,
    16: 12,
    17: 13,
    18: 14,
    19: 15,
    20: 16,
    21: 17,
}


THAI_NAMES = {
    "HF-VH221(2.8mm)": "กล้องตรวจจับความร้อน HIKFIRE 2MP รุ่น HF-VH221 เลนส์ 2.8mm",
    "HF-VH243(2.8mm)": "กล้องตรวจจับความร้อน HIKFIRE 4MP รุ่น HF-VH243 เลนส์ 2.8mm",
    "HF-VR343(4mm)": "กล้องตรวจจับเปลวไฟ Dual-IR HIKFIRE 4MP รุ่น HF-VR343 เลนส์ 4mm",
    "HF-VT343(3mm)": "กล้อง Thermal Imaging ตรวจจับไฟ AI HIKFIRE 4MP แบบ Turret รุ่น HF-VT343 เลนส์ 3mm",
    "HF-VTR344(3mm)": "กล้อง Thermal Imaging ตรวจจับไฟ AI Triple-IR HIKFIRE 4MP แบบ Bullet รุ่น HF-VTR344 เลนส์ 3mm",
    "NP-FP105-B": "ขายึดผนัง HIKFIRE รุ่น NP-FP105-B สำหรับติดตั้งกล้องและอุปกรณ์",
    "HF-S2": "เครื่องตรวจจับควัน HIKFIRE รุ่น HF-S2 แบตเตอรี่ AA 3 ปี พร้อม I/O",
    "HF-S2E": "เครื่องตรวจจับควัน HIKFIRE รุ่น HF-S2E แบตเตอรี่ AA 3 ปี",
    "HF-S3E": "เครื่องตรวจจับควัน HIKFIRE รุ่น HF-S3E แบตเตอรี่ลิเธียม 10 ปี",
    "HF-S3E-R": "เครื่องตรวจจับควันไร้สาย 433MHz HIKFIRE รุ่น HF-S3E-R",
    "HF-S3-WF": "เครื่องตรวจจับควัน Wi-Fi HIKFIRE รุ่น HF-S3-WF ดูแจ้งเตือนผ่าน Hik-Connect",
    "HF-SG3-WF/R": "เครื่องตรวจจับควัน Wi-Fi และ 433MHz HIKFIRE รุ่น HF-SG3-WF/R",
    "HF-C104": "ตู้ควบคุมแจ้งเหตุเพลิงไหม้ Conventional HIKFIRE รุ่น HF-C104 รองรับ 128 อุปกรณ์",
    "HF-C108": "ตู้ควบคุมแจ้งเหตุเพลิงไหม้ Conventional HIKFIRE รุ่น HF-C108 รองรับ 256 อุปกรณ์",
    "HF-CS1": "อุปกรณ์ตรวจจับควัน Conventional HIKFIRE รุ่น HF-CS1 เทคโนโลยี Dual Light พร้อมฐาน",
    "HF-CH1": "อุปกรณ์ตรวจจับความร้อน Conventional HIKFIRE รุ่น HF-CH1 พร้อมฐาน",
    "HF-CMP1": "ปุ่มกดแจ้งเหตุเพลิงไหม้ Conventional HIKFIRE รุ่น HF-CMP1",
    "HF-CSS1": "อุปกรณ์ Sounder Strobe Conventional HIKFIRE รุ่น HF-CSS1",
    "HF-CAB1": "กระดิ่งแจ้งเหตุเพลิงไหม้ Conventional HIKFIRE รุ่น HF-CAB1",
}


ENGLISH_NAMES = {
    "HF-VH221(2.8mm)": "HIKFIRE HF-VH221 2MP Heat Sensing Camera 2.8mm",
    "HF-VH243(2.8mm)": "HIKFIRE HF-VH243 4MP Heat Sensing Camera 2.8mm",
    "HF-VR343(4mm)": "HIKFIRE HF-VR343 4MP Dual-IR Flame Detection Camera 4mm",
    "HF-VT343(3mm)": "HIKFIRE HF-VT343 4MP Thermal Imaging AI Fire Detection Turret Camera 3mm",
    "HF-VTR344(3mm)": "HIKFIRE HF-VTR344 4MP Thermal Imaging AI Fire Detection Bullet Camera 3mm",
    "NP-FP105-B": "HIKFIRE NP-FP105-B Wall Mount Bracket",
    "HF-S2": "HIKFIRE HF-S2 3-Year AA Battery Smoke Alarm with I/O",
    "HF-S2E": "HIKFIRE HF-S2E 3-Year AA Battery Smoke Alarm",
    "HF-S3E": "HIKFIRE HF-S3E 10-Year Lithium Battery Smoke Alarm",
    "HF-S3E-R": "HIKFIRE HF-S3E-R 433MHz Wireless Interlinked Smoke Alarm",
    "HF-S3-WF": "HIKFIRE HF-S3-WF Wi-Fi Smoke Alarm with Hik-Connect",
    "HF-SG3-WF/R": "HIKFIRE HF-SG3-WF/R Wi-Fi and 433MHz Smoke Alarm",
    "HF-C104": "HIKFIRE HF-C104 Conventional Fire Alarm Control Panel",
    "HF-C108": "HIKFIRE HF-C108 Conventional Fire Alarm Control Panel",
    "HF-CS1": "HIKFIRE HF-CS1 Conventional Smoke Detector with Base",
    "HF-CH1": "HIKFIRE HF-CH1 Conventional Heat Detector with Base",
    "HF-CMP1": "HIKFIRE HF-CMP1 Conventional Manual Call Point",
    "HF-CSS1": "HIKFIRE HF-CSS1 Conventional Sounder Strobe",
    "HF-CAB1": "HIKFIRE HF-CAB1 Conventional Fire Alarm Bell",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        text = text[1:-1]
    return text.strip()


def money(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return clean(value)


def normalize_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


def sku_for(model: str) -> str:
    sku = re.sub(r"[^A-Za-z0-9]+", "-", model).strip("-").upper()
    return f"HIKFIRE-{sku}"


def slug_for(model: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", model.lower()).strip("-")
    return f"hikfire-{slug}"


def classify(model: str, description: str) -> tuple[str, str]:
    text = f"{model} {description}".lower()
    if model.startswith(("HF-VH", "HF-VR", "HF-VT", "HF-VTR")):
        return "กล้องตรวจจับไฟและความร้อน", ""
    if model.startswith("HF-S"):
        return "เครื่องตรวจจับควัน", ""
    if "bracket" in text or model.endswith("FP105-B"):
        return "อุปกรณ์ระบบแจ้งเหตุเพลิงไหม้", ""
    return "อุปกรณ์ระบบแจ้งเหตุเพลิงไหม้", ""


def extract_images(ws) -> dict[int, Path]:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    images_by_index: dict[int, Path] = {}
    for idx, image in enumerate(ws._images, start=1):
        if idx not in IMAGE_BY_ROW.values():
            continue
        ext = (getattr(image, "format", None) or "png").lower()
        out_path = IMAGE_DIR / f"hikfire_source_image_{idx:02d}.{ext}"
        out_path.write_bytes(image._data())
        images_by_index[idx] = out_path
    return images_by_index


def build_stock_map(ws) -> dict[str, int]:
    stock_map: dict[str, int] = {}
    for row in range(34, ws.max_row + 1):
        model = clean(ws.cell(row, 2).value)
        qty = ws.cell(row, 4).value
        if not model or qty is None:
            continue
        try:
            stock_map[normalize_code(model)] = int(qty)
        except (TypeError, ValueError):
            continue
    return stock_map


def stock_for(model: str, stock_map: dict[str, int]) -> tuple[int, str]:
    return 0, "ตั้ง stock เป็น 0 ตามคำสั่ง"


def source_specs(description: str, certificates: str, warranty: str, dealer: Any, si: Any, msrp: Any) -> str:
    lines = [
        f"Description: {description}" if description else "",
        f"Certificates: {certificates}" if certificates else "",
        f"Warranty: {warranty}" if warranty else "",
        "Price fields in prepared workbook: 0 บาท",
    ]
    return "\n".join(line for line in lines if line)


def thai_detail(thai_name: str, model: str, description: str, certificates: str, warranty: str, stock: int) -> str:
    pieces = [
        thai_name,
        f"รุ่น {model} จากแบรนด์ HIKFIRE สำหรับงานระบบแจ้งเหตุเพลิงไหม้ ความปลอดภัย อาคาร โรงงาน และงานติดตั้งเชิงอุตสาหกรรม",
    ]
    if description:
        pieces.append(f"รายละเอียดสินค้า: {description}")
    if certificates:
        pieces.append(f"มาตรฐาน/ใบรับรอง: {certificates}")
    if warranty:
        pieces.append(f"การรับประกัน: {warranty}")
    pieces.append("จำนวนสต๊อกในไฟล์เตรียมลงร้านตั้งไว้ที่ 0 ชิ้นตามคำสั่ง")
    pieces.append("ราคาบนไฟล์เตรียมลง LnwShop ตั้งไว้ที่ 0 บาทตาม workflow เดิม เพื่อให้ตรวจสอบก่อนเผยแพร่จริง")
    return "\n\n".join(pieces)


def short_detail(thai_name: str, model: str) -> str:
    return f"{thai_name} รุ่น {model} สำหรับระบบแจ้งเหตุเพลิงไหม้และงานความปลอดภัยจาก HIKFIRE"


def features(description: str, certificates: str, warranty: str) -> str:
    items = []
    for line in description.splitlines():
        line = clean(line)
        if line:
            items.append(f"- {line}")
    if certificates:
        items.append(f"- Certificates: {certificates.replace(chr(10), ', ')}")
    if warranty:
        items.append(f"- Warranty: {warranty}")
    return "\n".join(items)


def make_product_rows() -> list[dict[str, Any]]:
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)
    if not WATERMARK.exists():
        raise FileNotFoundError(WATERMARK)

    wb = load_workbook(SOURCE_EXCEL, data_only=True)
    ws = wb["HIKFIRE"]
    source_images = extract_images(ws)
    stock_map: dict[str, int] = {}
    WATERMARKED_DIR.mkdir(parents=True, exist_ok=True)
    mark = crop_watermark(Image.open(WATERMARK))

    rows: list[dict[str, Any]] = []
    for row_number in range(3, 22):
        item_no = ws.cell(row_number, 1).value
        model = clean(ws.cell(row_number, 2).value)
        if not model:
            continue
        description = clean(ws.cell(row_number, 3).value)
        certificates = clean(ws.cell(row_number, 4).value)
        warranty = clean(ws.cell(row_number, 5).value)
        dealer = ws.cell(row_number, 6).value
        si = ws.cell(row_number, 7).value
        msrp = ws.cell(row_number, 8).value
        main_category, sub_category = classify(model, description)
        category_chain = f"{BRAND} > {main_category}"
        stock, stock_note = stock_for(model, stock_map)
        source_img = source_images.get(IMAGE_BY_ROW[row_number])
        watermarked = ""
        original_rel = ""
        if source_img and source_img.exists():
            watermarked_path = WATERMARKED_DIR / f"hikfire_{slug_for(model)}_watermarked.png"
            apply_watermark(source_img, watermarked_path, mark)
            original_rel = str(source_img.relative_to(ROOT / "output" / "spreadsheet"))
            watermarked = str(watermarked_path.relative_to(ROOT / "output" / "spreadsheet"))

        thai_name = THAI_NAMES.get(model, f"สินค้า HIKFIRE รุ่น {model}")
        english_name = ENGLISH_NAMES.get(model, f"HIKFIRE {model}")
        source = source_specs(description, certificates, warranty, dealer, si, msrp)
        note = "ตั้งราคาเป็น 0 บาทเพื่อให้ตรวจสอบก่อนลงจริง"
        if stock_note:
            note = f"{note}; {stock_note}"
        seo_title_th = f"{thai_name} | One Tech Solution"
        seo_title_en = f"{english_name} | One Tech Solution"
        keywords = ", ".join(
            token
            for token in [
                model,
                "HIKFIRE",
                "Hikfire",
                "ระบบแจ้งเหตุเพลิงไหม้",
                "Fire Alarm",
                "Smoke Alarm",
                "Fire Detection",
                main_category,
                sub_category,
                "One Tech Solution",
            ]
            if token
        )
        meta = f"{thai_name} / {english_name} สำหรับระบบแจ้งเหตุเพลิงไหม้และงานความปลอดภัย ราคา 0 บาทในไฟล์ตรวจสอบก่อนลงร้าน"
        tags = ", ".join(token for token in [model, "HIKFIRE", "Fire Alarm", "Smoke Alarm", "Fire Detection", main_category] if token)
        rows.append(
            {
                "ลำดับ": item_no,
                "สถานะลงสินค้า": "รอตรวจสอบ",
                "เผยแพร่บนร้าน": "ปิด",
                "แบรนด์": BRAND,
                "หมวดหมู่หลัก": main_category,
                "หมวดหมู่ย่อย": sub_category,
                "หมวดหมู่แนะนำใน LnwShop": category_chain,
                "ชื่อสินค้า (ไทย)": thai_name,
                "Product Name (EN)": english_name,
                "SKU / Article Code": sku_for(model),
                "Vendor Part No.": model,
                "Model": model,
                "EAN/UPC": "",
                "ราคาปกติ (บาท)": 0,
                "ราคาขาย (บาท)": 0,
                "ราคาพิเศษ (บาท)": 0,
                "จำนวนสต๊อกตั้งต้น": stock,
                "สถานะสต๊อก": "รอตรวจสอบ",
                "น้ำหนักรวม (kg)": "",
                "กว้าง (cm)": "",
                "ยาว (cm)": "",
                "สูง (cm)": "",
                "ขนาดรวม": "",
                "การรับประกัน": warranty,
                "รายละเอียดสั้น (ไทย)": short_detail(thai_name, model),
                "รายละเอียดเต็ม (ไทย)": thai_detail(thai_name, model, description, certificates, warranty, stock),
                "จุดเด่นสินค้า (ไทย)": features(description, certificates, warranty),
                "Product Detail (EN)": description,
                "SEO Title (TH)": seo_title_th,
                "SEO Title (EN)": seo_title_en,
                "SEO Title (TH+EN)": f"{seo_title_th} / {seo_title_en}",
                "SEO Keywords (TH+EN)": keywords,
                "SEO Meta Description (TH+EN)": meta,
                "URL Slug": slug_for(model),
                "Search Tags": tags,
                "รูปภาพหลัก/ไฟล์ภาพ": watermarked,
                "หมายเหตุสำหรับลงสินค้า": note,
                "ข้อมูลสเปกต้นฉบับ": source,
                "แหล่งข้อมูลในไฟล์ต้นฉบับ": f"HIKFIRE row {row_number}; stock table rows 34-52",
                "รูปภาพต้นฉบับ": original_rel,
                "รูปภาพพร้อมลายน้ำ": watermarked,
                "หมวดหมู่แนะนำเดิม": main_category,
                "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                "ราคาต้นฉบับ Dealer": 0,
                "ราคาต้นฉบับ SI": 0,
                "ราคาต้นฉบับ MSRP": 0,
            }
        )
    return rows


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        1: 8,
        2: 16,
        3: 14,
        4: 16,
        5: 28,
        6: 28,
        7: 54,
        8: 54,
        9: 48,
        10: 28,
        11: 22,
        12: 22,
        14: 14,
        15: 14,
        16: 14,
        17: 16,
        24: 14,
        25: 54,
        26: 70,
        27: 54,
        28: 60,
        31: 70,
        32: 60,
        33: 70,
        35: 52,
        36: 48,
        37: 56,
        38: 60,
        39: 36,
        40: 48,
        41: 48,
        42: 46,
        43: 28,
        44: 40,
    }
    thin_gray = Side(style="thin", color="D9E2F3")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            border = copy(cell.border)
            border.bottom = thin_gray
            cell.border = border


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def category_description(name: str, parent: str) -> str:
    if parent == "ไม่มีหมวดหมู่หลัก":
        return "HIKFIRE รวมสินค้าและอุปกรณ์ระบบแจ้งเหตุเพลิงไหม้ กล้องตรวจจับไฟ เครื่องตรวจจับควัน และอุปกรณ์ประกอบ"
    return f"{name} ในหมวด {parent} สำหรับจัดกลุ่มสินค้า HIKFIRE บนร้าน One Tech Solution พร้อม SEO ไทยและอังกฤษ"


def build_categories(product_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: "OrderedDict[str, dict[str, Any]]" = OrderedDict()

    def add(name: str, parent: str, image: str, source: str) -> None:
        key = f"{parent}>{name}" if parent else name
        if key in categories:
            if not categories[key]["รูปภาพหมวดหมู่"] and image:
                categories[key]["รูปภาพหมวดหมู่"] = image
            return
        parent_text = parent or "ไม่มีหมวดหมู่หลัก"
        categories[key] = {
            "สถานะ": "รอตรวจสอบ",
            "ชื่อหมวดหมู่": name,
            "หมวดหมู่ย่อยของ": parent_text,
            "รายละเอียดหมวดหมู่": category_description(name, parent_text),
            "แสดงหมวดหมู่ที่หน้าร้าน": "เปิด",
            "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
            "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
            "รูปภาพหมวดหมู่": image,
            "SEO Title": f"{name} | One Tech Solution HIKFIRE",
            "SEO Keywords": ", ".join([name, parent, "HIKFIRE", "Fire Alarm", "Smoke Alarm", "One Tech Solution"]).strip(", "),
            "SEO Meta Description": f"{name} สินค้า HIKFIRE สำหรับระบบแจ้งเหตุเพลิงไหม้และงานความปลอดภัยจาก One Tech Solution",
            "แหล่งข้อมูล": source,
        }

    first_image = product_rows[0].get("รูปภาพหลัก/ไฟล์ภาพ", "") if product_rows else ""
    add(BRAND, "", first_image, "สร้างเป็นหมวดแบรนด์หลักจากไฟล์ HIKFIRE")
    for row in product_rows:
        main = row["หมวดหมู่หลัก"]
        image = row.get("รูปภาพหลัก/ไฟล์ภาพ", "")
        add(main, BRAND, image, "สร้างจากหมวดหมู่หลักแบบย่อในไฟล์สินค้า HIKFIRE")

    out = []
    for idx, item in enumerate(categories.values(), start=1):
        out.append({"ลำดับ": idx, **item})
    return out


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    add_table(ws, table_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_review_workbook(product_rows: list[dict[str, Any]], category_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Review_Summary"
    summary = [
        ["รายการ", "ค่า"],
        ["Source file", str(SOURCE_EXCEL)],
        ["Product rows", len(product_rows)],
        ["Category rows", len(category_rows)],
        ["Price rule", "ราคาปกติ/ขาย/พิเศษ = 0 บาททั้งหมด"],
        ["Stock rule", "จำนวนสต๊อกตั้งต้น = 0 ทุกสินค้า"],
        ["Category rule", "ลดหมวดหมู่เป็น HIKFIRE + 3 หมวดลูกหลัก"],
        ["Image rule", "ใช้รูปจาก Excel + ใส่ลายน้ำ One Tech Solution"],
        ["Ready for automation", "ยังไม่รันลงเว็บ ให้ตรวจไฟล์นี้ก่อน"],
    ]
    for row in summary:
        ws_summary.append(row)
    style_sheet(ws_summary)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 100

    ws_products = wb.create_sheet("LnwShop_AI_Ready")
    ws_products.append(PRODUCT_HEADERS)
    for row in product_rows:
        ws_products.append([row.get(header, "") for header in PRODUCT_HEADERS])
    style_sheet(ws_products)
    add_table(ws_products, "HIKFIREProductsReview")

    ws_categories = wb.create_sheet("Categories_AI_Ready")
    ws_categories.append(CATEGORY_HEADERS)
    for row in category_rows:
        ws_categories.append([row.get(header, "") for header in CATEGORY_HEADERS])
    style_sheet(ws_categories)
    add_table(ws_categories, "HIKFIRECategoriesReview")

    REVIEW_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(REVIEW_EXCEL)


def main() -> int:
    product_rows = make_product_rows()
    category_rows = build_categories(product_rows)
    write_workbook(PRODUCT_EXCEL, "LnwShop_AI_Ready", PRODUCT_HEADERS, product_rows, "HIKFIREProducts")
    write_workbook(CATEGORY_EXCEL, "Categories_AI_Ready", CATEGORY_HEADERS, category_rows, "HIKFIRECategories")
    write_review_workbook(product_rows, category_rows)
    print(f"products: {len(product_rows)}")
    print(f"categories: {len(category_rows)}")
    print(f"product excel: {PRODUCT_EXCEL}")
    print(f"category excel: {CATEGORY_EXCEL}")
    print(f"review excel: {REVIEW_EXCEL}")
    print(f"watermarked images: {WATERMARKED_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
