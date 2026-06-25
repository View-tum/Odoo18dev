from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Any
import fitz
import requests
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont

from prepare_assets import apply_watermark, crop_watermark
from prepare_hikfire_assets import CATEGORY_HEADERS, DEFAULT_PRODUCT_TYPE_1, DEFAULT_PRODUCT_TYPE_2, PRODUCT_HEADERS


ROOT = Path(__file__).resolve().parents[2]
SOURCE_EXCEL = Path(r"C:\Users\tumsu\Downloads\Huawei_Energy_Dealer_20261616.xls")
WATERMARK = Path(__file__).resolve().parent / "watermark_current_transparent_hq.png"
IMAGE_DIR = ROOT / "output" / "spreadsheet" / "huawei_product_images"
WATERMARKED_DIR = ROOT / "output" / "spreadsheet" / "huawei_product_images_watermarked"
PRODUCT_EXCEL = ROOT / "output" / "spreadsheet" / "HUAWEI_LnwShop_AI_ready_watermarked.xlsx"
CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "HUAWEI_LnwShop_categories.xlsx"
REVIEW_EXCEL = ROOT / "output" / "spreadsheet" / "HUAWEI_LnwShop_review.xlsx"
BRAND = "Huawei"
ROOT_CATEGORY = "Huawei"


CATEGORY_MAP = {
    "Inverter": "Inverter",
    "Inverter accessories": "Inverter Accessories",
    "Battery for Solar Cell": "Battery for Solar Cell",
}


CATEGORY_DESCRIPTION = {
    "Inverter": "อินเวอร์เตอร์ Huawei FusionSolar สำหรับระบบโซลาร์เซลล์บ้าน อาคารพาณิชย์ และโครงการ พร้อมข้อมูลไทยและอังกฤษ",
    "Inverter Accessories": "อุปกรณ์ประกอบ ระบบมอนิเตอร์ มิเตอร์ ออปติไมเซอร์ และอุปกรณ์ควบคุมสำหรับระบบ Huawei FusionSolar",
    "Battery for Solar Cell": "แบตเตอรี่และอุปกรณ์จัดเก็บพลังงาน Huawei LUNA สำหรับระบบโซลาร์เซลล์และ Smart ESS",
}


DIRECT_IMAGES: dict[str, tuple[str, str]] = {
    "HUW-2000-5K-LB0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/70a42d018f954b71be1922c013c43ce1.jpg",
        "https://solar.huawei.com/th/products/SUN2000-3-368-4-46-5-6K-LB0/",
    ),
    "HUW-2000-5K-MAP0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/1e2eb62222ec445cab39ef93df0fc944.jpg",
        "https://solar.huawei.com/en/products/sun2000-5-12k-map0/",
    ),
    "HUW-2000-10K-MAP0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/1e2eb62222ec445cab39ef93df0fc944.jpg",
        "https://solar.huawei.com/en/products/sun2000-5-12k-map0/",
    ),
    "HUW-2000-15K-MB0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/647bb291a0084b8f877b1af77cc16c75.jpg",
        "https://solar.huawei.com/ie/products/sun2000-12-15-17-20-25k-mb0/",
    ),
    "HUW-2000-20K-MB0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/647bb291a0084b8f877b1af77cc16c75.jpg",
        "https://solar.huawei.com/ie/products/sun2000-12-15-17-20-25k-mb0/",
    ),
    "HUW-2000-50KTLM3": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/1b1d9ed688fa4cdb92204543e7c90dbf.png",
        "https://solar.huawei.com/en/products/sun2000-50ktl-m3",
    ),
    "HUW-5000-150K-MG0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/92f63eff69094d20b57c07792cc6a21e.png",
        "https://solar.huawei.com/en/products/sun5000-150k-mg0/",
    ),
    "HUW-SDONGLEA-05": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/5d697fcaf2ae4814b85fd17716df92a6.png",
        "https://solar.huawei.com/th/products/SmartDongle-WLAN-FE/",
    ),
    "HUW-LUNA200010KWC1": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/6190bb25bc874610baeb9b18011bc433.png",
        "https://solar.huawei.com/th/products/LUNA2000-7-14-21-S1/specs/",
    ),
    "HUW-LUNA2000-7-E1": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/6190bb25bc874610baeb9b18011bc433.png",
        "https://solar.huawei.com/th/products/LUNA2000-7-14-21-S1/specs/",
    ),
    "HUW-SMART-G-63A-S0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/d3a6abd7431648de84abc1eb9504a89b.png",
        "https://solar.huawei.com/en/products/HUAWEI-SmartGuard/specs/",
    ),
    "HUW-SMART-G-63A-T0": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/d740c7ffcd1c4b30a455e3d06bd5a35d.png",
        "https://solar.huawei.com/en/products/HUAWEI-SmartGuard-63A-T0-AUT0/specs/",
    ),
    "KIT-HUW-1300PL-25Y": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/ba9018537eaa48ac828e0dddedeb43b9.png",
        "https://solar.huawei.com/th/products/merc-1100-1300-p/",
    ),
    "HUW-JNZ-UMG96RM": (
        "https://images.janitza.com/ce18jq9ih0x6/7CGFzyUvuCb3YyYqaPVPLg/694207377d9fbaf7487ea1df24308e3d/Janitza_UMG_96RM_Front_S.png?w=1920",
        "https://www.janitza.com/en-us/products/umg-96-rm-e",
    ),
    "HUW-JNZ-UMG512PRO": (
        "https://images.janitza.com/ce18jq9ih0x6/3z32EsC4bnfrUhKTxaQ5cQ/74842a882c647a9991bec8d47b87e9fb/Janitza_UMG_512-PRO_Front_EN_1_S.png?w=1920",
        "https://www.janitza.com/en-us/products/umg-512-pro",
    ),
    "HUW-2000-241-2S1": (
        "https://solar.huawei.com/admin/asset/v1/pro/view/de4bdb8635984897a8c459898f32aba7.png",
        "https://solar.huawei.com/en/",
    ),
}


PDF_IMAGES: dict[str, str] = {
    "HUW-2000-100KTLM2": "https://solar.huawei.com/admin/asset/v1/pro/view/0d3b86d0da694d88b8a9871688cda48d.pdf",
    "HUW-EMMA-A02": "https://solar.huawei.com/admin/asset/v1/pro/view/bb755a8d336a43f1b972c9d1830b619b.pdf",
}


SERVICE_CODES = {
    "HUW-EXTEND5Y-30K",
    "HUW-EXTEND5Y-36K",
    "HUW-EXTEND5Y-40K",
    "HUW-EXTEND5Y-50K",
    "HUW-EXTEND5Y-100K",
    "HUW-EXT5Y150K-2000",
    "HUW-EXT10YMERC600",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def one_line(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value).replace("\n", " / ")).strip()


def safe_slug(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").lower()


def file_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path(r"C:\Windows\Fonts\tahomabd.ttf" if bold else r"C:\Windows\Fonts\tahoma.ttf"),
        Path(r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def normalized_canvas(source: Path, target: Path) -> None:
    with Image.open(source) as raw:
        image = raw.convert("RGBA")
        base = Image.new("RGBA", (1400, 1400), (255, 255, 255, 255))
        scale = min(1240 / image.width, 1120 / image.height)
        image = image.resize((max(1, int(image.width * scale)), max(1, int(image.height * scale))), Image.Resampling.LANCZOS)
        x = (base.width - image.width) // 2
        y = max(55, (base.height - image.height) // 2 - 35)
        base.alpha_composite(image, (x, y))
        base.convert("RGB").save(target, "PNG")


def fetch(url: str) -> bytes:
    response = requests.get(url, timeout=45, headers={"User-Agent": "Mozilla/5.0"})
    response.raise_for_status()
    return response.content


def download_image(code: str, url: str) -> Path:
    raw_path = IMAGE_DIR / f"{safe_slug(code)}_download"
    raw_path.write_bytes(fetch(url))
    output = IMAGE_DIR / f"{safe_slug(code)}_source.png"
    normalized_canvas(raw_path, output)
    raw_path.unlink(missing_ok=True)
    return output


def render_pdf_image(code: str, url: str) -> Path:
    pdf_path = IMAGE_DIR / f"{safe_slug(code)}_datasheet.pdf"
    pdf_path.write_bytes(fetch(url))
    document = fitz.open(pdf_path)
    pixmap = document[0].get_pixmap(matrix=fitz.Matrix(2.2, 2.2), alpha=False)
    rendered = IMAGE_DIR / f"{safe_slug(code)}_pdf_page.png"
    pixmap.save(str(rendered))
    document.close()
    output = IMAGE_DIR / f"{safe_slug(code)}_source.png"
    normalized_canvas(rendered, output)
    rendered.unlink(missing_ok=True)
    return output


def make_service_image(code: str, description: str, warranty: str) -> Path:
    output = IMAGE_DIR / f"{safe_slug(code)}_service_source.png"
    canvas = Image.new("RGB", (1400, 1400), "#ffffff")
    draw = ImageDraw.Draw(canvas)
    green = "#23884d"
    dark = "#172a33"
    draw.rectangle((0, 0, 1400, 24), fill=green)
    draw.text((100, 145), "HUAWEI FUSIONSOLAR", fill=green, font=file_font(58, bold=True))
    draw.text((100, 302), "EXTENDED WARRANTY", fill=dark, font=file_font(76, bold=True))
    draw.text((100, 430), "บริการขยายการรับประกัน", fill=dark, font=file_font(54, bold=True))
    years = "10 YEARS" if "10Y" in code else "5 YEARS"
    draw.rounded_rectangle((100, 580, 550, 765), radius=28, fill=green)
    draw.text((150, 630), years, fill="#ffffff", font=file_font(52, bold=True))
    detail = one_line(description)
    draw.text((100, 900), detail[:50], fill=dark, font=file_font(36))
    draw.text((100, 970), f"SKU: {code}", fill="#475b63", font=file_font(34))
    draw.text((100, 1035), f"Warranty: {warranty or years.title()}", fill="#475b63", font=file_font(34))
    draw.text((100, 1205), "Service item - not equipment photo", fill="#6e7c83", font=file_font(28))
    canvas.save(output, "PNG")
    return output


def read_source() -> list[dict[str, Any]]:
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)
    workbook = xlrd.open_workbook(str(SOURCE_EXCEL), formatting_info=True)
    sheet = workbook.sheet_by_index(0)
    headers = [clean(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    hyperlinks = {item.frowx: item.url_or_path for item in sheet.hyperlink_list}
    products: list[dict[str, Any]] = []
    for row_index in range(1, sheet.nrows):
        row = dict(zip(headers, [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]))
        code = clean(row.get("Article Code"))
        if not code:
            continue
        row["_source_row"] = row_index + 1
        row["_source_url"] = hyperlinks.get(row_index, "")
        products.append(row)
    return products


def source_image_for(product: dict[str, Any]) -> tuple[Path, str, str]:
    code = clean(product["Article Code"])
    description = one_line(product["Article Description"])
    warranty = one_line(product["Warranty"])
    if code in SERVICE_CODES:
        return make_service_image(code, description, warranty), "service graphic", "สร้างภาพบริการแยกจากรูปอุปกรณ์"
    if code in DIRECT_IMAGES:
        image_url, page_url = DIRECT_IMAGES[code]
        return download_image(code, image_url), "manufacturer product image", page_url
    pdf_url = PDF_IMAGES.get(code) or clean(product.get("_source_url"))
    if pdf_url and (".pdf" in pdf_url.lower() or "/download?" in pdf_url.lower()):
        return render_pdf_image(code, pdf_url), "manufacturer datasheet page 1", pdf_url
    raise ValueError(f"ไม่มีแหล่งภาพที่ตรวจสอบได้สำหรับ {code}")


def category_for(source_category: str) -> str:
    try:
        return CATEGORY_MAP[clean(source_category)]
    except KeyError as exc:
        raise ValueError(f"ไม่พบ category mapping: {source_category}") from exc


def product_english_name(code: str, category: str) -> str:
    model = code.removeprefix("HUW-").replace("-", " ")
    descriptions = {
        "Inverter": "Huawei FusionSolar Inverter",
        "Inverter Accessories": "Huawei FusionSolar Accessory",
        "Battery for Solar Cell": "Huawei FusionSolar Energy Storage Product",
    }
    return f"{descriptions[category]} {model}".strip()


def product_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if not WATERMARK.exists():
        raise FileNotFoundError(WATERMARK)
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    WATERMARKED_DIR.mkdir(parents=True, exist_ok=True)
    watermark = crop_watermark(Image.open(WATERMARK))
    rows: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    for number, source in enumerate(read_source(), start=1):
        code = clean(source["Article Code"])
        thai_name = one_line(source["Article Description"])
        category = category_for(source["Category Level 3"])
        source_image, image_type, image_reference = source_image_for(source)
        watermarked_path = WATERMARKED_DIR / f"{safe_slug(code)}_watermarked.png"
        apply_watermark(source_image, watermarked_path, watermark)
        image_rel = str(watermarked_path.relative_to(PRODUCT_EXCEL.parent))
        source_rel = str(source_image.relative_to(PRODUCT_EXCEL.parent))
        english_name = product_english_name(code, category)
        warranty = one_line(source["Warranty"])
        category_chain = f"{ROOT_CATEGORY} > {category}"
        detail_th = (
            f"{thai_name}\n\n"
            f"แบรนด์: {BRAND}\n"
            f"รหัสสินค้า: {code}\n"
            f"หมวดหมู่: {category_chain}\n"
            f"การรับประกัน: {warranty or 'ตรวจสอบตามเงื่อนไขผู้ผลิต'}\n\n"
            "ข้อมูลสินค้าอ้างอิงจาก price list และเอกสาร/ภาพของผู้ผลิต กรุณาตรวจสอบราคาและสเปกก่อนเผยแพร่จำหน่าย"
        )
        detail_en = (
            f"{english_name}. Article code: {code}. Category: {category_chain}. "
            f"Warranty: {warranty or 'Refer to manufacturer terms'}. "
            "Product information prepared from the dealer price list and manufacturer reference."
        )
        seo_th = f"{thai_name} ราคาและสเปก | One Tech Solution"
        seo_en = f"{english_name} Price and Specification | One Tech Solution"
        keywords = ", ".join(
            [
                code,
                "Huawei",
                "FusionSolar",
                thai_name,
                category,
                "โซลาร์เซลล์",
                "อินเวอร์เตอร์โซลาร์เซลล์",
                "Solar Inverter",
                "Solar Energy",
                "One Tech Solution",
            ]
        )
        meta = (
            f"{thai_name} / {english_name} รหัส {code} สำหรับระบบ Huawei FusionSolar "
            "ตรวจสอบสเปกและสอบถามราคาได้ที่ One Tech Solution"
        )
        source_note = (
            f"Source Excel row {source['_source_row']}; Article Code {code}; Warranty {warranty}; "
            f"Image type: {image_type}; Image reference: {image_reference}"
        )
        rows.append(
            {
                "ลำดับ": number,
                "สถานะลงสินค้า": "รอตรวจสอบ",
                "เผยแพร่บนร้าน": "ปิด",
                "แบรนด์": BRAND,
                "หมวดหมู่หลัก": category,
                "หมวดหมู่ย่อย": "",
                "หมวดหมู่แนะนำใน LnwShop": category_chain,
                "ชื่อสินค้า (ไทย)": thai_name,
                "Product Name (EN)": english_name,
                "SKU / Article Code": code,
                "Vendor Part No.": code,
                "Model": code.removeprefix("HUW-"),
                "EAN/UPC": "",
                "ราคาปกติ (บาท)": 0,
                "ราคาขาย (บาท)": 0,
                "ราคาพิเศษ (บาท)": 0,
                "จำนวนสต๊อกตั้งต้น": 0,
                "สถานะสต๊อก": "หมดสต๊อก",
                "น้ำหนักรวม (kg)": "",
                "กว้าง (cm)": "",
                "ยาว (cm)": "",
                "สูง (cm)": "",
                "ขนาดรวม": "",
                "การรับประกัน": warranty,
                "รายละเอียดสั้น (ไทย)": f"{thai_name} รหัส {code} สำหรับระบบ Huawei FusionSolar",
                "รายละเอียดเต็ม (ไทย)": detail_th,
                "จุดเด่นสินค้า (ไทย)": f"- แบรนด์ Huawei FusionSolar\n- รหัสสินค้า {code}\n- การรับประกัน {warranty or 'ตรวจสอบตามเงื่อนไขผู้ผลิต'}",
                "Product Detail (EN)": detail_en,
                "SEO Title (TH)": seo_th,
                "SEO Title (EN)": seo_en,
                "SEO Title (TH+EN)": f"{seo_th} / {seo_en}",
                "SEO Keywords (TH+EN)": keywords,
                "SEO Meta Description (TH+EN)": meta,
                "URL Slug": f"huawei-{safe_slug(code)}",
                "Search Tags": f"Huawei, FusionSolar, {code}, {category}, Solar, โซลาร์เซลล์",
                "รูปภาพหลัก/ไฟล์ภาพ": image_rel,
                "หมายเหตุสำหรับลงสินค้า": "ตั้งราคาและจำนวนสต๊อกเป็น 0 เพื่อให้ตรวจสอบก่อนเปิดขาย",
                "ข้อมูลสเปกต้นฉบับ": source_note,
                "แหล่งข้อมูลในไฟล์ต้นฉบับ": f"{SOURCE_EXCEL.name} row {source['_source_row']}",
                "รูปภาพต้นฉบับ": source_rel,
                "รูปภาพพร้อมลายน้ำ": image_rel,
                "หมวดหมู่แนะนำเดิม": clean(source["Category Level 3"]),
                "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                "ราคาต้นฉบับ Dealer": 0,
                "ราคาต้นฉบับ SI": 0,
                "ราคาต้นฉบับ MSRP": 0,
            }
        )
        audit.append(
            {
                "Source Row": source["_source_row"],
                "SKU": code,
                "Description": thai_name,
                "Source Category": clean(source["Category Level 3"]),
                "LnwShop Category": category_chain,
                "Warranty": warranty,
                "Source SRP inc.vat": source.get("SRP inc.vat", ""),
                "Source NDP ex.vat": source.get("NDP ex.vat", ""),
                "Upload Price": 0,
                "Upload Stock": 0,
                "Image Type": image_type,
                "Image Reference": image_reference,
                "Watermarked Image": image_rel,
            }
        )
    return rows, audit


def build_categories(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = ["Inverter", "Inverter Accessories", "Battery for Solar Cell"]
    first_images = {name: "" for name in order}
    for row in rows:
        if not first_images[row["หมวดหมู่หลัก"]]:
            first_images[row["หมวดหมู่หลัก"]] = row["รูปภาพหลัก/ไฟล์ภาพ"]
    categories: list[dict[str, Any]] = []
    for number, name in enumerate(order, start=1):
        categories.append(
            {
                "ลำดับ": number,
                "สถานะ": "รอตรวจสอบ",
                "ชื่อหมวดหมู่": name,
                "หมวดหมู่ย่อยของ": ROOT_CATEGORY,
                "รายละเอียดหมวดหมู่": CATEGORY_DESCRIPTION[name],
                "แสดงหมวดหมู่ที่หน้าร้าน": "เปิด",
                "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                "รูปภาพหมวดหมู่": first_images[name],
                "SEO Title": f"Huawei {name} | One Tech Solution",
                "SEO Keywords": f"Huawei, FusionSolar, {name}, โซลาร์เซลล์, Solar Energy, One Tech Solution",
                "SEO Meta Description": CATEGORY_DESCRIPTION[name],
                "แหล่งข้อมูล": "สรุปจาก Category Level 3 ใน Huawei dealer price list; หมวดหลัก Huawei มีอยู่แล้วใน LnwShop",
            }
        )
    return categories


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            border = copy(cell.border)
            border.bottom = thin_gray
            cell.border = border
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20
    for col in [7, 8, 9, 25, 26, 27, 28, 31, 32, 33, 36, 38, 39, 40, 41]:
        if col <= sheet.max_column:
            sheet.column_dimensions[get_column_letter(col)].width = 52
    sheet.row_dimensions[1].height = 36


def add_table(sheet, name: str) -> None:
    if sheet.max_row < 2:
        return
    ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)


def write_sheet(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)
    add_table(sheet, table_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_review(rows: list[dict[str, Any]], categories: list[dict[str, Any]], audit: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Review_Summary"
    for item in [
        ["รายการ", "ค่า"],
        ["Source file", str(SOURCE_EXCEL)],
        ["Product rows", len(rows)],
        ["Category rows to create", len(categories)],
        ["Existing root category", ROOT_CATEGORY],
        ["Price rule", "ราคาที่ใช้ upload = 0 บาททุกสินค้า"],
        ["Stock rule", "จำนวนสต๊อกตั้งต้น = 0 ทุกสินค้า"],
        ["Image rule", "ใช้รูปหรือหน้า datasheet จากผู้ผลิตตรงสินค้า พร้อมลายน้ำ One Tech Solution"],
        ["Warranty rows", "ใช้ภาพบริการเฉพาะ ไม่แสดงรูปอุปกรณ์ผิดรายการ"],
        ["Ready for automation", "ให้สร้างหมวดย่อย 3 รายการก่อนรันสินค้า"],
    ]:
        summary.append(item)
    style_sheet(summary)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 108

    products = workbook.create_sheet("LnwShop_AI_Ready")
    products.append(PRODUCT_HEADERS)
    for row in rows:
        products.append([row.get(header, "") for header in PRODUCT_HEADERS])
    style_sheet(products)
    add_table(products, "HuaweiProductsReview")

    category_sheet = workbook.create_sheet("Categories_AI_Ready")
    category_sheet.append(CATEGORY_HEADERS)
    for row in categories:
        category_sheet.append([row.get(header, "") for header in CATEGORY_HEADERS])
    style_sheet(category_sheet)
    add_table(category_sheet, "HuaweiCategoriesReview")

    audit_sheet = workbook.create_sheet("Source_Image_Audit")
    audit_headers = list(audit[0].keys())
    audit_sheet.append(audit_headers)
    for row in audit:
        audit_sheet.append([row.get(header, "") for header in audit_headers])
    style_sheet(audit_sheet)
    add_table(audit_sheet, "HuaweiImageAudit")
    workbook.save(REVIEW_EXCEL)


def validate(rows: list[dict[str, Any]], categories: list[dict[str, Any]]) -> None:
    issues: list[str] = []
    if len(rows) != 34:
        issues.append(f"จำนวนสินค้าไม่เท่ากับ 34: {len(rows)}")
    if len({row["SKU / Article Code"] for row in rows}) != len(rows):
        issues.append("พบ SKU ซ้ำ")
    expected_categories = {f"{ROOT_CATEGORY} > {name}" for name in CATEGORY_MAP.values()}
    if {row["หมวดหมู่แนะนำใน LnwShop"] for row in rows} != expected_categories:
        issues.append("หมวดหมู่สินค้าไม่ครบตาม mapping")
    if {row["ชื่อหมวดหมู่"] for row in categories} != set(CATEGORY_MAP.values()):
        issues.append("ไฟล์หมวดหมู่ไม่ครบ 3 หมวด")
    for row in rows:
        code = row["SKU / Article Code"]
        if any(row[name] != 0 for name in ["ราคาปกติ (บาท)", "ราคาขาย (บาท)", "ราคาพิเศษ (บาท)", "จำนวนสต๊อกตั้งต้น"]):
            issues.append(f"{code} ราคา/stock ไม่เป็น 0")
        image = PRODUCT_EXCEL.parent / row["รูปภาพหลัก/ไฟล์ภาพ"]
        if not image.exists():
            issues.append(f"{code} ไม่มีไฟล์ภาพลายน้ำ")
        if not row["SEO Title (TH+EN)"] or not row["SEO Keywords (TH+EN)"]:
            issues.append(f"{code} ไม่มี SEO")
    if issues:
        raise ValueError("\n".join(issues))


def main() -> int:
    rows, audit = product_rows()
    categories = build_categories(rows)
    validate(rows, categories)
    write_sheet(PRODUCT_EXCEL, "LnwShop_AI_Ready", PRODUCT_HEADERS, rows, "HuaweiProducts")
    write_sheet(CATEGORY_EXCEL, "Categories_AI_Ready", CATEGORY_HEADERS, categories, "HuaweiCategories")
    write_review(rows, categories, audit)
    print(f"products: {len(rows)}")
    print(f"categories to create: {len(categories)}")
    print(f"product excel: {PRODUCT_EXCEL}")
    print(f"category excel: {CATEGORY_EXCEL}")
    print(f"review excel: {REVIEW_EXCEL}")
    print(f"watermarked images: {WATERMARKED_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
