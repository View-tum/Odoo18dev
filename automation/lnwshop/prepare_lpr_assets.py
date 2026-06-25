from __future__ import annotations

import re
from copy import copy
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image

from prepare_assets import apply_watermark, crop_watermark
from prepare_hikfire_assets import CATEGORY_HEADERS, DEFAULT_PRODUCT_TYPE_1, DEFAULT_PRODUCT_TYPE_2, PRODUCT_HEADERS


ROOT = Path(__file__).resolve().parents[2]
SOURCE_EXCEL = ROOT / "automation" / "Price LPR.xlsx"
WATERMARK = Path(__file__).resolve().parent / "watermark_current_transparent_hq.png"
IMAGE_DIR = ROOT / "output" / "spreadsheet" / "hip_product_images"
WATERMARKED_DIR = ROOT / "output" / "spreadsheet" / "hip_product_images_watermarked"
PRODUCT_EXCEL = ROOT / "output" / "spreadsheet" / "HIP_LnwShop_AI_ready_watermarked.xlsx"
CATEGORY_EXCEL = ROOT / "output" / "spreadsheet" / "HIP_LnwShop_categories.xlsx"
REVIEW_EXCEL = ROOT / "output" / "spreadsheet" / "HIP_LnwShop_review.xlsx"
BRAND = "HIP"
EXISTING_ROOT_CATEGORY = "ระบบจอดรถ Carpark"


THAI_NAMES = {
    "CMS9101": "เครื่องจ่ายสลิปที่จอดรถ HIP รุ่น CMS9101",
    "CMS9102": "เครื่องจ่ายสลิปที่จอดรถ HIP รุ่น CMS9102 อ่าน QR Code",
    "CMHB607": "ตู้จ่ายบัตร RFID อัตโนมัติ HIP รุ่น CMHB607",
    "CMHB608": "ตู้รับบัตร RFID อัตโนมัติ HIP รุ่น CMHB608",
    "LPR Professional Kit Set": "ชุดระบบอ่านป้ายทะเบียนอัจฉริยะ HIP LPR Professional Kit Set",
    "CMH88 Pro": "กล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 Pro",
    "CMH88 Plus": "กล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 Plus",
    "SET LPR CMH88 Plus": "ชุดกล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 Plus พร้อมป้ายแสดงผล LED",
    "เซ็ตรวม CMH88 Plus + Bracket + Fill light": "ชุดกล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 Plus (กล้อง + ขาตั้ง + ไฟส่อง)",
    "SET LPR CMH88": "ชุดกล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 พร้อมป้ายแสดงผล LED",
    "เซ็ตรวม CMH88 + Bracket + Fill light": "ชุดกล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMH88 (กล้อง + ขาตั้ง + ไฟส่อง)",
    "CMTV94 Pro": "กล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMTV94 Pro",
    "เซ็ตรวม CMTV94 Pro + Bracket + Fill light + Adaptor": "ชุดกล้องอ่านป้ายทะเบียนอัจฉริยะ HIP รุ่น CMTV94 Pro (กล้อง + ขาตั้ง + ไฟส่อง + อะแดปเตอร์)",
    "LED P4 Display RS485": "ป้ายไฟ LED P4 แสดงป้ายทะเบียน HIP แบบ RS485",
    "LED P4 Display TCP/IP": "ป้ายไฟ LED P4 แสดงป้ายทะเบียน HIP แบบ TCP/IP",
    "SEFLPR": "ไฟส่องป้ายทะเบียน LED Fill Light HIP รุ่น SEFLPR",
    "P30 POS": "ตู้ชำระเงินที่จอดรถอัจฉริยะ HIP รุ่น P30 POS",
    "P32 POS": "ตู้ชำระเงินที่จอดรถอัจฉริยะ HIP รุ่น P32 POS",
}


ENGLISH_NAMES = {
    "CMS9101": "HIP CMS9101 Carpark Entrance Ticket Dispenser",
    "CMS9102": "HIP CMS9102 Carpark Entrance Ticket Dispenser with QR Code Reader",
    "CMHB607": "HIP CMHB607 Automatic RFID Card Dispenser",
    "CMHB608": "HIP CMHB608 Automatic RFID Card Receiver",
    "LPR Professional Kit Set": "HIP LPR Professional Kit Set Parking System",
    "CMH88 Pro": "HIP CMH88 Pro License Plate Recognition Camera",
    "CMH88 Plus": "HIP CMH88 Plus License Plate Recognition Camera",
    "SET LPR CMH88 Plus": "HIP CMH88 Plus License Plate Recognition Camera Set with LED Display",
    "เซ็ตรวม CMH88 Plus + Bracket + Fill light": "HIP CMH88 Plus LPR Camera Set with Pole Bracket and Fill Light",
    "SET LPR CMH88": "HIP CMH88 License Plate Recognition Camera Set with LED Display",
    "เซ็ตรวม CMH88 + Bracket + Fill light": "HIP CMH88 LPR Camera Set with Pole Bracket and Fill Light",
    "CMTV94 Pro": "HIP CMTV94 Pro License Plate Recognition Camera",
    "เซ็ตรวม CMTV94 Pro + Bracket + Fill light + Adaptor": "HIP CMTV94 Pro LPR Camera Set with Bracket, Fill Light, and Power Adapter",
    "LED P4 Display RS485": "HIP LED P4 Plate Display Screen RS485 Connection",
    "LED P4 Display TCP/IP": "HIP LED P4 Plate Display Screen TCP/IP Connection",
    "SEFLPR": "HIP SEFLPR LED License Plate Fill Light",
    "P30 POS": "HIP P30 POS Smart Carpark Payment Machine",
    "P32 POS": "HIP P32 POS Smart Carpark Payment Machine",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        text = text[1:-1]
    return text.strip()


def normalize_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", value.upper())


def sku_for(model: str) -> str:
    sku = re.sub(r"[^A-Za-z0-9]+", "-", model).strip("-").upper()
    return f"HIP-{sku}"


def slug_for(model: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", model.lower()).strip("-")
    return f"hip-{slug}"


def classify(model: str, name: str) -> tuple[str, str]:
    text = f"{model} {name}".lower()
    if "ticket" in text or "dispenser" in text or "ตู้จ่ายบัตร" in text or "ตู้รับบัตร" in text or "card" in text or "rfid" in text or "cmhb" in model or "cms9" in model:
        return "ตู้จ่ายบัตร", ""
    if "lpr" in text or "อ่านป้ายทะเบียน" in text or "cmh88" in model or "cmtv" in model:
        return "License Plate Recognition", ""
    if "pos" in text or "payment" in text or re.match(r"^p\d+ pos", text):
        return "Car park", ""
    return "Car Parking Accessories", ""


def extract_images_dynamic(ws, sheet_prefix: str) -> dict[int, Path]:
    IMAGE_DIR.mkdir(parents=True, exist_ok=True)
    images_by_row: dict[int, Path] = {}
    for idx, image in enumerate(ws._images, start=1):
        anchor = image.anchor._from
        row = anchor.row + 1
        fmt = (getattr(image, "format", None) or "png").lower()
        if fmt == "jpeg":
            ext = "jpg"
        else:
            ext = fmt if re.fullmatch(r"[a-z0-9]+", fmt) else "png"
        out_path = IMAGE_DIR / f"hip_{sheet_prefix}_row_{row:04d}_image_{idx:03d}.{ext}"
        out_path.write_bytes(image._data())
        images_by_row[row] = out_path
    return images_by_row


def source_specs(specs: str, features: str, desc: str, warranty: str, price: Any) -> str:
    lines = [
        f"Specifications: {specs}" if specs else "",
        f"Features: {features}" if features else "",
        f"Descriptions: {desc}" if desc else "",
        f"Warranty: {warranty}" if warranty else "",
        f"Original Price: {price} THB" if price else "",
    ]
    return "\n".join(line for line in lines if line)


def thai_detail(name_th: str, model: str, main_category: str, specs: str, features: str, desc: str, warranty: str) -> str:
    parts = [
        name_th,
        f"สินค้าแบรนด์ HIP แท้ รุ่น {model} สำหรับการติดตั้งระบบลานจอดรถอัจฉริยะ (Smart Carpark System) และระบบอ่านป้ายทะเบียนรถยนต์ (LPR)",
        f"หมวดหมู่สินค้า: {main_category}",
    ]
    if specs:
        parts.append(f"ข้อมูลทางเทคนิค (Specifications):\n{specs}")
    if features:
        parts.append(f"คุณลักษณะเด่น (Features):\n{features}")
    if desc:
        parts.append(f"รายละเอียดสินค้า (Descriptions):\n{desc}")
    if warranty:
        parts.append(f"การรับประกันสินค้า (Warranty): {warranty}")
    parts.append("หมายเหตุ: ข้อมูลราคาขายและสต๊อกสินค้าในระบบ LnwShop จะแสดงผลเป็น 0 บาทชั่วคราว เพื่อผ่านการตรวจสอบข้อมูลของฝ่ายการเงินก่อนประกาศขาย")
    return "\n\n".join(parts)


def short_detail(name_th: str, model: str) -> str:
    return f"{name_th} รุ่น {model} คุณภาพสูง ประสิทธิภาพดีเยี่ยมสำหรับระบบงานจอดรถและ LPR แบรนด์ HIP"


def features_bullet(specs: str, features: str, desc: str) -> str:
    items = []
    text = f"{specs}\n{features}\n{desc}"
    lines = [clean(line) for line in text.split("\n") if clean(line)]
    for line in lines:
        if line.startswith(("-", "*", "•")):
            items.append(line)
        elif len(line) < 120 and not line.lower().startswith(("model", "specs", "feature")):
            items.append(f"- {line}")
    if not items:
        items.append("- สินค้าระบบบริหารลานจอดรถอัจฉริยะ HIP คุณภาพมาตรฐานสากล")
    return "\n".join(items[:8])


def make_product_rows() -> list[dict[str, Any]]:
    if not SOURCE_EXCEL.exists():
        raise FileNotFoundError(SOURCE_EXCEL)
    if not WATERMARK.exists():
        raise FileNotFoundError(WATERMARK)

    wb = load_workbook(SOURCE_EXCEL, data_only=True)
    mark = crop_watermark(Image.open(WATERMARK))
    WATERMARKED_DIR.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, Any]] = []
    overall_index = 1

    sheet_configs = [
        ("ตู้จ่ายบัตร , QR ,Slip ", "ticket"),
        ("LPR ", "lpr")
    ]

    for sheet_name, prefix in sheet_configs:
        ws = wb[sheet_name]
        images_by_row = extract_images_dynamic(ws, prefix)

        for row_num in range(5, ws.max_row + 1):
            erp_code = clean(ws.cell(row_num, 1).value)
            model = clean(ws.cell(row_num, 2).value)
            prod_name = clean(ws.cell(row_num, 3).value)
            specs = clean(ws.cell(row_num, 5).value)
            features = clean(ws.cell(row_num, 6).value)
            desc = clean(ws.cell(row_num, 7).value)
            price = ws.cell(row_num, 8).value
            warranty = clean(ws.cell(row_num, 9).value)
            remark = clean(ws.cell(row_num, 10).value)

            if not model:
                continue

            main_category, sub_category = classify(model, prod_name)
            category_chain = main_category

            source_img = images_by_row.get(row_num)
            if not source_img:
                if sheet_name == "ตู้จ่ายบัตร , QR ,Slip " and model == "CMS9102":
                    source_img = images_by_row.get(5)
                elif sheet_name == "LPR " and model == "P30 POS":
                    source_img = images_by_row.get(18)
                else:
                    if sheet_name == "ตู้จ่ายบัตร , QR ,Slip ":
                        source_img = images_by_row.get(5)
                    else:
                        source_img = images_by_row.get(5)

            watermarked = ""
            original_rel = ""
            if source_img and source_img.exists():
                watermarked_path = WATERMARKED_DIR / f"hip_{prefix}_{slug_for(model)}_watermarked.png"
                apply_watermark(source_img, watermarked_path, mark)
                original_rel = str(source_img.relative_to(ROOT / "output" / "spreadsheet"))
                watermarked = str(watermarked_path.relative_to(ROOT / "output" / "spreadsheet"))

            thai_name = THAI_NAMES.get(model, f"สินค้า HIP รุ่น {model}")
            english_name = ENGLISH_NAMES.get(model, f"HIP {model}")
            source_detail = source_specs(specs, features, desc, warranty, price)
            note_parts = ["ตั้งราคาขายเป็น 0 บาทตาม Flow เพื่อตรวจสอบก่อนเปิดตัว"]
            if remark:
                note_parts.append(f"Original Remark: {remark}")
            if erp_code:
                note_parts.append(f"ERP Code: {erp_code}")

            seo_title_th = f"{thai_name} | One Tech Solution"
            seo_title_en = f"{english_name} | One Tech Solution"
            keywords = ", ".join(
                token
                for token in [
                    model,
                    "HIP",
                    "ระบบจอดรถ",
                    "Carpark",
                    "LPR",
                    "กล้องอ่านป้ายทะเบียน",
                    main_category,
                    "One Tech Solution",
                    "ซ่อมติดตั้งสั่งซื้ออะไหล่",
                    "ขายส่ง",
                    "คอนโดโรงเรียนโรงแรมหมู่บ้าน"
                ]
                if token
            )
            meta_desc = f"{thai_name} / {english_name} สำหรับงานระบบลานจอดรถอัจฉริยะและ LPR คุณภาพสูงจาก One Tech Solution"

            rows.append(
                {
                    "ลำดับ": overall_index,
                    "สถานะลงสินค้า": "รอตรวจสอบ",
                    "เผยแพร่บนร้าน": "ปิด",
                    "แบรนด์": BRAND,
                    "หมวดหมู่หลัก": main_category,
                    "หมวดหมู่ย่อย": sub_category,
                    "หมวดหมู่แนะนำใน LnwShop": category_chain,
                    "ชื่อสินค้า (ไทย)": thai_name,
                    "Product Name (EN)": english_name,
                    "SKU / Article Code": sku_for(model),
                    "Vendor Part No.": erp_code or model,
                    "Model": model,
                    "EAN/UPC": "",
                    "ราคาปกติ (บาท)": 0,
                    "ราคาขาย (บาท)": 0,
                    "ราคาพิเศษ (บาท)": 0,
                    "จำนวนสต๊อกตั้งต้น": 0,
                    "สถานะสต๊อก": "รอตรวจสอบ",
                    "น้ำหนักรวม (kg)": "",
                    "กว้าง (cm)": "",
                    "ยาว (cm)": "",
                    "สูง (cm)": "",
                    "ขนาดรวม": "",
                    "การรับประกัน": warranty or "1 Year",
                    "รายละเอียดสั้น (ไทย)": short_detail(thai_name, model),
                    "รายละเอียดเต็ม (ไทย)": thai_detail(thai_name, model, main_category, specs, features, desc, warranty),
                    "จุดเด่นสินค้า (ไทย)": features_bullet(specs, features, desc),
                    "Product Detail (EN)": english_name,
                    "SEO Title (TH)": seo_title_th,
                    "SEO Title (EN)": seo_title_en,
                    "SEO Title (TH+EN)": f"{seo_title_th} / {seo_title_en}",
                    "SEO Keywords (TH+EN)": keywords,
                    "SEO Meta Description (TH+EN)": meta_desc,
                    "URL Slug": slug_for(model),
                    "Search Tags": ", ".join(token for token in [model, BRAND, main_category, "LPR", "Carpark", "ซ่อมติดตั้งสั่งซื้ออะไหล่", "ขายส่ง", "คอนโดโรงเรียนโรงแรมหมู่บ้าน"] if token),
                    "รูปภาพหลัก/ไฟล์ภาพ": watermarked,
                    "หมายเหตุสำหรับลงสินค้า": "; ".join(note_parts),
                    "ข้อมูลสเปกต้นฉบับ": source_detail,
                    "แหล่งข้อมูลในไฟล์ต้นฉบับ": f"Sheet {sheet_name} Row {row_num}",
                    "รูปภาพต้นฉบับ": original_rel,
                    "รูปภาพพร้อมลายน้ำ": watermarked,
                    "หมวดหมู่แนะนำเดิม": main_category,
                    "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
                    "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
                    "ราคาต้นฉบับ Dealer": price or 0,
                    "ราคาต้นฉบับ SI": price or 0,
                    "ราคาต้นฉบับ MSRP": price or 0,
                }
            )
            overall_index += 1

    return rows


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        1: 8,
        2: 16,
        3: 14,
        4: 16,
        5: 28,
        6: 28,
        7: 54,
        8: 54,
        9: 48,
        10: 28,
        11: 22,
        12: 22,
        14: 14,
        15: 14,
        16: 14,
        17: 16,
        24: 14,
        25: 54,
        26: 70,
        27: 54,
        28: 60,
        31: 70,
        32: 60,
        33: 70,
        35: 52,
        36: 48,
        37: 56,
        38: 60,
        39: 36,
        40: 48,
        41: 48,
        42: 46,
        43: 28,
        44: 40,
    }
    thin_gray = Side(style="thin", color="D9E2F3")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            border = copy(cell.border)
            border.bottom = thin_gray
            cell.border = border


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def build_categories(product_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories: "OrderedDict[str, dict[str, Any]]" = OrderedDict()

    def add(name: str, parent: str, image: str, source: str) -> None:
        key = name
        if key in categories:
            if not categories[key]["รูปภาพหมวดหมู่"] and image:
                categories[key]["รูปภาพหมวดหมู่"] = image
            return
        categories[key] = {
            "สถานะ": "ใช้งาน",
            "ชื่อหมวดหมู่": name,
            "หมวดหมู่ย่อยของ": "ไม่มีหมวดหมู่หลัก",
            "รายละเอียดหมวดหมู่": f"{name} อุปกรณ์และระบบจอดรถอัจฉริยะ HIP บนหน้าเว็บ One Tech Solution",
            "แสดงหมวดหมู่ที่หน้าร้าน": "เปิด",
            "Product Type ระดับ 1": DEFAULT_PRODUCT_TYPE_1,
            "Product Type ระดับ 2": DEFAULT_PRODUCT_TYPE_2,
            "รูปภาพหมวดหมู่": image,
            "SEO Title": f"{name} | One Tech Solution HIP",
            "SEO Keywords": ", ".join([name, "HIP", "ระบบจอดรถ", "Carpark", "LPR", "One Tech Solution"]),
            "SEO Meta Description": f"{name} สินค้าระบบจอดรถอัจฉริยะและกล้องอ่านป้ายทะเบียน LPR จากแบรนด์ HIP",
            "แหล่งข้อมูล": source,
        }

    for row in product_rows:
        main = row["หมวดหมู่หลัก"]
        image = row.get("รูปภาพหลัก/ไฟล์ภาพ", "")
        add(main, "", image, "อ้างอิงหมวดหมู่เดิมที่มีอยู่แล้วในระบบ LnwShop")

    out = []
    for idx, item in enumerate(categories.values(), start=1):
        out.append({"ลำดับ": idx, **item})
    return out


def write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]], table_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    add_table(ws, table_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_review_workbook(product_rows: list[dict[str, Any]], category_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Review_Summary"
    summary = [
        ["รายการ", "ค่า"],
        ["Source file", str(SOURCE_EXCEL)],
        ["Product rows", len(product_rows)],
        ["Category rows", len(category_rows)],
        ["Price rule", "ราคาปกติ/ขาย/พิเศษ = 0 บาททั้งหมด เพื่อการตรวจสอบก่อนประกาศขาย"],
        ["Stock rule", "จำนวนสต๊อกตั้งต้น = 0 เพื่อป้องกันออเดอร์ตกหล่น"],
        ["Image rule", "ใช้รูปภาพจาก Excel + ใส่ลายน้ำ One Tech Solution bottom-right"],
        ["Ready for automation", "พร้อมส่งให้ Playwright script รันเข้าระบบ"],
    ]
    for row in summary:
        ws_summary.append(row)
    style_sheet(ws_summary)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 100

    ws_products = wb.create_sheet("LnwShop_AI_Ready")
    ws_products.append(PRODUCT_HEADERS)
    for row in product_rows:
        ws_products.append([row.get(header, "") for header in PRODUCT_HEADERS])
    style_sheet(ws_products)
    add_table(ws_products, "HIPProductsReview")

    ws_categories = wb.create_sheet("Categories_AI_Ready")
    ws_categories.append(CATEGORY_HEADERS)
    for row in category_rows:
        ws_categories.append([row.get(header, "") for header in CATEGORY_HEADERS])
    style_sheet(ws_categories)
    add_table(ws_categories, "HIPCategoriesReview")

    REVIEW_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(REVIEW_EXCEL)


def main() -> int:
    product_rows = make_product_rows()
    category_rows = build_categories(product_rows)
    write_workbook(PRODUCT_EXCEL, "LnwShop_AI_Ready", PRODUCT_HEADERS, product_rows, "HIPProducts")
    write_workbook(CATEGORY_EXCEL, "Categories_AI_Ready", CATEGORY_HEADERS, category_rows, "HIPCategories")
    write_review_workbook(product_rows, category_rows)
    print(f"products: {len(product_rows)}")
    print(f"categories: {len(category_rows)}")
    print(f"product excel: {PRODUCT_EXCEL}")
    print(f"category excel: {CATEGORY_EXCEL}")
    print(f"review excel: {REVIEW_EXCEL}")
    print(f"watermarked images: {WATERMARKED_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
