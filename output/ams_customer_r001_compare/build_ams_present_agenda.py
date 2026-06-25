from pathlib import Path
from html import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

present_dir = Path(r"C:\365_project\TheCool18e\Dev\output\AMS_PRESENT_CUSTOMER_TH")
package_dir = Path(r"C:\365_project\TheCool18e\Dev\output\ams_customer_r001_compare\AMS_R001_COMPARE_PACKAGE")

html_path = present_dir / "00A_Agenda_AMS_Presentation.html"
xlsx_path = present_dir / "00A_Agenda_AMS_Presentation.xlsx"
package_html_path = package_dir / "00A_Agenda_AMS_Presentation.html"
package_xlsx_path = package_dir / "00A_Agenda_AMS_Presentation.xlsx"

agenda_rows = [
    ("1", "เปิดเป้าหมายการประชุม", "ยืนยันว่าเราจะดูภาพรวมงาน AMS, request R001, flow, standard/custom และ manday", "5 นาที", "00_เริ่มที่นี่_AMS_PRESENT_ลูกค้า.html"),
    ("2", "สรุปภาพรวมว่า Odoo ช่วยอะไร", "ชี้ให้เห็นจุดลดงาน manual, จุดที่ระบบทำ auto และจุดที่ยังต้อง confirm เพิ่ม", "10 นาที", "01_Dashboard_AMS_เราช่วยอะไร.html"),
    ("3", "ยืนยันขอบเขต request ลูกค้า", "อธิบายว่า customer request จริงมี 44 ข้อ และ supporting point ใช้ช่วย mapping เท่านั้น", "10 นาที", "03_Mapping_44_Request_กับ_Flow.html"),
    ("4", "Mapping Standard vs Custom", "ไล่ทีละ flow ว่า Odoo standard รองรับอะไร และจุดไหนต้อง custom/report/approval เพิ่ม", "20 นาที", "03A_Mapping_Standard_Custom_By_Flow.drawio"),
    ("5", "ดู Flow ลูกค้าเดิมแบบ Standard/Custom", "เปิด flow เดิมของลูกค้า layout เดิม แล้วใช้สี block อธิบายว่าอะไรเป็น Odoo standard และอะไรต้อง custom", "20 นาที", "04A_Client_Original_Flow_Standard_vs_Custom.drawio"),
    ("6", "อธิบาย Business Flow ใหญ่", "เปิดหน้า Overall ของ workflow แล้วอธิบายตั้งแต่ Sales ถึง Accounting", "15 นาที", "04_Workflow_Business_Flow_AMS.drawio"),
    ("7", "ลงรายละเอียด Flow ราย module", "ไล่ Sales, Purchase, RM Warehouse, Engineering, Quality, Planning, Production, Delivery, Accounting", "35 นาที", "04_Workflow_Business_Flow_AMS.drawio"),
    ("8", "ชี้จุด Confirm / Approve", "แยกให้ลูกค้าเห็นว่า step ไหน user ต้อง confirm/approve และ step ไหน Odoo ทำต่อให้", "10 นาที", "04_Workflow_Business_Flow_AMS.drawio"),
    ("9", "ยืนยัน DB Setup และ Test Flow ใน Odoo", "เปิดหลักฐานว่า database AMS port 8813 setup product, BOM, routing, quality และ test PO/SO/MO/DO/Invoice ผ่านครบ", "15 นาที", "06_หลักฐาน_Setup_DB_AMS_และ_Mapping_Standard_Custom.xlsx"),
    ("10", "สรุป Manday และ Phase", "อธิบาย effort โดยเริ่มจาก summary แล้วลงรายละเอียดเฉพาะข้อสำคัญ", "15 นาที", "02_Manday_และ_รายละเอียด_Request.xlsx"),
    ("11", "Open Issues / สิ่งที่ต้องตัดสินใจ", "เก็บประเด็นที่ต้อง confirm เช่น approval policy, report format, API, legacy document", "10 นาที", "05_สรุปปิดการนำเสนอ_และ_Action_Next.html"),
    ("12", "สรุป Next Step", "ปิดด้วยรายการ action, owner, due date และข้อมูลที่ต้องขอจากลูกค้า", "5 นาที", "05_สรุปปิดการนำเสนอ_และ_Action_Next.html"),
]

speaker_notes = [
    ("เปิดประชุม", "วันนี้เราจะไม่เริ่มจากหน้าจอระบบก่อน แต่จะเริ่มจาก business flow เพื่อให้ทุกฝ่ายเห็นภาพเดียวกันก่อนลงรายละเอียด"),
    ("Dashboard", "ให้ลูกค้าเห็นก่อนว่า Odoo จะช่วยลดงานซ้ำ ลดเอกสาร manual และทำให้ข้อมูลไหลจากต้นทางไปบัญชีได้อย่างไร"),
    ("Mapping", "ย้ำว่าตารางนี้ไม่ได้เพิ่ม request เอง แต่แยก customer request 44 ข้อออกจาก supporting mapping point"),
    ("Flow ลูกค้าเดิม", "ใช้ไฟล์ 04A เพื่อคุยจากภาพเดิมของลูกค้าก่อน สีเขียวคือ Odoo standard/config สีส้มคือ custom/report/API/approval เพิ่ม สีเทาคือเอกสารหรือ reference ภายนอก"),
    ("Workflow", "อ่านจากซ้ายไปขวาและดู lane ว่าแต่ละขั้นอยู่ทีมไหน สีเขียวคือระบบทำให้ สีฟ้าคือ user ทำ สีม่วงคือ confirm/approve"),
    ("DB Setup", "ใช้ไฟล์ 06 ยืนยันว่าระบบ demo สร้างข้อมูลจริงใน Odoo แล้ว เช่น product, BOM, 21 operation, quality point, PO, receipt, SO, MO, delivery และ invoice"),
    ("Manday", "อธิบายว่า manday เป็น initial estimate สำหรับวาง phase และใช้ยืนยัน scope ก่อนทำ quotation ขั้นสุดท้าย"),
    ("ปิดประชุม", "สรุปสิ่งที่เห็นตรงกัน สิ่งที่ต้องถามเพิ่ม และ action ที่ต้องได้จากลูกค้าก่อนเริ่ม setup/config/custom"),
]


def build_html():
    rows = "\n".join(
        "<tr>"
        f"<td>{escape(no)}</td>"
        f"<td>{escape(topic)}</td>"
        f"<td>{escape(goal)}</td>"
        f"<td>{escape(duration)}</td>"
        f"<td>{escape(file)}</td>"
        "</tr>"
        for no, topic, goal, duration, file in agenda_rows
    )
    notes = "\n".join(
        f"<li><strong>{escape(title)}</strong><span>{escape(note)}</span></li>"
        for title, note in speaker_notes
    )
    html = f"""<!doctype html>
<html lang="th">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>AMS Presentation Agenda</title>
<style>
body {{
  margin: 0;
  font-family: Arial, Tahoma, sans-serif;
  color: #172033;
  background: #f5f7fb;
}}
.wrap {{
  max-width: 1180px;
  margin: 0 auto;
  padding: 36px 28px 48px;
}}
.hero {{
  background: #5b1747;
  color: white;
  padding: 28px 32px;
  border-radius: 8px;
}}
h1 {{
  margin: 0 0 8px;
  font-size: 30px;
}}
.hero p {{
  margin: 0;
  color: #f7d9ec;
  font-size: 15px;
}}
.section {{
  margin-top: 22px;
  background: white;
  border: 1px solid #dbe3ef;
  border-radius: 8px;
  overflow: hidden;
}}
h2 {{
  margin: 0;
  padding: 18px 22px;
  font-size: 18px;
  background: #edf2f7;
  border-bottom: 1px solid #dbe3ef;
}}
table {{
  width: 100%;
  border-collapse: collapse;
}}
th, td {{
  padding: 13px 14px;
  border-bottom: 1px solid #e6edf5;
  vertical-align: top;
  font-size: 14px;
}}
th {{
  text-align: left;
  background: #f8fafc;
  color: #334155;
}}
td:first-child {{
  width: 44px;
  font-weight: 700;
  color: #5b1747;
}}
td:nth-child(4) {{
  width: 92px;
  white-space: nowrap;
}}
td:nth-child(5) {{
  width: 220px;
  color: #0f5e7a;
  word-break: break-word;
}}
ul {{
  margin: 0;
  padding: 18px 28px 22px 42px;
}}
li {{
  margin: 0 0 12px;
}}
li span {{
  display: block;
  margin-top: 4px;
  color: #475569;
}}
.footer {{
  margin-top: 18px;
  color: #64748b;
  font-size: 13px;
}}
</style>
</head>
<body>
<main class="wrap">
  <section class="hero">
    <h1>AMS Odoo Blueprint Presentation Agenda</h1>
    <p>ลำดับการนำเสนอสำหรับคุยกับลูกค้า: เริ่มจากภาพรวม ไป mapping, workflow, manday และปิดด้วย next step</p>
  </section>
  <section class="section">
    <h2>Agenda หลัก</h2>
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>หัวข้อ</th>
          <th>เป้าหมายที่ต้องคุยให้จบ</th>
          <th>เวลา</th>
          <th>ไฟล์ที่เปิด</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
  </section>
  <section class="section">
    <h2>Speaker Notes</h2>
    <ul>
      {notes}
    </ul>
  </section>
  <p class="footer">ไฟล์นี้มี version Excel ให้แก้ไขเวลา/ข้อความก่อนประชุม: 00A_Agenda_AMS_Presentation.xlsx</p>
</main>
</body>
</html>
"""
    html_path.write_text(html, encoding="utf-8")
    package_html_path.write_text(html, encoding="utf-8")


def build_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"
    ws.append(["AMS Odoo Blueprint Presentation Agenda"])
    ws.merge_cells("A1:E1")
    ws.append(["#", "หัวข้อ", "เป้าหมายที่ต้องคุยให้จบ", "เวลา", "ไฟล์ที่เปิด"])
    for row in agenda_rows:
        ws.append(list(row))

    ws_notes = wb.create_sheet("Speaker Notes")
    ws_notes.append(["ช่วงนำเสนอ", "คำอธิบาย/แนวพูด"])
    for row in speaker_notes:
        ws_notes.append(list(row))

    header_fill = PatternFill("solid", fgColor="5B1747")
    sub_fill = PatternFill("solid", fgColor="EDF2F7")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(color="334155", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].fill = header_fill
    ws["A1"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="center")
    for cell in ws[2]:
        cell.fill = sub_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [8, 28, 68, 14, 42]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    for cell in ws_notes[1]:
        cell.fill = sub_fill
        cell.font = header_font
        cell.border = border
    for row in ws_notes.iter_rows(min_row=2, max_row=ws_notes.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws_notes.column_dimensions["A"].width = 24
    ws_notes.column_dimensions["B"].width = 100
    ws_notes.freeze_panes = "A2"
    ws_notes.sheet_view.showGridLines = False

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    package_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    wb.save(package_xlsx_path)


build_html()
build_xlsx()
print({"html": str(html_path), "xlsx": str(xlsx_path), "package_html": str(package_html_path), "package_xlsx": str(package_xlsx_path)})
