from base64 import b64encode
from dataclasses import dataclass
from pathlib import Path
from shutil import copyfile
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

source_image = Path(r"C:\Users\tumsu\AppData\Local\Temp\codex-clipboard-42013340-be8a-4f38-8b3c-e01bcef3c602.png")
present_dir = Path(r"C:\365_project\TheCool18e\Dev\output\AMS_PRESENT_CUSTOMER_TH")
package_dir = Path(r"C:\365_project\TheCool18e\Dev\output\ams_customer_r001_compare\AMS_R001_COMPARE_PACKAGE")
download_dir = Path(r"C:\Users\tumsu\Downloads")

drawio_name = "04A_Client_Original_Flow_Standard_vs_Custom.drawio"
xlsx_name = "04A_Client_Original_Flow_Standard_vs_Custom.xlsx"
source_name = "04A_Client_Original_Flow_Source.png"
preview_name = "04A_Client_Original_Flow_Standard_vs_Custom_preview.png"


@dataclass(frozen=True)
class Block:
    no: int
    name: str
    area: str
    x: int
    y: int
    w: int
    h: int
    shape: str
    category: str
    module: str
    reason: str


blocks = [
    Block(1, "Request FA Sample", "Sales / Customer", 452, 30, 99, 59, "rect", "custom", "Sales + Custom", "ต้องมี workflow/ฟอร์มเฉพาะสำหรับ FA sample"),
    Block(2, "Request Raw Material", "Raw Material Request", 315, 30, 99, 77, "rect", "custom", "Inventory + Manufacturing", "รูปแบบ request/approve วัตถุดิบเป็น process เฉพาะ"),
    Block(3, "Issue Raw Material", "Raw Material Issue", 177, 30, 99, 59, "rect", "standard", "Inventory / MRP", "Odoo รองรับ stock move และ issue วัตถุดิบเข้า production"),
    Block(4, "Initial Raw Material Step", "Raw Material", 40, 30, 99, 59, "rect", "standard", "Inventory", "เป็น movement/operation ในคลัง รองรับด้วย Odoo Inventory"),
    Block(5, "PCC Temp", "Engineering / Quality", 40, 125, 99, 59, "doc", "custom", "Manufacturing + Quality", "Template PCC เป็นเอกสารเฉพาะลูกค้า ต้องทำ report/template"),
    Block(6, "Sample Production", "Production Sample", 177, 125, 99, 59, "rect", "standard", "Manufacturing", "ใช้ MO/WO หรือ sample production flow ได้"),
    Block(7, "Connector", "Connector", 480, 140, 43, 44, "ellipse", "external", "-", "เป็นจุดเชื่อม flow ไม่ใช่ function ที่ต้องพัฒนา"),
    Block(8, "Start", "Start", 721, 30, 99, 59, "start", "external", "-", "เป็นจุดเริ่ม flow ไม่ใช่ function"),
    Block(9, "BOM Option / Material & Process", "Engineering", 721, 133, 118, 93, "rect", "standard", "Manufacturing", "ใช้ product, BOM, routing และ configuration ใน Odoo ได้"),
    Block(10, "BOM", "Engineering", 855, 133, 99, 67, "rect", "standard", "Manufacturing", "Odoo Manufacturing รองรับ BOM standard"),
    Block(11, "Routing", "Engineering", 990, 133, 99, 93, "rect", "standard", "Manufacturing", "Odoo รองรับ routing/work center/operation"),
    Block(12, "Document Control", "Document Control", 1124, 40, 99, 59, "rect", "custom", "Documents + Custom", "ทะเบียนเอกสารเฉพาะ ต้องกำหนด workflow/report เพิ่ม"),
    Block(13, "Process Control Chart PCC", "Quality", 1124, 133, 99, 59, "doc", "custom", "Quality + Custom", "PCC เป็นเอกสาร/format เฉพาะ ต้อง custom template/report"),
    Block(14, "Quotation", "Sales", 721, 227, 99, 93, "rect", "standard", "Sales", "Odoo Sales รองรับ quotation standard"),
    Block(15, "Create Product Code", "Engineering", 855, 227, 134, 59, "rect", "standard", "Inventory / Manufacturing", "Odoo รองรับ product master และ internal reference"),
    Block(16, "Create BOM & Process", "Engineering", 990, 227, 133, 59, "rect", "standard", "Manufacturing", "Odoo รองรับ BOM/routing เป็น standard"),
    Block(17, "Create PCC", "Quality", 1124, 227, 133, 59, "rect", "custom", "Quality + Custom", "ต้องสร้างฟอร์ม/รายงาน PCC ตาม format ลูกค้า"),
    Block(18, "Run Cost", "Costing", 1258, 227, 99, 78, "rect", "standard", "Manufacturing / Accounting", "Odoo มี costing/valuation ใช้ได้โดย config และ process"),
    Block(19, "Account Payable", "Accounting", 177, 221, 109, 59, "rect", "standard", "Accounting", "Odoo Accounting รองรับ vendor bill/AP"),
    Block(20, "PO Purchase", "Purchase", 315, 221, 107, 59, "rect", "standard", "Purchase", "Odoo Purchase รองรับ PO standard"),
    Block(21, "Issue Raw Material", "Inventory", 452, 221, 99, 59, "rect", "standard", "Inventory / Manufacturing", "Odoo รองรับ issue raw material ด้วย stock move"),
    Block(22, "Receive Raw Material", "Inventory", 177, 317, 99, 59, "rect", "standard", "Inventory", "Odoo Inventory รองรับ receipt standard"),
    Block(23, "PR Purchase Request", "Purchase", 315, 317, 99, 59, "rect", "custom", "Purchase + Approval", "Odoo standard มี RFQ/PO แต่ PR approval ตามรูปมักต้อง config/custom"),
    Block(24, "Request Raw Material", "Inventory", 452, 317, 99, 59, "rect", "custom", "Inventory + Manufacturing", "ฟอร์ม request/approve วัตถุดิบเป็นขั้นตอนเฉพาะ"),
    Block(25, "Request PCC", "Quality", 855, 321, 99, 59, "rect", "custom", "Quality + Custom", "Request PCC เป็น workflow เอกสารเฉพาะ"),
    Block(26, "Approved by MD&GM", "Approval", 1081, 417, 99, 59, "rect", "custom", "Approvals + Custom", "ต้องกำหนด approval matrix/role เฉพาะ"),
    Block(27, "Customer PO", "Customer", 721, 414, 99, 59, "doc", "external", "Sales", "เป็นเอกสารลูกค้า ใช้เป็น reference ใน SO"),
    Block(28, "PO Raw Material", "Purchase", 177, 412, 99, 59, "rect", "standard", "Purchase", "Odoo รองรับ PO วัตถุดิบ"),
    Block(29, "Outside Process", "Subcontract", 315, 412, 99, 59, "rect", "standard", "Manufacturing / Purchase", "Odoo รองรับ subcontracting/operation ได้ด้วย config"),
    Block(30, "Semi WO2", "Manufacturing", 452, 412, 99, 59, "rect", "standard", "Manufacturing", "Odoo รองรับ semi-finished/MO หลายระดับ"),
    Block(31, "Stock Balance", "Inventory", 316, 509, 97, 58, "decision", "standard", "Inventory", "ตรวจ stock on hand/forecasted quantity ได้จาก Odoo"),
    Block(32, "MRP Calculation", "Planning", 452, 508, 99, 59, "rect", "standard", "MRP", "Odoo MRP รองรับ planning/procurement demand"),
    Block(33, "Production Planning", "Planning", 587, 508, 99, 59, "rect", "standard", "MRP", "ใช้ MPS/MRP/Manufacturing schedule ได้"),
    Block(34, "Sale Order SO", "Sales", 721, 508, 133, 59, "rect", "standard", "Sales", "Odoo Sales รองรับ sales order standard"),
    Block(35, "Delivery Order DO", "Delivery", 855, 508, 107, 59, "rect", "standard", "Inventory / Sales", "Odoo รองรับ delivery order standard"),
    Block(36, "Issue FG", "Inventory", 990, 508, 133, 59, "rect", "standard", "Inventory", "Odoo รองรับ delivery/issue finished goods"),
    Block(37, "Invoice", "Accounting", 1124, 508, 133, 59, "rect", "standard", "Accounting", "Odoo รองรับ customer invoice standard"),
    Block(38, "Account Receivable", "Accounting", 1258, 508, 99, 59, "rect", "standard", "Accounting", "Odoo Accounting รองรับ AR"),
    Block(39, "PR Raw Material", "Purchase", 177, 508, 99, 59, "rect", "custom", "Purchase + Approval", "PR ตาม flow ต้องมี approval/request เพิ่ม"),
    Block(40, "Work Order WO", "Manufacturing", 315, 601, 99, 59, "rect", "standard", "Manufacturing", "Odoo รองรับ work order standard"),
    Block(41, "Return Raw Material", "Inventory", 177, 601, 99, 59, "rect", "standard", "Inventory", "Odoo รองรับ return/internal transfer"),
    Block(42, "Production Operation", "Production", 452, 601, 99, 59, "rect", "standard", "Manufacturing", "บันทึกการผลิตผ่าน MO/WO ได้"),
    Block(43, "Production Daily", "Production Report", 587, 601, 99, 59, "rect", "custom", "Manufacturing + Report", "Daily production report/tracking format ต้องทำ report เพิ่ม"),
    Block(44, "Transfer FG", "Inventory", 721, 601, 133, 59, "rect", "standard", "Inventory", "Odoo รองรับ internal transfer/finished goods move"),
    Block(45, "Receive FG", "Inventory", 855, 601, 99, 120, "rect", "standard", "Inventory / Manufacturing", "Odoo รองรับ receipt of finished goods จาก production"),
    Block(46, "Stock FG", "Inventory", 976, 601, 127, 59, "parallelogram", "standard", "Inventory", "Odoo รองรับ stock on hand finished goods"),
    Block(47, "Receive Return Raw Material", "Inventory", 177, 700, 99, 59, "rect", "standard", "Inventory", "Odoo รองรับรับคืนวัตถุดิบ"),
    Block(48, "Request Raw Material", "Inventory", 315, 700, 99, 59, "rect", "custom", "Inventory + Manufacturing", "request material แบบฟอร์มเฉพาะควร custom/config approval"),
    Block(49, "Issue Raw Material", "Inventory", 452, 700, 99, 59, "rect", "standard", "Inventory / Manufacturing", "Odoo รองรับ issue raw material"),
    Block(50, "Stock RM", "Inventory", 26, 799, 127, 59, "parallelogram", "standard", "Inventory", "Odoo รองรับ stock raw material"),
    Block(51, "Connector", "Connector", 205, 798, 43, 44, "ellipse", "external", "-", "เป็นจุดเชื่อม flow ไม่ใช่ function ที่ต้องพัฒนา"),
    Block(52, "Quotation Decision", "Sales / Customer", 721, 321, 99, 59, "decision", "custom", "Sales + Approval", "จุดตัดสินใจรับ/แก้ไข quotation ต้องกำหนด policy"),
    Block(53, "Cost Approval Decision", "Costing / Approval", 1196, 321, 99, 59, "decision", "custom", "Accounting + Approval", "จุดอนุมัติ cost ต้องกำหนด approval policy"),
]

category_styles = {
    "standard": {
        "label": "Standard Odoo / Configuration",
        "fill": "#22C55E",
        "stroke": "#15803D",
    },
    "custom": {
        "label": "Custom / Report / API / Approval เพิ่ม",
        "fill": "#F97316",
        "stroke": "#C2410C",
    },
    "external": {
        "label": "External / Manual Reference",
        "fill": "#64748B",
        "stroke": "#334155",
    },
}


original_labels = {
    1: "Request<br>FA Sample",
    2: "Request<br>Raw Material<br>ขอเบิกวัตถุดิบ",
    3: "Issue<br>Raw Material<br>เบิกจ่ายวัตถุดิบ",
    4: "สั่งตัดชิ้นส่วนตัวอย่าง",
    5: "PCC Temp",
    6: "ผลิตชิ้นส่วนตัวอย่าง",
    7: "",
    8: "Start",
    9: "BOM Option<br>สูตรการผลิต<br>Material & Process",
    10: "BOM<br>สูตรการผลิต",
    11: "Routing<br>เส้นทางการผลิต",
    12: "Document Control<br>ขึ้นทะเบียน",
    13: "Process Control Chart<br>(PCC)",
    14: "Quotation<br>ใบเสนอราคา",
    15: "Create<br>Product Code",
    16: "Create<br>BOM & Process",
    17: "Create<br>Process Control Chart<br>(PCC)",
    18: "Run Cost",
    19: "Account Payable<br>บัญชีเจ้าหนี้",
    20: "PO ใบสั่งซื้อ<br>(งานจ้างทำ)",
    21: "Issue<br>Raw Material<br>เบิกจ่ายวัตถุดิบ",
    22: "Receive<br>Raw Material<br>รับวัตถุดิบเข้า",
    23: "PR ใบขอซื้อ<br>(งานจ้างทำ)",
    24: "Request<br>Raw Material<br>ขอเบิกวัตถุดิบ",
    25: "Request Process<br>Control Chart",
    26: "Approved By<br>MD&GM",
    27: "PO (Customer)<br>ใบสั่งซื้อจากลูกค้า",
    28: "PO ใบสั่งซื้อ<br>(สั่งซื้อวัตถุดิบ)",
    29: "Outside Process<br>(งานจ้างทำ)",
    30: "Semi (WO2)<br>ใบสั่งผลิตชิ้นงาน",
    31: "Stock Balance",
    32: "คำนวณวัตถุดิบการผลิต<br>(MRP)",
    33: "Production Planning<br>วางแผนการผลิต",
    34: "Sale Order (SO)<br>ใบสั่งขาย",
    35: "Delivery Order (DO)<br>ใบส่งสินค้า",
    36: "Issue FG<br>เบิกจ่ายสินค้า",
    37: "Invoice<br>ใบกำกับภาษี",
    38: "Account Receivable<br>บัญชีลูกหนี้",
    39: "PR ใบขอซื้อ<br>(จัดซื้อวัตถุดิบ)",
    40: "Work Order (WO)<br>ใบสั่งผลิตสินค้า",
    41: "Return<br>Raw Material ส่วนเกิน<br>วัตถุดิบ",
    42: "ดำเนินการผลิต",
    43: "Production Daily<br>บันทึกการผลิต",
    44: "Transfer FG<br>ส่งมอบสินค้า",
    45: "Receive FG<br>รับสินค้าเข้าคลังสำเร็จ",
    46: "Stock FG",
    47: "Receive Return<br>Raw Material<br>รับคืนวัตถุดิบ",
    48: "Request<br>Raw Material<br>ขอเบิกวัตถุดิบ",
    49: "Issue<br>Raw Material<br>เบิกจ่ายวัตถุดิบ",
    50: "Stock RM",
    51: "",
    52: "Quotation<br>ผ่าน?",
    53: "Cost<br>ผ่าน?",
}

lane_groups = [
    ("Sample / FA Flow", 20, 15, 560, 215, "#E0F2FE"),
    ("Purchase / AP / Outside Process", 160, 205, 330, 295, "#FEF3C7"),
    ("RM Planning / Production Prep", 160, 490, 430, 385, "#DCFCE7"),
    ("Engineering / BOM / PCC / Cost", 680, 15, 720, 460, "#EDE9FE"),
    ("Sales / Delivery / Accounting", 680, 485, 720, 210, "#DBEAFE"),
    ("FG / RM Stock Return", 20, 585, 1090, 290, "#F1F5F9"),
]

flow_edges = [
    ("b08", "b01", ""),
    ("b08", "b09", ""),
    ("b01", "b02", ""),
    ("b02", "b07", ""),
    ("b03", "b04", ""),
    ("b04", "b05", ""),
    ("b05", "b06", ""),
    ("b06", "b03", ""),
    ("b07", "b21", ""),
    ("b24", "b21", ""),
    ("b30", "b24", ""),
    ("b39", "b28", ""),
    ("b28", "b22", ""),
    ("b22", "b19", ""),
    ("b29", "b23", ""),
    ("b23", "b20", ""),
    ("b20", "b19", ""),
    ("b09", "b10", ""),
    ("b09", "b14", ""),
    ("b10", "b16", ""),
    ("b11", "b16", ""),
    ("b14", "b52", ""),
    ("b52", "b25", "Yes"),
    ("b52", "b27", "Yes"),
    ("b52", "b09", "No"),
    ("b25", "b15", ""),
    ("b15", "b16", ""),
    ("b16", "b17", ""),
    ("b17", "b18", ""),
    ("b18", "b53", ""),
    ("b53", "b13", "Yes"),
    ("b13", "b12", ""),
    ("b53", "b26", "No"),
    ("b26", "b16", "Rework"),
    ("b27", "b34", ""),
    ("b34", "b33", ""),
    ("b33", "b32", ""),
    ("b32", "b31", ""),
    ("b31", "b40", "Yes"),
    ("b31", "b39", "No"),
    ("b31", "b29", "No"),
    ("b40", "b42", ""),
    ("b40", "b48", ""),
    ("b48", "b51", ""),
    ("b51", "b49", ""),
    ("b49", "b42", ""),
    ("b42", "b43", ""),
    ("b42", "b44", ""),
    ("b44", "b45", ""),
    ("b45", "b46", ""),
    ("b34", "b35", ""),
    ("b35", "b36", ""),
    ("b46", "b36", ""),
    ("b36", "b37", ""),
    ("b37", "b38", ""),
    ("b40", "b41", ""),
    ("b41", "b47", ""),
    ("b47", "b50", ""),
    ("b50", "b51", ""),
]

POSITION_SCALE = 1.55
SIZE_SCALE = 1.15


def style_for(block):
    meta = category_styles[block.category]
    base = {
        "rect": "rounded=0;whiteSpace=wrap;html=1;",
        "doc": "shape=document;whiteSpace=wrap;html=1;boundedLbl=1;",
        "decision": "rhombus;whiteSpace=wrap;html=1;",
        "ellipse": "ellipse;whiteSpace=wrap;html=1;",
        "start": "rounded=1;arcSize=50;whiteSpace=wrap;html=1;",
        "parallelogram": "shape=parallelogram;whiteSpace=wrap;html=1;",
    }[block.shape]
    return (
        base
        + f"fillColor={meta['fill']};strokeColor={meta['stroke']};"
        + "fillOpacity=88;strokeOpacity=100;strokeWidth=3;fontColor=#FFFFFF;fontStyle=1;fontSize=9;align=center;verticalAlign=middle;spacing=4;labelBackgroundColor=none;"
    )


def mxcell(cell_id, value, style, x, y, w, h):
    return (
        f'<mxCell id="{cell_id}" value="{escape(value)}" style="{style}" vertex="1" parent="1">'
        f'<mxGeometry x="{x}" y="{y}" width="{w}" height="{h}" as="geometry" />'
        "</mxCell>"
    )


def mxedge(cell_id, source, target, value=""):
    style = (
        "edgeStyle=orthogonalEdgeStyle;rounded=1;orthogonalLoop=1;jettySize=auto;html=1;"
        "strokeColor=#0284C7;strokeWidth=2;endArrow=block;endFill=1;fontColor=#0F172A;"
        "fontSize=10;labelBackgroundColor=#FFFFFF;"
    )
    return (
        f'<mxCell id="{cell_id}" value="{escape(value)}" style="{style}" edge="1" parent="1" '
        f'source="{source}" target="{target}"><mxGeometry relative="1" as="geometry" /></mxCell>'
    )


def build_drawio():
    ox, oy = 40, 165
    page_w, page_h = 2320, 1620

    def sx(value):
        return ox + value * POSITION_SCALE

    def sy(value):
        return oy + value * POSITION_SCALE

    def sw(value):
        return value * SIZE_SCALE

    def lane_size(value):
        return value * POSITION_SCALE

    cells = ['<mxCell id="0" />', '<mxCell id="1" parent="0" />']
    cells.append(mxcell("title", "AMS Editable Customer Flow - Standard Odoo vs Custom", "rounded=1;whiteSpace=wrap;html=1;fillColor=#5B1747;strokeColor=#5B1747;fontColor=#FFFFFF;fontStyle=1;fontSize=22;align=center;verticalAlign=middle;", 40, 25, 1420, 48))
    x = 1520
    for index, key in enumerate(["standard", "custom", "external"]):
        meta = category_styles[key]
        cells.append(mxcell(f"legend_{key}", meta["label"], f"rounded=1;whiteSpace=wrap;html=1;fillColor={meta['fill']};strokeColor={meta['stroke']};fillOpacity=60;strokeWidth=2;fontColor=#111827;fontStyle=1;fontSize=12;align=center;verticalAlign=middle;", x, 20 + index * 38, 310, 30))
    for index, (label, x0, y0, w0, h0, fill) in enumerate(lane_groups, start=1):
        lane_style = (
            f"rounded=1;whiteSpace=wrap;html=1;fillColor={fill};strokeColor=#94A3B8;"
            "fillOpacity=18;strokeOpacity=70;dashed=1;dashPattern=8 5;strokeWidth=2;"
            "fontColor=#334155;fontStyle=1;fontSize=13;align=left;verticalAlign=top;spacing=10;"
        )
        cells.append(mxcell(f"lane_{index:02d}", label, lane_style, sx(x0), sy(y0), lane_size(w0), lane_size(h0)))
    for block in blocks:
        cells.append(mxcell(f"b{block.no:02d}", original_labels.get(block.no, block.name), style_for(block), sx(block.x), sy(block.y), sw(block.w), sw(block.h)))
    for index, (source, target, label) in enumerate(flow_edges, start=1):
        cells.append(mxedge(f"e{index:03d}", source, target, label))
    graph = (
        f'<mxGraphModel dx="1500" dy="1000" grid="1" gridSize="10" guides="1" tooltips="1" connect="1" arrows="1" '
        f'fold="1" page="1" pageScale="1" pageWidth="{page_w}" pageHeight="{page_h}" math="0" shadow="0">'
        "<root>"
        + "".join(cells)
        + "</root></mxGraphModel>"
    )
    mxfile = f'<mxfile host="app.diagrams.net" modified="2026-06-19T00:00:00.000Z" agent="Codex" version="24.7.17"><diagram id="original-flow-standard-custom" name="Original Flow Standard vs Custom">{graph}</diagram></mxfile>'
    for directory in [present_dir, package_dir, download_dir]:
        directory.mkdir(parents=True, exist_ok=True)
        (directory / drawio_name).write_text(mxfile, encoding="utf-8")
    for directory in [present_dir, package_dir]:
        copyfile(source_image, directory / source_name)


def build_mapping_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Block Mapping"
    headers = ["No.", "Block", "Original AMS Text", "Area", "Result", "Odoo Module", "Reason"]
    ws.append(headers)
    for block in blocks:
        ws.append([
            block.no,
            block.name,
            original_labels.get(block.no, "").replace("<br>", "\n"),
            block.area,
            category_styles[block.category]["label"],
            block.module,
            block.reason,
        ])
    fill_header = PatternFill("solid", fgColor="5B1747")
    font_header = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    fills = {
        "Standard Odoo / Configuration": PatternFill("solid", fgColor="DCFCE7"),
        "Custom / Report / API / Approval เพิ่ม": PatternFill("solid", fgColor="FFEDD5"),
        "External / Manual Reference": PatternFill("solid", fgColor="E2E8F0"),
    }
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        result = row[4].value
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[4].fill = fills[result]
    widths = [8, 28, 32, 24, 34, 28, 72]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for directory in [present_dir, package_dir, download_dir]:
        wb.save(directory / xlsx_name)


def rgba(hex_color, alpha):
    value = hex_color.lstrip("#")
    return tuple(int(value[i:i + 2], 16) for i in (0, 2, 4)) + (alpha,)


def build_preview_png():
    ox, oy = 40, 165
    width, height = 2320, 1620
    base = Image.new("RGBA", (width, height), (255, 255, 255, 255))
    draw = ImageDraw.Draw(base)

    def sx(value):
        return ox + value * POSITION_SCALE

    def sy(value):
        return oy + value * POSITION_SCALE

    def sw(value):
        return value * SIZE_SCALE

    def lane_size(value):
        return value * POSITION_SCALE

    try:
        font = ImageFont.truetype(r"C:\Windows\Fonts\tahoma.ttf", 11)
        small_font = ImageFont.truetype(r"C:\Windows\Fonts\tahoma.ttf", 10)
        title_font = ImageFont.truetype(r"C:\Windows\Fonts\tahomabd.ttf", 24)
    except OSError:
        font = ImageFont.load_default()
        small_font = font
        title_font = font

    draw.rounded_rectangle((40, 25, 1460, 73), radius=6, fill=(91, 23, 71, 255), outline=(91, 23, 71, 255), width=2)
    draw.text((750, 49), "AMS Editable Customer Flow - Standard Odoo vs Custom", font=title_font, fill=(255, 255, 255, 255), anchor="mm")

    legend_x = 1520
    for index, key in enumerate(["standard", "custom", "external"]):
        meta = category_styles[key]
        y = 20 + index * 38
        draw.rounded_rectangle((legend_x, y, legend_x + 310, y + 30), radius=4, fill=rgba(meta["fill"], 155), outline=rgba(meta["stroke"], 255), width=2)
        draw.text((legend_x + 155, y + 15), meta["label"], font=small_font, fill=(17, 24, 39, 255), anchor="mm")

    for label, x0, y0, w0, h0, fill in lane_groups:
        rgb = tuple(int(fill.lstrip("#")[i:i + 2], 16) for i in (0, 2, 4))
        box = (sx(x0), sy(y0), sx(x0) + lane_size(w0), sy(y0) + lane_size(h0))
        draw.rounded_rectangle(box, radius=8, fill=rgb + (40,), outline=(148, 163, 184, 180), width=2)
        draw.text((box[0] + 12, box[1] + 10), label, font=font, fill=(51, 65, 85, 255))

    block_by_id = {f"b{block.no:02d}": block for block in blocks}

    def center(block):
        return (sx(block.x) + sw(block.w) / 2, sy(block.y) + sw(block.h) / 2)

    for source, target, label in flow_edges:
        if source not in block_by_id or target not in block_by_id:
            continue
        s = block_by_id[source]
        t = block_by_id[target]
        x1, y1 = center(s)
        x2, y2 = center(t)
        mid_x = (x1 + x2) / 2
        points = [(x1, y1), (mid_x, y1), (mid_x, y2), (x2, y2)]
        draw.line(points, fill=(2, 132, 199, 210), width=2, joint="curve")
        angle = 0 if x2 >= points[-2][0] else 3.14159
        arrow = [(x2, y2), (x2 - 10, y2 - 5), (x2 - 10, y2 + 5)] if angle == 0 else [(x2, y2), (x2 + 10, y2 - 5), (x2 + 10, y2 + 5)]
        draw.polygon(arrow, fill=(2, 132, 199, 210))
        if label:
            draw.rounded_rectangle((mid_x - 18, (y1 + y2) / 2 - 10, mid_x + 35, (y1 + y2) / 2 + 10), radius=3, fill=(255, 255, 255, 230))
            draw.text((mid_x + 8, (y1 + y2) / 2), label, font=small_font, fill=(15, 23, 42, 255), anchor="mm")

    for block in blocks:
        meta = category_styles[block.category]
        fill = rgba(meta["fill"], 185)
        stroke = rgba(meta["stroke"], 230)
        box = (sx(block.x), sy(block.y), sx(block.x) + sw(block.w), sy(block.y) + sw(block.h))
        if block.shape == "decision":
            cx = sx(block.x) + sw(block.w) / 2
            cy = sy(block.y) + sw(block.h) / 2
            poly = [(cx, sy(block.y)), (sx(block.x) + sw(block.w), cy), (cx, sy(block.y) + sw(block.h)), (sx(block.x), cy)]
            draw.polygon(poly, fill=fill, outline=stroke)
            draw.line(poly + [poly[0]], fill=stroke, width=3)
        elif block.shape == "ellipse" or block.shape == "start":
            draw.ellipse(box, fill=fill, outline=stroke, width=3)
        elif block.shape == "parallelogram":
            slant = min(30, sw(block.w) // 4)
            poly = [(sx(block.x) + slant, sy(block.y)), (sx(block.x) + sw(block.w), sy(block.y)), (sx(block.x) + sw(block.w) - slant, sy(block.y) + sw(block.h)), (sx(block.x), sy(block.y) + sw(block.h))]
            draw.polygon(poly, fill=fill, outline=stroke)
            draw.line(poly + [poly[0]], fill=stroke, width=3)
        else:
            draw.rectangle(box, fill=fill, outline=stroke, width=3)
        label = original_labels.get(block.no, "").replace("<br>", "\n")
        if not label:
            continue
        text_box = draw.multiline_textbbox((0, 0), label, font=font, spacing=2, align="center")
        text_w = text_box[2] - text_box[0]
        text_h = text_box[3] - text_box[1]
        x = sx(block.x) + (sw(block.w) - text_w) / 2
        y = sy(block.y) + (sw(block.h) - text_h) / 2
        draw.multiline_text((x, y), label, font=font, fill=(255, 255, 255, 255), spacing=2, align="center")
    result = base.convert("RGB")
    for directory in [present_dir, package_dir, download_dir]:
        result.save(directory / preview_name, quality=95)


build_drawio()
build_mapping_xlsx()
build_preview_png()
print({
    "drawio": str(present_dir / drawio_name),
    "xlsx": str(present_dir / xlsx_name),
    "preview": str(present_dir / preview_name),
    "download_drawio": str(download_dir / drawio_name),
})
