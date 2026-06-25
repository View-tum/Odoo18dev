from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


R001_FILLED = Path(r"C:\Users\tumsu\Desktop\AMS_Present\ams_source_files\02_Requirement_on_New_System_R001_With_Solution.xlsx")
AWS_PRESENT = Path(r"C:\Users\tumsu\Desktop\AMS_Present\AWS_Present.xlsx")


def source_row_number(value):
    if not value:
        return None
    text = str(value)
    if "A" not in text:
        return None
    try:
        return int(text.rsplit("A", 1)[1])
    except ValueError:
        return None


aws = load_workbook(AWS_PRESENT, data_only=False)
aws_ws = aws["06 R001 Requirement Links"]

source_to_aws_row = {}
for row in range(6, aws_ws.max_row + 1):
    source_row = source_row_number(aws_ws[f"N{row}"].value)
    if source_row:
        source_to_aws_row[source_row] = row

r001 = load_workbook(R001_FILLED)
ws = r001["Requirement"]

for row in range(5, ws.max_row + 1):
    if row not in source_to_aws_row:
        continue
    aws_row = source_to_aws_row[row]
    cell = ws[f"E{row}"]
    cell.value = f"เปิด AWS Present Seq {aws_ws[f'A{aws_row}'].value}"
    cell.hyperlink = f"../AWS_Present.xlsx#'06 R001 Requirement Links'!A{aws_row}"
    cell.font = Font(color="0563C1", underline="single")

r001.save(R001_FILLED)
print(R001_FILLED)
print(f"linked_rows={len(source_to_aws_row)}")
