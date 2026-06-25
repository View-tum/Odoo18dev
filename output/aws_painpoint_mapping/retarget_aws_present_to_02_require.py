from pathlib import Path

from openpyxl import load_workbook


AWS_PRESENT = Path(r"C:\Users\tumsu\Desktop\AMS_Present\AWS_Present.xlsx")
OLD_NAME = "02_Requirement_on_New_System_R001.xlsx"
NEW_NAME = "02_Require.xlsx"


wb = load_workbook(AWS_PRESENT)
changed = 0

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and OLD_NAME in cell.value:
                cell.value = cell.value.replace(OLD_NAME, NEW_NAME)
                changed += 1
            if cell.hyperlink and cell.hyperlink.target and OLD_NAME in cell.hyperlink.target:
                cell.hyperlink.target = cell.hyperlink.target.replace(OLD_NAME, NEW_NAME)
                changed += 1

wb.save(AWS_PRESENT)
print(AWS_PRESENT)
print(f"changed={changed}")
